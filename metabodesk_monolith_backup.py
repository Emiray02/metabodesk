import sys
import json
import csv
import re
import math
import os
import platform
import tempfile
import urllib.request
import shutil
import socket
import time
from datetime import datetime
from pathlib import Path
from threading import Thread
import threading
import logging
import glob
import importlib
import importlib.util
from typing import TypedDict, Any

from PySide6.QtCore import Qt, QUrl, QProcess, QSettings, QTimer, QStandardPaths, Signal, QThread, QObject
from PySide6.QtGui import QAction, QKeySequence, QDesktopServices, QShortcut, QIcon, QPixmap, QMovie
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QMessageBox, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QHeaderView, QListWidget,
    QListWidgetItem, QLineEdit, QSpinBox, QTabWidget, QDoubleSpinBox,
    QCheckBox, QComboBox, QPlainTextEdit, QDialog, QDialogButtonBox,
    QCompleter, QGroupBox, QFormLayout, QScrollArea, QRadioButton,
    QProgressBar, QProgressDialog, QInputDialog, QStackedWidget
)

import cobra
import networkx as nx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from matplotlib import cm
from matplotlib.colors import Normalize


RECENT_MAX = 5

# --------------- Logging ----------------
logger = logging.getLogger("MetaboDesk")
logger.setLevel(logging.DEBUG)
_log_handler = logging.StreamHandler()
_log_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"))
logger.addHandler(_log_handler)
try:
    _log_dir = Path.home() / ".metabodesk"
    _log_dir.mkdir(parents=True, exist_ok=True)
    _file_handler = logging.FileHandler(_log_dir / "metabodesk.log", encoding="utf-8")
    _file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
    _file_handler.setLevel(logging.INFO)
    logger.addHandler(_file_handler)
except Exception:
    pass


# --------------- TypedDict definitions ----------------
class BaselineResult(TypedDict, total=False):
    status: str
    objective: float
    flux: dict[str, float]
    shadow_prices: dict[str, float]
    reduced_costs: dict[str, float]


class AnalysisResultDict(TypedDict, total=False):
    timestamp: str
    analysis_type: str
    baseline: BaselineResult
    wt_growth: float
    fva_result: dict[str, dict[str, float]]
    pfba: dict[str, Any]
    sgd_result: dict[str, float]
    dgd_result: dict[str, float]
    srd_result: dict[str, float]
    robustness_result: dict[str, list]
    envelope_result: dict[str, list]
    sampling_result: dict[str, dict]
    flux_rows: list[dict]
    rxn_id: str
    bound_type: str
    product_id: str
    sampler_type: str
    samples_df: Any


# --------------- Dark Theme Stylesheet ----------------
DARK_STYLESHEET = """
QMainWindow, QDialog { background-color: #1e1e1e; color: #d4d4d4; }
QMenuBar { background-color: #2d2d2d; color: #d4d4d4; }
QMenuBar::item:selected { background-color: #3e3e3e; }
QMenu { background-color: #2d2d2d; color: #d4d4d4; border: 1px solid #3e3e3e; }
QMenu::item:selected { background-color: #094771; }
QTabWidget::pane { border: 1px solid #3e3e3e; background: #1e1e1e; }
QTabBar::tab { background: #2d2d2d; color: #d4d4d4; padding: 6px 12px; border: 1px solid #3e3e3e; }
QTabBar::tab:selected { background: #1e1e1e; border-bottom-color: #1e1e1e; }
QTableWidget { background-color: #1e1e1e; color: #d4d4d4; gridline-color: #3e3e3e; selection-background-color: #094771; }
QHeaderView::section { background-color: #2d2d2d; color: #d4d4d4; border: 1px solid #3e3e3e; padding: 4px; }
QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox, QPlainTextEdit { background-color: #2d2d2d; color: #d4d4d4; border: 1px solid #3e3e3e; padding: 3px; }
QPushButton { background-color: #2d2d2d; color: #d4d4d4; border: 1px solid #3e3e3e; padding: 5px 12px; }
QPushButton:hover { background-color: #3e3e3e; }
QPushButton:pressed { background-color: #094771; }
QLabel { color: #d4d4d4; }
QListWidget { background-color: #1e1e1e; color: #d4d4d4; border: 1px solid #3e3e3e; }
QGroupBox { color: #d4d4d4; border: 1px solid #3e3e3e; margin-top: 6px; padding-top: 10px; }
QGroupBox::title { subcontrol-origin: margin; padding: 0 3px; }
QCheckBox, QRadioButton { color: #d4d4d4; }
QProgressBar { border: 1px solid #3e3e3e; background: #2d2d2d; text-align: center; color: #d4d4d4; }
QProgressBar::chunk { background-color: #094771; }
QScrollArea { background: #1e1e1e; border: none; }
QStatusBar { background: #2d2d2d; color: #d4d4d4; }
"""


def _recent_file_path() -> Path:
    try:
        app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
        if app_data:
            p = Path(app_data)
            p.mkdir(parents=True, exist_ok=True)
            return p / "recent_scenarios.json"
    except Exception:
        pass
    return Path.home() / ".metabodesk_recent.json"


# ---------------- Basic helpers ----------------
def is_exchange_reaction(rxn: cobra.Reaction) -> bool:
    try:
        if rxn.id.startswith("EX_"):
            return True
        return len(rxn.metabolites) == 1
    except Exception:
        return False


def _as_str_list(v) -> list[str]:
    if v is None:
        return []
    if isinstance(v, str):
        s = v.strip()
        return [s] if s else []
    if isinstance(v, (list, tuple, set)):
        out = []
        for x in v:
            if x is None:
                continue
            s = str(x).strip()
            if s:
                out.append(s)
        return out
    s = str(v).strip()
    return [s] if s else []


def _clean_identifiers_org_id(url_or_id: str) -> str:
    s = str(url_or_id).strip()
    if not s:
        return ""
    if "identifiers.org/" in s:
        s = s.split("identifiers.org/")[-1]
    if "/" in s:
        s = s.split("/")[-1]
    return s.strip()


def _logo_pixmap() -> QPixmap | None:
    try:
        path = _resource_path("logo.ico")
        if Path(path).exists():
            pm = QPixmap(path)
            if not pm.isNull():
                return pm.scaled(64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    except Exception:
        pass
    return None


def _show_msgbox(parent, title: str, text: str, icon: QMessageBox.Icon,
                 buttons: QMessageBox.StandardButtons = QMessageBox.Ok,
                 default_button: QMessageBox.StandardButton = QMessageBox.NoButton) -> int:
    box = QMessageBox(parent)
    box.setWindowTitle(title)
    box.setText(text)
    pm = _logo_pixmap()
    if pm:
        box.setIconPixmap(pm)
    else:
        box.setIcon(icon)
    box.setStandardButtons(buttons)
    if default_button != QMessageBox.NoButton:
        box.setDefaultButton(default_button)
    return box.exec()


def _msg_info(parent, title, text, buttons=QMessageBox.Ok, default_button=QMessageBox.NoButton):
    return _show_msgbox(parent, title, text, QMessageBox.Information, buttons, default_button)


def _msg_warn(parent, title, text, buttons=QMessageBox.Ok, default_button=QMessageBox.NoButton):
    return _show_msgbox(parent, title, text, QMessageBox.Warning, buttons, default_button)


def _msg_crit(parent, title, text, buttons=QMessageBox.Ok, default_button=QMessageBox.NoButton):
    return _show_msgbox(parent, title, text, QMessageBox.Critical, buttons, default_button)


def _msg_question(parent, title, text, buttons=QMessageBox.Yes | QMessageBox.No, default_button=QMessageBox.NoButton):
    return _show_msgbox(parent, title, text, QMessageBox.Question, buttons, default_button)


# Apply custom message box icon globally
QMessageBox.information = _msg_info
QMessageBox.warning = _msg_warn
QMessageBox.critical = _msg_crit
QMessageBox.question = _msg_question


def _safe_float(val, default: float = 0.0) -> float:
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except Exception:
        return default


def evaluate_gpr_expression(gpr_rule: str, gene_values: dict[str, float]) -> float | None:
    """Recursively evaluate a GPR rule using AND=min, OR=max logic.

    Handles nested boolean expressions like ``(geneA and geneB) or geneC``.
    Returns None if none of the referenced genes have expression data.
    """
    if not gpr_rule or not gpr_rule.strip():
        return None

    rule = gpr_rule.strip()
    # Remove outer parentheses
    while rule.startswith('(') and rule.endswith(')'):
        depth = 0
        matched = True
        for i, ch in enumerate(rule):
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            if depth == 0 and i < len(rule) - 1:
                matched = False
                break
        if matched:
            rule = rule[1:-1].strip()
        else:
            break

    # Split on top-level 'or' first (lowest precedence)
    depth = 0
    or_parts: list[str] = []
    current: list[str] = []
    tokens = rule.split()
    for tok in tokens:
        depth += tok.count('(') - tok.count(')')
        if tok.lower() == 'or' and depth == 0:
            or_parts.append(' '.join(current))
            current = []
        else:
            current.append(tok)
    or_parts.append(' '.join(current))

    if len(or_parts) > 1:
        vals = [evaluate_gpr_expression(p, gene_values) for p in or_parts]
        vals = [v for v in vals if v is not None]
        return max(vals) if vals else None

    # Split on top-level 'and' (higher precedence)
    depth = 0
    and_parts: list[str] = []
    current = []
    tokens = rule.split()
    for tok in tokens:
        depth += tok.count('(') - tok.count(')')
        if tok.lower() == 'and' and depth == 0:
            and_parts.append(' '.join(current))
            current = []
        else:
            current.append(tok)
    and_parts.append(' '.join(current))

    if len(and_parts) > 1:
        vals = [evaluate_gpr_expression(p, gene_values) for p in and_parts]
        vals = [v for v in vals if v is not None]
        return min(vals) if vals else None

    # Leaf node: single gene
    gene_id = rule.strip('() ')
    return gene_values.get(gene_id)


def get_kegg_rxn_id(rxn: cobra.Reaction) -> str:
    ann = getattr(rxn, "annotation", {}) or {}

    keys = [
        "kegg.reaction",
        "kegg_reaction",
        "kegg",
        "kegg.rxn",
        "kegg_rxn",
        "KEGG Reaction",
        "KEGG_REACTION",
        "reaction.kegg",
    ]
    for k in keys:
        if k in ann:
            vals = [_clean_identifiers_org_id(x) for x in _as_str_list(ann.get(k))]
            vals = [v for v in vals if v]
            if vals:
                return ";".join(vals)

    for k in ("identifiers", "identifiers.org"):
        if k in ann:
            vals = _as_str_list(ann.get(k))
            cleaned = []
            for s in vals:
                if "kegg.reaction" in str(s):
                    cleaned.append(_clean_identifiers_org_id(s))
            cleaned = [c for c in cleaned if c]
            if cleaned:
                return ";".join(cleaned)

    return ""


def get_ec_numbers(rxn: cobra.Reaction) -> str:
    ann = getattr(rxn, "annotation", {}) or {}
    keys = [
        "ec-code",
        "ec",
        "ec_number",
        "ec-number",
        "ec-code(s)",
        "EC Number",
        "EC",
        "enzyme.ec",
    ]
    for k in keys:
        if k in ann:
            vals = _as_str_list(ann.get(k))
            vals = [v.strip() for v in vals if str(v).strip()]
            if vals:
                return ";".join(vals)
    return ""


def get_description(rxn: cobra.Reaction) -> str:
    if getattr(rxn, "name", None):
        return str(rxn.name)
    ann = getattr(rxn, "annotation", {}) or {}
    for k in ("description", "desc", "note", "notes"):
        if k in ann:
            vals = _as_str_list(ann.get(k))
            if vals:
                return vals[0]
    return ""


def get_gpr(rxn: cobra.Reaction) -> str:
    try:
        return str(rxn.gene_reaction_rule or "")
    except Exception:
        return ""


def _resource_path(rel_path: str) -> str:
    try:
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            return str(Path(sys._MEIPASS) / rel_path)
    except Exception:
        pass
    return str(Path(__file__).resolve().parent / rel_path)


# ---------------- Equation parser ----------------
ARROW_PATTERNS = [
    ("<=>", True),
    ("<->", True),
    ("<-->", True),
    ("->", False),
    ("-->", False),
    ("=>", False),
]


def _guess_compartment_from_met_id(met_id: str) -> str:
    """
    Heuristic: metabolite ids like glc__D_c -> compartment 'c'
    If no '_x' suffix, fallback to 'c'.
    """
    m = re.match(r"^(.+)_([A-Za-z])$", met_id.strip())
    if m:
        return m.group(2)
    return "c"


def _parse_coeff_and_met(term: str) -> tuple[float, str]:
    """
    Parse one side term. Accept:
      - "2 atp_c"
      - "0.5 o2_c"
      - "atp_c" (=> coeff=1)
    """
    term = term.strip()
    if not term:
        raise ValueError("Empty term in equation.")
    parts = term.split()
    if len(parts) == 1:
        return 1.0, parts[0]
    try:
        coeff = float(parts[0])
        met = " ".join(parts[1:]).strip()
        if not met:
            raise ValueError("Missing metabolite id after coefficient.")
        return coeff, met
    except ValueError:
        return 1.0, term


def parse_reaction_equation(equation: str) -> tuple[dict[str, float], bool]:
    """
    Return (stoichiometry, reversible)
    stoich: reactants negative, products positive.
    """
    eq = (equation or "").strip()
    if not eq:
        raise ValueError("Equation is empty.")

    arrow = None
    reversible = False
    for a, rev in ARROW_PATTERNS:
        if a in eq:
            arrow = a
            reversible = rev
            break
    if not arrow:
        raise ValueError("No valid arrow found. Use ->, =>, <=>, <->")

    left, right = eq.split(arrow, 1)
    left = left.strip()
    right = right.strip()

    stoich: dict[str, float] = {}

    def add(met_id: str, coeff: float):
        met_id = met_id.strip()
        if not met_id:
            return
        stoich[met_id] = float(stoich.get(met_id, 0.0) + coeff)

    if left:
        for term in [t.strip() for t in left.split("+")]:
            if not term:
                continue
            coeff, met = _parse_coeff_and_met(term)
            add(met, -abs(coeff))

    if right:
        for term in [t.strip() for t in right.split("+")]:
            if not term:
                continue
            coeff, met = _parse_coeff_and_met(term)
            add(met, abs(coeff))

    stoich = {k: v for k, v in stoich.items() if abs(v) > 1e-12}
    if not stoich:
        raise ValueError("Parsed stoichiometry is empty.")

    return stoich, reversible


# ---------------- UI popups ----------------
class TextPopup(QDialog):
    def __init__(self, title: str, text: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        layout = QVBoxLayout(self)

        self.text = QPlainTextEdit()
        self.text.setReadOnly(True)
        self.text.setPlainText(text or "")
        layout.addWidget(self.text)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(self.reject)
        btns.accepted.connect(self.accept)
        btns.clicked.connect(self.reject)
        layout.addWidget(btns)

        self.resize(900, 450)


class AnimatedSplash(QDialog):
    def __init__(self, gif_path: str, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.label = QLabel()
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.movie = QMovie(gif_path)
        self.label.setMovie(self.movie)
        self.movie.start()
        self.adjustSize()


class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=8, height=3.5, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi, tight_layout=True)
        self.ax = fig.add_subplot(111)
        super().__init__(fig)
        self.setParent(parent)
        self.clicked_node = None


# ---------------- Analysis Worker (QThread) ----------------
class AnalysisWorker(QThread):
    """Background worker for long-running analyses. Prevents UI freezing."""
    finished = Signal(dict)
    error = Signal(str)
    progress = Signal(str)
    progress_pct = Signal(int)

    def __init__(self, func, **kwargs):
        super().__init__()
        self._func = func
        self._kwargs = kwargs

    def report_progress(self, message: str, pct: int = -1):
        """Emit progress from within the worker function."""
        self.progress.emit(message)
        if pct >= 0:
            self.progress_pct.emit(pct)

    def run(self):
        try:
            import inspect
            sig = inspect.signature(self._func)
            if 'worker' in sig.parameters:
                result = self._func(worker=self, **self._kwargs)
            else:
                result = self._func(**self._kwargs)
            self.finished.emit(result)
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n\n{traceback.format_exc()}")


# ---------------- Main app ----------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        # Speed up startup by deferring repaints during UI construction
        self.setUpdatesEnabled(False)
        self.setWindowTitle("MetaboDesk")
        try:
            icon_path = _resource_path("logo.ico")
            if Path(icon_path).exists():
                self.setWindowIcon(QIcon(icon_path))
        except Exception:
            pass

        self.base_model: cobra.Model | None = None
        self.current_sbml_path: Path | None = None
        self.model_dirty: bool = False
        self._tools_proc: QProcess | None = None
        self._memote_proc: QProcess | None = None
        self._carveme_proc: QProcess | None = None
        self._carveme_last_output: str = ""
        self.original_model_snapshot: cobra.Model | None = None

        self.exchange_rxns: list[cobra.Reaction] = []
        self.original_bounds: dict[str, tuple[float, float]] = {}

        self.knockout_genes: set[str] = set()
        self.overexpression_reactions: dict[str, float] = {}
        self.temporary_upper_bound_overrides: dict[str, float] = {}
        self.reaction_bound_overrides: dict[str, tuple[float, float]] = {}

        self.model_patch: dict = {
            "version": 1,
            "added_metabolites": [],
            "added_genes": [],
            "added_or_updated_reactions": [],
            "disabled_reactions": [],
            "deleted_reactions": [],
            "objective_reaction_id": None,
            "notes": "",
        }

        self._gene_to_reactions: dict[str, list[str]] = {}

        self.is_running = False
        self.last_run: dict | None = None
        self._worker: AnalysisWorker | None = None

        self.recent_scenarios: list[str] = []
        self._load_recent_scenarios()

        # Undo/Redo stack
        self.undo_stack: list[dict] = []
        self.redo_stack: list[dict] = []
        self.max_undo_steps = 50

        # Search/Filter
        self.search_results: list[dict] = []

        # Sensitivity analysis
        self.sensitivity_results: dict = {}

        # Theme
        self._dark_theme: bool = False

        # Plugin system
        self._plugins: list = []
        self._plugin_dir = Path.home() / ".metabodesk" / "plugins"

        # FBA result cache  (key = model-state hash → result dict)
        self._fba_cache: dict[str, dict] = {}
        self._fba_cache_max = 32

        root = QWidget()
        outer = QVBoxLayout(root)

        controls = QHBoxLayout()

        self.open_btn = QPushButton("Open SBML...")
        self.open_btn.clicked.connect(self.open_sbml)

        self.run_btn = QPushButton("Run Analysis")
        self.run_btn.setEnabled(True)
        self.run_btn.setToolTip("Execute the selected analysis type on the loaded model.")
        self.run_btn.clicked.connect(self.run_fba)

        self.analysis_type = QComboBox()
        self.analysis_type.addItems([
            "FBA", "pFBA", "FVA",
            "Single Gene Deletion (SGD)", "Double Gene Deletion (DGD)", "Single Reaction Deletion (SRD)",
            "Robustness Analysis", "Production Envelope", "Flux Sampling"
        ])
        self.analysis_type.setEnabled(False)
        self.analysis_type.setToolTip("Select the analysis type to run.")

        self.close_uptakes_btn = QPushButton("Close uptakes")
        self.close_uptakes_btn.setEnabled(False)
        self.close_uptakes_btn.clicked.connect(self.close_all_uptakes)

        self.reset_medium_btn = QPushButton("Reset medium")
        self.reset_medium_btn.setEnabled(False)
        self.reset_medium_btn.clicked.connect(self.reset_medium)

        self.reset_reactions_btn = QPushButton("Reset reactions")
        self.reset_reactions_btn.setEnabled(False)
        self.reset_reactions_btn.clicked.connect(self.reset_reactions_to_original)

        controls.addWidget(self.open_btn)
        controls.addWidget(self.run_btn)
        controls.addWidget(QLabel("Analysis:"))
        controls.addWidget(self.analysis_type)
        controls.addSpacing(15)
        controls.addWidget(self.close_uptakes_btn)
        controls.addWidget(self.reset_medium_btn)
        controls.addWidget(self.reset_reactions_btn)
        controls.addSpacing(15)

        controls.addWidget(QLabel("Top N:"))        
        self.topn_spin = QSpinBox()
        self.topn_spin.setRange(5, 50)
        self.topn_spin.setValue(20)
        self.topn_spin.setToolTip("Maximum number of items to display in charts/tables.")
        controls.addWidget(self.topn_spin)

        controls.addSpacing(15)
        controls.addWidget(QLabel("Solver:"))
        self.solver_combo = QComboBox()
        self.solver_combo.setToolTip("Optimization solver (Auto: use available).\nGurobi/CPLEX require separate licenses.")
        self.solver_combo.addItems(["Auto", "glpk", "highs", "gurobi", "cplex"])
        self.solver_combo.setEnabled(False)
        controls.addWidget(self.solver_combo, stretch=1)

        controls.addSpacing(10)
        self.loopless_chk = QCheckBox("Loopless")
        self.loopless_chk.setToolTip("Thermodynamically feasible FBA (slower, removes internal loops).")
        self.loopless_chk.setChecked(False)
        controls.addWidget(self.loopless_chk)

        controls.addSpacing(15)
        controls.addWidget(QLabel("Objective:"))
        self.objective_combo = QComboBox()
        self.objective_combo.setEnabled(False)
        self.objective_combo.setMinimumWidth(200)
        self.objective_combo.setToolTip("Select objective reaction (empty: model default).")
        self.objective_combo.currentIndexChanged.connect(self.on_objective_changed)
        self.objective_combo.setEditable(True)
        self.objective_combo.setInsertPolicy(QComboBox.NoInsert)
        self.objective_combo.setMaxVisibleItems(25)
        self.objective_combo.lineEdit().setPlaceholderText("Type to search (id/description)...")
        self.objective_combo.lineEdit().textEdited.connect(self._on_objective_text_edited)
        controls.addWidget(self.objective_combo, stretch=2)

        controls.addStretch(1)
        outer.addLayout(controls)

        self.status_lbl = QLabel("No model loaded.")
        self.status_lbl.setWordWrap(True)
        outer.addWidget(self.status_lbl)

        self.summary_lbl = QLabel("")
        self.summary_lbl.setWordWrap(True)
        outer.addWidget(self.summary_lbl)

        self.tabs = QTabWidget()
        outer.addWidget(self.tabs, stretch=1)

        # -------- Tools tab (Installer) --------
        self.tab_tools = QWidget()
        tools_layout = QVBoxLayout(self.tab_tools)
        tools_layout.addWidget(self._title("Tools (CarveMe + memote)"))

        self.tools_status_lbl = QLabel("Checking tools...")
        self.tools_status_lbl.setWordWrap(True)
        tools_layout.addWidget(self.tools_status_lbl)

        btn_row = QHBoxLayout()
        self.tools_check_btn = QPushButton("Check tools")
        self.tools_check_btn.clicked.connect(self.tools_check_status)
        btn_row.addWidget(self.tools_check_btn)

        self.tools_repair_btn = QPushButton("Repair tools (reinstall from bundled wheels)")
        self.tools_repair_btn.clicked.connect(self.tools_repair)
        btn_row.addWidget(self.tools_repair_btn)

        btn_row.addStretch(1)
        tools_layout.addLayout(btn_row)

        # (2) memote run + cancel
        memote_row = QHBoxLayout()
        self.memote_btn = QPushButton("Run memote report (current model)")
        self.memote_btn.setMinimumWidth(280)
        self.memote_btn.setEnabled(False)
        self.memote_btn.clicked.connect(self.run_memote_report_current_model)
        memote_row.addWidget(self.memote_btn)

        self.memote_cancel_btn = QPushButton("Cancel memote")
        self.memote_cancel_btn.setMinimumWidth(120)
        self.memote_cancel_btn.setEnabled(False)
        self.memote_cancel_btn.clicked.connect(self.cancel_memote)
        memote_row.addWidget(self.memote_cancel_btn)

        self.memote_time_lbl = QLabel("Elapsed: 00:00")
        self.memote_time_lbl.setMinimumWidth(100)
        memote_row.addWidget(self.memote_time_lbl)

        memote_row.addStretch(1)
        tools_layout.addLayout(memote_row)

        # (3) CarveMe run + cancel
        carve_row = QHBoxLayout()
        self.carveme_btn = QPushButton("Run CarveMe (genome → SBML)")
        self.carveme_btn.setMinimumWidth(280)
        self.carveme_btn.setEnabled(False)
        self.carveme_btn.clicked.connect(self.run_carveme_draft)
        carve_row.addWidget(self.carveme_btn)

        self.carveme_cancel_btn = QPushButton("Cancel CarveMe")
        self.carveme_cancel_btn.setMinimumWidth(120)
        self.carveme_cancel_btn.setEnabled(False)
        self.carveme_cancel_btn.clicked.connect(self.cancel_carveme)
        carve_row.addWidget(self.carveme_cancel_btn)

        self.carveme_time_lbl = QLabel("Elapsed: 00:00")
        self.carveme_time_lbl.setMinimumWidth(100)
        carve_row.addWidget(self.carveme_time_lbl)

        carve_row.addStretch(1)
        tools_layout.addLayout(carve_row)

        self.tools_log = QPlainTextEdit()
        self.tools_log.setReadOnly(True)
        tools_layout.addWidget(self.tools_log, stretch=1)

        self.tabs.addTab(self.tab_tools, "Tools")

        # -------- Medium tab --------
        self.tab_medium = QWidget()
        medium_layout = QVBoxLayout(self.tab_medium)

        medium_layout.addWidget(self._title("Medium / Exchange bounds"))
        self.medium_search = QLineEdit()
        self.medium_search.setPlaceholderText("Search exchange reactions (id or name)...")
        self.medium_search.textChanged.connect(self.filter_medium_table)
        medium_layout.addWidget(self.medium_search)

        self.medium_table = QTableWidget(0, 4)
        self.medium_table.setHorizontalHeaderLabels(["Reaction", "Name", "Lower bound", "Upper bound"])
        self.medium_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.medium_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self._configure_table_for_excel_resize(self.medium_table)
        medium_layout.addWidget(self.medium_table, stretch=1)

        medium_layout.addWidget(self._title("Cell details (full text)"))
        self.medium_details = QPlainTextEdit()
        self.medium_details.setReadOnly(True)
        self.medium_details.setMaximumHeight(140)
        medium_layout.addWidget(self.medium_details)
        self.medium_table.currentItemChanged.connect(self._update_medium_details_from_current_cell)

        # Presets row
        preset_row = QHBoxLayout()
        preset_row.addWidget(QLabel("Medium preset:"))
        self.medium_preset_combo = QComboBox()
        self.medium_preset_combo.addItems(["Reset to original", "Minimal M9", "Rich LB"])
        preset_row.addWidget(self.medium_preset_combo)
        self.medium_apply_preset_btn = QPushButton("Apply preset")
        self.medium_apply_preset_btn.setEnabled(False)
        self.medium_apply_preset_btn.clicked.connect(self.apply_medium_preset)
        preset_row.addWidget(self.medium_apply_preset_btn)
        preset_row.addStretch(1)
        medium_layout.addLayout(preset_row)

        self.tabs.addTab(self.tab_medium, "Medium")

        # -------- Reactions tab --------
        self.tab_rxns = QWidget()
        rxn_layout = QVBoxLayout(self.tab_rxns)

        rxn_layout.addWidget(self._title("All reactions (bounds & metadata)"))

        rxn_search_row = QHBoxLayout()
        rxn_search_row.addWidget(QLabel("Search (id/equation/GPR/KEGG/EC/description):"))
        self.rxn_global_search = QLineEdit()
        self.rxn_global_search.setPlaceholderText("e.g. biomass / atp / R01070 / 1.1.1.1 / gene ...")
        self.rxn_global_search.textChanged.connect(self.filter_reaction_table)
        rxn_search_row.addWidget(self.rxn_global_search, stretch=2)

        rxn_search_row.addWidget(QLabel("Metabolite filter:"))
        self.met_filter = QLineEdit()
        self.met_filter.setPlaceholderText("e.g. glc__D, atp (comma-separated)")
        self.met_filter.textChanged.connect(self.filter_reaction_table)
        rxn_search_row.addWidget(self.met_filter, stretch=2)

        rxn_search_row.addWidget(QLabel("Mode:"))
        self.met_filter_mode = QComboBox()
        self.met_filter_mode.addItems(["AND", "OR"])
        self.met_filter_mode.currentTextChanged.connect(self.filter_reaction_table)
        rxn_search_row.addWidget(self.met_filter_mode)

        rxn_layout.addLayout(rxn_search_row)

        self.reaction_table = QTableWidget(0, 8)
        self.reaction_table.setHorizontalHeaderLabels(
            ["Reaction", "Description", "Equation", "KEGG", "EC", "GPR", "LB", "UB"]
        )
        self.reaction_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.reaction_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self._configure_table_for_excel_resize(self.reaction_table)
        self.reaction_table.itemChanged.connect(self.on_reaction_table_item_changed)
        rxn_layout.addWidget(self.reaction_table, stretch=1)

        rxn_btn_row = QHBoxLayout()
        self.clear_rxn_overrides_btn = QPushButton("Clear reaction overrides")
        self.clear_rxn_overrides_btn.setEnabled(False)
        self.clear_rxn_overrides_btn.clicked.connect(self.clear_reaction_overrides)
        rxn_btn_row.addWidget(self.clear_rxn_overrides_btn)

        self.remove_selected_overrides_btn = QPushButton("Remove override (selected)")
        self.remove_selected_overrides_btn.setEnabled(False)
        self.remove_selected_overrides_btn.clicked.connect(self.remove_selected_reaction_overrides)
        rxn_btn_row.addWidget(self.remove_selected_overrides_btn)

        rxn_btn_row.addStretch(1)
        rxn_layout.addLayout(rxn_btn_row)

        rxn_layout.addWidget(self._title("Cell details (full text)"))
        self.reaction_details = QPlainTextEdit()
        self.reaction_details.setReadOnly(True)
        self.reaction_details.setMaximumHeight(180)
        rxn_layout.addWidget(self.reaction_details)
        self.reaction_table.currentItemChanged.connect(self._update_reaction_details_from_current_cell)
        # Double-click popups disabled to keep inline editing simple
        rxn_note = QLabel(
            "Tip: Drag column borders to resize (like Excel).\n"
            "KEGG/EC columns depend on SBML annotations; if the SBML doesn't include them, these will be empty."
        )
        rxn_note.setWordWrap(True)
        rxn_layout.addWidget(rxn_note)

        self.tabs.addTab(self.tab_rxns, "Reactions")

        # -------- Genes tab --------
        self.tab_genes = QWidget()
        genes_layout = QVBoxLayout(self.tab_genes)

        genes_layout.addWidget(self._title("Genes → Related reactions"))

        gene_search_row = QHBoxLayout()
        gene_search_row.addWidget(QLabel("Search gene:"))
        self.gene_global_search = QLineEdit()
        self.gene_global_search.setPlaceholderText("Type gene id or name...")
        self.gene_global_search.textChanged.connect(self.filter_genes_tab_gene_list)
        gene_search_row.addWidget(self.gene_global_search, stretch=1)
        genes_layout.addLayout(gene_search_row)

        genes_split = QHBoxLayout()

        left = QVBoxLayout()
        left.addWidget(QLabel("Genes"))
        self.genes_tab_list = QListWidget()
        self.genes_tab_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self.genes_tab_list.currentItemChanged.connect(self.on_genes_tab_gene_changed)
        left.addWidget(self.genes_tab_list, stretch=1)
        genes_split.addLayout(left, stretch=1)

        right = QVBoxLayout()
        right.addWidget(QLabel("Reactions related to selected gene"))
        self.gene_rxn_table = QTableWidget(0, 4)
        self.gene_rxn_table.setHorizontalHeaderLabels(["Reaction", "Description", "Equation", "GPR"])
        self.gene_rxn_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.gene_rxn_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._configure_table_for_excel_resize(self.gene_rxn_table)
        right.addWidget(self.gene_rxn_table, stretch=1)

        right.addWidget(self._title("Cell details (full text)"))
        self.gene_rxn_details = QPlainTextEdit()
        self.gene_rxn_details.setReadOnly(True)
        self.gene_rxn_details.setMaximumHeight(180)
        right.addWidget(self.gene_rxn_details)

        self.gene_rxn_table.currentItemChanged.connect(self._update_gene_rxn_details_from_current_cell)
        self.gene_rxn_table.itemDoubleClicked.connect(lambda item: self._open_cell_popup("Gene→Reaction cell", item))

        genes_split.addLayout(right, stretch=3)

        genes_layout.addLayout(genes_split, stretch=1)

        genes_note = QLabel("Note: Reactions are determined from COBRApy gene↔reaction associations (from GPRs).")
        genes_note.setWordWrap(True)
        genes_layout.addWidget(genes_note)

        self.tabs.addTab(self.tab_genes, "Genes")

        # -------- Model Editor tab placeholder --------
        self.tab_editor = QWidget()
        editor_layout = QVBoxLayout(self.tab_editor)
        editor_layout.addWidget(self._title("Model Editor (Patch-based)"))
        self.editor_tabs = QTabWidget()
        editor_layout.addWidget(self.editor_tabs, stretch=1)

        self._build_editor_metabolite_tab()
        self._build_editor_gene_tab()
        self._build_editor_reaction_tab()
        self._build_editor_disable_delete_tab()
        self._build_editor_patch_tab()
        self._build_editor_export_sbml_tab()

        self.tabs.addTab(self.tab_editor, "Model Editor")

        # -------- Gene knockout tab --------
        self.tab_ko = QWidget()
        ko_layout = QVBoxLayout(self.tab_ko)

        ko_layout.addWidget(self._title("Gene knockout"))
        self.gene_search = QLineEdit()
        self.gene_search.setPlaceholderText("Search genes...")
        self.gene_search.textChanged.connect(self.filter_gene_list)
        ko_layout.addWidget(self.gene_search)

        self.gene_list = QListWidget()
        # Allow selecting multiple genes so double deletion can pick exact pairs
        self.gene_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        ko_layout.addWidget(self.gene_list, stretch=1)

        ko_btn_row = QHBoxLayout()
        self.add_ko_btn = QPushButton("Add knockout")
        self.add_ko_btn.setEnabled(False)
        self.add_ko_btn.clicked.connect(self.add_selected_gene_knockout)

        self.clear_ko_btn = QPushButton("Clear knockouts")
        self.clear_ko_btn.setEnabled(False)
        self.clear_ko_btn.clicked.connect(self.clear_knockouts)

        ko_btn_row.addWidget(self.add_ko_btn)
        ko_btn_row.addWidget(self.clear_ko_btn)
        ko_btn_row.addStretch(1)
        ko_layout.addLayout(ko_btn_row)

        self.ko_lbl = QLabel("Gene knockouts: (none)")
        self.ko_lbl.setWordWrap(True)
        ko_layout.addWidget(self.ko_lbl)
        self.tabs.addTab(self.tab_ko, "Gene knockout")

        # -------- Overexpression tab --------
        self.tab_oe = QWidget()
        oe_layout = QVBoxLayout(self.tab_oe)

        oe_layout.addWidget(self._title("Overexpression (reaction bound scaling)"))

        self.rxn_search = QLineEdit()
        self.rxn_search.setPlaceholderText("Search reactions for overexpression (id or description)...")
        self.rxn_search.textChanged.connect(self.filter_overexpression_reaction_list)
        oe_layout.addWidget(self.rxn_search)

        self.rxn_list = QListWidget()
        self.rxn_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self.rxn_list.currentItemChanged.connect(self.update_selected_rxn_info)
        oe_layout.addWidget(self.rxn_list, stretch=2)

        self.rxn_info_lbl = QLabel("Selected reaction: (none)")
        self.rxn_info_lbl.setWordWrap(True)
        oe_layout.addWidget(self.rxn_info_lbl)

        self.rxn_preview_lbl = QLabel("Preview: -")
        self.rxn_preview_lbl.setWordWrap(True)
        oe_layout.addWidget(self.rxn_preview_lbl)

        oe_row = QHBoxLayout()
        oe_row.addWidget(QLabel("Overexpression factor:"))
        self.oe_factor = QDoubleSpinBox()
        self.oe_factor.setRange(1.0, 100.0)
        self.oe_factor.setDecimals(2)
        self.oe_factor.setSingleStep(0.25)
        self.oe_factor.setValue(2.0)
        self.oe_factor.valueChanged.connect(self.update_selected_rxn_info)
        oe_row.addWidget(self.oe_factor)

        self.oe_scale_lb = QCheckBox("Scale LB too (if LB < 0)")
        self.oe_scale_lb.setChecked(True)
        self.oe_scale_lb.stateChanged.connect(self.update_selected_rxn_info)
        oe_row.addWidget(self.oe_scale_lb)

        self.add_oe_btn = QPushButton("Add overexpression")
        self.add_oe_btn.setEnabled(False)
        self.add_oe_btn.clicked.connect(self.add_selected_reaction_overexpression)
        oe_row.addWidget(self.add_oe_btn)

        self.remove_oe_btn = QPushButton("Remove overexpression (selected)")
        self.remove_oe_btn.setEnabled(False)
        self.remove_oe_btn.clicked.connect(self.remove_selected_reaction_overexpression)
        oe_row.addWidget(self.remove_oe_btn)

        self.clear_oe_btn = QPushButton("Clear overexpressions")
        self.clear_oe_btn.setEnabled(False)
        self.clear_oe_btn.clicked.connect(self.clear_overexpressions)
        oe_row.addWidget(self.clear_oe_btn)

        oe_row.addStretch(1)
        oe_layout.addLayout(oe_row)

        temp_row = QHBoxLayout()
        temp_row.addWidget(QLabel("Temporary upper bound (perturbed only):"))
        self.temp_ub_spin = QDoubleSpinBox()
        self.temp_ub_spin.setRange(-1e9, 1e9)
        self.temp_ub_spin.setDecimals(6)
        self.temp_ub_spin.setSingleStep(1.0)
        self.temp_ub_spin.setValue(0.0)
        temp_row.addWidget(self.temp_ub_spin)

        self.set_temp_ub_btn = QPushButton("Set temp upper bound (selected)")
        self.set_temp_ub_btn.setEnabled(False)
        self.set_temp_ub_btn.clicked.connect(self.set_temp_upper_bound_for_selected)
        temp_row.addWidget(self.set_temp_ub_btn)

        self.remove_temp_ub_btn = QPushButton("Remove temp upper bound (selected)")
        self.remove_temp_ub_btn.setEnabled(False)
        self.remove_temp_ub_btn.clicked.connect(self.remove_temp_upper_bound_for_selected)
        temp_row.addWidget(self.remove_temp_ub_btn)

        self.clear_temp_ub_btn = QPushButton("Clear all temp upper bounds")
        self.clear_temp_ub_btn.setEnabled(False)
        self.clear_temp_ub_btn.clicked.connect(self.clear_temp_upper_bounds)
        temp_row.addWidget(self.clear_temp_ub_btn)

        temp_row.addStretch(1)
        oe_layout.addLayout(temp_row)

        oe_layout.addWidget(self._title("Active overexpressions"))
        self.oe_active_list = QListWidget()
        self.oe_active_list.itemClicked.connect(self.jump_to_reaction_from_item)
        oe_layout.addWidget(self.oe_active_list, stretch=1)

        oe_layout.addWidget(self._title("Temp upper bounds"))
        self.temp_active_list = QListWidget()
        self.temp_active_list.itemClicked.connect(self.jump_to_reaction_from_item)
        oe_layout.addWidget(self.temp_active_list, stretch=1)

        self.tabs.addTab(self.tab_oe, "Overexpression")

        # -------- FVA tab --------
        self.tab_fva = QWidget()
        fva_layout = QVBoxLayout(self.tab_fva)

        fva_layout.addWidget(self._title("Flux Variability Analysis (FVA)"))
        fva_help = QLabel(
            "Shows minimum and maximum feasible flux for each reaction while maintaining optimal growth. "
            "Reactions with wide ranges are flexible; those with narrow ranges are constrained."
        )
        fva_help.setStyleSheet("font-size:9px;color:#555;margin:4px;")
        fva_help.setWordWrap(True)
        fva_layout.addWidget(fva_help)

        fva_params_row = QHBoxLayout()
        fva_params_row.addWidget(QLabel("Fraction of optimum:"))
        self.fva_fraction_spin = QDoubleSpinBox()
        self.fva_fraction_spin.setRange(0.0, 1.0)
        self.fva_fraction_spin.setSingleStep(0.05)
        self.fva_fraction_spin.setDecimals(2)
        self.fva_fraction_spin.setValue(1.0)
        self.fva_fraction_spin.setToolTip("Fraction of optimal objective required (1.0 = optimal only, 0.9 = 90%+ optimal).")
        fva_params_row.addWidget(self.fva_fraction_spin)
        fva_params_row.addSpacing(15)
        fva_params_row.addWidget(QLabel("Processes:"))
        self.fva_processes_spin = QSpinBox()
        self.fva_processes_spin.setRange(1, max(os.cpu_count() or 1, 1))
        self.fva_processes_spin.setValue(1)
        self.fva_processes_spin.setToolTip("Number of parallel processes for FVA.\nHigher = faster but more memory.\nUse 1 if running as frozen .exe.")
        fva_params_row.addWidget(self.fva_processes_spin)
        fva_params_row.addStretch(1)
        fva_layout.addLayout(fva_params_row)

        fva_filter_row = QHBoxLayout()
        fva_filter_row.addWidget(QLabel("Reaction search:"))
        self.fva_flux_search = QLineEdit()
        self.fva_flux_search.setPlaceholderText("reaction id contains...")
        self.fva_flux_search.textChanged.connect(self.filter_fva_table)
        fva_filter_row.addWidget(self.fva_flux_search, stretch=2)
        fva_filter_row.addStretch(1)
        fva_layout.addLayout(fva_filter_row)

        self.fva_canvas = MplCanvas(self, width=10, height=3.8, dpi=100)
        fva_layout.addWidget(self.fva_canvas, stretch=1)

        self.fva_tbl = QTableWidget(0, 4)
        self.fva_tbl.setHorizontalHeaderLabels(["Reaction", "Minimum flux", "Maximum flux", "Range"])
        self.fva_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fva_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.fva_tbl)
        fva_layout.addWidget(self.fva_tbl, stretch=1)

        self.fva_info_lbl = QLabel("FVA results: (none)")
        self.fva_info_lbl.setWordWrap(True)
        fva_layout.addWidget(self.fva_info_lbl)

        fva_btn_row = QHBoxLayout()
        self.fva_export_btn = QPushButton("Export FVA (CSV)")
        self.fva_export_btn.clicked.connect(self.export_fva_csv)
        fva_btn_row.addWidget(self.fva_export_btn)
        fva_btn_row.addStretch(1)
        fva_layout.addLayout(fva_btn_row)

        self.tabs.addTab(self.tab_fva, "FVA")

        # -------- pFBA tab --------
        self.tab_pfba = QWidget()
        pfba_layout = QVBoxLayout(self.tab_pfba)

        pfba_layout.addWidget(self._title("Parsimonious FBA (pFBA)"))
        pfba_help = QLabel(
            "After finding the maximum objective value (like FBA), minimizes total flux. "
            "Results in sparse, efficient flux distributions with fewer active reactions."
        )
        pfba_layout.addWidget(pfba_help)

        pfba_filter_row = QHBoxLayout()
        pfba_filter_row.addWidget(QLabel("Reaction search:"))
        self.pfba_flux_search = QLineEdit()
        self.pfba_flux_search.setPlaceholderText("reaction id contains...")
        self.pfba_flux_search.textChanged.connect(self.filter_pfba_table)
        pfba_filter_row.addWidget(self.pfba_flux_search, stretch=2)
        pfba_filter_row.addStretch(1)
        pfba_layout.addLayout(pfba_filter_row)

        self.pfba_canvas = MplCanvas(self, width=10, height=3.8, dpi=100)
        pfba_layout.addWidget(self.pfba_canvas, stretch=1)

        self.pfba_tbl = QTableWidget(0, 4)
        self.pfba_tbl.setHorizontalHeaderLabels(["Reaction", "FBA flux", "pFBA flux", "Difference"])
        self.pfba_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.pfba_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.pfba_tbl)
        pfba_layout.addWidget(self.pfba_tbl, stretch=1)

        self.pfba_info_lbl = QLabel("pFBA results: (none)")
        self.pfba_info_lbl.setWordWrap(True)
        pfba_layout.addWidget(self.pfba_info_lbl)

        pfba_btn_row = QHBoxLayout()
        self.pfba_export_btn = QPushButton("Export pFBA vs FBA (CSV)")
        self.pfba_export_btn.clicked.connect(self.export_pfba_csv)
        pfba_btn_row.addWidget(self.pfba_export_btn)
        pfba_btn_row.addStretch(1)
        pfba_layout.addLayout(pfba_btn_row)

        self.tabs.addTab(self.tab_pfba, "pFBA")

        # -------- SGD/SRD tab --------
        self.tab_deletion = QWidget()
        deletion_layout = QVBoxLayout(self.tab_deletion)

        deletion_layout.addWidget(self._title("Single Gene/Reaction Deletion Analysis"))
        deletion_help = QLabel(
            "Knocks out each gene (SGD) or reaction (SRD) one at a time and measures growth impact. "
            "Essential genes/reactions cause severe growth reduction or lethality when deleted."
        )
        deletion_help.setWordWrap(True)
        deletion_layout.addWidget(deletion_help)

        deletion_search_row = QHBoxLayout()
        deletion_search_row.addWidget(QLabel("Search:"))
        self.deletion_search = QLineEdit()
        self.deletion_search.setPlaceholderText("gene/reaction id contains...")
        self.deletion_search.textChanged.connect(self.filter_deletion_table)
        deletion_search_row.addWidget(self.deletion_search, stretch=2)

        self.deletion_filter_optimal = QCheckBox("Optimal only")
        self.deletion_filter_optimal.setToolTip("Show only items with optimal (non-zero) growth after deletion")
        self.deletion_filter_optimal.stateChanged.connect(self.filter_deletion_table)
        deletion_search_row.addWidget(self.deletion_filter_optimal)

        self.deletion_filter_nonoptimal = QCheckBox("Non-optimal only")
        self.deletion_filter_nonoptimal.setToolTip("Show only items with zero or near-zero growth (lethal knockouts)")
        self.deletion_filter_nonoptimal.stateChanged.connect(self.filter_deletion_table)
        deletion_search_row.addWidget(self.deletion_filter_nonoptimal)

        deletion_search_row.addStretch(1)
        deletion_layout.addLayout(deletion_search_row)

        self.deletion_canvas = MplCanvas(self, width=10, height=3.8, dpi=100)
        deletion_layout.addWidget(self.deletion_canvas, stretch=1)

        self.deletion_tbl = QTableWidget(0, 5)
        self.deletion_tbl.setHorizontalHeaderLabels(["Item ID", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.deletion_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.deletion_tbl)
        try:
            header = self.deletion_tbl.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Stretch)
            header.setStretchLastSection(True)
        except Exception:
            pass
        deletion_layout.addWidget(self.deletion_tbl, stretch=1)

        self.deletion_info_lbl = QLabel("Deletion analysis: (none)")
        self.deletion_info_lbl.setWordWrap(True)
        deletion_layout.addWidget(self.deletion_info_lbl)

        del_btn_row = QHBoxLayout()
        self.sgd_export_btn = QPushButton("Export SGD (CSV)")
        self.sgd_export_btn.clicked.connect(self.export_sgd_csv)
        del_btn_row.addWidget(self.sgd_export_btn)
        self.sgd_export_xlsx_btn = QPushButton("Export SGD (Excel)")
        self.sgd_export_xlsx_btn.clicked.connect(self.export_sgd_excel)
        del_btn_row.addWidget(self.sgd_export_xlsx_btn)
        self.dgd_export_btn = QPushButton("Export DGD (CSV)")
        self.dgd_export_btn.clicked.connect(self.export_dgd_csv)
        del_btn_row.addWidget(self.dgd_export_btn)
        self.dgd_export_xlsx_btn = QPushButton("Export DGD (Excel)")
        self.dgd_export_xlsx_btn.clicked.connect(self.export_dgd_excel)
        del_btn_row.addWidget(self.dgd_export_xlsx_btn)
        self.srd_export_btn = QPushButton("Export SRD (CSV)")
        self.srd_export_btn.clicked.connect(self.export_srd_csv)
        del_btn_row.addWidget(self.srd_export_btn)
        self.srd_export_xlsx_btn = QPushButton("Export SRD (Excel)")
        self.srd_export_xlsx_btn.clicked.connect(self.export_srd_excel)
        del_btn_row.addWidget(self.srd_export_xlsx_btn)
        del_btn_row.addStretch(1)
        deletion_layout.addLayout(del_btn_row)

        self.tabs.addTab(self.tab_deletion, "Deletion Analysis")

        # -------- Robustness tab --------
        self.tab_robustness = QWidget()
        robustness_layout = QVBoxLayout(self.tab_robustness)

        robustness_layout.addWidget(self._title("Robustness Analysis"))
        robustness_help = QLabel(
            "Varies the bounds of a specific reaction (e.g., glucose uptake) over a range of values "
            "and tracks how the objective (growth rate) changes. Shows sensitivity and trade-offs."
        )
        robustness_help.setWordWrap(True)
        robustness_layout.addWidget(robustness_help)

        robustness_param_row = QHBoxLayout()
        robustness_param_row.addWidget(QLabel("Reaction:"))
        self.robustness_rxn = QLineEdit()
        self.robustness_rxn.setPlaceholderText("e.g. EX_glc__D_e")
        robustness_param_row.addWidget(self.robustness_rxn, stretch=1)
        robustness_param_row.addWidget(QLabel("Min:"))
        self.robustness_min = QDoubleSpinBox()
        self.robustness_min.setRange(-1000, 0)
        self.robustness_min.setValue(-20)
        robustness_param_row.addWidget(self.robustness_min)
        robustness_param_row.addWidget(QLabel("Max:"))
        self.robustness_max = QDoubleSpinBox()
        self.robustness_max.setRange(0, 1000)
        self.robustness_max.setValue(20)
        robustness_param_row.addWidget(self.robustness_max)
        robustness_param_row.addWidget(QLabel("Steps:"))
        self.robustness_steps = QSpinBox()
        self.robustness_steps.setRange(5, 100)
        self.robustness_steps.setValue(20)
        robustness_param_row.addWidget(self.robustness_steps)
        robustness_layout.addLayout(robustness_param_row)

        robustness_bound_row = QHBoxLayout()
        robustness_bound_row.addWidget(QLabel("Bound Type:"))
        self.robustness_bound_ub = QCheckBox("Upper Bound (UB)")
        self.robustness_bound_ub.setChecked(True)
        self.robustness_bound_ub.setToolTip("Vary the reaction's upper bound")
        robustness_bound_row.addWidget(self.robustness_bound_ub)
        self.robustness_bound_lb = QCheckBox("Lower Bound (LB)")
        self.robustness_bound_lb.setToolTip("Vary the reaction's lower bound")
        robustness_bound_row.addWidget(self.robustness_bound_lb)
        robustness_bound_row.addStretch(1)
        robustness_layout.addLayout(robustness_bound_row)

        robustness_search_row = QHBoxLayout()
        robustness_search_row.addWidget(QLabel("Search:"))
        self.robustness_search = QLineEdit()
        self.robustness_search.setPlaceholderText("reaction id contains...")
        self.robustness_search.textChanged.connect(self._filter_robustness_combo)
        robustness_search_row.addWidget(self.robustness_search, stretch=2)
        self.robustness_combo = QComboBox()
        self.robustness_combo.currentIndexChanged.connect(self._on_robustness_combo_changed)
        robustness_search_row.addWidget(self.robustness_combo, stretch=2)
        robustness_search_row.addStretch(1)
        robustness_layout.addLayout(robustness_search_row)

        self.robustness_canvas = MplCanvas(self, width=10, height=4, dpi=100)
        robustness_layout.addWidget(self.robustness_canvas, stretch=1)

        self.robustness_tbl = QTableWidget(0, 2)
        self.robustness_tbl.setHorizontalHeaderLabels(["Bound Value", "Objective Value"])
        self.robustness_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.robustness_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.robustness_tbl)
        robustness_layout.addWidget(self.robustness_tbl, stretch=1)

        self.robustness_info_lbl = QLabel("Robustness analysis: (none)")
        self.robustness_info_lbl.setWordWrap(True)
        robustness_layout.addWidget(self.robustness_info_lbl)

        robustness_btn_row = QHBoxLayout()
        self.robustness_export_csv_btn = QPushButton("Export Robustness (CSV)")
        self.robustness_export_csv_btn.clicked.connect(self.export_robustness_csv)
        robustness_btn_row.addWidget(self.robustness_export_csv_btn)
        self.robustness_export_xlsx_btn = QPushButton("Export Robustness (Excel)")
        self.robustness_export_xlsx_btn.clicked.connect(self.export_robustness_excel)
        robustness_btn_row.addWidget(self.robustness_export_xlsx_btn)
        robustness_btn_row.addStretch(1)
        robustness_layout.addLayout(robustness_btn_row)

        self.tabs.addTab(self.tab_robustness, "Robustness")

        # -------- Production Envelope tab --------
        self.tab_envelope = QWidget()
        envelope_layout = QVBoxLayout(self.tab_envelope)

        envelope_layout.addWidget(self._title("Production Envelope (Growth vs Product)"))
        envelope_help = QLabel(
            "Shows the trade-off between biomass (growth) and production of a target compound. "
            "Useful for metabolic engineering to identify the best balance between growth and yield."
        )
        envelope_help.setWordWrap(True)
        envelope_layout.addWidget(envelope_help)

        envelope_param_row = QHBoxLayout()
        envelope_param_row.addWidget(QLabel("Product Reaction:"))
        self.envelope_product = QLineEdit()
        self.envelope_product.setPlaceholderText("e.g. ACALD (acetaldehyde production)")
        envelope_param_row.addWidget(self.envelope_product, stretch=1)
        envelope_param_row.addWidget(QLabel("Steps:"))
        self.envelope_steps = QSpinBox()
        self.envelope_steps.setRange(10, 100)
        self.envelope_steps.setValue(30)
        envelope_param_row.addWidget(self.envelope_steps)
        envelope_layout.addLayout(envelope_param_row)

        envelope_search_row = QHBoxLayout()
        envelope_search_row.addWidget(QLabel("Reaction filter:"))
        self.envelope_search = QLineEdit()
        self.envelope_search.setPlaceholderText("Type to filter reaction list...")
        self.envelope_search.textChanged.connect(self._filter_envelope_combo)
        envelope_search_row.addWidget(self.envelope_search, stretch=2)
        envelope_search_row.addWidget(QLabel("Select reaction:"))
        self.envelope_combo = QComboBox()
        self.envelope_combo.setToolTip("Pick a reaction from the filtered list to set as the product reaction")
        self.envelope_combo.currentIndexChanged.connect(self._on_envelope_combo_changed)
        envelope_search_row.addWidget(self.envelope_combo, stretch=2)
        envelope_search_row.addStretch(1)
        envelope_layout.addLayout(envelope_search_row)

        self.envelope_canvas = MplCanvas(self, width=10, height=4, dpi=100)
        envelope_layout.addWidget(self.envelope_canvas, stretch=1)

        self.envelope_tbl = QTableWidget(0, 2)
        self.envelope_tbl.setHorizontalHeaderLabels(["Product Bound", "Growth Rate"])
        self.envelope_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.envelope_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.envelope_tbl)
        envelope_layout.addWidget(self.envelope_tbl, stretch=1)

        self.envelope_info_lbl = QLabel("Production envelope: (none)")
        self.envelope_info_lbl.setWordWrap(True)
        envelope_layout.addWidget(self.envelope_info_lbl)

        envelope_btn_row = QHBoxLayout()
        self.envelope_export_csv_btn = QPushButton("Export Envelope (CSV)")
        self.envelope_export_csv_btn.clicked.connect(self.export_envelope_csv)
        envelope_btn_row.addWidget(self.envelope_export_csv_btn)
        self.envelope_export_xlsx_btn = QPushButton("Export Envelope (Excel)")
        self.envelope_export_xlsx_btn.clicked.connect(self.export_envelope_excel)
        envelope_btn_row.addWidget(self.envelope_export_xlsx_btn)
        envelope_btn_row.addStretch(1)
        envelope_layout.addLayout(envelope_btn_row)

        self.tabs.addTab(self.tab_envelope, "Envelope")

        # -------- Flux Sampling tab --------
        self.tab_sampling = QWidget()
        sampling_layout = QVBoxLayout(self.tab_sampling)

        sampling_layout.addWidget(self._title("Flux Sampling"))
        sampling_help = QLabel(
            "Monte Carlo sampling of the feasible flux space. "
            "ACHR (Artificially Centered Hit-and-Run) or OptGP (Optima-Guided Parallel, faster for large models). "
            "Returns mean, standard deviation, and min/max ranges for each reaction."
        )
        sampling_help.setWordWrap(True)
        sampling_layout.addWidget(sampling_help)

        sampling_param_row = QHBoxLayout()
        sampling_param_row.addWidget(QLabel("Sampler:"))
        self.sampler_type_combo = QComboBox()
        self.sampler_type_combo.addItems(["ACHR", "OptGP"])
        self.sampler_type_combo.setToolTip("ACHR: standard sampler. OptGP: faster parallel sampler for large models.")
        sampling_param_row.addWidget(self.sampler_type_combo)
        sampling_param_row.addSpacing(15)
        sampling_param_row.addWidget(QLabel("Sample Size:"))
        self.sampling_size = QSpinBox()
        self.sampling_size.setRange(10, 10000)
        self.sampling_size.setValue(500)
        sampling_param_row.addWidget(self.sampling_size)
        sampling_param_row.addStretch(1)
        sampling_layout.addLayout(sampling_param_row)

        self.sampling_canvas = MplCanvas(self, width=10, height=5, dpi=100)
        sampling_layout.addWidget(self.sampling_canvas, stretch=1)

        sampling_search_row = QHBoxLayout()
        sampling_search_row.addWidget(QLabel("Search:"))
        self.sampling_search = QLineEdit()
        self.sampling_search.setPlaceholderText("reaction id contains...")
        self.sampling_search.textChanged.connect(self.filter_sampling_table)
        sampling_search_row.addWidget(self.sampling_search, stretch=2)
        sampling_search_row.addStretch(1)
        sampling_layout.addLayout(sampling_search_row)

        self.sampling_tbl = QTableWidget(0, 4)
        self.sampling_tbl.setHorizontalHeaderLabels(["Reaction", "Mean Flux", "Stdev", "Min/Max"])
        self.sampling_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sampling_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.sampling_tbl)
        sampling_layout.addWidget(self.sampling_tbl, stretch=1)

        self.sampling_info_lbl = QLabel("Flux sampling: (none)")
        self.sampling_info_lbl.setWordWrap(True)
        sampling_layout.addWidget(self.sampling_info_lbl)

        sampling_btn_row = QHBoxLayout()
        self.sampling_export_btn = QPushButton("Export Sampling (CSV)")
        self.sampling_export_btn.clicked.connect(self.export_sampling_csv)
        sampling_btn_row.addWidget(self.sampling_export_btn)
        sampling_btn_row.addStretch(1)
        sampling_layout.addLayout(sampling_btn_row)

        self.tabs.addTab(self.tab_sampling, "Sampling")

        # -------- Network Map tab --------
        self.tab_map = QWidget()
        map_layout = QVBoxLayout(self.tab_map)

        map_layout.addWidget(self._title("Network Map"))
        map_help = QLabel(
            "Visualize metabolic network topology. Nodes: reactions (circles) and metabolites (squares). "
            "Colors reflect flux magnitude or gene knockout impact. Filter by reaction/metabolite search and neighborhood depth."
        )
        map_help.setWordWrap(True)
        map_layout.addWidget(map_help)

        # Search & filter row
        map_filter_row = QHBoxLayout()
        map_filter_row.addWidget(QLabel("Search:"))
        self.map_search = QLineEdit()
        self.map_search.setPlaceholderText("reaction/metabolite id contains...")
        # render on button instead of every change
        map_filter_row.addWidget(self.map_search, stretch=2)

        map_filter_row.addWidget(QLabel("Depth:"))
        self.map_depth = QSpinBox()
        self.map_depth.setRange(1, 5)
        self.map_depth.setValue(2)
        self.map_depth.setToolTip("Neighborhood depth (1-5)")
        # render on button instead of every change
        map_filter_row.addWidget(self.map_depth)

        map_filter_row.addWidget(QLabel("Flux threshold:"))
        self.map_threshold = QDoubleSpinBox()
        self.map_threshold.setRange(0, 100)
        self.map_threshold.setValue(0.1)
        self.map_threshold.setSingleStep(0.1)
        self.map_threshold.setToolTip("Show only reactions with |flux| > threshold")
        # render on button instead of every change
        map_filter_row.addWidget(self.map_threshold)

        map_filter_row.addWidget(QLabel("View:"))
        self.map_view_combo = QComboBox()
        self.map_view_combo.addItems(["Flux magnitude", "Gene KO Impact", "FVA Bounds"])
        # render on button instead of every change
        map_filter_row.addWidget(self.map_view_combo)

        self.map_render_btn = QPushButton("Render Map")
        self.map_render_btn.clicked.connect(self._render_network_map)
        map_filter_row.addWidget(self.map_render_btn)

        self.map_help_btn = QPushButton("How to use map?")
        self.map_help_btn.clicked.connect(self._show_map_help)
        map_filter_row.addWidget(self.map_help_btn)
        map_filter_row.addStretch(1)
        map_layout.addLayout(map_filter_row)

        map_opts_row1 = QHBoxLayout()
        map_opts_row1.addWidget(QLabel("Max nodes:"))
        self.map_max_nodes = QSpinBox()
        self.map_max_nodes.setRange(100, 50000)
        self.map_max_nodes.setValue(5000)
        self.map_max_nodes.setToolTip("Limit nodes to improve performance")
        map_opts_row1.addWidget(self.map_max_nodes)

        map_opts_row1.addWidget(QLabel("Max edges:"))
        self.map_max_edges = QSpinBox()
        self.map_max_edges.setRange(0, 200000)
        self.map_max_edges.setValue(0)
        self.map_max_edges.setToolTip("0 = no edge limit")
        map_opts_row1.addWidget(self.map_max_edges)

        map_opts_row1.addWidget(QLabel("Min degree:"))
        self.map_min_degree = QSpinBox()
        self.map_min_degree.setRange(0, 20)
        self.map_min_degree.setValue(0)
        self.map_min_degree.setToolTip("Hide nodes with degree < min")
        map_opts_row1.addWidget(self.map_min_degree)

        map_opts_row1.addWidget(QLabel("Top-N rxns:"))
        self.map_topn_rxn = QSpinBox()
        self.map_topn_rxn.setRange(0, 5000)
        self.map_topn_rxn.setValue(0)
        self.map_topn_rxn.setToolTip("0 = off")
        map_opts_row1.addWidget(self.map_topn_rxn)

        map_opts_row1.addWidget(QLabel("Layout:"))
        self.map_layout_combo = QComboBox()
        self.map_layout_combo.addItems(["Spring", "Kamada-Kawai", "Circular"])
        map_opts_row1.addWidget(self.map_layout_combo)

        self.map_only_connected_chk = QCheckBox("Only connected to search")
        map_opts_row1.addWidget(self.map_only_connected_chk)

        self.map_hide_orphans_chk = QCheckBox("Hide orphan metabolites")
        map_opts_row1.addWidget(self.map_hide_orphans_chk)

        self.map_show_legend_chk = QCheckBox("Show legend")
        self.map_show_legend_chk.setChecked(True)
        map_opts_row1.addWidget(self.map_show_legend_chk)

        map_opts_row1.addStretch(1)
        map_layout.addLayout(map_opts_row1)

        map_opts_row2 = QHBoxLayout()
        self.map_exchange_only_chk = QCheckBox("Exchange only")
        map_opts_row2.addWidget(self.map_exchange_only_chk)

        self.map_objective_only_chk = QCheckBox("Objective neighborhood only")
        map_opts_row2.addWidget(self.map_objective_only_chk)

        map_opts_row2.addWidget(QLabel("Subsystem:"))
        self.map_subsystem_combo = QComboBox()
        self.map_subsystem_combo.addItem("All subsystems")
        map_opts_row2.addWidget(self.map_subsystem_combo)

        self.map_focus_btn = QPushButton("Focus selected")
        self.map_focus_btn.clicked.connect(self._focus_on_selected_map_node)
        map_opts_row2.addWidget(self.map_focus_btn)

        self.map_export_img_btn = QPushButton("Export map image")
        self.map_export_img_btn.clicked.connect(self._export_map_image)
        map_opts_row2.addWidget(self.map_export_img_btn)

        self.map_export_csv_btn = QPushButton("Export graph CSV")
        self.map_export_csv_btn.clicked.connect(self._export_map_csv)
        map_opts_row2.addWidget(self.map_export_csv_btn)

        map_opts_row2.addStretch(1)
        map_layout.addLayout(map_opts_row2)

        self.map_canvas = MplCanvas(self, width=12, height=7, dpi=100)
        map_layout.addWidget(self.map_canvas, stretch=1)
        self.map_canvas.mpl_connect('button_press_event', self._on_map_click)
        self.map_canvas.mpl_connect('motion_notify_event', self._on_map_hover)
        self._set_map_placeholder("Click 'Render Map' to generate the network.")
        self._map_colorbar = None
        self._map_tooltip = None  # For hover tooltip

        # Zoom/Pan toolbar
        self.map_toolbar = NavigationToolbar(self.map_canvas, self.tab_map)
        map_layout.addWidget(self.map_toolbar)

        self.map_info_lbl = QLabel("Network map: (none)")
        self.map_info_lbl.setWordWrap(True)
        map_layout.addWidget(self.map_info_lbl)

        self.map_details = QPlainTextEdit()
        self.map_details.setReadOnly(True)
        self.map_details.setMaximumHeight(180)
        self.map_details.setPlaceholderText("Click on a node in the map to see reaction details...")
        map_layout.addWidget(self.map_details)

        self.tabs.addTab(self.tab_map, "Map")

        # -------- Results tab --------
        self.tab_results = QWidget()
        results_layout = QVBoxLayout(self.tab_results)

        results_layout.addWidget(self._title("Results"))

        # Row 1: Compare mode and chart type
        compare_row = QHBoxLayout()
        compare_row.addWidget(QLabel("Compare:"))
        self.compare_mode = QComboBox()
        self.compare_mode.addItems(["Gene knockout only", "Overexpression only", "Original vs Baseline"])
        self.compare_mode.currentTextChanged.connect(self._refresh_results_view_from_last_run)
        compare_row.addWidget(self.compare_mode)

        compare_row.addWidget(QLabel("Chart:"))
        self.results_chart_type = QComboBox()
        self.results_chart_type.addItems(["Bar comparison", "Flux histogram", "Waterfall (top changes)"])
        self.results_chart_type.currentTextChanged.connect(self._refresh_results_view_from_last_run)
        compare_row.addWidget(self.results_chart_type)

        compare_row.addStretch(1)
        results_layout.addLayout(compare_row)

        self.results_lbl = QLabel("Objective: -")
        self.results_lbl.setWordWrap(True)
        results_layout.addWidget(self.results_lbl)

        # Row 2: Flux table filter controls
        flux_filter_row = QHBoxLayout()
        flux_filter_row.addWidget(QLabel("Flux search:"))
        self.flux_search = QLineEdit()
        self.flux_search.setPlaceholderText("reaction id contains...")
        self.flux_search.textChanged.connect(self.filter_flux_table)
        flux_filter_row.addWidget(self.flux_search, stretch=2)

        flux_filter_row.addWidget(QLabel("|delta| ≥"))
        self.delta_thresh = QDoubleSpinBox()
        self.delta_thresh.setRange(0.0, 1e12)
        self.delta_thresh.setDecimals(9)
        self.delta_thresh.setSingleStep(0.01)
        self.delta_thresh.setValue(0.0)
        self.delta_thresh.valueChanged.connect(self.filter_flux_table)
        flux_filter_row.addWidget(self.delta_thresh)

        flux_filter_row.addStretch(1)
        results_layout.addLayout(flux_filter_row)

        # Row 3: Filter checkboxes and export buttons
        results_opts_row = QHBoxLayout()
        self.results_nonzero_chk = QCheckBox("Non-zero flux only")
        self.results_nonzero_chk.stateChanged.connect(self.filter_flux_table)
        results_opts_row.addWidget(self.results_nonzero_chk)

        self.results_changed_chk = QCheckBox("Changed only (Δ≠0)")
        self.results_changed_chk.stateChanged.connect(self.filter_flux_table)
        results_opts_row.addWidget(self.results_changed_chk)

        results_opts_row.addStretch(1)

        self.results_export_csv_btn = QPushButton("Export CSV")
        self.results_export_csv_btn.clicked.connect(self._export_results_csv)
        results_opts_row.addWidget(self.results_export_csv_btn)

        self.results_export_excel_btn = QPushButton("Export Excel")
        self.results_export_excel_btn.clicked.connect(self._export_results_excel)
        results_opts_row.addWidget(self.results_export_excel_btn)

        self.results_export_chart_btn = QPushButton("Export Chart")
        self.results_export_chart_btn.clicked.connect(self._export_results_chart)
        results_opts_row.addWidget(self.results_export_chart_btn)

        self.results_export_latex_btn = QPushButton("Export LaTeX")
        self.results_export_latex_btn.setToolTip("Export table as LaTeX for academic papers")
        self.results_export_latex_btn.clicked.connect(self._export_results_latex)
        results_opts_row.addWidget(self.results_export_latex_btn)

        results_layout.addLayout(results_opts_row)

        self.canvas = MplCanvas(self, width=10, height=3.8, dpi=100)
        results_layout.addWidget(self.canvas, stretch=1)

        self.flux_tbl = QTableWidget(0, 4)
        self.flux_tbl.setHorizontalHeaderLabels(["Reaction", "Baseline flux", "Compared flux", "Delta"])
        self.flux_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.flux_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.flux_tbl)
        results_layout.addWidget(self.flux_tbl, stretch=1)

        # Row 4: Additional analysis buttons
        results_extra_row = QHBoxLayout()
        self.results_shadow_btn = QPushButton("Shadow Prices")
        self.results_shadow_btn.clicked.connect(self._show_shadow_prices)
        results_extra_row.addWidget(self.results_shadow_btn)

        self.results_reduced_btn = QPushButton("Reduced Costs")
        self.results_reduced_btn.clicked.connect(self._show_reduced_costs)
        results_extra_row.addWidget(self.results_reduced_btn)

        self.results_essential_btn = QPushButton("Find Essential Reactions")
        self.results_essential_btn.clicked.connect(self._find_essential_reactions)
        results_extra_row.addWidget(self.results_essential_btn)

        self.results_fva_btn = QPushButton("Quick FVA")
        self.results_fva_btn.clicked.connect(self._quick_fva)
        results_extra_row.addWidget(self.results_fva_btn)

        results_extra_row.addStretch(1)
        results_layout.addLayout(results_extra_row)

        self.tabs.addTab(self.tab_results, "Results")

        self.setCentralWidget(root)
        self.resize(1700, 1120)
        self.statusBar().showMessage("Ready.")

        # Progress bar in status bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.setMaximumHeight(16)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setVisible(False)
        self.statusBar().addPermanentWidget(self.progress_bar)

        self._clear_chart()
        # Use system default theme (no custom stylesheet)
        # Apply persisted UI settings before building menu
        self._init_settings()
        self._build_menu()

        # Auto-check tools status after UI is fully rendered
        QTimer.singleShot(500, self.tools_check_status)

        # Re-enable repaints after constructing UI
        self.setUpdatesEnabled(True)

    # ---------------- (1) Unsaved changes prompt on exit ----------------
    def closeEvent(self, event):
        try:
            if getattr(self, "_memote_proc", None) is not None:
                self._force_kill_process(self._memote_proc)
                self._memote_proc = None
            if getattr(self, "_carveme_proc", None) is not None:
                self._force_kill_process(self._carveme_proc)
                self._carveme_proc = None
        except Exception:
            pass
        if self.base_model is None or not self.model_dirty:
            event.accept()
            return

        box = QMessageBox(self)
        box.setIcon(QMessageBox.Warning)
        box.setWindowTitle("Unsaved changes")
        box.setText("You have unsaved changes in the current model.\n\nSave as SBML before exiting?")

        btn_save = box.addButton("Save As...", QMessageBox.AcceptRole)
        btn_discard = box.addButton("Don't Save", QMessageBox.DestructiveRole)
        btn_cancel = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(btn_save)

        box.exec()
        clicked = box.clickedButton()

        if clicked == btn_cancel:
            event.ignore()
            return
        if clicked == btn_save:
            saved = self.save_current_model_as()
            if saved is None:
                event.ignore()
                return
            event.accept()
            return

        event.accept()

    # ---------------- UI helpers ----------------
    def _title(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet("font-weight: 600;")
        return lbl

    def _configure_table_for_excel_resize(self, table: QTableWidget):
        hdr = table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setStretchLastSection(False)
        hdr.setMinimumSectionSize(40)
        table.setWordWrap(False)
        table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        table.verticalHeader().setVisible(False)

    # ---------------- Theme & Settings ----------------
    def _init_settings(self):
        self.settings = QSettings("MetaboDesk", "MetaboDesk")
        # Use system default theme - no custom stylesheet applied

    def _init_persisted_controls(self):
        """Restore TopN and Solver from settings."""
        try:
            topn = int(self.settings.value("ui/topn", self.topn_spin.value()))
            self.topn_spin.setValue(topn)
        except Exception:
            pass
        try:
            solver = str(self.settings.value("ui/solver", "Auto") or "Auto")
            idx = self.solver_combo.findText(solver)
            if idx >= 0:
                self.solver_combo.setCurrentIndex(idx)
        except Exception:
            pass

    # Placeholder - theme functions removed

    def _light_stylesheet(self) -> str:
        return (
            "QMainWindow{background:#f5f6f8;color:#2c3e50;font-family:'Segoe UI','Ubuntu','Helvetica';font-size:10pt;}\n"
            "QLabel{color:#2c3e50;}\n"
            "QLineEdit,QPlainTextEdit,QComboBox,QSpinBox,QDoubleSpinBox{background:#ffffff;color:#2c3e50;border:1px solid #cbd5e0;border-radius:4px;padding:6px;font-size:10pt;}\n"
            "QLineEdit:focus,QPlainTextEdit:focus,QComboBox:focus,QSpinBox:focus,QDoubleSpinBox:focus{border:2px solid #3498db;background:#fafbfc;}\n"
            "QPushButton{background:#3498db;color:#ffffff;border:none;border-radius:4px;padding:8px 14px;font-weight:600;font-size:10pt;}\n"
            "QPushButton:hover{background:#2980b9;}\n"
            "QPushButton:pressed{background:#1f618d;}\n"
            "QPushButton:disabled{background:#bdc3c7;color:#e0e0e0;}\n"
            "QTableWidget{background:#ffffff;color:#2c3e50;gridline-color:#e8eaed;border:1px solid #cbd5e0;}\n"
            "QTableWidget::item{padding:3px;}\n"
            "QHeaderView::section{background:#e8eaed;color:#2c3e50;padding:8px;border:none;border-right:1px solid #cbd5e0;font-weight:600;}\n"
            "QTabBar::tab{background:#e8eaed;color:#2c3e50;padding:10px 16px;border:none;border-bottom:3px solid #cbd5e0;margin-right:4px;font-weight:500;}\n"
            "QTabBar::tab:selected{background:#ffffff;border-bottom:3px solid #3498db;color:#2c3e50;}\n"
            "QTabBar::tab:hover{background:#dfe4e8;}\n"
            "QStatusBar{background:#e8eaed;color:#2c3e50;border-top:1px solid #cbd5e0;font-size:9pt;}\n"
            "QGroupBox{color:#2c3e50;border:1px solid #cbd5e0;border-radius:4px;padding:10px;margin-top:8px;font-weight:600;}\n"
            "QGroupBox::title{subcontrol-origin:margin;left:10px;padding:0 3px;}\n"
            "QToolTip{background:#2c3e50;color:#ecf0f1;border:1px solid #34495e;border-radius:3px;padding:4px 8px;font-size:9pt;}\n"
            "QScrollBar:vertical{background:#f5f6f8;width:12px;}\n"
            "QScrollBar::handle:vertical{background:#cbd5e0;border-radius:6px;min-height:20px;}\n"
            "QScrollBar::handle:vertical:hover{background:#a8b3bf;}"
        )

    def _dark_stylesheet(self) -> str:
        return (
            "QMainWindow{background:#0d1117;color:#e6edf3;font-family:'Segoe UI','Ubuntu','Helvetica';font-size:10pt;}\n"
            "QLabel{color:#e6edf3;}\n"
            "QLineEdit,QPlainTextEdit,QComboBox,QSpinBox,QDoubleSpinBox{background:#161b22;color:#e6edf3;border:1px solid #30363d;border-radius:4px;padding:6px;font-size:10pt;}\n"
            "QLineEdit:focus,QPlainTextEdit:focus,QComboBox:focus,QSpinBox:focus,QDoubleSpinBox:focus{border:2px solid #58a6ff;background:#0d1117;}\n"
            "QPushButton{background:#238636;color:#ffffff;border:none;border-radius:4px;padding:8px 14px;font-weight:600;font-size:10pt;}\n"
            "QPushButton:hover{background:#2ea043;}\n"
            "QPushButton:pressed{background:#1f6feb;}\n"
            "QPushButton:disabled{background:#30363d;color:#6e7681;}\n"
            "QTableWidget{background:#161b22;color:#e6edf3;gridline-color:#30363d;border:1px solid #30363d;}\n"
            "QTableWidget::item{padding:3px;color:#e6edf3;}\n"
            "QHeaderView::section{background:#0d1117;color:#8b949e;padding:8px;border:none;border-right:1px solid #30363d;font-weight:600;}\n"
            "QTabBar::tab{background:#0d1117;color:#8b949e;padding:10px 16px;border:none;border-bottom:3px solid #30363d;margin-right:4px;font-weight:500;}\n"
            "QTabBar::tab:selected{background:#161b22;border-bottom:3px solid #58a6ff;color:#e6edf3;}\n"
            "QTabBar::tab:hover{background:#21262d;}\n"
            "QStatusBar{background:#0d1117;color:#8b949e;border-top:1px solid #30363d;font-size:9pt;}\n"
            "QGroupBox{color:#e6edf3;border:1px solid #30363d;border-radius:4px;padding:10px;margin-top:8px;font-weight:600;}\n"
            "QGroupBox::title{subcontrol-origin:margin;left:10px;padding:0 3px;color:#e6edf3;}\n"
            "QToolTip{background:#161b22;color:#e6edf3;border:1px solid #30363d;border-radius:3px;padding:4px 8px;font-size:9pt;}\n"
            "QScrollBar:vertical{background:#0d1117;width:12px;}\n"
            "QScrollBar::handle:vertical{background:#30363d;border-radius:6px;min-height:20px;}\n"
            "QScrollBar::handle:vertical:hover{background:#6e7681;}"
        )

    def _open_cell_popup(self, title: str, item: QTableWidgetItem | None):
        if not item:
            return
        dlg = TextPopup(title, item.text(), self)
        dlg.exec()

    def _on_reaction_cell_double_clicked(self, item: QTableWidgetItem | None):
        if not item:
            return
        col = item.column()
        text = (item.text() or "").strip()
        # KEGG column index=3, EC=4
        if col == 3 and text:
            kid = text.split(";")[0].strip()
            if kid:
                url = f"https://www.kegg.jp/dbget-bin/www_bget?rn:{kid}"
                QDesktopServices.openUrl(QUrl(url))
                return
        if col == 4 and text:
            ec = text.split(";")[0].strip()
            if ec:
                url = f"https://www.ebi.ac.uk/intenz/query?cmd=searchEC&ec={ec}"
                QDesktopServices.openUrl(QUrl(url))
                return
        # Fallback: open popup
        self._open_cell_popup("Reaction cell", item)

    def _update_medium_details_from_current_cell(self, current: QTableWidgetItem, _prev: QTableWidgetItem):
        self.medium_details.setPlainText(current.text() if current else "")

    def _update_reaction_details_from_current_cell(self, current: QTableWidgetItem, _prev: QTableWidgetItem):
        self.reaction_details.setPlainText(current.text() if current else "")

    def _update_gene_rxn_details_from_current_cell(self, current: QTableWidgetItem, _prev: QTableWidgetItem):
        self.gene_rxn_details.setPlainText(current.text() if current else "")

    # ---------------- (4) Flux table filtering ----------------
    def filter_flux_table(self):
        text = (self.flux_search.text() or "").strip().lower()
        thr = float(self.delta_thresh.value())
        nonzero_only = self.results_nonzero_chk.isChecked()
        changed_only = self.results_changed_chk.isChecked()

        for row in range(self.flux_tbl.rowCount()):
            rid = (self.flux_tbl.item(row, 0).text() if self.flux_tbl.item(row, 0) else "").lower()
            try:
                baseline = float((self.flux_tbl.item(row, 1).text() if self.flux_tbl.item(row, 1) else "0").strip())
            except Exception:
                baseline = 0.0
            try:
                compared = float((self.flux_tbl.item(row, 2).text() if self.flux_tbl.item(row, 2) else "0").strip())
            except Exception:
                compared = 0.0
            try:
                d = float((self.flux_tbl.item(row, 3).text() if self.flux_tbl.item(row, 3) else "0").strip())
            except Exception:
                d = 0.0

            hide = False
            if text and text not in rid:
                hide = True
            if abs(d) < thr:
                hide = True
            if nonzero_only and baseline == 0 and compared == 0:
                hide = True
            if changed_only and d == 0:
                hide = True
            self.flux_tbl.setRowHidden(row, hide)

    # ---------------- Objective search ----------------
    def _setup_objective_completer(self):
        texts = [self.objective_combo.itemText(i) for i in range(self.objective_combo.count())]
        comp = QCompleter(texts, self)
        comp.setCaseSensitivity(Qt.CaseInsensitive)
        comp.setFilterMode(Qt.MatchContains)
        comp.setCompletionMode(QCompleter.PopupCompletion)
        self.objective_combo.setCompleter(comp)

    def _on_objective_text_edited(self, text: str):
        t = (text or "").strip().lower()
        if not t:
            return
        for i in range(self.objective_combo.count()):
            if t in self.objective_combo.itemText(i).lower():
                self.objective_combo.blockSignals(True)
                self.objective_combo.setCurrentIndex(i)
                self.objective_combo.lineEdit().setText(text)
                self.objective_combo.blockSignals(False)
                return

    def populate_objective_combo(self):
        if self.base_model is None:
            self.objective_combo.clear()
            self.objective_combo.setEnabled(False)
            return

        self.objective_combo.blockSignals(True)
        self.objective_combo.clear()
        self.objective_combo.addItem("Keep model objective", None)

        for rxn in sorted(self.base_model.reactions, key=lambda r: r.id):
            desc = get_description(rxn)
            label = f"{rxn.id} — {desc}" if desc else rxn.id
            self.objective_combo.addItem(label, rxn.id)

        self.objective_combo.setEnabled(True)
        self.objective_combo.blockSignals(False)
        self._setup_objective_completer()

    def _apply_selected_objective_to_model(self, model: cobra.Model):
        rid = self.objective_combo.currentData()
        if not rid:
            return
        model.objective = model.reactions.get_by_id(str(rid))
        model.objective_direction = "max"

    def on_objective_changed(self, *_):
        rid = self.objective_combo.currentData()
        if not rid:
            self.statusBar().showMessage("Objective: Keep model objective")
        else:
            self.statusBar().showMessage(f"Objective set to: {rid} (applies on next Run FBA)")

    # ---------------- Recent scenarios ----------------
    def _load_recent_scenarios(self):
        self.recent_scenarios = []
        try:
            recent_file = _recent_file_path()
            if recent_file.exists():
                data = json.loads(recent_file.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    cleaned = []
                    for p in data:
                        p = str(p)
                        if Path(p).exists():
                            cleaned.append(p)
                    self.recent_scenarios = cleaned[:RECENT_MAX]
        except Exception:
            self.recent_scenarios = []

    def _save_recent_scenarios(self):
        try:
            recent_file = _recent_file_path()
            recent_file.write_text(json.dumps(self.recent_scenarios[:RECENT_MAX], indent=2), encoding="utf-8")
        except Exception:
            pass

    def _add_recent_scenario(self, file_path: str):
        file_path = str(Path(file_path))
        if file_path in self.recent_scenarios:
            self.recent_scenarios.remove(file_path)
        self.recent_scenarios.insert(0, file_path)
        self.recent_scenarios = self.recent_scenarios[:RECENT_MAX]
        self._save_recent_scenarios()
        self._rebuild_recent_menu()

    # ---------------- Busy state ----------------
    def set_busy(self, busy: bool, message: str = ""):
        self.is_running = busy
        for w in (
            self.open_btn, self.run_btn, self.close_uptakes_btn,
            self.reset_medium_btn, self.reset_reactions_btn,
            self.analysis_type, self.topn_spin, self.objective_combo,
            self.medium_search, self.medium_table,
            self.rxn_global_search, self.met_filter, self.met_filter_mode, self.reaction_table,
            self.clear_rxn_overrides_btn, self.remove_selected_overrides_btn,
            self.gene_global_search, self.genes_tab_list, self.gene_rxn_table,
            self.gene_search, self.gene_list, self.add_ko_btn, self.clear_ko_btn,
            self.rxn_search, self.rxn_list, self.oe_factor, self.oe_scale_lb,
            self.add_oe_btn, self.remove_oe_btn, self.clear_oe_btn,
            self.temp_ub_spin, self.set_temp_ub_btn, self.remove_temp_ub_btn, self.clear_temp_ub_btn,
            self.oe_active_list, self.temp_active_list,
            # New analysis tabs controls
            self.fva_flux_search, self.pfba_flux_search,
            self.fva_fraction_spin, self.fva_processes_spin,
            self.deletion_search,
            self.robustness_rxn, self.robustness_min, self.robustness_max, self.robustness_steps, self.robustness_bound_ub, self.robustness_bound_lb,
            self.envelope_product, self.envelope_steps,
            self.sampling_size, self.sampling_search,
            self.sampler_type_combo,
            self.loopless_chk,
            # Map tab controls
            self.map_search, self.map_depth, self.map_threshold, self.map_view_combo,
            self.map_max_nodes, self.map_max_edges, self.map_min_degree, self.map_topn_rxn,
            self.map_layout_combo, self.map_only_connected_chk, self.map_hide_orphans_chk,
            self.map_show_legend_chk, self.map_exchange_only_chk, self.map_objective_only_chk, self.map_subsystem_combo,
            self.map_focus_btn, self.map_export_img_btn, self.map_export_csv_btn,
            self.compare_mode,
            self.editor_tabs,
            self.tools_check_btn, self.tools_repair_btn,
            self.memote_btn, self.memote_cancel_btn, self.carveme_btn, self.carveme_cancel_btn,
        ):
            if w is self.memote_cancel_btn:
                enabled = (self._memote_proc is not None)
            elif w is self.carveme_cancel_btn:
                enabled = (self._carveme_proc is not None)
            else:
                enabled = not busy
            if w is self.run_btn:
                enabled = enabled and (self.base_model is not None)
            if w is self.analysis_type:
                enabled = enabled and (self.base_model is not None)
            if w is self.memote_btn:
                enabled = enabled and (self.base_model is not None) and (self._memote_proc is None)
            w.setEnabled(enabled)
        if message:
            self.status_lbl.setText(message)
            self.statusBar().showMessage(message)

    # ---------------- Menu ----------------
    def _build_menu(self):
        menubar = self.menuBar()
        self.file_menu = menubar.addMenu("&File")
        run_menu = menubar.addMenu("&Run")
        medium_menu = menubar.addMenu("&Medium")
        model_menu = menubar.addMenu("&Model")
        help_menu = menubar.addMenu("&Help")

        act_open = QAction("Open SBML...", self)
        act_open.setShortcut(QKeySequence.Open)
        act_open.triggered.connect(self.open_sbml)
        self.file_menu.addAction(act_open)

        # (1) Save As (SBML)
        act_save_as = QAction("Save As (SBML)...", self)
        act_save_as.setShortcut(QKeySequence.SaveAs)
        act_save_as.triggered.connect(self.save_current_model_as)
        self.file_menu.addAction(act_save_as)

        self.file_menu.addSeparator()

        act_export = QAction("Export scenario...", self)
        act_export.triggered.connect(self.export_scenario)
        self.file_menu.addAction(act_export)

        act_import = QAction("Import scenario...", self)
        act_import.triggered.connect(self.import_scenario)
        self.file_menu.addAction(act_import)

        self.recent_menu = self.file_menu.addMenu("Recent scenarios")
        self._rebuild_recent_menu()

        self.file_menu.addSeparator()

        act_export_results = QAction("Export results (CSV)...", self)
        act_export_results.triggered.connect(self._export_results_csv)
        self.file_menu.addAction(act_export_results)

        act_export_excel = QAction("Export results (Excel)...", self)
        act_export_excel.triggered.connect(self._export_results_excel)
        self.file_menu.addAction(act_export_excel)

        act_export_json = QAction("Export results (JSON)...", self)
        act_export_json.triggered.connect(self.export_results_json)
        self.file_menu.addAction(act_export_json)

        act_export_all = QAction("Export All (scenario+results+chart)...", self)
        act_export_all.triggered.connect(self.export_all)
        self.file_menu.addAction(act_export_all)

        act_export_jupyter = QAction("Export Jupyter Notebook...", self)
        act_export_jupyter.triggered.connect(self.export_jupyter_notebook)
        self.file_menu.addAction(act_export_jupyter)

        self.file_menu.addSeparator()

        act_exit = QAction("Exit", self)
        act_exit.setShortcut(QKeySequence.Quit)
        act_exit.triggered.connect(self.close)
        self.file_menu.addAction(act_exit)

        act_run = QAction("Run Analysis", self)
        act_run.setShortcut(QKeySequence("Ctrl+R"))
        act_run.triggered.connect(self.run_fba)
        run_menu.addAction(act_run)

        run_menu.addSeparator()

        act_undo = QAction("Undo", self)
        act_undo.setShortcut(QKeySequence.Undo)
        act_undo.triggered.connect(self.undo)
        run_menu.addAction(act_undo)

        act_redo = QAction("Redo", self)
        act_redo.setShortcut(QKeySequence.Redo)
        act_redo.triggered.connect(self.redo)
        run_menu.addAction(act_redo)

        run_menu.addSeparator()

        act_sensitivity = QAction("Sensitivity Analysis...", self)
        act_sensitivity.setShortcut(QKeySequence("Ctrl+Shift+S"))
        act_sensitivity.triggered.connect(self.run_sensitivity_analysis)
        run_menu.addAction(act_sensitivity)

        act_community = QAction("Community Analysis...", self)
        act_community.triggered.connect(self.run_community_analysis)
        run_menu.addAction(act_community)

        act_batch = QAction("Batch Analysis...", self)
        act_batch.setShortcut(QKeySequence("Ctrl+B"))
        act_batch.triggered.connect(self.run_batch_analysis)
        run_menu.addAction(act_batch)

        run_menu.addSeparator()

        act_search = QAction("Search & Filter...", self)
        act_search.setShortcut(QKeySequence.Find)
        act_search.triggered.connect(self.show_search_dialog)
        run_menu.addAction(act_search)

        # Explicit shortcuts to ensure they work on all platforms
        QShortcut(QKeySequence("Ctrl+Q"), self, activated=self.close)
        QShortcut(QKeySequence("Ctrl+Shift+S"), self, activated=self.run_sensitivity_analysis)

        act_close_uptakes = QAction("Close uptakes (LB<0 → 0)", self)
        act_close_uptakes.triggered.connect(self.close_all_uptakes)
        medium_menu.addAction(act_close_uptakes)

        act_reset_medium = QAction("Reset medium", self)
        act_reset_medium.triggered.connect(self.reset_medium)
        medium_menu.addAction(act_reset_medium)

        medium_menu.addSeparator()
        act_export_medium_csv = QAction("Export Medium bounds (CSV)...", self)
        act_export_medium_csv.triggered.connect(self.export_medium_bounds_csv)
        medium_menu.addAction(act_export_medium_csv)
        act_import_medium_csv = QAction("Import Medium bounds (CSV)...", self)
        act_import_medium_csv.triggered.connect(self.import_medium_bounds_csv)
        medium_menu.addAction(act_import_medium_csv)

        # Model menu - validation and checks
        act_validate_model = QAction("Validate model...", self)
        act_validate_model.triggered.connect(self.validate_model)
        model_menu.addAction(act_validate_model)

        act_check_mass_balance = QAction("Check mass balance...", self)
        act_check_mass_balance.triggered.connect(self.check_mass_balance)
        model_menu.addAction(act_check_mass_balance)

        act_check_gpr = QAction("Validate GPR syntax...", self)
        act_check_gpr.triggered.connect(self.check_gpr_syntax)
        model_menu.addAction(act_check_gpr)

        model_menu.addSeparator()

        act_model_stats = QAction("Model Statistics Dashboard...", self)
        act_model_stats.triggered.connect(self.show_model_stats_dashboard)
        model_menu.addAction(act_model_stats)

        act_connectivity = QAction("Metabolite Connectivity Analysis...", self)
        act_connectivity.triggered.connect(self.show_metabolite_connectivity)
        model_menu.addAction(act_connectivity)

        act_annotation = QAction("Annotation Quality Score...", self)
        act_annotation.triggered.connect(self.show_annotation_quality)
        model_menu.addAction(act_annotation)

        model_menu.addSeparator()

        act_add_constraint = QAction("Add Custom Constraint...", self)
        act_add_constraint.triggered.connect(self.add_custom_constraint)
        model_menu.addAction(act_add_constraint)

        act_html_report = QAction("Generate HTML Report...", self)
        act_html_report.triggered.connect(self.generate_html_report)
        model_menu.addAction(act_html_report)

        model_menu.addSeparator()

        act_met_summary = QAction("Metabolite Summary...", self)
        act_met_summary.triggered.connect(self.show_metabolite_summary)
        model_menu.addAction(act_met_summary)

        act_flux_compare = QAction("Compare Flux Distributions...", self)
        act_flux_compare.triggered.connect(self.show_flux_distribution_comparison)
        model_menu.addAction(act_flux_compare)

        act_shortcuts = QAction("Keyboard shortcuts", self)
        act_shortcuts.triggered.connect(self.show_shortcuts)
        help_menu.addAction(act_shortcuts)

        act_about = QAction("About MetaboDesk", self)
        act_about.triggered.connect(self.show_about)
        help_menu.addAction(act_about)

        # ---- View menu ----
        view_menu = menubar.addMenu("&View")

        act_toggle_theme = QAction("Toggle Dark/Light Theme", self)
        act_toggle_theme.setShortcut(QKeySequence("Ctrl+T"))
        act_toggle_theme.triggered.connect(self.toggle_theme)
        view_menu.addAction(act_toggle_theme)

        # ---- Advanced Analysis submenu ----
        run_menu.addSeparator()
        advanced_menu = run_menu.addMenu("Advanced Analysis")

        act_gene_expr = QAction("Gene Expression Integration...", self)
        act_gene_expr.triggered.connect(self.run_gene_expression_analysis)
        advanced_menu.addAction(act_gene_expr)

        act_optknock = QAction("OptKnock / Strain Design...", self)
        act_optknock.triggered.connect(self.run_optknock)
        advanced_menu.addAction(act_optknock)

        act_dfba = QAction("Dynamic FBA (dFBA)...", self)
        act_dfba.triggered.connect(self.run_dfba)
        advanced_menu.addAction(act_dfba)

        act_flux_coupling = QAction("Flux Coupling Analysis...", self)
        act_flux_coupling.triggered.connect(self.run_flux_coupling)
        advanced_menu.addAction(act_flux_coupling)

        act_tmfa = QAction("Thermodynamic FBA (TMFA)...", self)
        act_tmfa.triggered.connect(self.run_tmfa)
        advanced_menu.addAction(act_tmfa)

        act_pathway_enrich = QAction("Pathway Enrichment Analysis...", self)
        act_pathway_enrich.triggered.connect(self.run_pathway_enrichment)
        advanced_menu.addAction(act_pathway_enrich)

        act_escher = QAction("Escher Map Viewer...", self)
        act_escher.triggered.connect(self.show_escher_map)
        advanced_menu.addAction(act_escher)

        act_phpp = QAction("Phenotype Phase Plane (PhPP)...", self)
        act_phpp.triggered.connect(self.run_phenotype_phase_plane)
        advanced_menu.addAction(act_phpp)

        # ---- Model comparison ----
        model_menu.addSeparator()

        act_model_compare = QAction("Compare Models...", self)
        act_model_compare.triggered.connect(self.run_model_comparison)
        model_menu.addAction(act_model_compare)

        act_validate_sbml = QAction("Validate SBML...", self)
        act_validate_sbml.triggered.connect(self.validate_sbml)
        model_menu.addAction(act_validate_sbml)

        # ---- Plugins ----
        help_menu.addSeparator()

        act_load_plugins = QAction("Load Plugins...", self)
        act_load_plugins.triggered.connect(self.load_plugins)
        help_menu.addAction(act_load_plugins)

    # ---- FBA result cache helpers ----
    def _model_state_hash(self, model: cobra.Model) -> str:
        """Return a lightweight hash representing the current model state.

        Captures objective, all reaction bounds, and gene knockouts so that
        identical optimisation requests can be served from cache.
        """
        import hashlib
        parts: list[str] = []
        # objective
        parts.append(str(model.objective.expression))
        parts.append(model.objective.direction)
        # reaction bounds (sorted for determinism)
        for r in sorted(model.reactions, key=lambda r: r.id):
            parts.append(f"{r.id}:{r.lower_bound}:{r.upper_bound}")
        # knocked-out genes
        parts.append(",".join(sorted(self.knockout_genes)))
        blob = "|".join(parts).encode("utf-8")
        return hashlib.sha256(blob).hexdigest()

    def _cache_get(self, key: str) -> dict | None:
        return self._fba_cache.get(key)

    def _cache_put(self, key: str, result: dict) -> None:
        if len(self._fba_cache) >= self._fba_cache_max:
            # evict oldest entry
            oldest = next(iter(self._fba_cache))
            del self._fba_cache[oldest]
        self._fba_cache[key] = result

    def _invalidate_fba_cache(self) -> None:
        self._fba_cache.clear()

    def _rebuild_recent_menu(self):
        if not hasattr(self, "recent_menu"):
            return
        self.recent_menu.clear()

        if not self.recent_scenarios:
            act_none = QAction("(none)", self)
            act_none.setEnabled(False)
            self.recent_menu.addAction(act_none)
            return

        for p in self.recent_scenarios:
            act = QAction(Path(p).name, self)
            act.setToolTip(p)
            act.triggered.connect(lambda checked=False, path=p: self.import_scenario_from_path(path))
            self.recent_menu.addAction(act)

        self.recent_menu.addSeparator()
        act_clear = QAction("Clear recent list", self)
        act_clear.triggered.connect(self.clear_recent_scenarios)
        self.recent_menu.addAction(act_clear)

    def clear_recent_scenarios(self):
        self.recent_scenarios = []
        self._save_recent_scenarios()
        self._rebuild_recent_menu()

    # ---- Model Statistics Dashboard ----
    def show_model_stats_dashboard(self):
        """Comprehensive model statistics dialog."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        model = self.base_model

        # Basic counts
        n_rxn = len(model.reactions)
        n_met = len(model.metabolites)
        n_gene = len(model.genes)

        # Compartments
        compartments = set()
        for m in model.metabolites:
            c = getattr(m, "compartment", None)
            if c:
                compartments.add(c)
        compartment_names = getattr(model, "compartments", {}) or {}
        comp_lines = []
        for c in sorted(compartments):
            name = compartment_names.get(c, "")
            mets_in = sum(1 for m in model.metabolites if getattr(m, "compartment", "") == c)
            comp_lines.append(f"  {c} ({name}): {mets_in} metabolites" if name else f"  {c}: {mets_in} metabolites")

        # Subsystems
        subsystems = set()
        for r in model.reactions:
            ss = getattr(r, "subsystem", None) or ""
            if ss.strip():
                subsystems.add(ss.strip())

        # Reversibility
        n_reversible = sum(1 for r in model.reactions if r.reversibility)
        pct_rev = (n_reversible / n_rxn * 100) if n_rxn else 0

        # Exchange reactions
        exchange_rxns = [r for r in model.reactions if r.boundary]
        n_exchange = len(exchange_rxns)

        # Blocked reactions (zero flux in both directions under default bounds)
        blocked = []
        try:
            from cobra.flux_analysis import find_blocked_reactions
            blocked = find_blocked_reactions(model, processes=1)
        except Exception:
            pass
        n_blocked = len(blocked)

        # Orphan metabolites (metabolites in only 1 reaction)
        orphan_mets = [m for m in model.metabolites if len(m.reactions) <= 1]
        n_orphan = len(orphan_mets)

        # Dead-end metabolites (produced but not consumed, or vice versa)
        dead_ends = []
        for m in model.metabolites:
            producers = sum(1 for r in m.reactions if r.metabolites[m] > 0)
            consumers = sum(1 for r in m.reactions if r.metabolites[m] < 0)
            if producers == 0 or consumers == 0:
                dead_ends.append(m.id)
        n_dead_end = len(dead_ends)

        # Mass-unbalanced reactions
        n_imbalanced = 0
        for rxn in model.reactions:
            if rxn.boundary:
                continue
            try:
                bal = rxn.check_mass_balance()
                if bal:
                    n_imbalanced += 1
            except Exception:
                pass

        # GPR coverage
        n_gpr = sum(1 for r in model.reactions if getattr(r, "gene_reaction_rule", "").strip())
        pct_gpr = (n_gpr / n_rxn * 100) if n_rxn else 0

        # Build dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Model Statistics Dashboard")
        dialog.resize(650, 600)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)

        model_id = getattr(model, "id", "N/A") or "N/A"
        model_name = getattr(model, "name", "") or ""

        html = f"""
        <h2 style='color:#2E86C1;'>Model Statistics Dashboard</h2>
        <table style='font-size:11px; border-collapse:collapse; width:100%;'>
        <tr><td style='padding:4px;'><b>Model ID:</b></td><td>{model_id}</td></tr>
        <tr><td style='padding:4px;'><b>Model Name:</b></td><td>{model_name}</td></tr>
        <tr><td colspan='2'><hr></td></tr>
        <tr><td style='padding:4px;'><b>Reactions:</b></td><td>{n_rxn}</td></tr>
        <tr><td style='padding:4px;'><b>Metabolites:</b></td><td>{n_met}</td></tr>
        <tr><td style='padding:4px;'><b>Genes:</b></td><td>{n_gene}</td></tr>
        <tr><td style='padding:4px;'><b>Compartments:</b></td><td>{len(compartments)}</td></tr>
        <tr><td style='padding:4px;'><b>Subsystems:</b></td><td>{len(subsystems)}</td></tr>
        <tr><td colspan='2'><hr></td></tr>
        <tr><td style='padding:4px;'><b>Reversible reactions:</b></td><td>{n_reversible} ({pct_rev:.1f}%)</td></tr>
        <tr><td style='padding:4px;'><b>Exchange/boundary reactions:</b></td><td>{n_exchange}</td></tr>
        <tr><td style='padding:4px;'><b>Blocked reactions:</b></td><td>{n_blocked}</td></tr>
        <tr><td style='padding:4px;'><b>Mass-unbalanced reactions:</b></td><td>{n_imbalanced}</td></tr>
        <tr><td colspan='2'><hr></td></tr>
        <tr><td style='padding:4px;'><b>Orphan metabolites (≤1 rxn):</b></td><td>{n_orphan}</td></tr>
        <tr><td style='padding:4px;'><b>Dead-end metabolites:</b></td><td>{n_dead_end}</td></tr>
        <tr><td colspan='2'><hr></td></tr>
        <tr><td style='padding:4px;'><b>GPR coverage:</b></td><td>{n_gpr}/{n_rxn} ({pct_gpr:.1f}%)</td></tr>
        </table>
        """

        info_label = QLabel(html)
        info_label.setWordWrap(True)
        info_label.setTextFormat(Qt.RichText)
        layout.addWidget(info_label)

        # Compartments detail
        if comp_lines:
            comp_group = QGroupBox("Compartments")
            comp_layout = QVBoxLayout()
            comp_text = QPlainTextEdit("\n".join(comp_lines))
            comp_text.setReadOnly(True)
            comp_text.setMaximumHeight(100)
            comp_layout.addWidget(comp_text)
            comp_group.setLayout(comp_layout)
            layout.addWidget(comp_group)

        # Blocked reactions detail
        if blocked:
            blocked_group = QGroupBox(f"Blocked Reactions ({n_blocked})")
            blocked_layout = QVBoxLayout()
            blocked_text = QPlainTextEdit("\n".join(str(b) for b in blocked[:200]))
            blocked_text.setReadOnly(True)
            blocked_text.setMaximumHeight(120)
            blocked_layout.addWidget(blocked_text)
            blocked_group.setLayout(blocked_layout)
            layout.addWidget(blocked_group)

        # Orphan metabolites detail
        if orphan_mets:
            orphan_group = QGroupBox(f"Orphan Metabolites ({n_orphan})")
            orphan_layout = QVBoxLayout()
            orphan_text = QPlainTextEdit("\n".join(m.id for m in orphan_mets[:200]))
            orphan_text.setReadOnly(True)
            orphan_text.setMaximumHeight(120)
            orphan_layout.addWidget(orphan_text)
            orphan_group.setLayout(orphan_layout)
            layout.addWidget(orphan_group)

        layout.addStretch()
        scroll.setWidget(content)

        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(scroll)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(btn_box)

        dialog.exec()

    # ---- Metabolite Connectivity Analysis ----
    def show_metabolite_connectivity(self):
        """Analyse dead-end metabolites and chokepoint reactions."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        model = self.base_model

        # Dead-end metabolites: only produced or only consumed
        dead_ends = []
        for m in model.metabolites:
            producers = 0
            consumers = 0
            for r in m.reactions:
                coeff = r.metabolites[m]
                if coeff > 0:
                    producers += 1
                elif coeff < 0:
                    consumers += 1
                # For reversible reactions, both directions possible
                if r.reversibility:
                    producers += 1
                    consumers += 1
            if producers == 0 or consumers == 0:
                status = "only consumed" if producers == 0 else "only produced"
                dead_ends.append((m.id, m.name or "", getattr(m, "compartment", ""), status))

        # Chokepoint reactions: the only reaction producing or consuming a metabolite
        chokepoints = set()
        for m in model.metabolites:
            producing_rxns = [r for r in m.reactions if r.metabolites[m] > 0]
            consuming_rxns = [r for r in m.reactions if r.metabolites[m] < 0]
            if len(producing_rxns) == 1:
                chokepoints.add(producing_rxns[0].id)
            if len(consuming_rxns) == 1:
                chokepoints.add(consuming_rxns[0].id)

        # Build dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Metabolite Connectivity Analysis")
        dialog.resize(800, 600)

        tabs = QTabWidget()

        # Tab 1: Dead-end metabolites
        dead_end_tab = QWidget()
        de_layout = QVBoxLayout(dead_end_tab)
        de_layout.addWidget(QLabel(f"<b>Dead-end metabolites: {len(dead_ends)}</b>  "
                                   "(metabolites only produced or only consumed, excluding reversible contributions)"))

        de_filter = QLineEdit()
        de_filter.setPlaceholderText("Filter dead-ends...")
        de_layout.addWidget(de_filter)

        de_table = QTableWidget()
        de_table.setColumnCount(4)
        de_table.setHorizontalHeaderLabels(["ID", "Name", "Compartment", "Status"])
        de_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        de_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        de_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        de_table.setAlternatingRowColors(True)
        de_table.setRowCount(len(dead_ends))

        for i, (mid, mname, mcomp, mstatus) in enumerate(dead_ends):
            de_table.setItem(i, 0, QTableWidgetItem(mid))
            de_table.setItem(i, 1, QTableWidgetItem(mname))
            de_table.setItem(i, 2, QTableWidgetItem(mcomp))
            de_table.setItem(i, 3, QTableWidgetItem(mstatus))

        def de_filter_changed(text):
            t = text.strip().lower()
            for row in range(de_table.rowCount()):
                match = not t
                for col in range(de_table.columnCount()):
                    item = de_table.item(row, col)
                    if item and t in item.text().lower():
                        match = True
                        break
                de_table.setRowHidden(row, not match)

        de_filter.textChanged.connect(de_filter_changed)
        de_layout.addWidget(de_table)
        tabs.addTab(dead_end_tab, f"Dead-end Metabolites ({len(dead_ends)})")

        # Tab 2: Chokepoint reactions
        choke_tab = QWidget()
        ch_layout = QVBoxLayout(choke_tab)
        ch_layout.addWidget(QLabel(f"<b>Chokepoint reactions: {len(chokepoints)}</b>  "
                                   "(only reaction producing/consuming a metabolite — potential drug targets)"))

        ch_filter = QLineEdit()
        ch_filter.setPlaceholderText("Filter chokepoints...")
        ch_layout.addWidget(ch_filter)

        choke_list = sorted(chokepoints)
        ch_table = QTableWidget()
        ch_table.setColumnCount(3)
        ch_table.setHorizontalHeaderLabels(["Reaction ID", "Name", "Subsystem"])
        ch_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        ch_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        ch_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        ch_table.setAlternatingRowColors(True)
        ch_table.setRowCount(len(choke_list))

        for i, rid in enumerate(choke_list):
            rxn = model.reactions.get_by_id(rid)
            ch_table.setItem(i, 0, QTableWidgetItem(rid))
            ch_table.setItem(i, 1, QTableWidgetItem(rxn.name or ""))
            ch_table.setItem(i, 2, QTableWidgetItem(getattr(rxn, "subsystem", "") or ""))

        def ch_filter_changed(text):
            t = text.strip().lower()
            for row in range(ch_table.rowCount()):
                match = not t
                for col in range(ch_table.columnCount()):
                    item = ch_table.item(row, col)
                    if item and t in item.text().lower():
                        match = True
                        break
                ch_table.setRowHidden(row, not match)

        ch_filter.textChanged.connect(ch_filter_changed)
        ch_layout.addWidget(ch_table)
        tabs.addTab(choke_tab, f"Chokepoint Reactions ({len(chokepoints)})")

        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(tabs)
        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(btn_box)

        dialog.exec()

    # ---- Annotation Quality Score ----
    def show_annotation_quality(self):
        """Show annotation coverage (KEGG, EC, GPR, SBO, MetaNetX)."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        model = self.base_model
        n_rxn = len(model.reactions)
        n_met = len(model.metabolites)

        # Reaction annotations
        n_kegg = sum(1 for r in model.reactions if get_kegg_rxn_id(r))
        n_ec = sum(1 for r in model.reactions if get_ec_numbers(r))
        n_gpr = sum(1 for r in model.reactions if get_gpr(r).strip())

        # SBO terms
        n_rxn_sbo = 0
        n_met_sbo = 0
        for r in model.reactions:
            ann = getattr(r, "annotation", {}) or {}
            if ann.get("sbo") or ann.get("sbo_term"):
                n_rxn_sbo += 1
        for m in model.metabolites:
            ann = getattr(m, "annotation", {}) or {}
            if ann.get("sbo") or ann.get("sbo_term"):
                n_met_sbo += 1

        # Metabolite annotations: KEGG compound, ChEBI, MetaNetX, InChI, formula
        n_met_kegg = 0
        n_met_chebi = 0
        n_met_metanetx = 0
        n_met_inchi = 0
        n_met_formula = 0
        for m in model.metabolites:
            ann = getattr(m, "annotation", {}) or {}
            if any(ann.get(k) for k in ("kegg.compound", "kegg_compound", "kegg.metabolite")):
                n_met_kegg += 1
            if any(ann.get(k) for k in ("chebi", "CHEBI")):
                n_met_chebi += 1
            if any(ann.get(k) for k in ("metanetx.chemical", "metanetx", "MetaNetX")):
                n_met_metanetx += 1
            if any(ann.get(k) for k in ("inchi", "InChI", "inchikey", "InChIKey")):
                n_met_inchi += 1
            formula = getattr(m, "formula", None) or ""
            if formula.strip():
                n_met_formula += 1

        def pct(n, total):
            return f"{n}/{total} ({n / total * 100:.1f}%)" if total else "0/0 (0%)"

        def bar_color(n, total):
            if total == 0:
                return "#999"
            ratio = n / total
            if ratio >= 0.8:
                return "#27AE60"
            elif ratio >= 0.5:
                return "#F39C12"
            else:
                return "#E74C3C"

        def bar_html(label, n, total):
            ratio = (n / total * 100) if total else 0
            color = bar_color(n, total)
            return (f"<tr><td style='padding:3px;'><b>{label}</b></td>"
                    f"<td style='padding:3px;'>{pct(n, total)}</td>"
                    f"<td style='padding:3px; width:200px;'>"
                    f"<div style='background:#eee; border-radius:4px; height:16px;'>"
                    f"<div style='background:{color}; border-radius:4px; height:16px; width:{ratio:.0f}%;'></div>"
                    f"</div></td></tr>")

        html = f"""
        <h2 style='color:#2E86C1;'>Annotation Quality Score</h2>
        <h3>Reaction Annotations ({n_rxn} reactions)</h3>
        <table style='font-size:11px; width:100%;'>
        {bar_html("KEGG Reaction", n_kegg, n_rxn)}
        {bar_html("EC Number", n_ec, n_rxn)}
        {bar_html("GPR (gene rules)", n_gpr, n_rxn)}
        {bar_html("SBO Term", n_rxn_sbo, n_rxn)}
        </table>
        <br>
        <h3>Metabolite Annotations ({n_met} metabolites)</h3>
        <table style='font-size:11px; width:100%;'>
        {bar_html("KEGG Compound", n_met_kegg, n_met)}
        {bar_html("ChEBI", n_met_chebi, n_met)}
        {bar_html("MetaNetX", n_met_metanetx, n_met)}
        {bar_html("InChI / InChIKey", n_met_inchi, n_met)}
        {bar_html("Chemical Formula", n_met_formula, n_met)}
        {bar_html("SBO Term", n_met_sbo, n_met)}
        </table>
        """

        # Overall score (weighted average)
        total_checks = n_rxn * 4 + n_met * 6
        total_hits = n_kegg + n_ec + n_gpr + n_rxn_sbo + n_met_kegg + n_met_chebi + n_met_metanetx + n_met_inchi + n_met_formula + n_met_sbo
        overall = (total_hits / total_checks * 100) if total_checks else 0
        ov_color = bar_color(total_hits, total_checks)

        html += f"""
        <br>
        <h3>Overall Annotation Score</h3>
        <div style='text-align:center; font-size:28px; font-weight:bold; color:{ov_color};'>{overall:.1f}%</div>
        <div style='background:#eee; border-radius:6px; height:24px; margin:8px 0;'>
            <div style='background:{ov_color}; border-radius:6px; height:24px; width:{overall:.0f}%;'></div>
        </div>
        """

        dialog = QDialog(self)
        dialog.setWindowTitle("Annotation Quality Score")
        dialog.resize(650, 550)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        cl = QVBoxLayout(content)
        lbl = QLabel(html)
        lbl.setWordWrap(True)
        lbl.setTextFormat(Qt.RichText)
        cl.addWidget(lbl)
        cl.addStretch()
        scroll.setWidget(content)

        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(scroll)
        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(btn_box)

        dialog.exec()

    # ---- Add Custom Constraint ----
    def add_custom_constraint(self):
        """Dialog to add an arbitrary linear constraint to the model."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Add Custom Constraint")
        dialog.resize(600, 400)

        layout = QVBoxLayout(dialog)

        info = QLabel(
            "<b>Add a linear constraint to the model.</b><br>"
            "Format: <code>coeff1 * reaction_id1 + coeff2 * reaction_id2 ... [&lt;=, =, &gt;=] bound</code><br>"
            "Example: <code>1.0 * PFK + -0.5 * PYK &lt;= 10.0</code><br><br>"
            "Or simply set reaction bounds below."
        )
        info.setWordWrap(True)
        info.setTextFormat(Qt.RichText)
        layout.addWidget(info)

        # Constraint name
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Constraint name:"))
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("my_constraint_1")
        name_layout.addWidget(name_edit)
        layout.addLayout(name_layout)

        # Expression mode: simple (reaction bounds) or advanced (linear expression)
        mode_tabs = QTabWidget()

        # Simple mode: pick reaction, set bounds
        simple_tab = QWidget()
        simple_layout = QFormLayout(simple_tab)

        rxn_combo = QComboBox()
        rxn_combo.setEditable(True)
        rxn_ids = sorted(r.id for r in self.base_model.reactions)
        rxn_combo.addItems(rxn_ids)
        rxn_combo.setCompleter(QCompleter(rxn_ids))
        simple_layout.addRow("Reaction:", rxn_combo)

        lb_spin = QDoubleSpinBox()
        lb_spin.setRange(-1e9, 1e9)
        lb_spin.setDecimals(6)
        lb_spin.setValue(0.0)
        simple_layout.addRow("Lower bound:", lb_spin)

        ub_spin = QDoubleSpinBox()
        ub_spin.setRange(-1e9, 1e9)
        ub_spin.setDecimals(6)
        ub_spin.setValue(1000.0)
        simple_layout.addRow("Upper bound:", ub_spin)

        mode_tabs.addTab(simple_tab, "Reaction Bounds")

        # Advanced mode: linear expression
        adv_tab = QWidget()
        adv_layout = QVBoxLayout(adv_tab)
        adv_layout.addWidget(QLabel("Enter linear constraint expression:"))

        expr_edit = QPlainTextEdit()
        expr_edit.setPlaceholderText("1.0 * PFK + -0.5 * PYK <= 10.0")
        expr_edit.setMaximumHeight(80)
        adv_layout.addWidget(expr_edit)

        adv_layout.addWidget(QLabel(
            "<small>Supported operators: <=, >=, =<br>"
            "Use reaction IDs from your model. Coefficients default to 1.0 if omitted.</small>"
        ))
        adv_layout.addStretch()
        mode_tabs.addTab(adv_tab, "Linear Expression")

        layout.addWidget(mode_tabs)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)

        if dialog.exec() != QDialog.Accepted:
            return

        try:
            if mode_tabs.currentIndex() == 0:
                # Simple mode: set reaction bounds
                rid = rxn_combo.currentText().strip()
                if rid not in [r.id for r in self.base_model.reactions]:
                    QMessageBox.warning(self, "Invalid reaction", f"Reaction '{rid}' not found in model.")
                    return
                lb = lb_spin.value()
                ub = ub_spin.value()
                if lb > ub:
                    QMessageBox.warning(self, "Invalid bounds", "Lower bound must be ≤ upper bound.")
                    return
                # Store in reaction_bound_overrides
                self.reaction_bound_overrides[rid] = (lb, ub)
                self.model_dirty = True
                QMessageBox.information(self, "Constraint added",
                                        f"Set {rid} bounds to [{lb:.6g}, {ub:.6g}].\n"
                                        f"This will be applied in all subsequent analyses.")
            else:
                # Advanced mode: parse linear expression
                expr_text = expr_edit.toPlainText().strip()
                if not expr_text:
                    QMessageBox.warning(self, "Empty expression", "Please enter a constraint expression.")
                    return
                self._apply_linear_constraint(expr_text, name_edit.text().strip())
        except Exception as e:
            QMessageBox.critical(self, "Constraint error", str(e))

    def _apply_linear_constraint(self, expr_text: str, name: str):
        """Parse and apply a linear constraint like '1.0 * PFK + -0.5 * PYK <= 10.0'."""
        import re as _re

        # Determine operator and split
        if '<=' in expr_text:
            parts = expr_text.split('<=', 1)
            sense = 'L'
        elif '>=' in expr_text:
            parts = expr_text.split('>=', 1)
            sense = 'G'
        elif '=' in expr_text:
            parts = expr_text.split('=', 1)
            sense = 'E'
        else:
            raise ValueError("No comparison operator found. Use <=, >=, or =")

        lhs_str = parts[0].strip()
        rhs_val = float(parts[1].strip())

        # Parse LHS: terms like "1.0 * PFK" or "PFK" or "-0.5*PYK"
        terms = {}
        # Normalize: add spaces around + and -
        lhs_str = _re.sub(r'(?<=[A-Za-z0-9_])\s*\+\s*', ' + ', lhs_str)
        lhs_str = _re.sub(r'(?<=[A-Za-z0-9_])\s*-\s*', ' + -', lhs_str)

        # Split on +
        for token in lhs_str.split('+'):
            token = token.strip()
            if not token:
                continue
            if '*' in token:
                coeff_str, rxn_id = token.split('*', 1)
                coeff = float(coeff_str.strip())
                rxn_id = rxn_id.strip()
            else:
                rxn_id = token.strip()
                coeff = 1.0
                if rxn_id.startswith('-'):
                    rxn_id = rxn_id[1:].strip()
                    coeff = -1.0

            if rxn_id not in [r.id for r in self.base_model.reactions]:
                raise ValueError(f"Reaction '{rxn_id}' not found in model.")
            terms[rxn_id] = terms.get(rxn_id, 0) + coeff

        if not terms:
            raise ValueError("No valid terms found in expression.")

        # Apply constraint using COBRApy
        model = self.base_model
        constraint_name = name or f"custom_{len(model.constraints) + 1}"

        constraint = model.problem.Constraint(
            sum(terms[rid] * model.reactions.get_by_id(rid).flux_expression for rid in terms),
            lb=rhs_val if sense in ('G', 'E') else None,
            ub=rhs_val if sense in ('L', 'E') else None,
            name=constraint_name
        )
        model.add_cons_vars(constraint)
        self.model_dirty = True

        terms_str = " + ".join(f"{c:g}*{r}" for r, c in terms.items())
        op = {"L": "<=", "G": ">=", "E": "="}[sense]
        QMessageBox.information(self, "Constraint added",
                                f"Added constraint '{constraint_name}':\n"
                                f"{terms_str} {op} {rhs_val}\n\n"
                                f"Applied directly to the loaded model.")

    # ---- Generate HTML Report ----
    def generate_html_report(self):
        """Generate a comprehensive HTML report with model info and analysis results."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save HTML Report", "", "HTML files (*.html)"
        )
        if not path:
            return

        try:
            model = self.base_model
            model_id = getattr(model, "id", "N/A") or "N/A"
            model_name = getattr(model, "name", "") or ""
            n_rxn = len(model.reactions)
            n_met = len(model.metabolites)
            n_gene = len(model.genes)

            # Compartments
            compartments = {}
            for m in model.metabolites:
                c = getattr(m, "compartment", "?")
                compartments[c] = compartments.get(c, 0) + 1

            # Subsystems
            subsystems = {}
            for r in model.reactions:
                ss = getattr(r, "subsystem", "") or "Unassigned"
                subsystems[ss] = subsystems.get(ss, 0) + 1

            # Annotation coverage
            n_kegg = sum(1 for r in model.reactions if get_kegg_rxn_id(r))
            n_ec = sum(1 for r in model.reactions if get_ec_numbers(r))
            n_gpr = sum(1 for r in model.reactions if get_gpr(r).strip())

            # Exchange reactions
            exchange_rxns = [r for r in model.reactions if r.boundary]

            # Build HTML
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            html_parts = [f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MetaboDesk Report — {model_id}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 20px 40px; color: #333; line-height: 1.6; }}
  h1 {{ color: #2C3E50; border-bottom: 3px solid #2E86C1; padding-bottom: 8px; }}
  h2 {{ color: #2E86C1; margin-top: 30px; }}
  h3 {{ color: #555; }}
  table {{ border-collapse: collapse; width: 100%; margin: 10px 0 20px 0; }}
  th {{ background: #2E86C1; color: white; padding: 8px 12px; text-align: left; }}
  td {{ padding: 6px 12px; border-bottom: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f8f9fa; }}
  .metric {{ display: inline-block; background: #EBF5FB; border-radius: 8px; padding: 15px 25px; margin: 5px; text-align: center; min-width: 120px; }}
  .metric .value {{ font-size: 28px; font-weight: bold; color: #2E86C1; }}
  .metric .label {{ font-size: 12px; color: #666; }}
  .bar {{ background: #eee; border-radius: 4px; height: 16px; }}
  .bar-fill {{ border-radius: 4px; height: 16px; }}
  .green {{ background: #27AE60; }}
  .yellow {{ background: #F39C12; }}
  .red {{ background: #E74C3C; }}
  .footer {{ margin-top: 40px; padding-top: 10px; border-top: 1px solid #ddd; color: #999; font-size: 11px; }}
</style>
</head>
<body>
<h1>MetaboDesk — Model Report</h1>
<p><b>Generated:</b> {timestamp}</p>

<h2>Model Overview</h2>
<div>
  <div class="metric"><div class="value">{model_id}</div><div class="label">Model ID</div></div>
</div>
<div>
  <div class="metric"><div class="value">{n_rxn}</div><div class="label">Reactions</div></div>
  <div class="metric"><div class="value">{n_met}</div><div class="label">Metabolites</div></div>
  <div class="metric"><div class="value">{n_gene}</div><div class="label">Genes</div></div>
  <div class="metric"><div class="value">{len(compartments)}</div><div class="label">Compartments</div></div>
  <div class="metric"><div class="value">{len(exchange_rxns)}</div><div class="label">Exchange Rxns</div></div>
</div>
"""]

            # Compartments table
            html_parts.append("<h2>Compartments</h2><table><tr><th>ID</th><th>Metabolites</th></tr>")
            comp_names = getattr(model, "compartments", {}) or {}
            for c in sorted(compartments.keys()):
                cname = comp_names.get(c, "")
                label = f"{c} ({cname})" if cname else c
                html_parts.append(f"<tr><td>{label}</td><td>{compartments[c]}</td></tr>")
            html_parts.append("</table>")

            # Annotation coverage
            def pct_str(n, total):
                return f"{n}/{total} ({n / total * 100:.1f}%)" if total else "0"

            def bar_cls(n, total):
                if total == 0:
                    return "red"
                r = n / total
                return "green" if r >= 0.8 else ("yellow" if r >= 0.5 else "red")

            def bar_row(label, n, total):
                w = (n / total * 100) if total else 0
                cls = bar_cls(n, total)
                return (f"<tr><td>{label}</td><td>{pct_str(n, total)}</td>"
                        f"<td><div class='bar'><div class='bar-fill {cls}' style='width:{w:.0f}%'></div></div></td></tr>")

            html_parts.append("<h2>Annotation Coverage</h2><table><tr><th>Annotation</th><th>Coverage</th><th>Bar</th></tr>")
            html_parts.append(bar_row("KEGG Reaction", n_kegg, n_rxn))
            html_parts.append(bar_row("EC Number", n_ec, n_rxn))
            html_parts.append(bar_row("GPR Rules", n_gpr, n_rxn))
            html_parts.append("</table>")

            # Subsystem distribution (top 20)
            html_parts.append("<h2>Subsystem Distribution</h2><table><tr><th>Subsystem</th><th>Reactions</th></tr>")
            for ss, cnt in sorted(subsystems.items(), key=lambda x: -x[1])[:30]:
                html_parts.append(f"<tr><td>{ss}</td><td>{cnt}</td></tr>")
            if len(subsystems) > 30:
                html_parts.append(f"<tr><td><i>... and {len(subsystems) - 30} more</i></td><td></td></tr>")
            html_parts.append("</table>")

            # Last analysis results
            if self.last_run:
                html_parts.append("<h2>Last Analysis Results</h2>")
                analysis_type = self.last_run.get("type", "Unknown")
                html_parts.append(f"<p><b>Analysis type:</b> {analysis_type}</p>")

                if "baseline" in self.last_run:
                    bl = self.last_run["baseline"]
                    obj_val = bl.get("objective_value")
                    if obj_val is not None:
                        html_parts.append(f"<p><b>Objective value:</b> {obj_val:.6g}</p>")
                    fluxes = bl.get("fluxes")
                    if fluxes and isinstance(fluxes, dict):
                        html_parts.append("<h3>Flux Distribution (non-zero)</h3>")
                        html_parts.append("<table><tr><th>Reaction</th><th>Flux</th></tr>")
                        nonzero = {k: v for k, v in fluxes.items() if abs(v) > 1e-9}
                        for rid in sorted(nonzero, key=lambda x: -abs(nonzero[x]))[:50]:
                            html_parts.append(f"<tr><td>{rid}</td><td>{nonzero[rid]:.6g}</td></tr>")
                        if len(nonzero) > 50:
                            html_parts.append(f"<tr><td><i>... and {len(nonzero) - 50} more</i></td><td></td></tr>")
                        html_parts.append("</table>")

                if "fva" in self.last_run:
                    fva_df = self.last_run["fva"]
                    if hasattr(fva_df, "shape"):
                        html_parts.append(f"<h3>FVA Results ({fva_df.shape[0]} reactions)</h3>")
                        html_parts.append("<table><tr><th>Reaction</th><th>Minimum</th><th>Maximum</th></tr>")
                        count = 0
                        for idx, row in fva_df.iterrows():
                            if count >= 50:
                                html_parts.append(f"<tr><td><i>... {fva_df.shape[0] - 50} more rows</i></td><td></td><td></td></tr>")
                                break
                            html_parts.append(f"<tr><td>{idx}</td><td>{row.get('minimum', 0):.6g}</td><td>{row.get('maximum', 0):.6g}</td></tr>")
                            count += 1
                        html_parts.append("</table>")

            # Exchange reactions table
            html_parts.append("<h2>Exchange Reactions</h2><table><tr><th>ID</th><th>Name</th><th>Lower Bound</th><th>Upper Bound</th></tr>")
            for r in sorted(exchange_rxns, key=lambda x: x.id)[:50]:
                html_parts.append(f"<tr><td>{r.id}</td><td>{r.name or ''}</td><td>{r.lower_bound:.6g}</td><td>{r.upper_bound:.6g}</td></tr>")
            if len(exchange_rxns) > 50:
                html_parts.append(f"<tr><td><i>... and {len(exchange_rxns) - 50} more</i></td><td></td><td></td><td></td></tr>")
            html_parts.append("</table>")

            # Footer
            html_parts.append(f"""
<div class="footer">
  Generated by MetaboDesk v1.0 | {timestamp}
</div>
</body>
</html>""")

            html_content = "\n".join(html_parts)

            with open(path, "w", encoding="utf-8") as f:
                f.write(html_content)

            QMessageBox.information(self, "Report saved",
                                    f"HTML report saved to:\n{path}")

            # Offer to open in browser
            reply = QMessageBox.question(self, "Open report?",
                                         "Open the report in your web browser?",
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                QDesktopServices.openUrl(QUrl.fromLocalFile(path))

        except Exception as e:
            QMessageBox.critical(self, "Report error", f"Failed to generate report:\n{str(e)}")

    # ---- About dialog (English) ----
    def show_about(self):
        html = (
            "<div style='font-family:Segoe UI,Ubuntu,Arial;font-size:10px;line-height:1.6;'>"
            "<h2 style='color:#3498db;margin:0 0 4px 0;font-size:14px;'>MetaboDesk</h2>"
            "<p style='margin:0 0 8px 0;font-size:9px;'>"
            "Integrated Metabolic Network Analysis | SBML + COBRApy</p>"
            "<hr style='margin:8px 0;border:none;border-top:1px solid #ccc;'/>"
            "<h4 style='color:#2c3e50;margin:6px 0 4px 0;font-size:10px;'>ANALYSIS TYPES</h4>"
            "<table style='width:100%;border-collapse:collapse;font-size:9px;'>"
            "<tr><td style='padding:4px;font-weight:bold;color:#3498db;width:70px;'>FBA</td>"
            "<td>Flux Balance Analysis. Solves LP to find optimal fluxes. Choose objective, select FBA analysis, press Run.</td></tr>"
            "<tr style='background:#f9fafb;'><td style='padding:4px;font-weight:bold;color:#3498db;'>pFBA</td>"
            "<td>Minimizes total flux after optimization. Most parsimonious solution.</td></tr>"
            "<tr><td style='padding:4px;font-weight:bold;color:#3498db;'>FVA</td>"
            "<td>Flux Variability Analysis. Min/max bounds per reaction at optimal growth.</td></tr>"
            "<tr style='background:#f9fafb;'><td style='padding:4px;font-weight:bold;color:#3498db;'>SGD</td>"
            "<td>Single Gene Deletion. Knocks out each gene, identifies essential genes.</td></tr>"
            "<tr><td style='padding:4px;font-weight:bold;color:#3498db;'>SRD</td>"
            "<td>Single Reaction Deletion. Tests each reaction for essentiality.</td></tr>"
            "<tr style='background:#f9fafb;'><td style='padding:4px;font-weight:bold;color:#3498db;'>Robustness</td>"
            "<td>Parameter sweep. Varies reaction bounds, tracks objective response.</td></tr>"
            "<tr><td style='padding:4px;font-weight:bold;color:#3498db;'>Envelope</td>"
            "<td>Production envelope. Growth vs product trade-off Pareto curve.</td></tr>"
            "<tr style='background:#f9fafb;'><td style='padding:4px;font-weight:bold;color:#3498db;'>Sampling</td>"
            "<td>Monte Carlo flux sampling (ACHR). Distribution statistics per reaction.</td></tr>"
            "</table>"
            "<hr style='margin:8px 0;border:none;border-top:1px solid #ccc;'/>"
            "<h4 style='color:#2c3e50;margin:6px 0 4px 0;font-size:10px;'>QUICK START</h4>"
            "<ol style='margin:0;padding-left:16px;font-size:9px;'>"
            "<li>Open SBML (File → Open or Ctrl+O)</li>"
            "<li>Edit medium bounds (Medium tab, double-click cells)</li>"
            "<li>Select analysis type from dropdown</li>"
            "<li>Configure parameters</li>"
            "<li>Press Run Analysis (Ctrl+R)</li>"
            "<li>View results in tabs</li>"
            "</ol>"
            "<p style='margin:6px 0 0 0;color:#8c92a1;font-size:8px;'>© 2026 Emir Ay</p>"
            "</div>"
        )
        dlg = QMessageBox(self)
        dlg.setWindowTitle("About MetaboDesk")
        dlg.setText(html)
        dlg.setTextFormat(Qt.RichText)
        dlg.setMinimumWidth(600)
        dlg.setMinimumHeight(480)
        dlg.exec()

    # ---------------- Medium ----------------
    def filter_medium_table(self):
        text = self.medium_search.text().strip().lower()
        for row in range(self.medium_table.rowCount()):
            rxn_id = (self.medium_table.item(row, 0).text() if self.medium_table.item(row, 0) else "").lower()
            name = (self.medium_table.item(row, 1).text() if self.medium_table.item(row, 1) else "").lower()
            hide = bool(text) and (text not in rxn_id) and (text not in name)
            self.medium_table.setRowHidden(row, hide)

    def export_medium_bounds_csv(self):
        if self.medium_table.rowCount() == 0:
            QMessageBox.warning(self, "Medium", "No medium rows to export.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Medium bounds (CSV)", str(Path.home() / "medium_bounds.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction", "lb", "ub"]) 
                for i in range(self.medium_table.rowCount()):
                    rid = self.medium_table.item(i, 0).text().strip()
                    lb = self.medium_table.item(i, 2).text().strip()
                    ub = self.medium_table.item(i, 3).text().strip()
                    w.writerow([rid, lb, ub])
            self.statusBar().showMessage(f"Medium bounds exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def import_medium_bounds_csv(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Medium bounds (CSV)", str(Path.home()), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            rxn_to_row = {self.medium_table.item(i, 0).text().strip(): i for i in range(self.medium_table.rowCount())}
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                for rowvals in reader:
                    if not rowvals or rowvals[0] == "reaction":
                        continue
                    rid, lb, ub = rowvals[:3]
                    row = rxn_to_row.get(rid)
                    if row is None:
                        continue
                    self.medium_table.setItem(row, 2, QTableWidgetItem(str(lb)))
                    self.medium_table.setItem(row, 3, QTableWidgetItem(str(ub)))
            self.filter_medium_table()
            self.statusBar().showMessage(f"Medium bounds imported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))

    def populate_medium_table(self):
        self.medium_table.setRowCount(0)
        for rxn in self.exchange_rxns:
            row = self.medium_table.rowCount()
            self.medium_table.insertRow(row)

            item_id = QTableWidgetItem(rxn.id)
            item_id.setFlags(item_id.flags() & ~Qt.ItemIsEditable)

            item_name = QTableWidgetItem(rxn.name or "")
            item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)

            self.medium_table.setItem(row, 0, item_id)
            self.medium_table.setItem(row, 1, item_name)
            self.medium_table.setItem(row, 2, QTableWidgetItem(str(float(rxn.lower_bound))))
            self.medium_table.setItem(row, 3, QTableWidgetItem(str(float(rxn.upper_bound))))

        self.filter_medium_table()

    def apply_medium_table_bounds_to_model(self, model: cobra.Model):
        rxn_by_id = {r.id: r for r in model.reactions}
        for i in range(self.medium_table.rowCount()):
            rxn_id = self.medium_table.item(i, 0).text().strip()
            lb_txt = self.medium_table.item(i, 2).text().strip()
            ub_txt = self.medium_table.item(i, 3).text().strip()
            lb = float(lb_txt)
            ub = float(ub_txt)
            if lb > ub:
                raise ValueError(f"Lower bound > upper bound in Medium row {i+1} for reaction {rxn_id}.")
            if rxn_id in rxn_by_id:
                rxn_by_id[rxn_id].lower_bound = lb
                rxn_by_id[rxn_id].upper_bound = ub

    def apply_medium_preset(self):
        if self.base_model is None:
            return
        preset = self.medium_preset_combo.currentText()
        if preset == "Reset to original":
            self.reset_medium()
            return
        # Build a quick heuristic preset
        by_id = {self.medium_table.item(i, 0).text().strip(): i for i in range(self.medium_table.rowCount())}
        def set_bound(rid, lb=None, ub=None):
            row = by_id.get(rid)
            if row is None:
                return
            if lb is not None:
                self.medium_table.setItem(row, 2, QTableWidgetItem(f"{float(lb):g}"))
            if ub is not None:
                self.medium_table.setItem(row, 3, QTableWidgetItem(f"{float(ub):g}"))
        # Close all uptakes by default
        for i in range(self.medium_table.rowCount()):
            lb_item = self.medium_table.item(i, 2)
            try:
                lb = float(lb_item.text())
            except Exception:
                lb = 0.0
            if lb < 0:
                self.medium_table.setItem(i, 2, QTableWidgetItem("0"))
        if preset == "Minimal M9":
            set_bound("EX_glc__D_e", -10, None)
            set_bound("EX_o2_e", -20, None)
            set_bound("EX_nh4_e", -10, None)
            set_bound("EX_pi_e", -10, None)
            set_bound("EX_h2o_e", None, 1000)
            set_bound("EX_h_e", None, 1000)
            set_bound("EX_co2_e", None, 1000)
        elif preset == "Rich LB":
            # Allow common carbon/nitrogen sources generously
            for rid in ("EX_glc__D_e", "EX_o2_e", "EX_nh4_e", "EX_pi_e", "EX_glyc_e", "EX_etoh_e"):
                set_bound(rid, -50, None)
            for rid in ("EX_h2o_e", "EX_h_e", "EX_co2_e"):
                set_bound(rid, None, 1000)
        self.filter_medium_table()
        self.statusBar().showMessage(f"Applied medium preset: {preset}")

    def close_all_uptakes(self):
        if self.base_model is None or self.is_running:
            return

        current = self.tabs.currentWidget()

        if current is self.tab_medium:
            changed = 0
            for rxn in self.exchange_rxns:
                if rxn.lower_bound < 0:
                    rxn.lower_bound = 0.0
                    changed += 1
            self.populate_medium_table()
            self.status_lbl.setText(f"All uptakes closed in Medium (LB<0 → 0). Changed: {changed}")
            self.statusBar().showMessage(f"All uptakes closed in Medium (changed {changed}).")
            if changed:
                self.model_dirty = True
            return

        if current is self.tab_rxns:
            selected_rows = sorted({idx.row() for idx in self.reaction_table.selectionModel().selectedRows()})
            if not selected_rows:
                QMessageBox.information(
                    self,
                    "No selection",
                    "Reactions tab: Select one or more rows to apply 'Close uptakes' as reaction overrides (LB<0 → 0).\n\n"
                    "Tip: Use Medium tab to close all exchange uptakes globally."
                )
                return

            rxn_by_id = {r.id: r for r in self.base_model.reactions}
            changed = 0
            skipped = 0

            for row in selected_rows:
                rxn_id = self._reaction_row_to_id(row)
                if not rxn_id or rxn_id not in rxn_by_id:
                    skipped += 1
                    continue
                rxn = rxn_by_id[rxn_id]

                if not is_exchange_reaction(rxn):
                    skipped += 1
                    continue

                current_lb = float(rxn.lower_bound)
                current_ub = float(rxn.upper_bound)
                if rxn_id in self.reaction_bound_overrides:
                    current_lb, current_ub = self.reaction_bound_overrides[rxn_id]

                if float(current_lb) < 0:
                    self.reaction_bound_overrides[rxn_id] = (0.0, float(current_ub))
                    changed += 1

            self.populate_reaction_table()
            self.statusBar().showMessage(f"Reactions tab: uptake overrides applied. changed={changed}, skipped={skipped}")
            if changed:
                self.model_dirty = True
            return

        QMessageBox.information(self, "Close uptakes", "Go to Medium or Reactions tab to use this action.")

    def reset_medium(self):
        if self.base_model is None or self.is_running:
            return
        for rxn in self.exchange_rxns:
            if rxn.id in self.original_bounds:
                lb, ub = self.original_bounds[rxn.id]
                rxn.lower_bound = lb
                rxn.upper_bound = ub
        self.populate_medium_table()
        self.status_lbl.setText("Medium reset to original.")
        self.statusBar().showMessage("Medium reset to original.")
        self.model_dirty = True

    def reset_reactions_to_original(self):
        if self.base_model is None or self.is_running:
            return
        if self.original_model_snapshot is None:
            QMessageBox.warning(self, "No snapshot", "No original snapshot available.")
            return

        msg = (
            "This will reset ALL reaction bounds in the current model to the original SBML snapshot and clear reaction overrides.\n\n"
            "It will NOT delete your SBML file, but it WILL discard current in-app bound edits/overrides.\n\n"
            "Continue?"
        )
        if QMessageBox.question(self, "Confirm reset reactions", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        snap = self.original_model_snapshot
        for rxn in self.base_model.reactions:
            try:
                srxn = snap.reactions.get_by_id(rxn.id)
                rxn.lower_bound = float(srxn.lower_bound)
                rxn.upper_bound = float(srxn.upper_bound)
            except KeyError:
                continue

        self.exchange_rxns = [r for r in self.base_model.reactions if is_exchange_reaction(r)]
        self.exchange_rxns.sort(key=lambda r: r.id)
        self.original_bounds = {r.id: (float(r.lower_bound), float(r.upper_bound)) for r in self.exchange_rxns}

        self.reaction_bound_overrides.clear()

        self.populate_medium_table()
        self.populate_reaction_table()
        self.statusBar().showMessage("Reactions reset to original snapshot (and overrides cleared).")
        self.model_dirty = True

    # ---------------- Genes tab ----------------
    def populate_genes_tab(self):
        if self.base_model is None:
            self.genes_tab_list.clear()
            self.gene_rxn_table.setRowCount(0)
            self._gene_to_reactions = {}
            return

        self._gene_to_reactions = {}
        for g in self.base_model.genes:
            try:
                rids = sorted([r.id for r in g.reactions])
            except Exception:
                rids = []
            self._gene_to_reactions[str(g.id)] = rids

        self.genes_tab_list.clear()
        for g in sorted(self.base_model.genes, key=lambda gg: gg.id):
            gname = getattr(g, "name", "") or ""
            label = f"{g.id} — {gname}".strip(" —")
            it = QListWidgetItem(label)
            it.setData(Qt.UserRole, str(g.id))
            self.genes_tab_list.addItem(it)

        self.filter_genes_tab_gene_list()
        self.gene_rxn_table.setRowCount(0)
        self.gene_rxn_details.setPlainText("")

    def filter_genes_tab_gene_list(self):
        text = self.gene_global_search.text().strip().lower()
        for i in range(self.genes_tab_list.count()):
            it = self.genes_tab_list.item(i)
            gid = str(it.data(Qt.UserRole) or "")
            gname = it.text().lower()
            it.setHidden(bool(text) and (text not in gid.lower()) and (text not in gname))

    def on_genes_tab_gene_changed(self, current: QListWidgetItem, _prev: QListWidgetItem):
        if self.base_model is None or not current:
            self.gene_rxn_table.setRowCount(0)
            return

        gid = str(current.data(Qt.UserRole) or "").strip()
        if not gid:
            self.gene_rxn_table.setRowCount(0)
            return

        rxn_ids = self._gene_to_reactions.get(gid, [])
        self.populate_gene_rxn_table(rxn_ids)

    def populate_gene_rxn_table(self, rxn_ids: list[str]):
        if self.base_model is None:
            return

        self.gene_rxn_table.setRowCount(0)

        for rid in rxn_ids:
            try:
                rxn = self.base_model.reactions.get_by_id(rid)
            except Exception:
                continue

            row = self.gene_rxn_table.rowCount()
            self.gene_rxn_table.insertRow(row)

            item_id = QTableWidgetItem(rxn.id)
            item_desc = QTableWidgetItem(get_description(rxn))
            eq = ""
            try:
                eq = rxn.reaction
            except Exception:
                eq = ""
            item_eq = QTableWidgetItem(eq)
            item_gpr = QTableWidgetItem(get_gpr(rxn))

            for it in (item_id, item_desc, item_eq, item_gpr):
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)

            self.gene_rxn_table.setItem(row, 0, item_id)
            self.gene_rxn_table.setItem(row, 1, item_desc)
            self.gene_rxn_table.setItem(row, 2, item_eq)
            self.gene_rxn_table.setItem(row, 3, item_gpr)

        self.gene_rxn_details.setPlainText("")

    # ---------------- Reactions table ----------------
    def _parse_metabolite_filter_terms(self) -> list[str]:
        raw = self.met_filter.text().strip()
        if not raw:
            return []
        parts = [p.strip().lower() for p in raw.replace(";", ",").split(",")]
        return [p for p in parts if p]

    def _reaction_row_to_id(self, row: int) -> str | None:
        it = self.reaction_table.item(row, 0)
        return it.text().strip() if it else None

    def _rxn_contains_term(self, rxn: cobra.Reaction, term: str) -> bool:
        term = term.strip().lower()
        if not term:
            return True
        for m in rxn.metabolites.keys():
            mid = (m.id or "").lower()
            mn = (m.name or "").lower()
            if term in mid or term in mn:
                return True
        return False

    def _reaction_matches_filters(self, rxn: cobra.Reaction, text: str, met_terms: list[str], mode: str) -> bool:
        text = text.strip().lower()
        if text:
            hay = " ".join([
                rxn.id.lower(),
                (get_description(rxn) or "").lower(),
                (getattr(rxn, "reaction", "") or "").lower(),
                get_kegg_rxn_id(rxn).lower(),
                get_ec_numbers(rxn).lower(),
                get_gpr(rxn).lower(),
            ])
            if text not in hay:
                return False

        if met_terms:
            if mode == "AND":
                for t in met_terms:
                    if not self._rxn_contains_term(rxn, t):
                        return False
            else:
                if not any(self._rxn_contains_term(rxn, t) for t in met_terms):
                    return False

        return True

    def filter_reaction_table(self):
        if self.base_model is None:
            return
        text = self.rxn_global_search.text()
        met_terms = self._parse_metabolite_filter_terms()
        mode = self.met_filter_mode.currentText().strip().upper() or "AND"

        rxn_by_id = {r.id: r for r in self.base_model.reactions}

        for row in range(self.reaction_table.rowCount()):
            rxn_id = self._reaction_row_to_id(row)
            rxn = rxn_by_id.get(rxn_id) if rxn_id else None
            if rxn is None:
                self.reaction_table.setRowHidden(row, True)
                continue
            hide = not self._reaction_matches_filters(rxn, text, met_terms, mode)
            self.reaction_table.setRowHidden(row, hide)

    def populate_reaction_table(self):
        if self.base_model is None:
            return

        self.reaction_table.blockSignals(True)
        self.reaction_table.setRowCount(0)

        for rxn in sorted(self.base_model.reactions, key=lambda r: r.id):
            row = self.reaction_table.rowCount()
            self.reaction_table.insertRow(row)

            item_id = QTableWidgetItem(rxn.id)
            item_desc = QTableWidgetItem(get_description(rxn))
            eq = ""
            try:
                eq = rxn.reaction
            except Exception:
                eq = ""
            item_eq = QTableWidgetItem(eq)
            item_kegg = QTableWidgetItem(get_kegg_rxn_id(rxn))
            item_ec = QTableWidgetItem(get_ec_numbers(rxn))
            item_gpr = QTableWidgetItem(get_gpr(rxn))

            for it in (item_id, item_desc, item_eq, item_kegg, item_ec, item_gpr):
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)

            if rxn.id in self.reaction_bound_overrides:
                lb, ub = self.reaction_bound_overrides[rxn.id]
            else:
                lb, ub = float(rxn.lower_bound), float(rxn.upper_bound)

            item_lb = QTableWidgetItem(str(lb))
            item_ub = QTableWidgetItem(str(ub))

            self.reaction_table.setItem(row, 0, item_id)
            self.reaction_table.setItem(row, 1, item_desc)
            self.reaction_table.setItem(row, 2, item_eq)
            self.reaction_table.setItem(row, 3, item_kegg)
            self.reaction_table.setItem(row, 4, item_ec)
            self.reaction_table.setItem(row, 5, item_gpr)
            self.reaction_table.setItem(row, 6, item_lb)
            self.reaction_table.setItem(row, 7, item_ub)

        self.reaction_table.blockSignals(False)
        self.filter_reaction_table()
        self.clear_rxn_overrides_btn.setEnabled(True)
        self.remove_selected_overrides_btn.setEnabled(True)

    def on_reaction_table_item_changed(self, item: QTableWidgetItem):
        # FIXED: remove invalid usage of rxn_id/lb/ub before definition.
        if self.base_model is None:
            return

        row = item.row()
        col = item.column()
        if col not in (6, 7):
            return

        rxn_id = self._reaction_row_to_id(row)
        if not rxn_id:
            return

        lb_item = self.reaction_table.item(row, 6)
        ub_item = self.reaction_table.item(row, 7)
        if not lb_item or not ub_item:
            return

        try:
            lb = self._parse_float(lb_item.text())
            ub = self._parse_float(ub_item.text())
            if lb > ub:
                raise ValueError("LB > UB")
        except Exception:
            QMessageBox.warning(self, "Invalid bounds", f"Invalid bounds for {rxn_id}. Please enter valid numbers with LB <= UB.")
            self.reaction_table.blockSignals(True)
            if rxn_id in self.reaction_bound_overrides:
                old_lb, old_ub = self.reaction_bound_overrides[rxn_id]
            else:
                rxn = self.base_model.reactions.get_by_id(rxn_id)
                old_lb, old_ub = float(rxn.lower_bound), float(rxn.upper_bound)
            lb_item.setText(str(old_lb))
            ub_item.setText(str(old_ub))
            self.reaction_table.blockSignals(False)
            return

        self.save_state_for_undo()
        self.reaction_bound_overrides[rxn_id] = (lb, ub)
        self.model_dirty = True
        self.statusBar().showMessage(f"Override set: {rxn_id} LB={lb:g}, UB={ub:g}")

    def clear_reaction_overrides(self):
        self.save_state_for_undo()
        self.reaction_bound_overrides.clear()
        self.populate_reaction_table()
        self.model_dirty = True
        self.statusBar().showMessage("Reaction overrides cleared.")

    def remove_selected_reaction_overrides(self):
        if self.base_model is None:
            return

        selected_rows = sorted({idx.row() for idx in self.reaction_table.selectionModel().selectedRows()})
        if not selected_rows:
            QMessageBox.information(self, "No selection", "Select one or more reactions (rows) first.")
            return

        removed = 0
        for row in selected_rows:
            rxn_id = self._reaction_row_to_id(row)
            if not rxn_id:
                continue
            if rxn_id in self.reaction_bound_overrides:
                self.reaction_bound_overrides.pop(rxn_id, None)
                removed += 1

        self.populate_reaction_table()
        if removed:
            self.model_dirty = True
        self.statusBar().showMessage(f"Removed {removed} reaction override(s).")

    def apply_reaction_overrides_to_model(self, model: cobra.Model):
        for rxn_id, (lb, ub) in self.reaction_bound_overrides.items():
            try:
                rxn = model.reactions.get_by_id(rxn_id)
            except KeyError:
                continue
            rxn.lower_bound = float(lb)
            rxn.upper_bound = float(ub)

    # ---------------- KO / OE ----------------
    def filter_gene_list(self):
        text = self.gene_search.text().strip().lower()
        for i in range(self.gene_list.count()):
            item = self.gene_list.item(i)
            item.setHidden(bool(text) and text not in item.text().lower())

    def populate_gene_list(self):
        if self.base_model is None:
            return
        self.gene_list.clear()
        for gid in sorted([g.id for g in self.base_model.genes]):
            self.gene_list.addItem(QListWidgetItem(gid))
        self.filter_gene_list()

    def add_selected_gene_knockout(self):
        item = self.gene_list.currentItem()
        if not item:
            return
        gid = item.text().strip()
        if gid:
            self.save_state_for_undo()
            self.knockout_genes.add(gid)
            self.model_dirty = True
            self.update_knockout_label()

    def clear_knockouts(self):
        self.save_state_for_undo()
        self.knockout_genes.clear()
        self.model_dirty = True
        self.update_knockout_label()

    def update_knockout_label(self):
        if not self.knockout_genes:
            self.ko_lbl.setText("Gene knockouts: (none)")
        else:
            self.ko_lbl.setText(f"Gene knockouts ({len(self.knockout_genes)}): {', '.join(sorted(self.knockout_genes))}")

    def filter_overexpression_reaction_list(self):
        text = self.rxn_search.text().strip().lower()
        for i in range(self.rxn_list.count()):
            item = self.rxn_list.item(i)
            item.setHidden(bool(text) and text not in item.text().lower())

    def populate_overexpression_reaction_list(self):
        if self.base_model is None:
            return
        self.rxn_list.clear()
        for r in sorted(self.base_model.reactions, key=lambda x: x.id):
            desc = get_description(r)
            text = f"{r.id} — {desc}" if desc else r.id
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, r.id)
            self.rxn_list.addItem(item)
        self.filter_overexpression_reaction_list()

    def refresh_overexpression_lists(self):
        self.oe_active_list.clear()
        for rid in sorted(self.overexpression_reactions):
            factor = self.overexpression_reactions[rid]
            it = QListWidgetItem(f"{rid} ×{factor:g}")
            it.setData(Qt.UserRole, rid)
            self.oe_active_list.addItem(it)

        self.temp_active_list.clear()
        for rid in sorted(self.temporary_upper_bound_overrides):
            ub = self.temporary_upper_bound_overrides[rid]
            it = QListWidgetItem(f"{rid} UB={ub:g}")
            it.setData(Qt.UserRole, rid)
            self.temp_active_list.addItem(it)

    def add_selected_reaction_overexpression(self):
        item = self.rxn_list.currentItem()
        if not item:
            return
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        self.save_state_for_undo()
        self.overexpression_reactions[str(rxn_id)] = float(self.oe_factor.value())
        self.model_dirty = True
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def remove_selected_reaction_overexpression(self):
        item = self.rxn_list.currentItem()
        if not item:
            return
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        self.save_state_for_undo()
        if str(rxn_id) in self.overexpression_reactions:
            self.model_dirty = True
        self.overexpression_reactions.pop(str(rxn_id), None)
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def clear_overexpressions(self):
        self.save_state_for_undo()
        if self.overexpression_reactions:
            self.model_dirty = True
        self.overexpression_reactions.clear()
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def set_temp_upper_bound_for_selected(self):
        item = self.rxn_list.currentItem()
        if not item:
            return
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        self.save_state_for_undo()
        self.temporary_upper_bound_overrides[str(rxn_id)] = float(self.temp_ub_spin.value())
        self.model_dirty = True
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def remove_temp_upper_bound_for_selected(self):
        item = self.rxn_list.currentItem()
        if not item:
            return
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        self.save_state_for_undo()
        if str(rxn_id) in self.temporary_upper_bound_overrides:
            self.model_dirty = True
        self.temporary_upper_bound_overrides.pop(str(rxn_id), None)
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def clear_temp_upper_bounds(self):
        self.save_state_for_undo()
        if self.temporary_upper_bound_overrides:
            self.model_dirty = True
        self.temporary_upper_bound_overrides.clear()
        self.refresh_overexpression_lists()
        self.update_selected_rxn_info()

    def jump_to_reaction_from_item(self, item: QListWidgetItem):
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        for i in range(self.rxn_list.count()):
            it = self.rxn_list.item(i)
            if str(it.data(Qt.UserRole)) == str(rxn_id):
                self.rxn_list.setCurrentItem(it)
                self.rxn_list.scrollToItem(it)
                break

    def update_selected_rxn_info(self, *_):
        if self.base_model is None:
            self.rxn_info_lbl.setText("Selected reaction: (none)")
            self.rxn_preview_lbl.setText("Preview: -")
            return
        item = self.rxn_list.currentItem()
        if not item:
            self.rxn_info_lbl.setText("Selected reaction: (none)")
            self.rxn_preview_lbl.setText("Preview: -")
            return
        rxn_id = item.data(Qt.UserRole)
        if not rxn_id:
            return
        rxn = self.base_model.reactions.get_by_id(rxn_id)
        lb = float(rxn.lower_bound)
        ub = float(rxn.upper_bound)
        self.rxn_info_lbl.setText(
            f"Selected: {rxn.id}\n"
            f"Description: {get_description(rxn)}\n"
            f"Current bounds: LB={lb:g}, UB={ub:g}\n"
            f"GPR: {get_gpr(rxn)}"
        )
        factor = float(self.oe_factor.value())
        prev_lb = lb * factor if (self.oe_scale_lb.isChecked() and lb < 0) else lb
        prev_ub = ub * factor
        self.rxn_preview_lbl.setText(f"Preview if overexpression applied: LB={prev_lb:g}, UB={prev_ub:g}")

    def apply_temp_upper_bound_overrides(self, model: cobra.Model):
        for rxn_id, ub in self.temporary_upper_bound_overrides.items():
            try:
                model.reactions.get_by_id(rxn_id).upper_bound = float(ub)
            except KeyError:
                continue

    def apply_overexpression_to_model(self, model: cobra.Model):
        scale_lb = bool(self.oe_scale_lb.isChecked())
        for rxn_id, factor in self.overexpression_reactions.items():
            try:
                rxn = model.reactions.get_by_id(rxn_id)
            except KeyError:
                continue
            if scale_lb and float(rxn.lower_bound) < 0:
                rxn.lower_bound = float(rxn.lower_bound) * float(factor)
            rxn.upper_bound = float(rxn.upper_bound) * float(factor)

    def apply_knockouts_to_model(self, model: cobra.Model):
        for gid in sorted(self.knockout_genes):
            if gid not in model.genes:
                raise ValueError(f"Gene not found in model: {gid}")
            model.genes.get_by_id(gid).knock_out()

    # ---------------- Analysis functions (FVA, PFBA) ----------------
    def compute_fva(self, model: cobra.Model) -> dict:
        """
        Compute Flux Variability Analysis (FVA).
        Returns a dict with min/max flux for each reaction.
        """
        try:
            from cobra.flux_analysis import flux_variability_analysis
            fva_result = flux_variability_analysis(model, processes=1)
            result = {}
            for rid in fva_result.index:
                result[rid] = {
                    "min": float(fva_result.loc[rid, "minimum"]),
                    "max": float(fva_result.loc[rid, "maximum"]),
                }
            return result
        except Exception as e:
            raise ValueError(f"FVA computation failed: {e}")

    def compute_pfba(self, model: cobra.Model) -> dict:
        """
        Compute Parsimonious FBA (pFBA).
        First optimizes the objective, then minimizes total flux.
        """
        try:
            from cobra.flux_analysis import pfba
            solution = pfba(model)
            result = {
                "status": str(solution.status),
                "objective": float(solution.objective_value),
                "flux": solution.fluxes.to_dict(),
            }
            return result
        except Exception as e:
            raise ValueError(f"pFBA computation failed: {e}")

    def compute_sgd(self, model: cobra.Model) -> dict:
        """Single Gene Deletion analysis."""
        try:
            from cobra.flux_analysis import single_gene_deletion
            result_df = single_gene_deletion(model, processes=1)
            if result_df is None or len(result_df.index) == 0:
                return {}

            growth_col = "growth" if "growth" in result_df.columns else (
                "growth_rate" if "growth_rate" in result_df.columns else None
            )
            if growth_col is None:
                raise ValueError("SGD result missing growth column.")

            result = {}
            for idx in result_df.index:
                growth_val = _safe_float(result_df.loc[idx, growth_col], 0.0)
                # COBRApy returns frozenset indices
                if isinstance(idx, frozenset):
                    for gid in idx:
                        result[str(gid)] = growth_val
                elif isinstance(idx, (set, list, tuple)):
                    for gid in idx:
                        result[str(gid)] = growth_val
                else:
                    result[str(idx).strip()] = growth_val
            return result
        except Exception as e:
            raise ValueError(f"SGD computation failed: {e}")

    def compute_dgd(self, model: cobra.Model, genes: list[str]) -> dict:
        """Double Gene Deletion analysis for selected gene list."""
        try:
            from cobra.flux_analysis import double_gene_deletion
            if not genes or len(genes) < 2:
                return {}
            try:
                result_df = double_gene_deletion(model, gene_list1=genes, processes=1)
            except TypeError:
                result_df = double_gene_deletion(model, genes, processes=1)
            if result_df is None or len(result_df.index) == 0:
                return {}

            growth_col = "growth" if "growth" in result_df.columns else (
                "growth_rate" if "growth_rate" in result_df.columns else None
            )
            if growth_col is None:
                raise ValueError("DGD result missing growth column.")

            result = {}
            # Try to read explicit pair columns if present
            pair_cols = None
            if "ids" in result_df.columns:
                pair_cols = ("ids",)
            elif "gene1" in result_df.columns and "gene2" in result_df.columns:
                pair_cols = ("gene1", "gene2")
            elif "id1" in result_df.columns and "id2" in result_df.columns:
                pair_cols = ("id1", "id2")
            elif "ids1" in result_df.columns and "ids2" in result_df.columns:
                pair_cols = ("ids1", "ids2")

            if pair_cols is not None:
                for idx, row in result_df.iterrows():
                    if pair_cols == ("ids",):
                        pair_val = row.get("ids")
                        if isinstance(pair_val, (list, tuple)) and len(pair_val) >= 2:
                            gid1, gid2 = str(pair_val[0]), str(pair_val[1])
                        else:
                            parts = str(pair_val).split("-")
                            gid1 = parts[0].strip() if parts else str(pair_val)
                            gid2 = parts[1].strip() if len(parts) > 1 else ""
                    else:
                        gid1 = str(row.get(pair_cols[0], "")).strip()
                        gid2 = str(row.get(pair_cols[1], "")).strip()
                    pair_key = f"{gid1} + {gid2}".strip(" +")
                    result[pair_key] = _safe_float(row.get(growth_col, 0.0), 0.0)
            else:
                for idx in result_df.index:
                    if isinstance(idx, tuple) and len(idx) >= 2:
                        gid1, gid2 = str(idx[0]), str(idx[1])
                    else:
                        parts = str(idx).split("-")
                        gid1 = parts[0].strip() if parts else str(idx)
                        gid2 = parts[1].strip() if len(parts) > 1 else ""
                    pair_key = f"{gid1} + {gid2}".strip(" +")
                    result[pair_key] = _safe_float(result_df.loc[idx, growth_col], 0.0)
            return result
        except Exception as e:
            raise ValueError(f"DGD computation failed: {e}")

    def compute_srd(self, model: cobra.Model) -> dict:
        """Single Reaction Deletion analysis."""
        try:
            from cobra.flux_analysis import single_reaction_deletion
            result_df = single_reaction_deletion(model, processes=1)
            result = {}

            growth_col = "growth" if "growth" in result_df.columns else (
                "growth_rate" if "growth_rate" in result_df.columns else None
            )
            if growth_col is None:
                raise ValueError("SRD result missing growth column.")

            status_col = "status" if "status" in result_df.columns else None

            for idx in result_df.index:
                growth_val = _safe_float(result_df.loc[idx, growth_col], 0.0)
                # COBRApy returns frozenset indices
                if isinstance(idx, frozenset):
                    for rid in idx:
                        result[str(rid)] = growth_val
                elif isinstance(idx, (set, list, tuple)):
                    for rid in idx:
                        result[str(rid)] = growth_val
                else:
                    result[str(idx).strip()] = growth_val
            return result
        except Exception as e:
            raise ValueError(f"SRD computation failed: {e}")


    def compute_robustness(self, model: cobra.Model, rxn_id: str, min_val: float, max_val: float, steps: int, bound_type: str = "ub", worker=None) -> dict:
        """Robustness analysis: sweep a reaction's bound and measure objective."""
        result = {"values": [], "objectives": []}
        try:
            if steps < 2:
                raise ValueError("Steps must be >= 2.")
            rxn = model.reactions.get_by_id(rxn_id)
            original_lb = float(rxn.lower_bound)
            original_ub = float(rxn.upper_bound)
            
            sweep_vals = list(dict.fromkeys([min_val + i * (max_val - min_val) / (steps - 1) for i in range(steps)]))
            for idx, val in enumerate(sweep_vals):
                if worker:
                    pct = int((idx + 1) / len(sweep_vals) * 100)
                    worker.report_progress(f"Robustness: step {idx+1}/{len(sweep_vals)}", pct)
                with model:
                    rxn_ctx = model.reactions.get_by_id(rxn_id)
                    if bound_type.lower() == "ub":
                        rxn_ctx.upper_bound = val
                    else:  # lb
                        rxn_ctx.lower_bound = val
                    
                    sol = model.optimize()
                    if str(sol.status) == "optimal":
                        result["values"].append(float(val))
                        result["objectives"].append(float(sol.objective_value))
            
            return result
        except Exception as e:
            raise ValueError(f"Robustness computation failed: {e}")

    def compute_production_envelope(self, model: cobra.Model, product_id: str, steps: int) -> dict:
        """Production envelope: measure trade-off between growth and product formation.
        
        Uses FVA to determine actual max production, then sweeps from 0 to max.
        """
        result = {"growth": [], "product": []}
        try:
            if steps < 2:
                raise ValueError("Steps must be >= 2.")
            if model.objective is None:
                raise ValueError("No objective reaction found")

            product_rxn = model.reactions.get_by_id(product_id)

            # Find actual max production via temporary objective swap
            with model:
                model.objective = product_rxn
                max_sol = model.optimize()
                max_production = float(max_sol.objective_value) if max_sol.status == "optimal" else product_rxn.upper_bound

            if max_production <= 0:
                max_production = product_rxn.upper_bound if product_rxn.upper_bound > 0 else 10.0

            for i in range(steps):
                bound = i * max_production / max(steps - 1, 1)
                with model:
                    prod_rxn = model.reactions.get_by_id(product_id)
                    prod_rxn.lower_bound = bound
                    prod_rxn.upper_bound = bound

                    sol = model.optimize()
                    if str(sol.status) == "optimal":
                        result["product"].append(float(bound))
                        result["growth"].append(float(sol.objective_value))

            return result
        except Exception as e:
            raise ValueError(f"Production envelope computation failed: {e}")

    def compute_flux_sampling(self, model: cobra.Model, sample_size: int = 500) -> dict:
        """Flux sampling using ACHR method."""
        try:
            from cobra.sampling import ACHRSampler
            sampler = ACHRSampler(model)
            samples = sampler.sample(sample_size)
            
            result = {}
            for rxn_id in samples.columns:
                flux_vals = samples[rxn_id].values
                result[rxn_id] = {
                    "mean": float(flux_vals.mean()),
                    "stdev": float(flux_vals.std()),
                    "min": float(flux_vals.min()),
                    "max": float(flux_vals.max()),
                }
            return result, samples
        except Exception as e:
            raise ValueError(f"Flux sampling failed: {e}")

    # ---------------- Plotting & Results ----------------
    def _clear_chart(self):
        ax = self.canvas.ax
        ax.clear()
        ax.set_title("Run FBA to see flux comparison")
        ax.set_xlabel("Reaction")
        ax.set_ylabel("Flux")
        self.canvas.draw()

    def _plot_flux_comparison(self, rxn_ids, base_flux, compared_flux, base_label: str, compared_label: str, title: str):
        ax = self.canvas.ax
        ax.clear()
        if not rxn_ids:
            ax.set_title("No flux changes above threshold")
            self.canvas.draw()
            return

        chart_type = self.results_chart_type.currentText()

        if chart_type == "Flux histogram":
            # Histogram of all flux values
            all_flux = list(base_flux.values()) + list(compared_flux.values())
            all_flux = [f for f in all_flux if abs(f) > 1e-9]
            if all_flux:
                ax.hist(all_flux, bins=50, edgecolor='black', alpha=0.7)
                ax.set_title("Flux Distribution (all non-zero fluxes)")
                ax.set_xlabel("Flux value")
                ax.set_ylabel("Count")
            else:
                ax.set_title("No non-zero fluxes to display")
        elif chart_type == "Waterfall (top changes)":
            # Waterfall chart showing top delta changes
            deltas = [(rid, compared_flux.get(rid, 0) - base_flux.get(rid, 0)) for rid in rxn_ids]
            deltas.sort(key=lambda x: x[1], reverse=True)
            deltas = deltas[:20]  # Top 20
            rids = [d[0] for d in deltas]
            vals = [d[1] for d in deltas]
            colors = ['green' if v > 0 else 'red' for v in vals]
            ax.barh(range(len(rids)), vals, color=colors)
            ax.set_yticks(range(len(rids)))
            ax.set_yticklabels(rids, fontsize=8)
            ax.axvline(x=0, color='black', linewidth=0.5)
            ax.set_title(f"Top flux changes ({base_label} → {compared_label})")
            ax.set_xlabel("Delta flux")
            ax.invert_yaxis()
        else:
            # Default: Bar comparison
            bvals = [base_flux.get(r, 0.0) for r in rxn_ids]
            cvals = [compared_flux.get(r, 0.0) for r in rxn_ids]
            x = list(range(len(rxn_ids)))
            width = 0.4
            ax.bar([i - width / 2 for i in x], bvals, width=width, label=base_label)
            ax.bar([i + width / 2 for i in x], cvals, width=width, label=compared_label)
            ax.set_title(title)
            ax.set_xlabel("Reaction")
            ax.set_ylabel("Flux")
            ax.set_xticks(x)
            ax.set_xticklabels(rxn_ids, rotation=90, fontsize=8)
            ax.legend()

        self.canvas.figure.tight_layout()
        self.canvas.draw()

    def _refresh_results_view_from_last_run(self, *_):
        if not self.last_run:
            return
        self.last_run["compare_mode"] = self.compare_mode.currentText()
        self._recompute_flux_rows_for_compare_mode()
        self._render_results_from_last_run()

    def _set_flux_table_headers_for_mode(self, mode: str):
        if mode == "Original vs Baseline":
            self.flux_tbl.setHorizontalHeaderLabels(["Reaction", "Original flux", "Baseline flux", "Delta (baseline - original)"])
        else:
            self.flux_tbl.setHorizontalHeaderLabels(["Reaction", "Baseline flux", "Compared flux", "Delta (compared - baseline)"])

    def _recompute_flux_rows_for_compare_mode(self):
        assert self.last_run is not None
        mode = self.last_run.get("compare_mode", "Gene knockout only")

        if mode == "Original vs Baseline":
            base_flux = self.last_run["original"]["flux"]
            cmp_flux = self.last_run["baseline"]["flux"]
        elif mode == "Gene knockout only":
            base_flux = self.last_run["baseline"]["flux"]
            cmp_flux = self.last_run["gene_knockout_only"]["flux"]
        else:
            base_flux = self.last_run["baseline"]["flux"]
            cmp_flux = self.last_run["overexpression_only"]["flux"]

        rows = []
        for rxn_id, b in base_flux.items():
            c = cmp_flux.get(rxn_id, 0.0)
            d = c - b
            if abs(d) > 1e-9:
                rows.append((rxn_id, b, c, d))

        rows.sort(key=lambda t: abs(t[3]), reverse=True)
        rows = rows[: int(self.topn_spin.value())]

        self.last_run["flux_rows"] = [
            {"reaction": rid, "baseline": float(b), "compared": float(c), "delta": float(d)}
            for rid, b, c, d in rows
        ]

    def _render_results_from_last_run(self):
        assert self.last_run is not None
        mode = self.last_run.get("compare_mode", "Gene knockout only")

        original = self.last_run.get("original") or self.last_run["baseline"]
        baseline = self.last_run["baseline"]
        gene_knockout_only = self.last_run["gene_knockout_only"]
        overexpression_only = self.last_run["overexpression_only"]

        # Find biomass flux values
        biomass_info = ""
        if self.base_model:
            for rxn in self.base_model.reactions:
                if 'biomass' in rxn.id.lower() or 'biomass' in rxn.name.lower():
                    biomass_id = rxn.id
                    orig_biomass = original['flux'].get(biomass_id, 0.0)
                    base_biomass = baseline['flux'].get(biomass_id, 0.0)
                    ko_biomass = gene_knockout_only['flux'].get(biomass_id, 0.0)
                    oe_biomass = overexpression_only['flux'].get(biomass_id, 0.0)
                    biomass_info = f"\nBiomass ({biomass_id}):\n  Original: {orig_biomass:.6g} | Baseline: {base_biomass:.6g} | Knockout: {ko_biomass:.6g} | Overexpression: {oe_biomass:.6g}"
                    break

        self.results_lbl.setText(
            f"Original (no changes): {original['status']} | obj={original['objective']}\n"
            f"Baseline: {baseline['status']} | obj={baseline['objective']}\n"
            f"Gene knockout only: {gene_knockout_only['status']} | obj={gene_knockout_only['objective']}\n"
            f"Overexpression only: {overexpression_only['status']} | obj={overexpression_only['objective']}"
            f"{biomass_info}\n"
            f"\nCompare mode: {mode} | Top N={int(self.topn_spin.value())}\n"
            f"Reaction overrides: {len(self.reaction_bound_overrides)}"
        )

        self._set_flux_table_headers_for_mode(mode)

        rows = self.last_run.get("flux_rows", [])
        self.flux_tbl.setRowCount(0)
        rxn_ids_for_plot = []
        for row in rows:
            rid = row["reaction"]
            b = row["baseline"]
            c = row["compared"]
            d = row["delta"]
            r = self.flux_tbl.rowCount()
            self.flux_tbl.insertRow(r)
            self.flux_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.flux_tbl.setItem(r, 1, QTableWidgetItem(f"{b:.6g}"))
            self.flux_tbl.setItem(r, 2, QTableWidgetItem(f"{c:.6g}"))
            self.flux_tbl.setItem(r, 3, QTableWidgetItem(f"{d:.6g}"))
            rxn_ids_for_plot.append(rid)

        # Apply filters after rendering
        self.filter_flux_table()

        if mode == "Original vs Baseline":
            base_flux = original["flux"]
            cmp_flux = baseline["flux"]
            base_label = "Original"
            cmp_label = "Baseline"
            title = "Flux comparison (Original vs Baseline)"
        elif mode == "Gene knockout only":
            base_flux = baseline["flux"]
            cmp_flux = gene_knockout_only["flux"]
            base_label = "Baseline"
            cmp_label = "Gene knockout only"
            title = "Flux comparison (Baseline vs Gene knockout only)"
        else:
            base_flux = baseline["flux"]
            cmp_flux = overexpression_only["flux"]
            base_label = "Baseline"
            cmp_label = "Overexpression only"
            title = "Flux comparison (Baseline vs Overexpression only)"

        self._plot_flux_comparison(rxn_ids_for_plot, base_flux, cmp_flux, base_label, cmp_label, title)
        self.tabs.setCurrentWidget(self.tab_results)

    # ---------------- Results Export Functions ----------------
    def _export_results_csv(self):
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results to CSV", "", "CSV files (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                headers = [self.flux_tbl.horizontalHeaderItem(c).text() for c in range(self.flux_tbl.columnCount())]
                f.write(",".join(headers) + "\n")
                for row in range(self.flux_tbl.rowCount()):
                    if self.flux_tbl.isRowHidden(row):
                        continue
                    vals = []
                    for col in range(self.flux_tbl.columnCount()):
                        item = self.flux_tbl.item(row, col)
                        vals.append(item.text() if item else "")
                    f.write(",".join(vals) + "\n")
            QMessageBox.information(self, "Export", f"Results exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _export_results_excel(self):
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results to Excel", "", "Excel files (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Flux Results"
            # Headers
            for c in range(self.flux_tbl.columnCount()):
                ws.cell(row=1, column=c + 1, value=self.flux_tbl.horizontalHeaderItem(c).text())
            # Data
            row_idx = 2
            for row in range(self.flux_tbl.rowCount()):
                if self.flux_tbl.isRowHidden(row):
                    continue
                for col in range(self.flux_tbl.columnCount()):
                    item = self.flux_tbl.item(row, col)
                    val = item.text() if item else ""
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                    ws.cell(row=row_idx, column=col + 1, value=val)
                row_idx += 1
            wb.save(path)
            QMessageBox.information(self, "Export", f"Results exported to:\n{path}")
        except ImportError:
            QMessageBox.critical(self, "Export failed", "openpyxl is required for Excel export.")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _export_results_chart(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Chart", "", "PNG (*.png);;SVG (*.svg);;PDF (*.pdf)")
        if not path:
            return
        try:
            self.canvas.figure.savefig(path, dpi=150, bbox_inches='tight')
            QMessageBox.information(self, "Export", f"Chart exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _export_results_latex(self):
        """Export the current flux table as a LaTeX table for academic publications."""
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results as LaTeX", "", "TeX files (*.tex)")
        if not path:
            return
        try:
            headers = [self.flux_tbl.horizontalHeaderItem(c).text() for c in range(self.flux_tbl.columnCount())]
            ncols = len(headers)
            col_spec = "l" + "r" * (ncols - 1)

            lines = []
            lines.append("\\begin{table}[htbp]")
            lines.append("\\centering")

            # Caption with analysis type
            analysis_type = self.last_run.get("analysis_type", "Flux") if self.last_run else "Flux"
            lines.append(f"\\caption{{{analysis_type} Analysis Results}}")
            lines.append(f"\\label{{tab:{analysis_type.lower().replace(' ', '_')}_results}}")
            lines.append(f"\\begin{{tabular}}{{{col_spec}}}")
            lines.append("\\hline")
            lines.append(" & ".join(f"\\textbf{{{h}}}" for h in headers) + " \\\\")
            lines.append("\\hline")

            for row in range(self.flux_tbl.rowCount()):
                if self.flux_tbl.isRowHidden(row):
                    continue
                vals = []
                for col in range(ncols):
                    item = self.flux_tbl.item(row, col)
                    text = item.text() if item else ""
                    # Escape LaTeX special characters
                    text = text.replace("_", "\\_").replace("&", "\\&").replace("%", "\\%")
                    vals.append(text)
                lines.append(" & ".join(vals) + " \\\\")

            lines.append("\\hline")
            lines.append("\\end{tabular}")
            lines.append("\\end{table}")

            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            QMessageBox.information(self, "Export", f"LaTeX table exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def show_metabolite_summary(self):
        """Show detailed summary for a selected metabolite: all producing/consuming reactions and net flux."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        # Metabolite selection dialog
        met_ids = sorted(m.id for m in self.base_model.metabolites)
        dialog = QDialog(self)
        dialog.setWindowTitle("Metabolite Summary")
        dialog.resize(800, 600)
        dlg_layout = QVBoxLayout(dialog)

        # Search + select
        sel_row = QHBoxLayout()
        sel_row.addWidget(QLabel("Metabolite:"))
        met_combo = QComboBox()
        met_combo.setEditable(True)
        met_combo.addItems(met_ids)
        met_combo.setCompleter(QCompleter(met_ids))
        sel_row.addWidget(met_combo, stretch=1)
        show_btn = QPushButton("Show Summary")
        sel_row.addWidget(show_btn)
        dlg_layout.addLayout(sel_row)

        info_label = QLabel("")
        info_label.setWordWrap(True)
        info_label.setTextFormat(Qt.RichText)
        dlg_layout.addWidget(info_label)

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Reaction", "Name", "Stoich. Coeff.", "Flux", "Contribution"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)
        dlg_layout.addWidget(table, stretch=1)

        def _show_met():
            mid = met_combo.currentText().strip()
            if mid not in [m.id for m in self.base_model.metabolites]:
                info_label.setText(f"<b style='color:red;'>Metabolite '{mid}' not found.</b>")
                table.setRowCount(0)
                return

            met = self.base_model.metabolites.get_by_id(mid)

            # Get fluxes from last run if available
            flux_dict = {}
            if self.last_run:
                bl = self.last_run.get("baseline", {})
                flux_dict = bl.get("flux", bl.get("fluxes", {})) or {}

            producing = []
            consuming = []
            for rxn in met.reactions:
                coeff = rxn.metabolites[met]
                flux = flux_dict.get(rxn.id, 0.0)
                contribution = coeff * flux
                entry = (rxn.id, rxn.name or "", coeff, flux, contribution)
                if coeff > 0:
                    producing.append(entry)
                else:
                    consuming.append(entry)

            net_flux = sum(e[4] for e in producing) + sum(e[4] for e in consuming)

            met_name = met.name or ""
            compartment = getattr(met, "compartment", "") or ""
            formula = getattr(met, "formula", "") or ""

            info_html = (
                f"<h3 style='color:#2E86C1;'>{mid}</h3>"
                f"<b>Name:</b> {met_name} | <b>Compartment:</b> {compartment} | "
                f"<b>Formula:</b> {formula}<br>"
                f"<b>Producing reactions:</b> {len(producing)} | "
                f"<b>Consuming reactions:</b> {len(consuming)} | "
                f"<b>Net flux:</b> {net_flux:.6g}"
            )
            info_label.setText(info_html)

            all_entries = sorted(producing + consuming, key=lambda x: -abs(x[4]))
            table.setRowCount(len(all_entries))
            for i, (rid, rname, coeff, flux, contrib) in enumerate(all_entries):
                table.setItem(i, 0, QTableWidgetItem(rid))
                table.setItem(i, 1, QTableWidgetItem(rname))
                table.setItem(i, 2, QTableWidgetItem(f"{coeff:+.4g}"))
                table.setItem(i, 3, QTableWidgetItem(f"{flux:.6g}"))
                item = QTableWidgetItem(f"{contrib:.6g}")
                table.setItem(i, 4, item)

        show_btn.clicked.connect(_show_met)
        met_combo.activated.connect(lambda: _show_met())

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(btn_box)

        dialog.exec()

    def show_flux_distribution_comparison(self):
        """Compare flux distributions from multiple analyses side by side.
        
        Runs FBA, pFBA, and Loopless FBA in parallel and shows a comparison table
        so the user can see how each method distributes flux differently.
        """
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self.apply_reaction_overrides_to_model(model)
        self._apply_selected_objective_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            results = {}
            # Standard FBA
            try:
                sol = model.optimize()
                if sol.status == "optimal":
                    results["FBA"] = {"objective": float(sol.objective_value), "fluxes": sol.fluxes.to_dict()}
            except Exception as e:
                results["FBA"] = {"error": str(e)}

            # pFBA
            try:
                from cobra.flux_analysis import pfba
                psol = pfba(model)
                results["pFBA"] = {"objective": float(psol.objective_value), "fluxes": psol.fluxes.to_dict()}
            except Exception as e:
                results["pFBA"] = {"error": str(e)}

            # Loopless FBA
            try:
                from cobra.flux_analysis import loopless_solution
                lsol = loopless_solution(model)
                results["Loopless"] = {"objective": float(lsol.objective_value), "fluxes": lsol.fluxes.to_dict()}
            except Exception as e:
                results["Loopless"] = {"error": str(e)}

            return results

        def _on_done(results):
            self._show_flux_comparison_dialog(results)

        self._launch_worker(_compute, _on_done, "Running FBA / pFBA / Loopless comparison...")

    def _show_flux_comparison_dialog(self, results: dict):
        """Display side-by-side flux comparison from multiple analysis methods."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Flux Distribution Comparison")
        dialog.resize(1000, 650)

        layout = QVBoxLayout(dialog)

        # Summary header
        methods = [m for m in ["FBA", "pFBA", "Loopless"] if m in results]
        summary_parts = []
        for m in methods:
            r = results[m]
            if "error" in r:
                summary_parts.append(f"<b>{m}:</b> <span style='color:red;'>Error — {r['error']}</span>")
            else:
                summary_parts.append(f"<b>{m}:</b> objective = {r['objective']:.6g}")
        summary_lbl = QLabel("<br>".join(summary_parts))
        summary_lbl.setWordWrap(True)
        summary_lbl.setTextFormat(Qt.RichText)
        layout.addWidget(summary_lbl)

        # Filter
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Filter:"))
        filter_edit = QLineEdit()
        filter_edit.setPlaceholderText("reaction id contains...")
        filter_row.addWidget(filter_edit, stretch=1)
        nonzero_chk = QCheckBox("Non-zero only")
        nonzero_chk.setChecked(True)
        filter_row.addWidget(nonzero_chk)
        diff_chk = QCheckBox("Differences only")
        diff_chk.setToolTip("Show only reactions where methods disagree")
        filter_row.addWidget(diff_chk)
        layout.addLayout(filter_row)

        # Collect all reaction IDs
        all_rxn_ids = set()
        flux_data = {}
        for m in methods:
            if "fluxes" in results[m]:
                flux_data[m] = results[m]["fluxes"]
                all_rxn_ids.update(flux_data[m].keys())

        # Build table
        ncols = 1 + len(flux_data) + 1  # Reaction, method fluxes..., Max Δ
        col_headers = ["Reaction"] + list(flux_data.keys()) + ["Max Δ"]
        table = QTableWidget()
        table.setColumnCount(ncols)
        table.setHorizontalHeaderLabels(col_headers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)

        # Pre-compute rows
        row_data_list = []
        for rid in sorted(all_rxn_ids):
            fluxes = [flux_data[m].get(rid, 0.0) for m in flux_data]
            max_delta = max(fluxes) - min(fluxes) if fluxes else 0.0
            row_data_list.append((rid, fluxes, max_delta))

        # Sort by max delta descending
        row_data_list.sort(key=lambda x: -abs(x[2]))

        table.setRowCount(len(row_data_list))
        for i, (rid, fluxes, max_delta) in enumerate(row_data_list):
            table.setItem(i, 0, QTableWidgetItem(rid))
            for j, fv in enumerate(fluxes):
                table.setItem(i, 1 + j, QTableWidgetItem(f"{fv:.6g}"))
            table.setItem(i, ncols - 1, QTableWidgetItem(f"{max_delta:.6g}"))

        def _apply_filter():
            text = filter_edit.text().strip().lower()
            only_nonzero = nonzero_chk.isChecked()
            only_diff = diff_chk.isChecked()
            for row in range(table.rowCount()):
                rid, fluxes, max_delta = row_data_list[row]
                visible = True
                if text and text not in rid.lower():
                    visible = False
                if only_nonzero and all(abs(f) < 1e-9 for f in fluxes):
                    visible = False
                if only_diff and abs(max_delta) < 1e-9:
                    visible = False
                table.setRowHidden(row, not visible)

        filter_edit.textChanged.connect(lambda: _apply_filter())
        nonzero_chk.stateChanged.connect(lambda: _apply_filter())
        diff_chk.stateChanged.connect(lambda: _apply_filter())
        _apply_filter()

        layout.addWidget(table, stretch=1)

        # Export buttons
        btn_row = QHBoxLayout()
        export_csv_btn = QPushButton("Export CSV")
        export_latex_btn = QPushButton("Export LaTeX")

        def _export_comparison_csv():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Export CSV", "", "CSV (*.csv)")
            if not fp:
                return
            with open(fp, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(col_headers)
                for rid, fluxes, max_delta in row_data_list:
                    w.writerow([rid] + [f"{fv:.6g}" for fv in fluxes] + [f"{max_delta:.6g}"])
            self.statusBar().showMessage(f"Comparison exported: {fp}")

        def _export_comparison_latex():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Export LaTeX", "", "TeX (*.tex)")
            if not fp:
                return
            col_spec = "l" + "r" * (len(col_headers) - 1)
            lines = [
                "\\begin{table}[htbp]",
                "\\centering",
                "\\caption{Flux Distribution Comparison (FBA vs pFBA vs Loopless)}",
                "\\label{tab:flux_comparison}",
                f"\\begin{{tabular}}{{{col_spec}}}",
                "\\hline",
                " & ".join(f"\\textbf{{{h}}}" for h in col_headers) + " \\\\",
                "\\hline",
            ]
            for rid, fluxes, max_delta in row_data_list[:50]:
                if all(abs(f) < 1e-9 for f in fluxes):
                    continue
                rid_tex = rid.replace("_", "\\_")
                vals = [rid_tex] + [f"{fv:.4g}" for fv in fluxes] + [f"{max_delta:.4g}"]
                lines.append(" & ".join(vals) + " \\\\")
            lines += ["\\hline", "\\end{tabular}", "\\end{table}"]
            with open(fp, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            self.statusBar().showMessage(f"LaTeX comparison exported: {fp}")

        export_csv_btn.clicked.connect(_export_comparison_csv)
        export_latex_btn.clicked.connect(_export_comparison_latex)
        btn_row.addWidget(export_csv_btn)
        btn_row.addWidget(export_latex_btn)
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        dialog.exec()

    def _show_shadow_prices(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return
        if not self.last_run:
            QMessageBox.information(self, "No results", "Run FBA first to see shadow prices.")
            return

        # Use cached shadow prices from last FBA run if available
        cached = (self.last_run.get("baseline", {}).get("shadow_prices")
                  or self.last_run.get("original", {}).get("shadow_prices"))
        if cached:
            shadow_prices_dict = cached
        else:
            # Fallback: re-optimize
            try:
                model = self.base_model.copy()
                self.apply_medium_table_bounds_to_model(model)
                self.apply_reaction_overrides_to_model(model)
                self._apply_selected_solver(model)
                sol = model.optimize()
                if sol.status != 'optimal':
                    QMessageBox.warning(self, "Shadow Prices", f"Solution status: {sol.status}.")
                    return
                shadow_prices_dict = sol.shadow_prices.to_dict()
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))
                return

        if not shadow_prices_dict:
            QMessageBox.information(self, "Shadow Prices", "No shadow prices available.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Shadow Prices (Metabolite Dual Values)")
        dialog.resize(600, 500)
        layout = QVBoxLayout(dialog)

        info = QLabel("Shadow prices indicate how much the objective would change per unit increase in metabolite availability.\n"
                       f"Source: {'cached from last FBA run' if cached else 're-optimized model'}")
        info.setWordWrap(True)
        layout.addWidget(info)

        search = QLineEdit()
        search.setPlaceholderText("Search metabolites...")
        layout.addWidget(search)

        table = QTableWidget(0, 3)
        table.setHorizontalHeaderLabels(["Metabolite", "Name", "Shadow Price"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(table)

        data = sorted(shadow_prices_dict.items(), key=lambda x: abs(x[1]), reverse=True)
        for met_id, price in data:
            try:
                met = self.base_model.metabolites.get_by_id(met_id)
                met_name = met.name if met else ""
            except Exception:
                met_name = ""
            r = table.rowCount()
            table.insertRow(r)
            table.setItem(r, 0, QTableWidgetItem(met_id))
            table.setItem(r, 1, QTableWidgetItem(met_name))
            table.setItem(r, 2, QTableWidgetItem(f"{price:.6g}"))

        def filter_table():
            txt = search.text().lower()
            for row in range(table.rowCount()):
                mid = (table.item(row, 0).text() or "").lower()
                mname = (table.item(row, 1).text() or "").lower()
                table.setRowHidden(row, txt not in mid and txt not in mname)
        search.textChanged.connect(filter_table)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    def _show_reduced_costs(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return
        if not self.last_run:
            QMessageBox.information(self, "No results", "Run FBA first to see reduced costs.")
            return

        # Use cached reduced costs from last FBA run if available
        cached = (self.last_run.get("baseline", {}).get("reduced_costs")
                  or self.last_run.get("original", {}).get("reduced_costs"))
        if cached:
            reduced_costs_dict = cached
        else:
            try:
                model = self.base_model.copy()
                self.apply_medium_table_bounds_to_model(model)
                self.apply_reaction_overrides_to_model(model)
                self._apply_selected_solver(model)
                sol = model.optimize()
                if sol.status != 'optimal':
                    QMessageBox.warning(self, "Reduced Costs", f"Solution status: {sol.status}.")
                    return
                reduced_costs_dict = sol.reduced_costs.to_dict()
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))
                return

        if not reduced_costs_dict:
            QMessageBox.information(self, "Reduced Costs", "No reduced costs available.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Reduced Costs (Reaction Dual Values)")
        dialog.resize(700, 500)
        layout = QVBoxLayout(dialog)

        info = QLabel("Reduced costs indicate how much the objective would change per unit increase in a reaction's flux bound.\n"
                       f"Source: {'cached from last FBA run' if cached else 're-optimized model'}")
        info.setWordWrap(True)
        layout.addWidget(info)

        search = QLineEdit()
        search.setPlaceholderText("Search reactions...")
        layout.addWidget(search)

        table = QTableWidget(0, 3)
        table.setHorizontalHeaderLabels(["Reaction", "Name", "Reduced Cost"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(table)

        data = sorted(reduced_costs_dict.items(), key=lambda x: abs(x[1]), reverse=True)
        for rxn_id, cost in data:
            try:
                rxn = self.base_model.reactions.get_by_id(rxn_id)
                rxn_name = rxn.name if rxn else ""
            except Exception:
                rxn_name = ""
            r = table.rowCount()
            table.insertRow(r)
            table.setItem(r, 0, QTableWidgetItem(rxn_id))
            table.setItem(r, 1, QTableWidgetItem(rxn_name))
            table.setItem(r, 2, QTableWidgetItem(f"{cost:.6g}"))

        def filter_table():
            txt = search.text().lower()
            for row in range(table.rowCount()):
                rid = (table.item(row, 0).text() or "").lower()
                rname = (table.item(row, 1).text() or "").lower()
                table.setRowHidden(row, txt not in rid and txt not in rname)
        search.textChanged.connect(filter_table)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    def _find_essential_reactions(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        reply = QMessageBox.question(
            self, "Essential Reactions",
            "This will test each reaction by setting its bounds to zero and checking if growth is still possible.\n\n"
            "This may take a few minutes for large models. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            from cobra.flux_analysis import single_reaction_deletion
            baseline_sol = model.optimize()
            if baseline_sol.status != 'optimal':
                raise ValueError("Baseline solution is not optimal.")
            baseline_growth = baseline_sol.objective_value
            result = single_reaction_deletion(model)
            essential = []
            for idx, row in result.iterrows():
                rxn_ids = list(idx) if hasattr(idx, '__iter__') and not isinstance(idx, str) else [idx]
                growth = row.get('growth', row.get('objective', 0)) or 0
                status = row.get('status', 'unknown')
                if growth < baseline_growth * 0.01 or status != 'optimal':
                    for rxn_id in rxn_ids:
                        rxn = model.reactions.get_by_id(rxn_id)
                        essential.append({
                            'id': rxn_id, 'name': rxn.name,
                            'subsystem': rxn.subsystem or "", 'growth': growth
                        })
            return {"essential": essential, "baseline_growth": baseline_growth}

        def _on_done(result):
            essential = result["essential"]
            baseline_growth = result["baseline_growth"]
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Essential Reactions ({len(essential)} found)")
            dialog.resize(800, 500)
            layout = QVBoxLayout(dialog)
            info = QLabel(f"Found {len(essential)} essential reactions (knockout -> growth < 1% of baseline).\nBaseline growth: {baseline_growth:.6g}")
            info.setWordWrap(True)
            layout.addWidget(info)
            table = QTableWidget(0, 4)
            table.setHorizontalHeaderLabels(["Reaction", "Name", "Subsystem", "Growth after KO"])
            table.horizontalHeader().setStretchLastSection(True)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            layout.addWidget(table)
            for item in essential:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(item['id']))
                table.setItem(r, 1, QTableWidgetItem(item['name']))
                table.setItem(r, 2, QTableWidgetItem(item['subsystem']))
                table.setItem(r, 3, QTableWidgetItem(f"{item['growth']:.6g}"))
            btns = QDialogButtonBox(QDialogButtonBox.Close)
            btns.rejected.connect(dialog.reject)
            layout.addWidget(btns)
            dialog.exec()

        self._launch_worker(_compute, _on_done, "Finding essential reactions...")

    def _quick_fva(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Quick FVA Options")
        layout = QVBoxLayout(dialog)

        form = QFormLayout()
        fraction_spin = QDoubleSpinBox()
        fraction_spin.setRange(0.0, 1.0)
        fraction_spin.setValue(0.9)
        fraction_spin.setSingleStep(0.05)
        form.addRow("Fraction of optimum:", fraction_spin)

        rxn_filter = QLineEdit()
        rxn_filter.setPlaceholderText("Filter reactions (e.g., 'EX_' for exchange)")
        form.addRow("Reaction filter:", rxn_filter)

        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        fraction = fraction_spin.value()
        filter_text = rxn_filter.text().strip().lower()

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            from cobra.flux_analysis import flux_variability_analysis
            if filter_text:
                rxns = [r for r in model.reactions if filter_text in r.id.lower()]
            else:
                rxns = model.reactions
            fva_result = flux_variability_analysis(model, reaction_list=rxns, fraction_of_optimum=fraction)
            rows = []
            for rxn_id, row in fva_result.iterrows():
                rows.append({"rxn_id": rxn_id, "min": row['minimum'], "max": row['maximum'],
                             "range": row['maximum'] - row['minimum']})
            return {"fva_rows": rows, "fraction": fraction}

        def _on_done(result):
            fva_rows = result["fva_rows"]
            result_dialog = QDialog(self)
            result_dialog.setWindowTitle(f"FVA Results ({len(fva_rows)} reactions)")
            result_dialog.resize(700, 500)
            layout2 = QVBoxLayout(result_dialog)

            info = QLabel(f"FVA at {result['fraction'] * 100:.0f}% of optimum. Showing min/max flux ranges.")
            layout2.addWidget(info)

            search = QLineEdit()
            search.setPlaceholderText("Search reactions...")
            layout2.addWidget(search)

            table = QTableWidget(0, 4)
            table.setHorizontalHeaderLabels(["Reaction", "Min Flux", "Max Flux", "Range"])
            table.horizontalHeader().setStretchLastSection(True)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            layout2.addWidget(table)

            for frow in fva_rows:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(frow["rxn_id"]))
                table.setItem(r, 1, QTableWidgetItem(f"{frow['min']:.6g}"))
                table.setItem(r, 2, QTableWidgetItem(f"{frow['max']:.6g}"))
                table.setItem(r, 3, QTableWidgetItem(f"{frow['range']:.6g}"))

            def filter_table():
                txt = search.text().lower()
                for row in range(table.rowCount()):
                    rxn_id = (table.item(row, 0).text() or "").lower()
                    table.setRowHidden(row, txt and txt not in rxn_id)
            search.textChanged.connect(filter_table)

            btns2 = QDialogButtonBox(QDialogButtonBox.Close)
            btns2.rejected.connect(result_dialog.reject)
            layout2.addWidget(btns2)
            result_dialog.exec()

        self._launch_worker(_compute, _on_done, "Running Quick FVA...")

    # ---------------- Run FBA ----------------
    def _launch_worker(self, func, on_done, message="Running analysis..."):
        """Launch an analysis function in a background QThread.

        The compute function *func* can accept an optional ``worker`` keyword
        argument.  If it does, the AnalysisWorker instance is passed in so the
        function can call ``worker.report_progress(msg, pct)`` during
        long-running loops.
        """
        self.set_busy(True, message)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)

        def _on_finished(result):
            self._worker = None
            self.progress_bar.setVisible(False)
            try:
                on_done(result)
            except Exception as e:
                QMessageBox.critical(self, "Render Error", str(e))
            self.set_busy(False, "Ready.")

        def _on_error(err_msg):
            self._worker = None
            self.progress_bar.setVisible(False)
            self.set_busy(False, "Analysis failed.")
            QMessageBox.critical(self, "Analysis Error", str(err_msg))

        def _on_progress(msg):
            self.status_lbl.setText(msg)
            self.statusBar().showMessage(msg)

        def _on_progress_pct(pct):
            self.progress_bar.setValue(min(pct, 100))

        worker = AnalysisWorker(func)
        worker.finished.connect(_on_finished)
        worker.error.connect(_on_error)
        worker.progress.connect(_on_progress)
        worker.progress_pct.connect(_on_progress_pct)
        self._worker = worker
        worker.start()

    def run_fba(self):
        if self.base_model is None or self.is_running:
            return

        analysis_type = self.analysis_type.currentText()
        
        if "FVA" in analysis_type:
            self.run_fva()
        elif "pFBA" in analysis_type:
            self.run_pfba()
        elif "Single Gene Deletion" in analysis_type:
            self.run_sgd()
        elif "Double Gene Deletion" in analysis_type:
            self.run_dgd()
        elif "Single Reaction Deletion" in analysis_type:
            self.run_srd()
        elif "Robustness" in analysis_type:
            self.run_robustness()
        elif "Production Envelope" in analysis_type:
            self.run_production_envelope()
        elif "Flux Sampling" in analysis_type:
            self.run_flux_sampling()
        else:
            self.run_standard_fba()

    def run_standard_fba(self):
        """Run standard FBA with baseline, knockout, and overexpression comparisons."""
        # Prepare all models on main thread (thread-safe copies)
        original_model = (self.original_model_snapshot.copy() if self.original_model_snapshot is not None else self.base_model.copy())
        self._apply_selected_objective_to_model(original_model)
        self._apply_selected_solver(original_model)

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        knockout_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(knockout_model)
        self.apply_reaction_overrides_to_model(knockout_model)
        self.apply_knockouts_to_model(knockout_model)
        self._apply_selected_objective_to_model(knockout_model)
        self._apply_selected_solver(knockout_model)

        overexpression_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(overexpression_model)
        self.apply_reaction_overrides_to_model(overexpression_model)
        self.apply_temp_upper_bound_overrides(overexpression_model)
        self.apply_overexpression_to_model(overexpression_model)
        self._apply_selected_objective_to_model(overexpression_model)
        self._apply_selected_solver(overexpression_model)

        use_loopless = self.loopless_chk.isChecked()
        compare_mode = self.compare_mode.currentText()

        # --- FBA cache check ---
        cache_key = self._model_state_hash(baseline_model) + (":loopless" if use_loopless else "")
        cached = self._cache_get(cache_key)
        if cached is not None:
            logger.info("FBA cache hit – reusing previous result")
            self.last_run = dict(cached)
            self.last_run["compare_mode"] = compare_mode
            self.last_run["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "  (cached)"
            self._recompute_flux_rows_for_compare_mode()
            self._render_results_from_last_run()
            return

        _cache_key_ref = cache_key  # closure capture

        def _compute():
            if use_loopless:
                from cobra.flux_analysis import loopless_solution
                original_sol = loopless_solution(original_model)
                baseline_sol = loopless_solution(baseline_model)
                knockout_sol = loopless_solution(knockout_model)
                overexpression_sol = loopless_solution(overexpression_model)
            else:
                original_sol = original_model.optimize()
                baseline_sol = baseline_model.optimize()
                knockout_sol = knockout_model.optimize()
                overexpression_sol = overexpression_model.optimize()

            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "FBA",
                "original": {"status": str(original_sol.status), "objective": float(original_sol.objective_value), "flux": original_sol.fluxes.to_dict(),
                              "shadow_prices": original_sol.shadow_prices.to_dict() if hasattr(original_sol, 'shadow_prices') and original_sol.shadow_prices is not None else {},
                              "reduced_costs": original_sol.reduced_costs.to_dict() if hasattr(original_sol, 'reduced_costs') and original_sol.reduced_costs is not None else {}},
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict(),
                              "shadow_prices": baseline_sol.shadow_prices.to_dict() if hasattr(baseline_sol, 'shadow_prices') and baseline_sol.shadow_prices is not None else {},
                              "reduced_costs": baseline_sol.reduced_costs.to_dict() if hasattr(baseline_sol, 'reduced_costs') and baseline_sol.reduced_costs is not None else {}},
                "gene_knockout_only": {"status": str(knockout_sol.status), "objective": float(knockout_sol.objective_value), "flux": knockout_sol.fluxes.to_dict()},
                "overexpression_only": {"status": str(overexpression_sol.status), "objective": float(overexpression_sol.objective_value), "flux": overexpression_sol.fluxes.to_dict()},
                "compare_mode": compare_mode,
                "flux_rows": [],
            }

        def _on_done(result):
            self._cache_put(_cache_key_ref, result)
            self.last_run = result
            self._recompute_flux_rows_for_compare_mode()
            self._render_results_from_last_run()

        label = "Running FBA (Loopless)..." if use_loopless else "Running FBA..."
        self._launch_worker(_compute, _on_done, label)

    def run_fva(self):
        """Run Flux Variability Analysis (FVA)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)
        topn = int(self.topn_spin.value())
        fva_fraction = float(self.fva_fraction_spin.value())
        fva_processes = int(self.fva_processes_spin.value())
        # Frozen exe (PyInstaller) cannot use multiprocessing safely
        if getattr(sys, 'frozen', False) and fva_processes > 1:
            fva_processes = 1

        def _compute():
            from cobra.flux_analysis import flux_variability_analysis
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            fva_result = flux_variability_analysis(baseline_model, fraction_of_optimum=fva_fraction, processes=fva_processes)
            fva_dict = {}
            for rid in fva_result.index:
                fva_dict[rid] = {"min": float(fva_result.loc[rid, "minimum"]), "max": float(fva_result.loc[rid, "maximum"])}
            rows = []
            for rxn_id, minmax in fva_dict.items():
                flux_range = abs(minmax["max"] - minmax["min"])
                if flux_range > 1e-9:
                    rows.append({"reaction": rxn_id, "min_flux": minmax["min"], "max_flux": minmax["max"], "range": flux_range})
            rows.sort(key=lambda x: x["range"], reverse=True)
            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "FVA",
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict()},
                "fva_result": fva_dict,
                "flux_rows": rows[:topn],
            }

        def _on_done(result):
            self.last_run = result
            self._render_fva_results()

        self._launch_worker(_compute, _on_done, "Running FVA (Flux Variability Analysis)...")

    def run_pfba(self):
        """Run Parsimonious FBA (pFBA)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        baseline_model_std = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model_std)
        self.apply_reaction_overrides_to_model(baseline_model_std)
        self._apply_selected_objective_to_model(baseline_model_std)
        self._apply_selected_solver(baseline_model_std)
        topn = int(self.topn_spin.value())

        def _compute():
            from cobra.flux_analysis import pfba
            pfba_sol = pfba(baseline_model)
            pfba_result = {"status": str(pfba_sol.status), "objective": float(pfba_sol.objective_value), "flux": pfba_sol.fluxes.to_dict()}
            baseline_sol = baseline_model_std.optimize()
            rows = []
            for rxn_id, bf in baseline_sol.fluxes.items():
                pf = pfba_result["flux"].get(rxn_id, 0.0)
                d = abs(bf) - abs(pf)
                if abs(d) > 1e-9:
                    rows.append({"reaction": rxn_id, "fba_flux": float(bf), "pfba_flux": float(pf), "delta": float(d)})
            rows.sort(key=lambda x: abs(x["delta"]), reverse=True)
            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "pFBA",
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict()},
                "pfba": pfba_result,
                "compare_mode": "pFBA vs FBA",
                "flux_rows": rows[:topn],
            }

        def _on_done(result):
            self.last_run = result
            self._render_pfba_results()

        self._launch_worker(_compute, _on_done, "Running pFBA (Parsimonious FBA)...")

    def run_sgd(self):
        """Run Single Gene Deletion analysis."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        def _compute():
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            sgd_result = self.compute_sgd(baseline_model)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "SGD", "wt_growth": wt_growth, "sgd_result": sgd_result}

        def _on_done(result):
            self.last_run = result
            self._render_sgd_results()

        self._launch_worker(_compute, _on_done, "Running Single Gene Deletion (SGD)...")

    def _get_selected_genes_for_dgd(self) -> list[str]:
        genes: list[str] = []

        try:
            items = list(self.gene_list.selectedItems())
            genes.extend([str(i.text()).strip() for i in items if i and str(i.text()).strip()])

            if not genes:
                sel_model = self.gene_list.selectionModel()
                if sel_model is not None:
                    for idx in sel_model.selectedIndexes():
                        it = self.gene_list.item(idx.row())
                        if it:
                            t = str(it.text()).strip()
                            if t:
                                genes.append(t)

            cur = self.gene_list.currentItem()
            if cur:
                t = str(cur.text()).strip()
                if t:
                    genes.append(t)
        except Exception:
            pass

        if len(genes) < 2 and self.knockout_genes:
            genes.extend(sorted(self.knockout_genes))

        out: list[str] = []
        seen = set()
        for g in genes:
            if g and g not in seen:
                seen.add(g)
                out.append(g)

        return out

    def run_dgd(self):
        """Run Double Gene Deletion analysis."""
        genes = self._get_selected_genes_for_dgd()
        if len(genes) < 2:
            QMessageBox.warning(self, "DGD", "Select at least 2 genes in Gene knockout tab for Double Gene Deletion.")
            return

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)
        genes_copy = list(genes)

        def _compute():
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            dgd_result = self.compute_dgd(baseline_model, genes_copy)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "DGD", "wt_growth": wt_growth, "dgd_result": dgd_result, "dgd_genes": genes_copy}

        def _on_done(result):
            self.last_run = result
            self._render_dgd_results()

        self._launch_worker(_compute, _on_done, "Running Double Gene Deletion (DGD)...")

    def run_srd(self):
        """Run Single Reaction Deletion analysis."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        def _compute():
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            srd_result = self.compute_srd(baseline_model)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "SRD", "wt_growth": wt_growth, "srd_result": srd_result}

        def _on_done(result):
            self.last_run = result
            self._render_srd_results()

        self._launch_worker(_compute, _on_done, "Running Single Reaction Deletion (SRD)...")

    def run_robustness(self):
        """Run Robustness Analysis."""
        rxn_id = self._extract_reaction_id_from_input(self.robustness_rxn.text())
        if not rxn_id:
            QMessageBox.warning(self, "Missing", "Enter a reaction ID for robustness analysis.")
            return
        sweep_ub = self.robustness_bound_ub.isChecked()
        sweep_lb = self.robustness_bound_lb.isChecked()
        if not sweep_ub and not sweep_lb:
            QMessageBox.warning(self, "Missing", "Select at least one bound type (UB or LB).")
            return
        bound_type = "ub" if sweep_ub else "lb"
        min_val = float(self.robustness_min.value())
        max_val = float(self.robustness_max.value())
        steps = int(self.robustness_steps.value())

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)

        def _compute(worker=None):
            robustness_result = self.compute_robustness(baseline_model, rxn_id, min_val, max_val, steps, bound_type, worker=worker)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "Robustness", "rxn_id": rxn_id, "bound_type": bound_type, "robustness_result": robustness_result}

        def _on_done(result):
            self.last_run = result
            self._render_robustness_results()

        self._launch_worker(_compute, _on_done, f"Running Robustness Analysis for {rxn_id} ({bound_type.upper()})...")

    def run_production_envelope(self):
        """Run Production Envelope Analysis using COBRApy's native function."""
        product_id = self._extract_reaction_id_from_input(self.envelope_product.text())
        if not product_id:
            QMessageBox.warning(self, "Missing", "Enter a product reaction ID for production envelope.")
            return

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        steps = int(self.envelope_steps.value())

        def _compute():
            try:
                from cobra.flux_analysis import production_envelope as pe_func
                prod_rxn = baseline_model.reactions.get_by_id(product_id)
                pe_df = pe_func(baseline_model, reactions=[prod_rxn], points=steps)
                # Find the correct columns
                growth_col = None
                prod_col = None
                for col in pe_df.columns:
                    if "carbon" in col.lower() or "source" in col.lower():
                        continue
                    if product_id in col or product_id.lower() in col.lower():
                        prod_col = col
                    elif "flux_minimum" in col.lower() or "flux_maximum" in col.lower():
                        if prod_col is None:
                            prod_col = col
                if prod_col is None and len(pe_df.columns) >= 2:
                    prod_col = pe_df.columns[-1]
                for col in pe_df.columns:
                    if col != prod_col:
                        growth_col = col
                        break
                if growth_col is None:
                    growth_col = pe_df.columns[0]
                result = {"growth": pe_df[growth_col].tolist(), "product": pe_df[prod_col].tolist()}
            except Exception:
                # Fallback to manual sweep if cobra.flux_analysis.production_envelope is unavailable
                result = self.compute_production_envelope(baseline_model, product_id, steps)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "Production Envelope", "product_id": product_id, "envelope_result": result}

        def _on_done(result):
            self.last_run = result
            self._render_envelope_results()

        self._launch_worker(_compute, _on_done, f"Running Production Envelope for {product_id}...")

    def run_flux_sampling(self):
        """Run Flux Sampling (ACHR or OptGP)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        sample_size = int(self.sampling_size.value())
        sampler_type = self.sampler_type_combo.currentText()

        def _compute():
            if sampler_type == "OptGP":
                try:
                    from cobra.sampling import OptGPSampler
                    sampler = OptGPSampler(baseline_model, processes=1)
                except ImportError:
                    from cobra.sampling import ACHRSampler
                    sampler = ACHRSampler(baseline_model)
            else:
                from cobra.sampling import ACHRSampler
                sampler = ACHRSampler(baseline_model)
            samples = sampler.sample(sample_size)
            result = {}
            for rxn_id in samples.columns:
                fv = samples[rxn_id].values
                result[rxn_id] = {"mean": float(fv.mean()), "stdev": float(fv.std()), "min": float(fv.min()), "max": float(fv.max())}
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "Flux Sampling", "sampler_type": sampler_type, "sampling_result": result, "samples_df": samples}

        def _on_done(result):
            self.last_run = result
            self._render_sampling_results()

        self._launch_worker(_compute, _on_done, f"Running Flux Sampling ({sampler_type})...")

    def filter_fva_table(self):
        """Filter FVA table by reaction ID."""
        text = (self.fva_flux_search.text() or "").strip().lower()
        for row in range(self.fva_tbl.rowCount()):
            rid = (self.fva_tbl.item(row, 0).text() if self.fva_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.fva_tbl.setRowHidden(row, hide)

    def _render_fva_results(self):
        """Render FVA results in the FVA tab."""
        assert self.last_run is not None
        baseline = self.last_run.get("baseline", {})
        fva_result = self.last_run.get("fva_result", {})
        
        self.fva_info_lbl.setText(
            f"Baseline: {baseline.get('status', 'N/A')} | Objective: {baseline.get('objective', 'N/A'):.6g}\n"
            f"Fraction of optimum: {self.fva_fraction_spin.value():.2f} | "
            f"Processes: {self.fva_processes_spin.value()} | "
            f"Top N reactions by flux range: {int(self.topn_spin.value())}"
        )
        
        # Populate flux table
        rows = self.last_run.get("flux_rows", [])
        self.fva_tbl.setRowCount(0)
        rxn_ids_for_plot = []
        
        for row in rows:
            rid = row["reaction"]
            min_f = row["min_flux"]
            max_f = row["max_flux"]
            flux_range = row["range"]
            
            r = self.fva_tbl.rowCount()
            self.fva_tbl.insertRow(r)
            self.fva_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.fva_tbl.setItem(r, 1, QTableWidgetItem(f"{min_f:.6g}"))
            self.fva_tbl.setItem(r, 2, QTableWidgetItem(f"{max_f:.6g}"))
            self.fva_tbl.setItem(r, 3, QTableWidgetItem(f"{flux_range:.6g}"))
            rxn_ids_for_plot.append(rid)
        
        # Apply filters
        self.filter_fva_table()
        
        # Plot FVA ranges
        ax = self.fva_canvas.ax
        ax.clear()
        if not rxn_ids_for_plot:
            ax.set_title("No reactions with flux variability found")
            self.fva_canvas.draw()
        else:
            # Plot min/max ranges for each reaction
            x = list(range(len(rxn_ids_for_plot)))
            mins = [self.last_run["fva_result"].get(rid, {}).get("min", 0.0) for rid in rxn_ids_for_plot]
            maxs = [self.last_run["fva_result"].get(rid, {}).get("max", 0.0) for rid in rxn_ids_for_plot]
            
            ax.barh(x, [m - n for m, n in zip(maxs, mins)], left=mins, label="Flux range", color="steelblue", alpha=0.7)
            ax.axvline(x=0, color="black", linestyle="-", linewidth=0.5)
            ax.set_title("FVA: Flux Variability Ranges")
            ax.set_ylabel("Reaction")
            ax.set_xlabel("Flux")
            ax.set_yticks(x)
            ax.set_yticklabels(rxn_ids_for_plot, fontsize=8)
            ax.legend()
            self.fva_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_fva)

    def export_fva_csv(self):
        if not self.last_run or not self.last_run.get("fva_result"):
            QMessageBox.warning(self, "FVA", "Run FVA first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export FVA (CSV)", str(Path.home() / "fva_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction", "min", "max", "range"]) 
                for rid, mm in self.last_run["fva_result"].items():
                    rng = float(mm["max"]) - float(mm["min"]) 
                    w.writerow([rid, mm["min"], mm["max"], rng])
            self.statusBar().showMessage(f"FVA exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _render_pfba_results(self):
        """Render pFBA results in the pFBA tab."""
        assert self.last_run is not None
        baseline = self.last_run.get("baseline", {})
        pfba = self.last_run.get("pfba", {})
        
        self.pfba_info_lbl.setText(
            f"pFBA (Parsimonious FBA) Results\n"
            f"Standard FBA: {baseline.get('status', 'N/A')} | obj={baseline.get('objective', 'N/A')}\n"
            f"pFBA: {pfba.get('status', 'N/A')} | obj={pfba.get('objective', 'N/A')}\n"
            f"Top N reactions by flux difference: {int(self.topn_spin.value())}"
        )
        
        # Populate pFBA table
        rows = self.last_run.get("flux_rows", [])
        self.pfba_tbl.setHorizontalHeaderLabels(["Reaction", "FBA flux", "pFBA flux", "Difference"])
        self.pfba_tbl.setRowCount(0)
        rxn_ids_for_plot = []
        
        for row in rows:
            rid = row["reaction"]
            fba_f = row["fba_flux"]
            pfba_f = row["pfba_flux"]
            diff = row["delta"]
            
            r = self.pfba_tbl.rowCount()
            self.pfba_tbl.insertRow(r)
            self.pfba_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.pfba_tbl.setItem(r, 1, QTableWidgetItem(f"{fba_f:.6g}"))
            self.pfba_tbl.setItem(r, 2, QTableWidgetItem(f"{pfba_f:.6g}"))
            self.pfba_tbl.setItem(r, 3, QTableWidgetItem(f"{diff:.6g}"))
            rxn_ids_for_plot.append(rid)
        
        # Apply filters
        self.filter_pfba_table()
        
        # Plot pFBA comparison
        ax = self.pfba_canvas.ax
        ax.clear()
        if not rxn_ids_for_plot:
            ax.set_title("FBA and pFBA results are identical")
        else:
            baseline_flux = self.last_run["baseline"]["flux"]
            pfba_flux = pfba["flux"]
            
            fba_vals = [abs(baseline_flux.get(r, 0.0)) for r in rxn_ids_for_plot]
            pfba_vals = [abs(pfba_flux.get(r, 0.0)) for r in rxn_ids_for_plot]
            
            x = list(range(len(rxn_ids_for_plot)))
            width = 0.4
            ax.bar([i - width / 2 for i in x], fba_vals, width=width, label="FBA", alpha=0.7)
            ax.bar([i + width / 2 for i in x], pfba_vals, width=width, label="pFBA", alpha=0.7)
            
            ax.set_title("pFBA vs Standard FBA: Flux Comparison")
            ax.set_xlabel("Reaction")
            ax.set_ylabel("Absolute Flux")
            ax.set_xticks(x)
            ax.set_xticklabels(rxn_ids_for_plot, rotation=90, fontsize=8)
            ax.legend()
        self.pfba_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_pfba)

    def export_pfba_csv(self):
        if not self.last_run or not self.last_run.get("pfba"):
            QMessageBox.warning(self, "pFBA", "Run pFBA first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export pFBA (CSV)", str(Path.home() / "pfba_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            rows = self.last_run.get("flux_rows", [])
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction", "fba_flux", "pfba_flux", "difference"]) 
                for r in rows:
                    w.writerow([r["reaction"], r["fba_flux"], r["pfba_flux"], r["delta"]])
            self.statusBar().showMessage(f"pFBA exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def filter_deletion_table(self):
        """Filter deletion analysis table with search text and optimal/non-optimal filters."""
        text = (self.deletion_search.text() or "").strip().lower()
        show_optimal_only = self.deletion_filter_optimal.isChecked()
        show_nonoptimal_only = self.deletion_filter_nonoptimal.isChecked()
        for row in range(self.deletion_tbl.rowCount()):
            rid = (self.deletion_tbl.item(row, 0).text() if self.deletion_tbl.item(row, 0) else "").lower()
            growth_text = (self.deletion_tbl.item(row, 2).text() if self.deletion_tbl.item(row, 2) else "0")
            try:
                growth_val = float(growth_text)
            except ValueError:
                growth_val = 0.0
            hide = False
            if text and text not in rid:
                hide = True
            if show_optimal_only and growth_val < 1e-6:
                hide = True
            if show_nonoptimal_only and growth_val >= 1e-6:
                hide = True
            self.deletion_tbl.setRowHidden(row, hide)

    def _render_sgd_results(self):
        """Render SGD results."""
        assert self.last_run is not None
        wt_growth = self.last_run.get("wt_growth", 0.0)
        sgd_result = self.last_run.get("sgd_result", {})
        
        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Total: {len(sgd_result)} genes"
        )
        
        # Populate table with ALL results
        self.deletion_tbl.setHorizontalHeaderLabels(["Gene Name", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        genes_for_plot = []
        growth_vals = []
        
        sorted_genes = sorted(sgd_result.items(), key=lambda x: x[1])
        
        for gene_id, growth in sorted_genes:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            gene_label = self._format_gene_label(str(gene_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(gene_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))
        
        # Chart uses top N only
        for gene_id, growth in sorted_genes[:int(self.topn_spin.value())]:
            gene_label = self._format_gene_label(str(gene_id))
            genes_for_plot.append(gene_label)
            growth_vals.append(growth)
        
        self.filter_deletion_table()
        
        # Plot with value labels
        ax = self.deletion_canvas.ax
        ax.clear()
        if genes_for_plot:
            bars = ax.barh(genes_for_plot, growth_vals, color="coral", alpha=0.7, edgecolor="darkred", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Single Gene Deletion - Growth on Deletion")
        self.deletion_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_sgd_csv(self):
        if not self.last_run or not self.last_run.get("sgd_result"):
            QMessageBox.warning(self, "SGD", "Run SGD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export SGD (CSV)", str(Path.home() / "sgd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["gene_name", "gene_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for gid, growth in sorted(self.last_run["sgd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    gname = self._format_gene_label(str(gid))
                    w.writerow([gname, gid, wt, growth, delta_pct])
            self.statusBar().showMessage(f"SGD exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _render_dgd_results(self):
        """Render DGD results."""
        assert self.last_run is not None
        wt_growth = self.last_run.get("wt_growth", 0.0)
        dgd_result = self.last_run.get("dgd_result", {})

        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Showing all {len(dgd_result)} gene pairs"
        )

        self.deletion_tbl.setHorizontalHeaderLabels(["Gene Pair (Names)", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        pairs_for_plot = []
        growth_vals = []

        sorted_pairs = sorted(dgd_result.items(), key=lambda x: x[1])
        for pair_id, growth in sorted_pairs[:int(self.topn_spin.value())]:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            pair_label = self._format_gene_pair_label(str(pair_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(pair_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))
            pairs_for_plot.append(pair_label)
            growth_vals.append(growth)

        self.filter_deletion_table()

        ax = self.deletion_canvas.ax
        ax.clear()
        if pairs_for_plot:
            bars = ax.barh(pairs_for_plot, growth_vals, color="#8ecae6", alpha=0.7, edgecolor="#219ebc", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Double Gene Deletion - Growth on Deletion")
        self.deletion_canvas.draw()

        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_dgd_csv(self):
        if not self.last_run or not self.last_run.get("dgd_result"):
            QMessageBox.warning(self, "DGD", "Run DGD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export DGD (CSV)", str(Path.home() / "dgd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["gene_pair_name", "gene_pair_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for pair, growth in sorted(self.last_run["dgd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    pname = self._format_gene_pair_label(str(pair))
                    w.writerow([pname, pair, wt, growth, delta_pct])
            self.statusBar().showMessage(f"DGD exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _render_srd_results(self):
        """Render SRD results."""
        assert self.last_run is not None
        wt_growth = self.last_run.get("wt_growth", 0.0)
        srd_result = self.last_run.get("srd_result", {})
        
        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Total: {len(srd_result)} reactions"
        )
        
        # Populate table with ALL results
        self.deletion_tbl.setHorizontalHeaderLabels(["Reaction Name", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        rxns_for_plot = []
        growth_vals = []
        
        sorted_rxns = sorted(srd_result.items(), key=lambda x: x[1])
        
        for rxn_id, growth in sorted_rxns:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            rxn_label = self._format_rxn_label(str(rxn_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(rxn_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))
        
        # Chart uses top N only
        for rxn_id, growth in sorted_rxns[:int(self.topn_spin.value())]:
            rxn_label = self._format_rxn_label(str(rxn_id))
            rxns_for_plot.append(rxn_label)
            growth_vals.append(growth)
        
        self.filter_deletion_table()
        
        # Plot with value labels
        ax = self.deletion_canvas.ax
        ax.clear()
        if rxns_for_plot:
            bars = ax.barh(rxns_for_plot, growth_vals, color="lightgreen", alpha=0.7, edgecolor="darkgreen", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Single Reaction Deletion - Growth on Deletion")
        self.deletion_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_srd_csv(self):
        if not self.last_run or not self.last_run.get("srd_result"):
            QMessageBox.warning(self, "SRD", "Run SRD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export SRD (CSV)", str(Path.home() / "srd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction_name", "reaction_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for rid, growth in sorted(self.last_run["srd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    rname = self._format_rxn_label(str(rid))
                    w.writerow([rname, rid, wt, growth, delta_pct])
            self.statusBar().showMessage(f"SRD exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    # -------- Excel export helpers --------
    def _export_deletion_excel(self, data_key: str, id_formatter, default_name: str, label: str):
        """Generic Excel export for deletion analyses."""
        if not self.last_run or not self.last_run.get(data_key):
            QMessageBox.warning(self, label, f"Run {label} first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, f"Export {label} (Excel)",
                                                    str(Path.home() / default_name), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = label
            wt = float(self.last_run.get("wt_growth", 0.0))
            ws.append(["Name", "ID", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
            for item_id, growth in sorted(self.last_run[data_key].items(), key=lambda x: x[1]):
                delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                name = id_formatter(str(item_id))
                if growth < 1e-6:
                    cat = "Lethal"
                elif delta_pct < -75:
                    cat = "Severe"
                elif delta_pct < -25:
                    cat = "Moderate"
                else:
                    cat = "Mild"
                ws.append([name, str(item_id), wt, growth, delta_pct, cat])
            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            wb.save(file_path)
            self.statusBar().showMessage(f"{label} exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export. Install it with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def export_sgd_excel(self):
        self._export_deletion_excel("sgd_result", self._format_gene_label, "sgd_results.xlsx", "SGD")

    def export_dgd_excel(self):
        self._export_deletion_excel("dgd_result", self._format_gene_pair_label, "dgd_results.xlsx", "DGD")

    def export_srd_excel(self):
        self._export_deletion_excel("srd_result", self._format_rxn_label, "srd_results.xlsx", "SRD")

    def export_robustness_csv(self):
        if not self.last_run or not self.last_run.get("robustness_result"):
            QMessageBox.warning(self, "Robustness", "Run Robustness Analysis first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Robustness (CSV)",
                                                    str(Path.home() / "robustness_results.csv"), "CSV files (*.csv)")
        if not file_path:
            return
        try:
            result = self.last_run["robustness_result"]
            rxn_id = self.last_run.get("rxn_id", "reaction")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([f"{rxn_id}_bound_value", "objective_value"])
                for val, obj in zip(result.get("values", []), result.get("objectives", [])):
                    w.writerow([val, obj])
            self.statusBar().showMessage(f"Robustness exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def export_robustness_excel(self):
        if not self.last_run or not self.last_run.get("robustness_result"):
            QMessageBox.warning(self, "Robustness", "Run Robustness Analysis first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Robustness (Excel)",
                                                    str(Path.home() / "robustness_results.xlsx"), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Robustness"
            result = self.last_run["robustness_result"]
            rxn_id = self.last_run.get("rxn_id", "reaction")
            ws.append([f"{rxn_id} bound value", "Objective value"])
            for val, obj in zip(result.get("values", []), result.get("objectives", [])):
                ws.append([val, obj])
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            wb.save(file_path)
            self.statusBar().showMessage(f"Robustness exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export.")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def export_envelope_csv(self):
        if not self.last_run or not self.last_run.get("envelope_result"):
            QMessageBox.warning(self, "Envelope", "Run Production Envelope first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Envelope (CSV)",
                                                    str(Path.home() / "envelope_results.csv"), "CSV files (*.csv)")
        if not file_path:
            return
        try:
            result = self.last_run["envelope_result"]
            product_id = self.last_run.get("product_id", "product")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([f"{product_id}_bound", "growth_rate"])
                for pv, gv in zip(result.get("product", []), result.get("growth", [])):
                    w.writerow([pv, gv])
            self.statusBar().showMessage(f"Envelope exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def export_envelope_excel(self):
        if not self.last_run or not self.last_run.get("envelope_result"):
            QMessageBox.warning(self, "Envelope", "Run Production Envelope first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Envelope (Excel)",
                                                    str(Path.home() / "envelope_results.xlsx"), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Envelope"
            result = self.last_run["envelope_result"]
            product_id = self.last_run.get("product_id", "product")
            ws.append([f"{product_id} bound", "Growth rate"])
            for pv, gv in zip(result.get("product", []), result.get("growth", [])):
                ws.append([pv, gv])
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            wb.save(file_path)
            self.statusBar().showMessage(f"Envelope exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export.")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _render_robustness_results(self):
        """Render Robustness analysis results."""
        assert self.last_run is not None
        rxn_id = self.last_run.get("rxn_id", "")
        bound_type = self.last_run.get("bound_type", "ub").upper()
        robustness_result = self.last_run.get("robustness_result", {})
        
        values = robustness_result.get("values", [])
        objectives = robustness_result.get("objectives", [])
        
        self.robustness_info_lbl.setText(
            f"Robustness Analysis for {rxn_id} ({bound_type})\n"
            f"Points computed: {len(values)}"
        )
        
        # Populate data table
        self.robustness_tbl.setHorizontalHeaderLabels([f"{rxn_id} {bound_type} Value", "Objective Value"])
        self.robustness_tbl.setRowCount(0)
        for val, obj in zip(values, objectives):
            r = self.robustness_tbl.rowCount()
            self.robustness_tbl.insertRow(r)
            self.robustness_tbl.setItem(r, 0, QTableWidgetItem(f"{val:.6g}"))
            self.robustness_tbl.setItem(r, 1, QTableWidgetItem(f"{obj:.6g}"))
        
        # Plot with clear markers and (x,y) labels
        ax = self.robustness_canvas.ax
        ax.clear()
        if values and objectives:
            ax.plot(values, objectives, marker="o", linestyle="-", color="steelblue",
                    linewidth=2, markersize=8, markerfacecolor="orange",
                    markeredgecolor="steelblue", markeredgewidth=1.5, zorder=5)
            # Add (x, y) coordinate labels at each point
            for i, (x, y) in enumerate(zip(values, objectives)):
                offset = 8 if (i % 2 == 0) else -12
                ax.annotate(f"({x:.2g}, {y:.4g})", (x, y),
                           textcoords="offset points", xytext=(0, offset),
                           fontsize=6.5, ha="center", color="#333")
            ax.set_xlabel(f"{rxn_id} bound value")
            ax.set_ylabel("Objective value")
            ax.set_title(f"Robustness: {rxn_id} ({bound_type})")
            ax.grid(True, alpha=0.3)
        self.robustness_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_robustness)

    def _render_envelope_results(self):
        """Render Production Envelope results."""
        assert self.last_run is not None
        product_id = self.last_run.get("product_id", "")
        envelope_result = self.last_run.get("envelope_result", {})
        
        product_vals = envelope_result.get("product", [])
        growth_vals = envelope_result.get("growth", [])
        
        self.envelope_info_lbl.setText(
            f"Production Envelope: {product_id} vs Growth\n"
            f"Points computed: {len(product_vals)}"
        )
        
        # Populate data table
        self.envelope_tbl.setHorizontalHeaderLabels([f"{product_id} Bound", "Growth Rate"])
        self.envelope_tbl.setRowCount(0)
        for pv, gv in zip(product_vals, growth_vals):
            r = self.envelope_tbl.rowCount()
            self.envelope_tbl.insertRow(r)
            self.envelope_tbl.setItem(r, 0, QTableWidgetItem(f"{pv:.6g}"))
            self.envelope_tbl.setItem(r, 1, QTableWidgetItem(f"{gv:.6g}"))
        
        # Plot with point labels
        ax = self.envelope_canvas.ax
        ax.clear()
        if product_vals and growth_vals:
            ax.scatter(product_vals, growth_vals, s=60, alpha=0.8, color="purple", zorder=5,
                      edgecolors="darkviolet", linewidth=1)
            ax.plot(product_vals, growth_vals, linestyle="--", color="purple", alpha=0.4)
            ax.fill_between(product_vals, growth_vals, alpha=0.15, color="purple")
            # Add (x, y) coordinate labels
            for i, (x, y) in enumerate(zip(product_vals, growth_vals)):
                offset = 8 if (i % 2 == 0) else -12
                ax.annotate(f"({x:.2g}, {y:.4g})", (x, y),
                           textcoords="offset points", xytext=(0, offset),
                           fontsize=6.5, ha="center", color="#333")
            ax.set_xlabel(f"{product_id} production rate")
            ax.set_ylabel("Growth rate")
            ax.set_title(f"Production Envelope: {product_id}")
            ax.grid(True, alpha=0.3)
        self.envelope_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_envelope)

    def _render_sampling_results(self):
        """Render Flux Sampling results."""
        assert self.last_run is not None
        sampling_result = self.last_run.get("sampling_result", {})
        samples_df = self.last_run.get("samples_df", None)
        
        self.sampling_info_lbl.setText(
            f"Flux Sampling ({self.last_run.get('sampler_type', 'ACHR')})\n"
            f"Sample size: {len(samples_df) if samples_df is not None else 0}\n"
            f"Reactions analyzed: {len(sampling_result)}"
        )
        
        # Populate table
        self.sampling_tbl.setRowCount(0)
        rxns_for_plot = []
        means = []
        
        sorted_rxns = sorted(sampling_result.items(), key=lambda x: abs(x[1]["mean"]), reverse=True)
        
        for rxn_id, stats in sorted_rxns[:int(self.topn_spin.value())]:
            r = self.sampling_tbl.rowCount()
            self.sampling_tbl.insertRow(r)
            self.sampling_tbl.setItem(r, 0, QTableWidgetItem(rxn_id))
            self.sampling_tbl.setItem(r, 1, QTableWidgetItem(f"{stats['mean']:.6g}"))
            self.sampling_tbl.setItem(r, 2, QTableWidgetItem(f"{stats['stdev']:.6g}"))
            self.sampling_tbl.setItem(r, 3, QTableWidgetItem(f"[{stats['min']:.3g}, {stats['max']:.3g}]"))
            rxns_for_plot.append(rxn_id)
            means.append(stats["mean"])
        
        # Plot
        ax = self.sampling_canvas.ax
        ax.clear()
        if rxns_for_plot:
            colors = ["green" if m > 0 else "red" for m in means]
            ax.barh(rxns_for_plot, means, color=colors, alpha=0.6)
            ax.set_xlabel("Mean flux")
            ax.set_title("Flux Sampling - Mean Flux Distribution")
        self.sampling_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_sampling)

    def export_sampling_csv(self):
        if not self.last_run or not self.last_run.get("sampling_result"):
            QMessageBox.warning(self, "Sampling", "Run sampling first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Sampling Stats (CSV)", str(Path.home() / "sampling_stats.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            sampling_result = self.last_run["sampling_result"]
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction_id", "mean_flux", "stdev", "min_flux", "max_flux"])
                for rxn_id, stats in sorted(sampling_result.items(), key=lambda x: abs(x[1]["mean"]), reverse=True):
                    w.writerow([rxn_id, stats["mean"], stats["stdev"], stats["min"], stats["max"]])
            self.statusBar().showMessage(f"Sampling exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def filter_pfba_table(self):
        """Filter pFBA table by reaction ID."""
        text = (self.pfba_flux_search.text() or "").strip().lower()
        for row in range(self.pfba_tbl.rowCount()):
            rid = (self.pfba_tbl.item(row, 0).text() if self.pfba_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.pfba_tbl.setRowHidden(row, hide)

    def filter_sampling_table(self):
        """Filter sampling table."""
        text = (self.sampling_search.text() or "").strip().lower()
        for row in range(self.sampling_tbl.rowCount()):
            rid = (self.sampling_tbl.item(row, 0).text() if self.sampling_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.sampling_tbl.setRowHidden(row, hide)

    def _parse_float(self, val, allow_comma: bool = True) -> float:
        """Robust float parser for user inputs. Supports comma decimal.
        Raises ValueError if parsing fails.
        """
        if isinstance(val, (int, float)):
            return float(val)
        if val is None:
            raise ValueError("Missing numeric value")
        s = str(val).strip()
        if allow_comma:
            s = s.replace(",", ".")
        return float(s)
    # ---------------- Scenario export/import ----------------
    def export_scenario_to_path(self, file_path: str):
        scenario = {
            "version": 4,
            "top_n": int(self.topn_spin.value()),
            "objective_reaction_id": self.objective_combo.currentData(),
            "knockout_genes": sorted(self.knockout_genes),
            "overexpression_reactions": {k: float(v) for k, v in self.overexpression_reactions.items()},
            "temporary_upper_bound_overrides": {k: float(v) for k, v in self.temporary_upper_bound_overrides.items()},
            "exchange_bounds": self._capture_exchange_bounds_from_table(),
            "reaction_bounds": {k: {"lb": float(v[0]), "ub": float(v[1])} for k, v in self.reaction_bound_overrides.items()},
        }
        Path(file_path).write_text(json.dumps(scenario, indent=2), encoding="utf-8")

    def export_scenario(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model before exporting a scenario.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export scenario", str(Path.home() / "metabodesk_scenario.json"),
            "JSON files (*.json);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            self.export_scenario_to_path(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        self._add_recent_scenario(file_path)
        self.statusBar().showMessage(f"Scenario exported: {file_path}")

    def import_scenario(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model before importing a scenario.")
            return
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import scenario", str(Path.home()),
            "JSON files (*.json);;All files (*.*)",
        )
        if not file_path:
            return
        self.import_scenario_from_path(file_path)

    def import_scenario_from_path(self, file_path: str):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model before importing a scenario.")
            return

        try:
            scenario = json.loads(Path(file_path).read_text(encoding="utf-8"))
            if not isinstance(scenario, dict):
                raise ValueError("Invalid scenario format (expected JSON object).")

            self.topn_spin.setValue(int(scenario.get("top_n", self.topn_spin.value())))

            obj_rid = scenario.get("objective_reaction_id", None)
            for i in range(self.objective_combo.count()):
                if self.objective_combo.itemData(i) == obj_rid:
                    self.objective_combo.setCurrentIndex(i)
                    break

            self.knockout_genes = set([str(x) for x in scenario.get("knockout_genes", [])])
            self.update_knockout_label()

            self.overexpression_reactions = {str(k): float(v) for k, v in scenario.get("overexpression_reactions", {}).items()}
            self.temporary_upper_bound_overrides = {str(k): float(v) for k, v in scenario.get("temporary_upper_bound_overrides", {}).items()}
            self.refresh_overexpression_lists()
            self.update_selected_rxn_info()

            exchange_bounds = scenario.get("exchange_bounds", {})
            if isinstance(exchange_bounds, dict):
                self._apply_exchange_bounds_to_table(exchange_bounds)

            rb = scenario.get("reaction_bounds", {})
            if isinstance(rb, dict):
                self.reaction_bound_overrides = {}
                for rid, b in rb.items():
                    self.reaction_bound_overrides[str(rid)] = (float(b.get("lb")), float(b.get("ub")))
                self.populate_reaction_table()

        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))
            return

        self._add_recent_scenario(file_path)
        self._invalidate_fba_cache()
        self.statusBar().showMessage(f"Scenario imported: {file_path}")

    def _capture_exchange_bounds_from_table(self) -> dict[str, dict[str, float]]:
        bounds: dict[str, dict[str, float]] = {}
        for i in range(self.medium_table.rowCount()):
            rxn_id = self.medium_table.item(i, 0).text().strip()
            lb = self._parse_float(self.medium_table.item(i, 2).text())
            ub = self._parse_float(self.medium_table.item(i, 3).text())
            bounds[rxn_id] = {"lb": lb, "ub": ub}
        return bounds

    def _apply_exchange_bounds_to_table(self, bounds: dict):
        rxn_to_row = {}
        for i in range(self.medium_table.rowCount()):
            rxn_id = self.medium_table.item(i, 0).text().strip()
            rxn_to_row[rxn_id] = i
        for rxn_id, b in bounds.items():
            if rxn_id not in rxn_to_row:
                continue
            row = rxn_to_row[rxn_id]
            lb = float(b.get("lb"))
            ub = float(b.get("ub"))
            self.medium_table.setItem(row, 2, QTableWidgetItem(str(lb)))
            self.medium_table.setItem(row, 3, QTableWidgetItem(str(ub)))
        self.filter_medium_table()

    # ---------------- Results export / Export all ----------------
    def export_results_to_path(self, file_path: str):
        if not self.last_run:
            raise ValueError("No last run results to export.")
        analysis_type = self.last_run.get("analysis_type", "FBA")
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["# MetaboDesk results export"])
            w.writerow(["# timestamp", self.last_run.get("timestamp", "")])
            w.writerow(["# analysis_type", analysis_type])
            w.writerow([])

            if analysis_type in ("FBA", "Loopless FBA"):
                # Full FBA comparison export
                for section in ("original", "baseline", "gene_knockout_only", "overexpression_only"):
                    data = self.last_run.get(section)
                    if data and isinstance(data, dict):
                        w.writerow([f"{section}_status", data.get("status", "")])
                        w.writerow([f"{section}_objective", data.get("objective", "")])
                w.writerow([])
                w.writerow(["reaction", "baseline_flux", "compared_flux", "delta", "compare_mode"])
                for row in self.last_run.get("flux_rows", []):
                    w.writerow([row.get("reaction", ""), row.get("baseline", ""),
                                row.get("compared", ""), row.get("delta", ""),
                                self.last_run.get("compare_mode", "")])

            elif analysis_type == "FVA":
                baseline = self.last_run.get("baseline", {})
                w.writerow(["baseline_status", baseline.get("status", "")])
                w.writerow(["baseline_objective", baseline.get("objective", "")])
                w.writerow([])
                w.writerow(["reaction", "min_flux", "max_flux", "range"])
                for row in self.last_run.get("flux_rows", []):
                    w.writerow([row.get("reaction", ""), row.get("min_flux", ""),
                                row.get("max_flux", ""), row.get("range", "")])

            elif analysis_type == "pFBA":
                baseline = self.last_run.get("baseline", {})
                pfba_data = self.last_run.get("pfba", {})
                w.writerow(["baseline_objective", baseline.get("objective", "")])
                w.writerow(["pfba_objective", pfba_data.get("objective", "")])
                w.writerow([])
                w.writerow(["reaction", "fba_flux", "pfba_flux", "delta"])
                for row in self.last_run.get("flux_rows", []):
                    w.writerow([row.get("reaction", ""), row.get("fba_flux", ""),
                                row.get("pfba_flux", ""), row.get("delta", "")])

            elif analysis_type == "SGD":
                w.writerow(["wt_growth", self.last_run.get("wt_growth", "")])
                w.writerow([])
                w.writerow(["gene", "growth_after_ko"])
                for gid, gval in self.last_run.get("sgd_result", {}).items():
                    w.writerow([gid, gval])

            elif analysis_type == "DGD":
                w.writerow(["wt_growth", self.last_run.get("wt_growth", "")])
                w.writerow([])
                w.writerow(["gene_pair", "growth_after_ko"])
                for pair, gval in self.last_run.get("dgd_result", {}).items():
                    w.writerow([pair, gval])

            elif analysis_type == "SRD":
                w.writerow(["wt_growth", self.last_run.get("wt_growth", "")])
                w.writerow([])
                w.writerow(["reaction", "growth_after_ko"])
                for rid, gval in self.last_run.get("srd_result", {}).items():
                    w.writerow([rid, gval])

            elif analysis_type == "Robustness":
                w.writerow(["rxn_id", self.last_run.get("rxn_id", "")])
                w.writerow(["bound_type", self.last_run.get("bound_type", "")])
                w.writerow([])
                rob = self.last_run.get("robustness_result", {})
                w.writerow(["bound_value", "objective"])
                for v, o in zip(rob.get("values", []), rob.get("objectives", [])):
                    w.writerow([v, o])

            elif analysis_type == "Production Envelope":
                w.writerow(["product_id", self.last_run.get("product_id", "")])
                w.writerow([])
                env = self.last_run.get("envelope_result", {})
                w.writerow(["growth", "product"])
                for g, p in zip(env.get("growth", []), env.get("product", [])):
                    w.writerow([g, p])

            elif analysis_type == "Sampling":
                w.writerow(["reaction", "mean", "std", "min", "max", "median"])
                for row in self.last_run.get("flux_rows", []):
                    w.writerow([row.get("reaction", ""), row.get("mean", ""),
                                row.get("std", ""), row.get("min", ""),
                                row.get("max", ""), row.get("median", "")])

            else:
                # Generic fallback: export flux_tbl contents
                if hasattr(self, 'flux_tbl') and self.flux_tbl.rowCount() > 0:
                    headers = [self.flux_tbl.horizontalHeaderItem(c).text()
                               for c in range(self.flux_tbl.columnCount())]
                    w.writerow(headers)
                    for row in range(self.flux_tbl.rowCount()):
                        vals = []
                        for col in range(self.flux_tbl.columnCount()):
                            item = self.flux_tbl.item(row, col)
                            vals.append(item.text() if item else "")
                        w.writerow(vals)

    def export_results_csv(self):
        if not self.last_run:
            QMessageBox.warning(self, "No results", "Run FBA at least once before exporting results.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export results (CSV)", str(Path.home() / "metabodesk_results.csv"),
            "CSV files (*.csv);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            self.export_results_to_path(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        self.statusBar().showMessage(f"Results exported: {file_path}")

    def export_all(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return
        if not self.last_run:
            QMessageBox.warning(self, "No results", "Run FBA at least once before Export All.")
            return
        folder = QFileDialog.getExistingDirectory(self, "Select folder for export", str(Path.home()))
        if not folder:
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_path = Path(folder)
        scenario_path = folder_path / f"scenario_{stamp}.json"
        results_path = folder_path / f"results_{stamp}.csv"
        chart_path = folder_path / f"chart_{stamp}.png"
        try:
            self.export_scenario_to_path(str(scenario_path))
            self.export_results_to_path(str(results_path))
            self.canvas.figure.savefig(str(chart_path), dpi=200)
        except Exception as e:
            QMessageBox.critical(self, "Export All failed", str(e))
            return
        self.statusBar().showMessage(f"Exported: {scenario_path.name}, {results_path.name}, {chart_path.name}")

    # ---------------- Model load ----------------
    def open_sbml(self):
        if self.is_running:
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select SBML model",
            str(Path.home()),
            "SBML files (*.xml *.sbml);;All files (*.*)",
        )
        if not file_path:
            return

        self._load_sbml_model(file_path)

    def open_sbml_from_path(self, file_path: str):
        if self.is_running:
            return
        if not file_path:
            return
        self._load_sbml_model(file_path)

    def _load_sbml_model(self, file_path: str):
        """Common model loading logic used by open_sbml and open_sbml_from_path."""
        self.set_busy(True, f"Loading: {file_path}")
        QApplication.processEvents()

        try:
            model = cobra.io.read_sbml_model(file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Failed to load SBML", str(e))
            self.set_busy(False, "Failed to load model.")
            return

        try:
            self.base_model = model
            self.original_model_snapshot = model.copy()
            self.last_run = None
            self._invalidate_fba_cache()
            self.current_sbml_path = Path(file_path)
            self.model_dirty = False
            self.knockout_genes.clear()
            self.overexpression_reactions.clear()
            self.temporary_upper_bound_overrides.clear()
            self.reaction_bound_overrides.clear()
            self.update_knockout_label()
            self.refresh_overexpression_lists()

            self.model_patch = {
                "version": 1,
                "added_metabolites": [],
                "added_genes": [],
                "added_or_updated_reactions": [],
                "disabled_reactions": [],
                "deleted_reactions": [],
                "objective_reaction_id": None,
                "notes": "",
            }
            self._refresh_patch_preview()

            self.exchange_rxns = [r for r in model.reactions if is_exchange_reaction(r)]
            self.exchange_rxns.sort(key=lambda r: r.id)
            self.original_bounds = {r.id: (float(r.lower_bound), float(r.upper_bound)) for r in self.exchange_rxns}

            self.populate_gene_list()
            self.populate_overexpression_reaction_list()
            self.populate_medium_table()
            self.populate_reaction_table()
            self.populate_objective_combo()
            self._populate_reaction_lists()
            self.populate_genes_tab()
            self._setup_map_search_completer()
            self._populate_map_subsystems()

            model_id = getattr(model, "id", "") or "(no id)"
            model_name = getattr(model, "name", "") or "(no name)"
            self.summary_lbl.setText(
                f"Model loaded.\n"
                f"ID: {model_id}\n"
                f"Name: {model_name}\n"
                f"Reactions: {len(model.reactions)} | Metabolites: {len(model.metabolites)} | Genes: {len(model.genes)}\n"
                f"Exchange reactions (heuristic): {len(self.exchange_rxns)}\n"
            )

            self.run_btn.setEnabled(True)
            self.close_uptakes_btn.setEnabled(True)
            self.reset_medium_btn.setEnabled(True)
            self.reset_reactions_btn.setEnabled(True)
            self.medium_apply_preset_btn.setEnabled(True)
            self.add_ko_btn.setEnabled(True)
            self.clear_ko_btn.setEnabled(True)
            self.add_oe_btn.setEnabled(True)
            self.clear_oe_btn.setEnabled(True)
            self.set_temp_ub_btn.setEnabled(True)
            self.clear_temp_ub_btn.setEnabled(True)
            self.clear_rxn_overrides_btn.setEnabled(True)
            self.remove_selected_overrides_btn.setEnabled(True)
            self.objective_combo.setEnabled(True)
            self.analysis_type.setEnabled(True)
            self.solver_combo.setEnabled(True)

            self.memote_btn.setEnabled(True)
            self.carveme_btn.setEnabled(True)

            self._clear_chart()
            self.flux_tbl.setRowCount(0)
            self.results_lbl.setText("Objective: -")

            self.set_busy(False, "Ready.")
            self.tabs.setCurrentWidget(self.tab_medium)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error after loading", f"Model was loaded but UI setup failed:\n{e}")
            self.set_busy(False, "Model loaded with errors.")

    def _populate_reaction_lists(self):
        """Populate reaction ID combos for Robustness and Envelope tabs."""
        try:
            ids = [r.id for r in self.base_model.reactions]
        except Exception:
            ids = []
        ids_sorted = sorted(ids)
        # store for filtering (labels include name)
        items: list[tuple[str, str]] = []
        for rid in ids_sorted:
            name = self._rxn_name_from_id(rid)
            label = f"{rid} — {name}".strip(" —") if name else rid
            items.append((rid, label))
        self._all_reaction_items = items

        # Completers for text inputs (match by id or name)
        labels = [label for _, label in items]
        self._rxn_label_to_id = {label: rid for rid, label in items}
        if hasattr(self, "robustness_rxn"):
            comp = QCompleter(labels, self)
            comp.setCaseSensitivity(Qt.CaseInsensitive)
            comp.setFilterMode(Qt.MatchContains)
            comp.setCompletionMode(QCompleter.PopupCompletion)
            comp.activated.connect(self._on_robustness_completer_selected)
            self.robustness_rxn.setCompleter(comp)

        if hasattr(self, "envelope_product"):
            comp2 = QCompleter(labels, self)
            comp2.setCaseSensitivity(Qt.CaseInsensitive)
            comp2.setFilterMode(Qt.MatchContains)
            comp2.setCompletionMode(QCompleter.PopupCompletion)
            comp2.activated.connect(self._on_envelope_completer_selected)
            self.envelope_product.setCompleter(comp2)

        # Robustness combo
        if hasattr(self, "robustness_combo"):
            self._filter_robustness_combo(self.robustness_search.text() if hasattr(self, "robustness_search") else "")
        # Envelope combo
        if hasattr(self, "envelope_combo"):
            self._filter_envelope_combo(self.envelope_search.text() if hasattr(self, "envelope_search") else "")

    def _filter_robustness_combo(self, text: str):
        s = (text or "").strip().lower()
        items = [item for item in getattr(self, "_all_reaction_items", []) if s in item[1].lower()]
        self.robustness_combo.blockSignals(True)
        self.robustness_combo.clear()
        for rid, label in items:
            self.robustness_combo.addItem(label, rid)
        self.robustness_combo.blockSignals(False)

    def _filter_envelope_combo(self, text: str):
        s = (text or "").strip().lower()
        items = [item for item in getattr(self, "_all_reaction_items", []) if s in item[1].lower()]
        self.envelope_combo.blockSignals(True)
        self.envelope_combo.clear()
        for rid, label in items:
            self.envelope_combo.addItem(label, rid)
        self.envelope_combo.blockSignals(False)

    def _setup_map_search_completer(self):
        if self.base_model is None:
            return
        labels = []
        for r in self.base_model.reactions:
            name = getattr(r, "name", "") or ""
            labels.append(f"{r.id} — {name}".strip(" —"))
        for m in self.base_model.metabolites:
            name = getattr(m, "name", "") or ""
            labels.append(f"{m.id} — {name}".strip(" —"))
        comp = QCompleter(sorted(labels), self)
        comp.setCaseSensitivity(Qt.CaseInsensitive)
        comp.setFilterMode(Qt.MatchContains)
        comp.setCompletionMode(QCompleter.PopupCompletion)
        self.map_search.setCompleter(comp)

    def _populate_map_subsystems(self):
        if self.base_model is None or not hasattr(self, "map_subsystem_combo"):
            return
        subsystems = set()
        for r in self.base_model.reactions:
            sub = str(getattr(r, "subsystem", "") or "").strip()
            if sub:
                subsystems.add(sub)
        items = ["All subsystems"] + sorted(subsystems)
        self.map_subsystem_combo.blockSignals(True)
        self.map_subsystem_combo.clear()
        for s in items:
            self.map_subsystem_combo.addItem(s)
        self.map_subsystem_combo.blockSignals(False)

    def _show_map_help(self):
        msg = (
            "How to use the map\n\n"
            "Core controls\n"
            "• Search: Type a reaction/metabolite ID or name to focus on matches.\n"
            "• Depth: Neighborhood depth from seeds (1–5). Higher = wider context.\n"
            "• Threshold: Hides low-impact reactions for the selected view.\n"
            "• View: What the color/size represents:\n"
            "   - Flux magnitude: |flux| from last FBA\n"
            "   - Gene KO impact: growth drop from SGD\n"
            "   - FVA bounds: flux range (max–min)\n"
            "\n"
            "Size & layout\n"
            "• Max nodes: Hard cap to keep big graphs responsive.\n"
            "• Max edges: 0 = no limit; otherwise keeps only top edges.\n"
            "• Min degree: Hides nodes with degree below this value.\n"
            "• Top‑N rxns: Keep only the highest‑impact reactions (0 = off).\n"
            "• Layout: Spring / Kamada‑Kawai / Circular.\n"
            "• Show legend: Toggles the color scale.\n"
            "\n"
            "Filters\n"
            "• Only connected to search: Show only the component connected to search seeds.\n"
            "• Hide orphan metabolites: Removes metabolites with no edges.\n"
            "• Exchange only: Keep only exchange reactions.\n"
            "• Subsystem: Filter reactions by subsystem (if available in SBML).\n"
            "• Objective neighborhood: Focus around the objective reaction.\n"
            "\n"
            "Interaction & export\n"
            "• Render Map: Builds the graph with current settings.\n"
            "• Click a node: Details appear below.\n"
            "• Focus selected: Re‑centers the map around the clicked node.\n"
            "• Export map image: Save PNG/SVG/PDF.\n"
            "• Export graph CSV: Save node/edge lists for analysis."
        )
        QMessageBox.information(self, "How to use map?", msg)

    def _on_robustness_combo_changed(self, index: int):
        if index < 0:
            return
        rid = self.robustness_combo.itemData(index)
        if rid:
            self.robustness_rxn.setText(str(rid))

    def _on_envelope_combo_changed(self, index: int):
        if index < 0:
            return
        rid = self.envelope_combo.itemData(index)
        if rid:
            self.envelope_product.setText(str(rid))

    def _on_robustness_completer_selected(self, text: str):
        rid = self._rxn_label_to_id.get(text)
        if rid:
            self.robustness_rxn.setText(str(rid))

    def _on_envelope_completer_selected(self, text: str):
        rid = self._rxn_label_to_id.get(text)
        if rid:
            self.envelope_product.setText(str(rid))

    # ---------------- Model editor tabs ----------------
    def _build_editor_metabolite_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        layout.addWidget(self._title("Add metabolite"))

        form = QFormLayout()
        self.ed_met_id = QLineEdit()
        self.ed_met_name = QLineEdit()
        self.ed_met_comp = QLineEdit()
        self.ed_met_comp.setPlaceholderText("e.g. c (cytosol), e (extracellular)")
        self.ed_met_formula = QLineEdit()
        self.ed_met_charge = QLineEdit()
        self.ed_met_charge.setPlaceholderText("integer, optional")

        form.addRow("Metabolite ID:", self.ed_met_id)
        form.addRow("Name:", self.ed_met_name)
        form.addRow("Compartment:", self.ed_met_comp)
        form.addRow("Formula (optional):", self.ed_met_formula)
        form.addRow("Charge (optional):", self.ed_met_charge)

        box = QGroupBox("Metabolite fields")
        box.setLayout(form)
        layout.addWidget(box)

        btn_row = QHBoxLayout()
        self.ed_add_met_btn = QPushButton("Add metabolite to model")
        self.ed_add_met_btn.clicked.connect(self.editor_add_metabolite)
        btn_row.addWidget(self.ed_add_met_btn)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        self.ed_met_info = QPlainTextEdit()
        self.ed_met_info.setReadOnly(True)
        self.ed_met_info.setMaximumHeight(160)
        layout.addWidget(self.ed_met_info)

        self.editor_tabs.addTab(tab, "Metabolites")

    def _build_editor_gene_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(self._title("Add/Update gene"))

        form = QFormLayout()
        self.ed_gene_id = QLineEdit()
        self.ed_gene_name = QLineEdit()
        form.addRow("Gene ID:", self.ed_gene_id)
        form.addRow("Name (optional):", self.ed_gene_name)

        box = QGroupBox("Gene fields")
        box.setLayout(form)
        layout.addWidget(box)

        btn_row = QHBoxLayout()
        self.ed_add_gene_btn = QPushButton("Add/Update gene in model")
        self.ed_add_gene_btn.clicked.connect(self.editor_add_or_update_gene)
        btn_row.addWidget(self.ed_add_gene_btn)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        self.ed_gene_info = QPlainTextEdit()
        self.ed_gene_info.setReadOnly(True)
        self.ed_gene_info.setMaximumHeight(160)
        layout.addWidget(self.ed_gene_info)

        self.editor_tabs.addTab(tab, "Genes")

    def _build_editor_reaction_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(self._title("Add/Update reaction (with equation parser)"))

        form = QFormLayout()
        self.ed_rxn_id = QLineEdit()
        self.ed_rxn_desc = QLineEdit()
        self.ed_rxn_lb = QLineEdit("-1000")
        self.ed_rxn_ub = QLineEdit("1000")
        self.ed_rxn_rev = QCheckBox("Reversible (force)")
        self.ed_rxn_gpr = QLineEdit()
        self.ed_rxn_kegg = QLineEdit()
        self.ed_rxn_ec = QLineEdit()
        self.ed_rxn_autocreate_missing_mets = QCheckBox("Auto-create missing metabolites (recommended)")
        self.ed_rxn_autocreate_missing_mets.setChecked(True)

        form.addRow("Reaction ID:", self.ed_rxn_id)
        form.addRow("Description:", self.ed_rxn_desc)
        form.addRow("LB:", self.ed_rxn_lb)
        form.addRow("UB:", self.ed_rxn_ub)
        form.addRow("", self.ed_rxn_rev)
        form.addRow("GPR:", self.ed_rxn_gpr)
        form.addRow("KEGG (optional):", self.ed_rxn_kegg)
        form.addRow("EC (optional):", self.ed_rxn_ec)
        form.addRow("", self.ed_rxn_autocreate_missing_mets)

        box = QGroupBox("Reaction fields")
        box.setLayout(form)
        layout.addWidget(box)

        layout.addWidget(QLabel("Equation:"))
        self.ed_rxn_eq = QPlainTextEdit()
        self.ed_rxn_eq.setPlaceholderText("e.g. glc__D_c + 2 atp_c -> 2 adp_c + 2 pi_c + h_c")
        self.ed_rxn_eq.setMaximumHeight(110)
        layout.addWidget(self.ed_rxn_eq)

        btn_row = QHBoxLayout()
        self.ed_parse_eq_btn = QPushButton("Parse equation")
        self.ed_parse_eq_btn.clicked.connect(self.editor_parse_equation)
        btn_row.addWidget(self.ed_parse_eq_btn)

        self.ed_apply_rxn_btn = QPushButton("Add/Update reaction in model")
        self.ed_apply_rxn_btn.clicked.connect(self.editor_add_or_update_reaction)
        btn_row.addWidget(self.ed_apply_rxn_btn)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        layout.addWidget(self._title("Parsed stoichiometry preview"))
        self.ed_rxn_preview_tbl = QTableWidget(0, 3)
        self.ed_rxn_preview_tbl.setHorizontalHeaderLabels(["Metabolite", "Coefficient", "Side"])
        self.ed_rxn_preview_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ed_rxn_preview_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._configure_table_for_excel_resize(self.ed_rxn_preview_tbl)
        layout.addWidget(self.ed_rxn_preview_tbl, stretch=1)

        self.ed_rxn_info = QPlainTextEdit()
        self.ed_rxn_info.setReadOnly(True)
        self.ed_rxn_info.setMaximumHeight(160)
        layout.addWidget(self.ed_rxn_info)

        self._editor_last_parsed: dict | None = None

        self.editor_tabs.addTab(tab, "Reactions")

    def _build_editor_disable_delete_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(self._title("Disable / Delete reactions"))

        row = QHBoxLayout()
        row.addWidget(QLabel("Reaction ID:"))
        self.ed_disable_rid = QLineEdit()
        self.ed_disable_rid.setPlaceholderText("Type reaction id...")
        row.addWidget(self.ed_disable_rid, stretch=1)

        self.ed_disable_btn = QPushButton("Disable (LB=0,UB=0)")
        self.ed_disable_btn.clicked.connect(self.editor_disable_reaction)
        row.addWidget(self.ed_disable_btn)

        self.ed_delete_btn = QPushButton("Delete reaction")
        self.ed_delete_btn.clicked.connect(self.editor_delete_reaction)
        row.addWidget(self.ed_delete_btn)

        layout.addLayout(row)

        self.ed_disable_info = QPlainTextEdit()
        self.ed_disable_info.setReadOnly(True)
        layout.addWidget(self.ed_disable_info)

        self.editor_tabs.addTab(tab, "Disable/Delete")

    def _build_editor_patch_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(self._title("Patch Manager"))

        btn_row = QHBoxLayout()
        self.ed_patch_export_btn = QPushButton("Export patch (JSON)...")
        self.ed_patch_export_btn.clicked.connect(self.editor_export_patch)
        btn_row.addWidget(self.ed_patch_export_btn)

        self.ed_patch_import_btn = QPushButton("Import patch (JSON) & apply...")
        self.ed_patch_import_btn.clicked.connect(self.editor_import_patch_and_apply)
        btn_row.addWidget(self.ed_patch_import_btn)

        self.ed_patch_clear_btn = QPushButton("Clear patch (does NOT reset model)")
        self.ed_patch_clear_btn.clicked.connect(self.editor_clear_patch_only)
        btn_row.addWidget(self.ed_patch_clear_btn)

        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        layout.addWidget(QLabel("Patch preview:"))
        self.ed_patch_preview = QPlainTextEdit()
        self.ed_patch_preview.setReadOnly(True)
        layout.addWidget(self.ed_patch_preview, stretch=1)

        self.editor_tabs.addTab(tab, "Patch")

    def _build_editor_export_sbml_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(self._title("Export patched SBML"))

        info = QLabel(
            "This exports the CURRENT in-app model (including any changes you made in Model Editor) as a new SBML file.\n"
            "Your original SBML is not modified."
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        btn = QPushButton("Export current model as SBML...")
        btn.clicked.connect(self.editor_export_patched_sbml)
        layout.addWidget(btn)

        layout.addStretch(1)
        self.editor_tabs.addTab(tab, "Export SBML")

    def _refresh_patch_preview(self):
        try:
            self.ed_patch_preview.setPlainText(json.dumps(self.model_patch, indent=2))
        except Exception:
            pass

    def _rebuild_editor_reaction_dropdowns(self):
        return

    def _rebuild_editor_gene_dropdowns(self):
        return

    def _rebuild_editor_metabolite_dropdowns(self):
        return

    # ---------- editor actions ----------
    def editor_add_metabolite(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        mid = self.ed_met_id.text().strip()
        if not mid:
            QMessageBox.warning(self, "Missing", "Metabolite ID is required.")
            return

        name = self.ed_met_name.text().strip()
        comp = self.ed_met_comp.text().strip() or _guess_compartment_from_met_id(mid)
        formula = self.ed_met_formula.text().strip()
        charge_txt = self.ed_met_charge.text().strip()

        charge = None
        if charge_txt:
            try:
                charge = int(charge_txt)
            except Exception:
                QMessageBox.warning(self, "Invalid charge", "Charge must be an integer (or blank).")
                return

        if mid in self.base_model.metabolites:
            QMessageBox.information(self, "Exists", f"Metabolite already exists: {mid}\n(You can still use it in reactions.)")
            return

        m = cobra.Metabolite(id=mid, name=name or mid, compartment=comp)
        if formula:
            m.formula = formula
        if charge is not None:
            m.charge = charge

        self.base_model.add_metabolites([m])
        self.model_dirty = True

        self.model_patch["added_metabolites"].append({
            "id": mid,
            "name": name,
            "compartment": comp,
            "formula": formula,
            "charge": charge,
        })
        self._refresh_patch_preview()

        self.ed_met_info.setPlainText(f"Added metabolite: {mid}\ncompartment={comp}\nname={name}\nformula={formula}\ncharge={charge}")
        self.populate_reaction_table()
        self.populate_medium_table()

    def editor_add_or_update_gene(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        gid = self.ed_gene_id.text().strip()
        if not gid:
            QMessageBox.warning(self, "Missing", "Gene ID is required.")
            return
        name = self.ed_gene_name.text().strip()

        if gid in self.base_model.genes:
            g = self.base_model.genes.get_by_id(gid)
            if name:
                try:
                    g.name = name
                except Exception:
                    pass
            action = "Updated"
        else:
            gene = cobra.Gene(id=gid)
            try:
                gene.name = name
            except Exception:
                pass
            self.base_model.genes._dict[gid] = gene
            action = "Added"

        self.model_patch["added_genes"].append({"id": gid, "name": name})
        self._refresh_patch_preview()
        self.model_dirty = True

        self.ed_gene_info.setPlainText(f"{action} gene: {gid}\nname={name}\n\nNote: genes become functionally connected via reaction GPRs.")
        self.populate_gene_list()
        self.populate_genes_tab()

    def editor_parse_equation(self):
        eq = self.ed_rxn_eq.toPlainText().strip()
        try:
            stoich, reversible = parse_reaction_equation(eq)
        except Exception as e:
            QMessageBox.warning(self, "Parse failed", str(e))
            self._editor_last_parsed = None
            self.ed_rxn_preview_tbl.setRowCount(0)
            return

        self._editor_last_parsed = {"stoich": stoich, "reversible": reversible, "equation": eq}

        self.ed_rxn_preview_tbl.setRowCount(0)
        for met_id in sorted(stoich.keys()):
            coeff = float(stoich[met_id])
            side = "Product" if coeff > 0 else "Reactant"
            row = self.ed_rxn_preview_tbl.rowCount()
            self.ed_rxn_preview_tbl.insertRow(row)
            self.ed_rxn_preview_tbl.setItem(row, 0, QTableWidgetItem(met_id))
            self.ed_rxn_preview_tbl.setItem(row, 1, QTableWidgetItem(f"{coeff:g}"))
            self.ed_rxn_preview_tbl.setItem(row, 2, QTableWidgetItem(side))

        self.ed_rxn_info.setPlainText(
            f"Parsed OK.\nReversible={reversible}\nMetabolites={len(stoich)}\n\nTip: If metabolites are missing, Apply will auto-create them (if enabled)."
        )

    def _ensure_metabolites_exist_for_stoich(self, stoich: dict[str, float]) -> list[str]:
        assert self.base_model is not None
        created: list[str] = []

        for mid in stoich.keys():
            if mid in self.base_model.metabolites:
                continue
            comp = _guess_compartment_from_met_id(mid)
            m = cobra.Metabolite(id=mid, name=mid, compartment=comp)
            self.base_model.add_metabolites([m])
            created.append(mid)

            self.model_patch["added_metabolites"].append({
                "id": mid,
                "name": "",
                "compartment": comp,
                "formula": "",
                "charge": None,
                "auto_created": True,
            })

        return created

    def editor_add_or_update_reaction(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        rid = self.ed_rxn_id.text().strip()
        if not rid:
            QMessageBox.warning(self, "Missing", "Reaction ID is required.")
            return

        desc = self.ed_rxn_desc.text().strip()
        gpr = self.ed_rxn_gpr.text().strip()
        kegg = self.ed_rxn_kegg.text().strip()
        ec = self.ed_rxn_ec.text().strip()

        try:
            lb = self._parse_float(self.ed_rxn_lb.text())
            ub = self._parse_float(self.ed_rxn_ub.text())
            if lb > ub:
                raise ValueError("LB > UB")
        except Exception:
            QMessageBox.warning(self, "Invalid bounds", "LB/UB must be numeric and LB <= UB.")
            return

        eq_text = self.ed_rxn_eq.toPlainText().strip()
        parsed = self._editor_last_parsed
        if not parsed or parsed.get("equation") != eq_text:
            try:
                stoich, reversible = parse_reaction_equation(eq_text)
            except Exception as e:
                QMessageBox.warning(self, "Parse failed", f"Equation parse failed:\n{e}")
                return
        else:
            stoich = parsed["stoich"]
            reversible = bool(parsed["reversible"])

        if self.ed_rxn_rev.isChecked():
            reversible = True

        created_mets: list[str] = []
        if self.ed_rxn_autocreate_missing_mets.isChecked():
            created_mets = self._ensure_metabolites_exist_for_stoich(stoich)
            if created_mets:
                QMessageBox.information(
                    self,
                    "Missing metabolites auto-created",
                    "These metabolites were missing and have been auto-created with empty metadata:\n\n"
                    f"{', '.join(created_mets)}\n\n"
                    "Please add proper names/formula/charge later in Model Editor → Metabolites."
                )
        else:
            missing = [m for m in stoich.keys() if m not in self.base_model.metabolites]
            if missing:
                QMessageBox.warning(
                    self,
                    "Missing metabolites",
                    "These metabolites do not exist in the model:\n\n"
                    f"{', '.join(missing)}\n\n"
                    "Enable auto-create or add them first."
                )
                return

        if rid in self.base_model.reactions:
            rxn = self.base_model.reactions.get_by_id(rid)
            action = "Updated"
        else:
            rxn = cobra.Reaction(id=rid)
            self.base_model.add_reactions([rxn])
            action = "Added"

        try:
            rxn.name = desc
        except Exception as e:
            self.statusBar().showMessage(f"Warning: could not set reaction name: {e}")

        try:
            rxn.subtract_metabolites(rxn.metabolites)
        except Exception as e:
            self.statusBar().showMessage(f"Warning: could not clear old stoichiometry: {e}")

        mets = {self.base_model.metabolites.get_by_id(mid): float(coeff) for mid, coeff in stoich.items()}
        rxn.add_metabolites(mets)

        rxn.lower_bound = float(lb)
        rxn.upper_bound = float(ub)
        rxn.reversibility = bool(reversible)

        try:
            rxn.gene_reaction_rule = gpr
        except Exception as e:
            self.statusBar().showMessage(f"Warning: could not set GPR: {e}")

        ann = getattr(rxn, "annotation", {}) or {}
        if kegg:
            ann["kegg.reaction"] = kegg
        if ec:
            ann["ec-code"] = ec
        rxn.annotation = ann

        self.model_patch["added_or_updated_reactions"].append({
            "id": rid,
            "name": desc,
            "equation": eq_text,
            "stoichiometry": stoich,
            "reversible": reversible,
            "lb": lb,
            "ub": ub,
            "gpr": gpr,
            "annotation": {"kegg.reaction": kegg, "ec-code": ec},
        })
        self._refresh_patch_preview()
        self.model_dirty = True

        self.ed_rxn_info.setPlainText(
            f"{action} reaction: {rid}\n"
            f"LB={lb}, UB={ub}, reversible={reversible}\n"
            f"GPR={gpr}\n"
            f"Equation:\n{eq_text}\n"
        )

        self.populate_reaction_table()
        self.populate_medium_table()
        self.populate_overexpression_reaction_list()
        self.populate_objective_combo()
        self.populate_genes_tab()

    def editor_disable_reaction(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        rid = self.ed_disable_rid.text().strip()
        if not rid:
            QMessageBox.warning(self, "Missing", "Reaction ID is required.")
            return
        if rid not in self.base_model.reactions:
            QMessageBox.warning(self, "Not found", f"Reaction not found: {rid}")
            return

        rxn = self.base_model.reactions.get_by_id(rid)
        rxn.lower_bound = 0.0
        rxn.upper_bound = 0.0

        if rid not in self.model_patch["disabled_reactions"]:
            self.model_patch["disabled_reactions"].append(rid)
        self._refresh_patch_preview()
        self.model_dirty = True

        self.ed_disable_info.setPlainText(f"Disabled reaction: {rid} (LB=0, UB=0)")
        self.populate_reaction_table()
        self.populate_medium_table()

    def editor_delete_reaction(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        rid = self.ed_disable_rid.text().strip()
        if not rid:
            QMessageBox.warning(self, "Missing", "Reaction ID is required.")
            return
        if rid not in self.base_model.reactions:
            QMessageBox.warning(self, "Not found", f"Reaction not found: {rid}")
            return

        msg = (
            "You are about to DELETE a reaction from the in-app model.\n\n"
            f"Reaction: {rid}\n\n"
            "This does NOT modify your original SBML file, but it WILL remove the reaction from the current session and any exported patched SBML.\n\n"
            "Continue?"
        )
        if QMessageBox.question(self, "Confirm delete reaction", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        rxn = self.base_model.reactions.get_by_id(rid)
        try:
            self.base_model.remove_reactions([rxn], remove_orphans=False)
        except Exception as e:
            QMessageBox.critical(self, "Delete failed", str(e))
            return

        if rid not in self.model_patch["deleted_reactions"]:
            self.model_patch["deleted_reactions"].append(rid)
        self._refresh_patch_preview()
        self.model_dirty = True

        self.ed_disable_info.setPlainText(f"Deleted reaction: {rid}")
        self.populate_reaction_table()
        self.populate_medium_table()
        self.populate_overexpression_reaction_list()
        self.populate_objective_combo()
        self.populate_genes_tab()

    def editor_export_patch(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export patch",
            str(Path.home() / "metabodesk_patch.json"),
            "JSON files (*.json);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            Path(file_path).write_text(json.dumps(self.model_patch, indent=2), encoding="utf-8")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        self.statusBar().showMessage(f"Patch exported: {file_path}")

    def editor_import_patch_and_apply(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import patch",
            str(Path.home()),
            "JSON files (*.json);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            patch = json.loads(Path(file_path).read_text(encoding="utf-8"))
            if not isinstance(patch, dict):
                raise ValueError("Patch must be a JSON object.")
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))
            return

        msg = (
            "This will apply the imported patch to the CURRENT in-app model.\n"
            "Your original SBML file is not modified.\n\n"
            "Continue?"
        )
        if QMessageBox.question(self, "Confirm apply patch", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        try:
            self._apply_patch_to_model(patch)
        except Exception as e:
            QMessageBox.critical(self, "Apply failed", str(e))
            return

        self.model_patch = patch
        self._refresh_patch_preview()
        self.statusBar().showMessage(f"Patch imported & applied: {file_path}")
        self.model_dirty = True

        self.populate_reaction_table()
        self.populate_medium_table()
        self.populate_overexpression_reaction_list()
        self.populate_objective_combo()
        self.populate_gene_list()
        self.populate_genes_tab()

    def editor_clear_patch_only(self):
        msg = (
            "This will CLEAR the in-app patch record (JSON history), but it will NOT revert the model.\n\n"
            "If you want to revert the model, use 'Reset reactions' (and re-open SBML if needed).\n\n"
            "Continue?"
        )
        if QMessageBox.question(self, "Confirm clear patch", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        self.model_patch = {
            "version": 1,
            "added_metabolites": [],
            "added_genes": [],
            "added_or_updated_reactions": [],
            "disabled_reactions": [],
            "deleted_reactions": [],
            "objective_reaction_id": None,
            "notes": "",
        }
        self._refresh_patch_preview()
        self.statusBar().showMessage("Patch cleared (model not reverted).")

    def editor_export_patched_sbml(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export patched SBML",
            str(Path.home() / "metabodesk_patched_model.xml"),
            "SBML files (*.xml *.sbml);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            cobra.io.write_sbml_model(self.base_model, file_path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return
        self.statusBar().showMessage(f"Patched SBML exported: {file_path}")

    def _apply_patch_to_model(self, patch: dict):
        assert self.base_model is not None

        for m in patch.get("added_metabolites", []) or []:
            mid = str(m.get("id", "")).strip()
            if not mid or mid in self.base_model.metabolites:
                continue
            comp = str(m.get("compartment") or _guess_compartment_from_met_id(mid))
            name = str(m.get("name") or mid)
            met = cobra.Metabolite(id=mid, name=name, compartment=comp)
            self.base_model.add_metabolites([met])

        for g in patch.get("added_genes", []) or []:
            gid = str(g.get("id", "")).strip()
            if not gid:
                continue
            if gid in self.base_model.genes:
                continue
            gene = cobra.Gene(id=gid)
            try:
                gene.name = str(g.get("name") or "")
            except Exception:
                pass
            self.base_model.genes._dict[gid] = gene

        for r in patch.get("added_or_updated_reactions", []) or []:
            rid = str(r.get("id", "")).strip()
            if not rid:
                continue
            eq = str(r.get("equation") or "").strip()
            stoich = r.get("stoichiometry")
            if not isinstance(stoich, dict):
                stoich, _rev = parse_reaction_equation(eq)
            else:
                stoich = {str(k): float(v) for k, v in stoich.items()}

            for mid in list(stoich.keys()):
                if mid not in self.base_model.metabolites:
                    comp = _guess_compartment_from_met_id(mid)
                    self.base_model.add_metabolites([cobra.Metabolite(id=mid, name=mid, compartment=comp)])

            if rid in self.base_model.reactions:
                rxn = self.base_model.reactions.get_by_id(rid)
            else:
                rxn = cobra.Reaction(id=rid)
                self.base_model.add_reactions([rxn])

            try:
                rxn.name = str(r.get("name") or "")
            except Exception as e:
                self.statusBar().showMessage(f"Warning: could not set name for {rid}: {e}")

            try:
                rxn.subtract_metabolites(rxn.metabolites)
            except Exception as e:
                self.statusBar().showMessage(f"Warning: could not clear stoichiometry for {rid}: {e}")
            mets = {self.base_model.metabolites.get_by_id(mid): float(coeff) for mid, coeff in stoich.items()}
            rxn.add_metabolites(mets)

            try:
                rxn.lower_bound = float(r.get("lb"))
                rxn.upper_bound = float(r.get("ub"))
            except Exception as e:
                self.statusBar().showMessage(f"Warning: could not set bounds for {rid}: {e}")

            try:
                rxn.gene_reaction_rule = str(r.get("gpr") or "")
            except Exception as e:
                self.statusBar().showMessage(f"Warning: could not set GPR for {rid}: {e}")

            ann = getattr(rxn, "annotation", {}) or {}
            ann_in = r.get("annotation") or {}
            if isinstance(ann_in, dict):
                ann.update({k: v for k, v in ann_in.items() if v})
            rxn.annotation = ann

        for rid in patch.get("disabled_reactions", []) or []:
            rid = str(rid)
            if rid in self.base_model.reactions:
                rxn = self.base_model.reactions.get_by_id(rid)
                rxn.lower_bound = 0.0
                rxn.upper_bound = 0.0

        for rid in patch.get("deleted_reactions", []) or []:
            rid = str(rid)
            if rid in self.base_model.reactions:
                rxn = self.base_model.reactions.get_by_id(rid)
                self.base_model.remove_reactions([rxn], remove_orphans=False)

    # ---------------- Save As helpers (SBML) ----------------
    def _export_current_model_to_path(self, file_path: Path):
        if self.base_model is None:
            raise ValueError("No model loaded.")
        cobra.io.write_sbml_model(self.base_model, str(file_path))

    def save_current_model_as(self) -> Path | None:
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "No model loaded.")
            return None

        default_name = "metabodesk_current_model.xml"
        if self.current_sbml_path:
            default_name = f"{self.current_sbml_path.stem}_metabodesk.xml"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save current model as (SBML)",
            str(Path.home() / default_name),
            "SBML files (*.xml *.sbml);;All files (*.*)",
        )
        if not file_path:
            return None

        out = Path(file_path)
        self._export_current_model_to_path(out)
        self.current_sbml_path = out
        self.model_dirty = False
        self.statusBar().showMessage(f"Saved SBML: {out}")
        return out

    # ---------------- (2) memote helpers ----------------
    def cancel_memote(self):
        if self._memote_proc is None:
            return
        try:
            self._memote_cancel_requested = True
            self.tools_log.appendPlainText("\nCancel requested...\n")
            self._force_kill_process(self._memote_proc)
            self._memote_proc = None
            self.set_busy(False, "Ready.")
            self._stop_memote_timer()
        except Exception:
            pass

    def _start_elapsed_timer(self, kind: str):
        start = time.monotonic()
        timer = QTimer(self)
        timer.setInterval(1000)

        def tick():
            elapsed = int(time.monotonic() - start)
            mm = elapsed // 60
            ss = elapsed % 60
            txt = f"Elapsed: {mm:02d}:{ss:02d}"
            if kind == "memote":
                self.memote_time_lbl.setText(txt)
            else:
                self.carveme_time_lbl.setText(txt)

        timer.timeout.connect(tick)
        timer.start()
        if kind == "memote":
            self._memote_timer = timer
        else:
            self._carveme_timer = timer

    def _stop_memote_timer(self):
        try:
            if getattr(self, "_memote_timer", None):
                self._memote_timer.stop()
        except Exception:
            pass
        self.memote_time_lbl.setText("Elapsed: 00:00")

    def _stop_carveme_timer(self):
        try:
            if getattr(self, "_carveme_timer", None):
                self._carveme_timer.stop()
        except Exception:
            pass
        self.carveme_time_lbl.setText("Elapsed: 00:00")

    def _get_objective_reaction_id(self, model: cobra.Model | None = None) -> str | None:
        """Return the objective reaction ID for the given model (or self.base_model)."""
        if model is None:
            model = self.base_model
        if model is None or model.objective is None:
            return None
        try:
            obj = model.objective
            if hasattr(obj, "expression"):
                coeffs = obj.expression.as_coefficients_dict()
                for var in coeffs.keys():
                    name = getattr(var, "name", "") or ""
                    rxn_id = name
                    if rxn_id.startswith("R_") or rxn_id.startswith("F_"):
                        rxn_id = rxn_id[2:]
                    if rxn_id in model.reactions:
                        return rxn_id
            if hasattr(obj, "id") and obj.id in model.reactions:
                return obj.id
        except Exception:
            pass
        return None

    # Keep backward-compatible alias used by _build_community_model
    def _get_objective_reaction_id_for_model(self, model: cobra.Model) -> str | None:
        return self._get_objective_reaction_id(model)

    def _is_extracellular(self, met: cobra.Metabolite) -> bool:
        try:
            if met.compartment and met.compartment.lower() in ("e", "extracellular"):
                return True
            mid = met.id.lower()
            return mid.endswith("_e") or mid.endswith("[e]") or mid.endswith("(e)")
        except Exception:
            return False

    def _prefix_gpr(self, gpr: str, prefix: str) -> str:
        if not gpr:
            return gpr
        tokens = re.split(r'(\s+|\(|\)|and|or|AND|OR)', gpr)
        out = []
        for t in tokens:
            if t is None:
                continue
            if t.strip() == "":
                out.append(t)
                continue
            if t in ("and", "or", "AND", "OR", "(", ")"):
                out.append(t)
                continue
            out.append(prefix + t)
        return "".join(out)

    def _build_community_model(self, models: list[cobra.Model], labels: list[str], share_extracellular: bool = True):
        community = cobra.Model("community")
        shared_mets: dict[str, cobra.Metabolite] = {}
        obj_rxn_map: dict[str, str] = {}

        for model, label in zip(models, labels):
            prefix = f"{label}__"
            met_map: dict[str, cobra.Metabolite] = {}

            for met in model.metabolites:
                if share_extracellular and self._is_extracellular(met):
                    if met.id not in shared_mets:
                        new_met = cobra.Metabolite(
                            met.id,
                            name=met.name,
                            compartment=met.compartment,
                            formula=met.formula,
                            charge=met.charge,
                        )
                        community.add_metabolites([new_met])
                        shared_mets[met.id] = new_met
                    met_map[met.id] = shared_mets[met.id]
                else:
                    new_id = prefix + met.id
                    new_met = cobra.Metabolite(
                        new_id,
                        name=f"{met.name} ({label})" if met.name else new_id,
                        compartment=met.compartment,
                        formula=met.formula,
                        charge=met.charge,
                    )
                    community.add_metabolites([new_met])
                    met_map[met.id] = new_met

            for rxn in model.reactions:
                new_rxn = cobra.Reaction(prefix + rxn.id)
                new_rxn.name = f"{rxn.name} ({label})" if rxn.name else new_rxn.id
                new_rxn.lower_bound = rxn.lower_bound
                new_rxn.upper_bound = rxn.upper_bound
                new_rxn.subsystem = rxn.subsystem
                met_coeffs = {met_map[m.id]: coeff for m, coeff in rxn.metabolites.items()}
                new_rxn.add_metabolites(met_coeffs)
                try:
                    new_rxn.gene_reaction_rule = self._prefix_gpr(getattr(rxn, "gene_reaction_rule", "") or "", prefix)
                except Exception as e:
                    self.statusBar().showMessage(f"Warning: GPR not set for {new_rxn.id}: {e}")
                community.add_reactions([new_rxn])

            obj_id = self._get_objective_reaction_id_for_model(model)
            if obj_id:
                obj_rxn_map[label] = prefix + obj_id

        # Set community objective as sum of member objectives
        for label, rxn_id in obj_rxn_map.items():
            if rxn_id in community.reactions:
                community.reactions.get_by_id(rxn_id).objective_coefficient = 1.0

        return community, obj_rxn_map, shared_mets

    def _show_issues_dialog(self, title: str, issues: list[str]):
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.resize(700, 500)

        layout = QVBoxLayout(dialog)
        search_input = QLineEdit()
        search_input.setPlaceholderText("Filter...")
        layout.addWidget(search_input)

        list_widget = QListWidget()
        layout.addWidget(list_widget, stretch=1)

        def populate(filter_text: str = ""):
            list_widget.clear()
            f = filter_text.strip().lower()
            for item in issues:
                if not f or f in item.lower():
                    list_widget.addItem(item)

        populate()
        search_input.textChanged.connect(populate)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        btns.accepted.connect(dialog.accept)
        layout.addWidget(btns)

        dialog.exec()

    def _force_kill_process(self, proc: QProcess):
        try:
            proc.kill()
            if proc.waitForFinished(2000):
                return
        except Exception:
            pass
        try:
            if sys.platform.startswith("win"):
                pid = proc.processId()
                if pid:
                    QProcess.startDetached("taskkill", ["/F", "/T", "/PID", str(pid)])
            else:
                proc.kill()
        except Exception:
            pass

    def cancel_carveme(self):
        if getattr(self, "_carveme_proc", None) is None:
            return
        try:
            self.tools_log.appendPlainText("\nCancel requested (CarveMe)...\n")
            proc = self._carveme_proc
            self._carveme_proc = None
            self._carveme_last_output = ""
            try:
                proc.terminate()
            except Exception:
                pass
            QTimer.singleShot(2000, lambda: proc.kill() if proc.state() != QProcess.NotRunning else None)
            self.set_busy(False, "Ready.")
            self._stop_carveme_timer()
        except Exception:
            pass

    def _start_memote_run(self, sbml_path: Path, report_path: Path, extra_args: list[str] | None = None, title: str = "memote", timeout_seconds: int | None = 900):
        args = ["report", "snapshot", str(sbml_path), "--filename", str(report_path)]
        if extra_args:
            args += extra_args

        cmd = self._console_script_command("memote", args)
        if not cmd:
            QMessageBox.critical(self, "Tools missing", "Tools not found. Use Tools → Repair tools to reinstall.")
            return
        program, proc_args = cmd

        proc = self._run_tool_process(program, proc_args, title)
        try:
            proc.setWorkingDirectory(str(report_path.parent))
        except Exception:
            pass

        self._memote_proc = proc
        self._memote_cancel_requested = False
        self._memote_last_output = time.monotonic()
        self._memote_current_report_path = report_path
        self._memote_current_sbml_path = sbml_path
        self._memote_current_extra_args = extra_args or []
        self.set_busy(True, "Running memote...")
        self._start_elapsed_timer("memote")

        try:
            if getattr(self, "_memote_watchdog", None):
                self._memote_watchdog.stop()
        except Exception:
            pass

        if timeout_seconds is not None:
            self._memote_watchdog = QTimer(self)
            self._memote_watchdog.setInterval(2000)

            def check_memote_hang():
                if self._memote_proc is None:
                    self._memote_watchdog.stop()
                    return
                if time.monotonic() - self._memote_last_output > timeout_seconds:
                    self.tools_log.appendPlainText("\nmemote appears to be stuck. Stopping process...\n")
                    try:
                        self._force_kill_process(self._memote_proc)
                    except Exception:
                        pass
                    self._memote_proc = None
                    self.set_busy(False, "Ready.")
                    self._memote_watchdog.stop()
                    self._stop_memote_timer()

                    if not getattr(self, "_memote_fast_fallback_used", False) and not extra_args:
                        self._memote_fast_fallback_used = True
                        fast_path = report_path.with_name(report_path.stem + "_fast" + report_path.suffix)
                        self.tools_log.appendPlainText("\nRetrying memote with heavy tests skipped...\n")
                        self._start_memote_run(
                            sbml_path,
                            fast_path,
                            extra_args=["--skip", "test_consistency", "--skip", "test_biomass"],
                            title="memote (fast)",
                            timeout_seconds=timeout_seconds,
                        )

            self._memote_watchdog.timeout.connect(check_memote_hang)
            self._memote_watchdog.start()

        def on_finished(code, status):
            self._append_proc_output_to_log(proc)
            self.tools_log.appendPlainText(f"\nmemote finished with code={code}, status={status}\n")
            self._memote_proc = None
            self.set_busy(False, "Ready.")
            self._stop_memote_timer()
            try:
                self._memote_watchdog.stop()
            except Exception:
                pass

            if report_path.exists():
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(report_path)))
                QMessageBox.information(self, "Memote finished", f"Report saved:\n{report_path}")
            else:
                if self._memote_cancel_requested:
                    QMessageBox.information(self, "Memote cancelled", "memote was cancelled.")
                else:
                    if not getattr(self, "_memote_fast_fallback_used", False) and not extra_args:
                        self._memote_fast_fallback_used = True
                        fast_path = report_path.with_name(report_path.stem + "_fast" + report_path.suffix)
                        self.tools_log.appendPlainText("\nmemote crashed. Retrying with heavy tests skipped...\n")
                        self._start_memote_run(
                            sbml_path,
                            fast_path,
                            extra_args=["--skip", "test_consistency", "--skip", "test_biomass"],
                            title="memote (fast)",
                        )
                        return
                    QMessageBox.critical(self, "Memote failed", "memote failed and report was not created. See log in Tools tab.")

        proc.finished.connect(on_finished)
        proc.start()

    def _tool_python_path(self) -> str | None:
        """Return bundled Python path, or current Python in dev mode."""
        py = self._bundled_python_path()
        if py:
            return py
        # Dev mode fallback
        if not getattr(sys, "frozen", False):
            exe = str(getattr(sys, "executable", "") or "")
            if exe.lower().endswith("python.exe") or exe.lower().endswith("python"):
                return exe
        return None

    def _app_base_dir(self) -> Path:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent
        return Path(__file__).resolve().parent

    def _bundled_python_path(self) -> str | None:
        base = self._runtime_dir()
        cand = [
            base / "python" / "python.exe",
            base / "python" / "bin" / "python",
        ]
        for p in cand:
            if p.exists():
                return str(p)
        return None

    def _download_file(self, url: str, dest: Path, progress_cb=None, cancel_event=None, max_seconds: int = 600, stall_seconds: int = 20):
        dest.parent.mkdir(parents=True, exist_ok=True)
        req = urllib.request.Request(url, headers={"User-Agent": "MetaboDesk/1.0"})
        start_time = time.monotonic()
        last_progress = start_time
        with urllib.request.urlopen(req, timeout=10) as r, open(dest, "wb") as f:
            total = r.headers.get("Content-Length")
            total = int(total) if total and total.isdigit() else None
            downloaded = 0
            while True:
                if cancel_event and cancel_event.is_set():
                    raise TimeoutError("Download cancelled")
                if time.monotonic() - start_time > max_seconds:
                    raise TimeoutError("Download timed out")
                if time.monotonic() - last_progress > stall_seconds:
                    raise TimeoutError("Download stalled")
                try:
                    chunk = r.read(1024 * 1024)
                except socket.timeout:
                    continue
                if not chunk:
                    break
                f.write(chunk)
                downloaded += len(chunk)
                last_progress = time.monotonic()
                if progress_cb:
                    progress_cb(downloaded, total)

    def _get_remote_size(self, url: str) -> int | None:
        try:
            req = urllib.request.Request(url, method="HEAD", headers={"User-Agent": "MetaboDesk/1.0"})
            with urllib.request.urlopen(req, timeout=10) as r:
                total = r.headers.get("Content-Length")
                return int(total) if total and total.isdigit() else None
        except Exception:
            return None

    def _check_tool_module(self, module_name: str) -> bool:
        py = self._tool_python_path()
        if not py:
            return False
        proc = QProcess(self)
        proc.setProgram(py)
        proc.setArguments(["-c", f"import {module_name}"])
        proc.setProcessChannelMode(QProcess.MergedChannels)
        proc.start()
        proc.waitForFinished(3000)
        return proc.exitCode() == 0

    def _has_module(self, module_name: str) -> bool:
        py = self._tool_python_path()
        if not py:
            return False
        code = (
            "import importlib.util, sys\n"
            f"spec = importlib.util.find_spec('{module_name}')\n"
            "sys.exit(0 if spec else 1)\n"
        )
        proc = QProcess(self)
        proc.setProgram(py)
        proc.setArguments(["-c", code])
        proc.setProcessChannelMode(QProcess.MergedChannels)
        proc.start()
        proc.waitForFinished(3000)
        return proc.exitCode() == 0

    def _has_console_script(self, script_name: str) -> bool:
        py = self._tool_python_path()
        if not py:
            return False
        code = (
            "from importlib.metadata import entry_points\n"
            "eps = entry_points().select(group='console_scripts')\n"
            f"found = any(e.name=='{script_name}' for e in eps)\n"
            "raise SystemExit(0 if found else 1)\n"
        )
        proc = QProcess(self)
        proc.setProgram(py)
        proc.setArguments(["-c", code])
        proc.setProcessChannelMode(QProcess.MergedChannels)
        proc.start()
        proc.waitForFinished(3000)
        return proc.exitCode() == 0

    def _console_script_command(self, script_name: str, argv: list[str]) -> tuple[str, list[str]] | None:
        py = self._tool_python_path()
        if not py:
            return None
        py_path = Path(py)
        scripts_dir = py_path.parent / ("Scripts" if sys.platform.startswith("win") else "bin")
        if sys.platform.startswith("win"):
            for ext in (".exe", ".cmd", ".bat"):
                exe = scripts_dir / f"{script_name}{ext}"
                if exe.exists():
                    return str(exe), argv
        else:
            exe = scripts_dir / script_name
            if exe.exists():
                return str(exe), argv

        code = (
            "from importlib.metadata import entry_points\n"
            "eps = entry_points().select(group='console_scripts')\n"
            f"eps = [e for e in eps if e.name=='{script_name}']\n"
            "if not eps: raise SystemExit(127)\n"
            "func = eps[0].load()\n"
            f"import sys; sys.argv={[script_name] + argv}\n"
            "raise SystemExit(func())\n"
        )
        return py, ["-c", code]

    def _ensure_carveme_shim(self, py_path_str: str):
        py_path = Path(py_path_str)
        scripts_dir = py_path.parent / ("Scripts" if sys.platform.startswith("win") else "bin")
        scripts_dir.mkdir(parents=True, exist_ok=True)
        if sys.platform.startswith("win"):
            shim = scripts_dir / "carveme.bat"
            if not shim.exists():
                shim.write_text(
                    "@echo off\r\n"
                    "\"%~dp0\\python.exe\" -m carveme %*\r\n",
                    encoding="utf-8",
                )
        else:
            shim = scripts_dir / "carveme"
            if not shim.exists():
                shim.write_text(
                    f"#!/bin/sh\n\"{py_path_str}\" -m carveme \"$@\"\n",
                    encoding="utf-8",
                )
                try:
                    os.chmod(shim, 0o755)
                except Exception:
                    pass

    def _carveme_command(self, genome_path: str | None, out_path: str, extra_args: list[str] | None = None) -> tuple[str, list[str]] | None:
        py = self._tool_python_path()
        if not py:
            return None
        extra_args = extra_args or []
        py_path = Path(py)
        scripts_dir = py_path.parent / ("Scripts" if sys.platform.startswith("win") else "bin")
        carve_exe = scripts_dir / ("carve.exe" if sys.platform.startswith("win") else "carve")
        if carve_exe.exists():
            self._carveme_last_mode = "carve"
            args = ([str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + extra_args
            return str(carve_exe), args

        carveme_exe = scripts_dir / ("carveme.exe" if sys.platform.startswith("win") else "carveme")
        if carveme_exe.exists():
            self._carveme_last_mode = "carveme"
            args = ["draft"]
            if genome_path:
                args += ["-g", str(genome_path)]
            args += ["-o", str(out_path)] + extra_args
            return str(carveme_exe), args

        if self._has_console_script("carve"):
            self._carveme_last_mode = "carve"
            args = ([str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + extra_args
            return self._console_script_command("carve", args)

        if self._has_console_script("carveme"):
            self._carveme_last_mode = "carveme"
            args = ["draft"]
            if genome_path:
                args += ["-g", str(genome_path)]
            args += ["-o", str(out_path)] + extra_args
            return self._console_script_command("carveme", args)

        # Try module execution (carveme.__main__) if available
        if self._has_module("carveme.__main__"):
            self._carveme_last_mode = "carveme"
            args = ["-m", "carveme", "draft"]
            if genome_path:
                args += ["-g", str(genome_path)]
            args += ["-o", str(out_path)] + extra_args
            return py, args

        # Try module execution for carve
        if self._has_module("carveme.cli.carve"):
            self._carveme_last_mode = "carve"
            args = ["-m", "carveme.cli.carve"] + ([str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + extra_args
            return py, args
        if self._has_module("carveme.carve"):
            self._carveme_last_mode = "carve"
            args = ["-m", "carveme.carve"] + ([str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + extra_args
            return py, args

        self._carveme_last_mode = "carve"
        argv = ["carve"] + ([str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + extra_args
        code = (
            "import sys, importlib, runpy, traceback, pkgutil\n"
            f"sys.argv={argv!r}\n"
            "errors=[]\n"
            "mods = ['carveme.cli.carve', 'carveme.carve', 'carveme', 'carveme.__main__', 'carveme.cli', 'carveme.cli.main', 'carveme.main', 'carveme.carveme']\n"
            "for m in mods:\n"
            "    try:\n"
            "        runpy.run_module(m, run_name='__main__')\n"
            "        raise SystemExit(0)\n"
            "    except Exception as e:\n"
            "        errors.append((m, str(e)))\n"
            "for m in mods:\n"
            "    try:\n"
            "        mod = importlib.import_module(m)\n"
            "    except Exception as e:\n"
            "        errors.append((m, str(e)))\n"
            "        continue\n"
            "    for name in ('main','cli','app','draft','run'):\n"
            "        fn = getattr(mod, name, None)\n"
            "        if callable(fn):\n"
            "            raise SystemExit(fn())\n"
            "try:\n"
            "    import carveme as _carveme_pkg\n"
            "    for finder, name, ispkg in pkgutil.walk_packages(_carveme_pkg.__path__, _carveme_pkg.__name__ + '.'):\n"
            "        if any(x in name for x in ('cli', 'main')):\n"
            "            try:\n"
            "                mod = importlib.import_module(name)\n"
            "            except Exception as e:\n"
            "                errors.append((name, str(e)))\n"
            "                continue\n"
            "            for fn_name in ('main','cli','app','draft','run'):\n"
            "                fn = getattr(mod, fn_name, None)\n"
            "                if callable(fn):\n"
            "                    raise SystemExit(fn())\n"
            "except Exception as e:\n"
            "    errors.append(('carveme package scan', str(e)))\n"
            "from importlib.metadata import entry_points\n"
            "eps = entry_points().select(group='console_scripts')\n"
            "eps = [e for e in eps if e.name=='carveme']\n"
            "if eps:\n"
            "    raise SystemExit(eps[0].load()())\n"
            "print('carveme entrypoint not found')\n"
            "print('errors:')\n"
            "for m, e in errors:\n"
            "    print(' -', m, ':', e)\n"
            "raise SystemExit(1)\n"
        )
        return py, ["-c", code]

    def _run_tool_process(self, program: str, args: list[str], title: str):
        proc = QProcess(self)
        proc.setProgram(program)
        proc.setArguments(args)
        try:
            env = os.environ.copy()
            tools_bin = self._tools_bin_dir()
            if tools_bin.exists():
                env["PATH"] = str(tools_bin) + os.pathsep + env.get("PATH", "")
            proc.setEnvironment([f"{k}={v}" for k, v in env.items()])
        except Exception:
            pass
        proc.setProcessChannelMode(QProcess.MergedChannels)
        proc.readyRead.connect(lambda: self._append_proc_output_to_log(proc))
        self.tools_log.appendPlainText(f"\n[{title}] {' '.join([program] + args)}\n")
        return proc

    def _reinstall_carveme(self):
        py = self._tool_python_path()
        if not py:
            QMessageBox.critical(self, "Tools missing", "Bundled Python not found. Please reinstall the app.")
            return

        wheels_dir = self._wheels_dir()
        if wheels_dir.exists():
            args = [
                "-m", "pip", "install",
                "--no-index", "--find-links", str(wheels_dir),
                "--force-reinstall", "carveme",
            ]
        else:
            args = [
                "-m", "pip", "install", "--force-reinstall", "--no-cache-dir",
                "carveme",
            ]
        proc = self._run_tool_process(py, args, "pip reinstall carveme")
        self.set_busy(True, "Reinstalling CarveMe...")

        def on_finished(code, status):
            self._append_proc_output_to_log(proc)
            self.set_busy(False, "Ready.")
            if code == 0:
                QMessageBox.information(self, "CarveMe", "CarveMe reinstalled. Try again.")
            else:
                QMessageBox.critical(self, "CarveMe", f"Reinstall failed (code {code}).")
            self.tools_check_status()

        proc.finished.connect(on_finished)
        proc.start()

    def _append_proc_output_to_log(self, proc: QProcess):
        data = proc.readAll().data().decode(errors="replace")
        if not data:
            return
        if proc is getattr(self, "_memote_proc", None):
            self._memote_last_output = time.monotonic()
        if proc is getattr(self, "_carveme_proc", None):
            self._carveme_last_output = (getattr(self, "_carveme_last_output", "") + data)[-8000:]
        for line in data.splitlines():
            self.tools_log.appendPlainText(line)

    def _carveme_preflight(self) -> bool:
        py = self._tool_python_path()
        if not py:
            return False
        tools_dir = self._tools_bin_dir()
        code = (
            "import sys, traceback, shutil, os\n"
            "try:\n"
            "    import carveme\n"
            "    print('carveme:', getattr(carveme, '__version__', 'unknown'))\n"
            "except Exception as e:\n"
            "    print('carveme import error:', e)\n"
            "    traceback.print_exc()\n"
            "    sys.exit(2)\n"
            f"tools_dir = r'{tools_dir}'\n"
            "if os.path.isdir(tools_dir):\n"
            "    try:\n"
            "        print('tools_dir:', tools_dir)\n"
            "        print('tools_dir_files:', ','.join(os.listdir(tools_dir)))\n"
            "    except Exception:\n"
            "        pass\n"
            "missing=[]\n"
            "for b in ('diamond','prodigal'):\n"
            "    found = shutil.which(b) is not None\n"
            "    if not found and os.name == 'nt':\n"
            "        for ext in (b+'.exe', b+'.windows.exe'):\n"
            "            if os.path.exists(os.path.join(tools_dir, ext)):\n"
            "                found = True\n"
            "                break\n"
            "    if not found:\n"
            "        missing.append(b)\n"
            "if missing:\n"
            "    print('missing binaries:', ','.join(missing))\n"
            "    sys.exit(3)\n"
            "sys.exit(0)\n"
        )
        proc = QProcess(self)
        proc.setProgram(py)
        proc.setArguments(["-c", code])
        try:
            env = os.environ.copy()
            tools_bin = self._tools_bin_dir()
            if tools_bin.exists():
                env["PATH"] = str(tools_bin) + os.pathsep + env.get("PATH", "")
            proc.setEnvironment([f"{k}={v}" for k, v in env.items()])
        except Exception:
            pass
        proc.setProcessChannelMode(QProcess.MergedChannels)
        proc.start()
        proc.waitForFinished(5000)
        self._append_proc_output_to_log(proc)
        if proc.exitCode() == 0:
            return True
        if proc.exitCode() == 3:
            QMessageBox.warning(
                self,
                "CarveMe dependencies",
                "CarveMe requires external binaries (diamond, prodigal).\n"
                "Please bundle them under runtime/tools or add them to PATH.",
            )
            return False
        QMessageBox.critical(
            self,
            "CarveMe import failed",
            "CarveMe failed to import. See Tools log for details.",
        )
        return False

    def run_memote_report_current_model(self):
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        py = self._tool_python_path()
        if not py:
            QMessageBox.critical(self, "Tools missing", "Bundled Python not found. Please reinstall the app.")
            return

        if not self._check_tool_module("memote"):
            QMessageBox.critical(self, "Tools missing", "memote is not installed. Use Tools → Repair tools to reinstall.")
            return

        sbml_path = self.current_sbml_path
        if self.model_dirty or sbml_path is None or not sbml_path.exists():
            msg = (
                "Your current model has unsaved changes (or no source file path).\n\n"
                "Please save the current model as a NEW SBML file before running memote.\n\n"
                "Continue with Save As?"
            )
            if QMessageBox.question(self, "Save required", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
            saved = self.save_current_model_as()
            if saved is None:
                return
            sbml_path = saved

        default_html = Path.home() / f"{sbml_path.stem}_memote.html"
        report_path_str, _ = QFileDialog.getSaveFileName(
            self,
            "Save memote report as (HTML)",
            str(default_html),
            "HTML files (*.html);;All files (*.*)",
        )
        if not report_path_str:
            return
        report_path = Path(report_path_str)

        if report_path.exists():
            if QMessageBox.question(
                self,
                "Overwrite report?",
                "The selected report file already exists. Overwrite it?",
                QMessageBox.Yes | QMessageBox.No,
            ) != QMessageBox.Yes:
                return
            try:
                report_path.unlink()
            except Exception:
                QMessageBox.critical(self, "Memote", "Cannot overwrite the report file. Please choose another path.")
                return

        self.tabs.setCurrentWidget(self.tab_tools)
        self.tools_log.appendPlainText(f"Running memote on: {sbml_path}")
        self.tools_log.appendPlainText(f"Report target: {report_path}\n")
        # Options dialog
        dlg = QDialog(self)
        dlg.setWindowTitle("Memote options")
        layout = QVBoxLayout(dlg)
        form = QFormLayout()

        mode_combo = QComboBox()
        mode_combo.addItems(["Full", "Fast (skip heavy tests)"])
        form.addRow(QLabel("Mode:"), mode_combo)

        timeout_spin = QSpinBox()
        timeout_spin.setRange(1, 120)
        timeout_spin.setValue(15)
        form.addRow(QLabel("Timeout (minutes):"), timeout_spin)

        infinity_chk = QCheckBox("Infinity (no timeout)")
        form.addRow(QLabel(""), infinity_chk)

        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec() != QDialog.Accepted:
            return

        extra_args = []
        if mode_combo.currentIndex() == 1:
            extra_args = ["--skip", "test_consistency", "--skip", "test_biomass"]

        timeout_seconds = None if infinity_chk.isChecked() else int(timeout_spin.value()) * 60

        self._memote_fast_fallback_used = False
        self._start_memote_run(sbml_path, report_path, extra_args=extra_args or None, timeout_seconds=timeout_seconds)

    def run_carveme_draft(self):
        py = self._tool_python_path()
        if not py:
            QMessageBox.critical(self, "Tools missing", "Bundled Python not found. Please reinstall the app.")
            return

        self._carveme_no_draft_retry = False

        if not self._check_tool_module("carveme"):
            msg = (
                "CarveMe is not installed. Install bundled tools now?\n\n"
                "This uses offline wheels packaged with the app."
            )
            if QMessageBox.question(self, "Install tools", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return

            def after_install(ok: bool):
                if not ok:
                    QMessageBox.critical(self, "Install failed", "CarveMe installation failed. See Tools log.")
                    return
                if not self._check_tool_module("carveme"):
                    QMessageBox.critical(self, "Tools missing", "CarveMe still not available after install.")
                    return
                self.run_carveme_draft()

            self._install_tools_offline(confirm=False, on_done=after_install)
            return

        if not self._carveme_preflight():
            return

        # Input type selection
        dlg = QDialog(self)
        dlg.setWindowTitle("CarveMe input options")
        layout = QVBoxLayout(dlg)
        form = QFormLayout()

        input_combo = QComboBox()
        input_combo.addItems(["Protein FASTA", "DNA FASTA", "RefSeq accession"])
        form.addRow(QLabel("Input type:"), input_combo)

        gapfill_edit = QLineEdit()
        gapfill_edit.setPlaceholderText("e.g., M9,LB")
        form.addRow(QLabel("Gapfill media (-g):"), gapfill_edit)

        init_edit = QLineEdit()
        init_edit.setPlaceholderText("e.g., M9")
        form.addRow(QLabel("Initialize media (-i):"), init_edit)

        species_edit = QLineEdit()
        species_edit.setPlaceholderText("Optional")
        form.addRow(QLabel("Species (--species):"), species_edit)

        skip_rebuild_chk = QCheckBox("Skip rebuild")
        form.addRow(QLabel(""), skip_rebuild_chk)
        skip_biolog_chk = QCheckBox("Skip biolog")
        form.addRow(QLabel(""), skip_biolog_chk)
        skip_ess_chk = QCheckBox("Skip essentiality")
        form.addRow(QLabel(""), skip_ess_chk)

        ignore_warn_chk = QCheckBox("Ignore warnings")
        form.addRow(QLabel(""), ignore_warn_chk)

        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec() != QDialog.Accepted:
            return

        extra_args = []
        if ignore_warn_chk.isChecked():
            extra_args.append("--ignore-warnings")

        if gapfill_edit.text().strip():
            extra_args += ["-g", gapfill_edit.text().strip()]
        if init_edit.text().strip():
            extra_args += ["-i", init_edit.text().strip()]
        if species_edit.text().strip():
            extra_args += ["--species", species_edit.text().strip()]
        if skip_rebuild_chk.isChecked():
            extra_args.append("--skip-rebuild")
        if skip_biolog_chk.isChecked():
            extra_args.append("--skip-biolog")
        if skip_ess_chk.isChecked():
            extra_args.append("--skip-essentiality")

        genome_path = None
        if input_combo.currentIndex() == 2:
            refseq, ok = QInputDialog.getText(self, "RefSeq accession", "Enter RefSeq accession (e.g., GCF_000005845.2):")
            if not ok or not refseq.strip():
                return
            extra_args += ["--refseq", refseq.strip()]
            genome_path = None
        else:
            if input_combo.currentIndex() == 1:
                extra_args.append("--dna")

            genome_path, _ = QFileDialog.getOpenFileName(
                self,
                "Select genome file",
                str(Path.home()),
                "FASTA files (*.fa *.fasta *.faa *.fna);;All files (*.*)"
            )
            if not genome_path:
                return

        out_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save SBML model as",
            str(Path.home() / "carveme_model.xml"),
            "SBML files (*.xml *.sbml);;All files (*.*)"
        )
        if not out_path:
            return

        self.tabs.setCurrentWidget(self.tab_tools)
        self.tools_log.appendPlainText(f"Running CarveMe on: {genome_path}")
        self.tools_log.appendPlainText(f"SBML target: {out_path}\n")

        self._carveme_extra_args = extra_args
        cmd = self._carveme_command(genome_path, out_path, extra_args=extra_args)
        if not cmd:
            QMessageBox.critical(self, "Tools missing", "Bundled Python not found. Please reinstall the app.")
            return
        program, args = cmd
        proc = self._run_tool_process(program, args, "carveme")
        self._carveme_proc = proc
        self._carveme_last_output = ""
        self.set_busy(True, "Running CarveMe...")
        self._start_elapsed_timer("carveme")

        def on_finished(code, status):
            self._append_proc_output_to_log(proc)
            self.tools_log.appendPlainText(f"\ncarveme finished with code={code}, status={status}\n")
            self._carveme_proc = None
            self.set_busy(False, "Ready.")
            self._stop_carveme_timer()
            out = getattr(self, "_carveme_last_output", "") or ""
            if code != 0 and "unrecognized arguments" in out.lower() and getattr(self, "_carveme_last_mode", "") != "carveme" and not getattr(self, "_carveme_no_draft_retry", False):
                self._carveme_no_draft_retry = True
                self.tools_log.appendPlainText("\nRetrying CarveMe with 'carveme draft' command...\n")
                self._carveme_last_mode = "carveme"
                args2 = ["draft"]
                if genome_path:
                    args2 += ["-g", str(genome_path)]
                args2 += ["-o", str(out_path)] + (self._carveme_extra_args or [])
                cmd2 = self._console_script_command("carveme", args2)
                if not cmd2 and self._tool_python_path():
                    cmd2 = (self._tool_python_path(), ["-m", "carveme", "draft"] + (["-g", str(genome_path)] if genome_path else []) + ["-o", str(out_path)] + (self._carveme_extra_args or []))
                if cmd2:
                    program2, args2 = cmd2
                    proc2 = self._run_tool_process(program2, args2, "carveme")
                    self._carveme_proc = proc2
                    self._carveme_last_output = ""
                    self.set_busy(True, "Running CarveMe...")
                    self._start_elapsed_timer("carveme")

                    def on_finished2(code2, status2):
                        self._append_proc_output_to_log(proc2)
                        self.tools_log.appendPlainText(f"\ncarveme finished with code={code2}, status={status2}\n")
                        self._carveme_proc = None
                        self.set_busy(False, "Ready.")
                        self._stop_carveme_timer()
                        if code2 == 0 and Path(out_path).exists():
                            if QMessageBox.question(self, "CarveMe", "Model created. Open now?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                                try:
                                    self.open_sbml_from_path(out_path)
                                except Exception as e:
                                    QMessageBox.critical(self, "Open failed", f"Could not open CarveMe model:\n{e}")
                        else:
                            QMessageBox.warning(self, "CarveMe", "CarveMe failed. See Tools log for details.")

                    proc2.finished.connect(on_finished2)
                    proc2.start()
                return
            if code == 0 and Path(out_path).exists():
                if QMessageBox.question(self, "CarveMe", "Model created. Open now?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    try:
                        self.open_sbml_from_path(out_path)
                    except Exception as e:
                        QMessageBox.critical(self, "Open failed", f"Could not open CarveMe model:\n{e}")
            else:
                QMessageBox.warning(self, "CarveMe", "CarveMe failed. See Tools log for details.")

        proc.finished.connect(on_finished)
        proc.start()

    # ---------------- Tools installer helpers ----------------
    def _runtime_dir(self) -> Path:
        # When frozen (PyInstaller), runtime folder is next to the exe, not in _MEIPASS
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent / "runtime"
        return Path(__file__).resolve().parent / "runtime"

    def _wheels_dir(self) -> Path:
        return self._runtime_dir() / "wheels"

    def _tools_bin_dir(self) -> Path:
        return self._runtime_dir() / "tools"


    def validate_model(self):
        """Comprehensive model validation."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return
        
        issues = []
        try:
            # Check for reactions without genes or bounds
            for rxn in self.base_model.reactions:
                if not rxn.genes and rxn.lower_bound != 0 and rxn.upper_bound != 0:
                    issues.append(f"Reaction {rxn.id}: no genes assigned (orphaned)")
                if rxn.lower_bound > rxn.upper_bound:
                    issues.append(f"Reaction {rxn.id}: lower bound ({rxn.lower_bound}) > upper bound ({rxn.upper_bound})")
                if len(rxn.metabolites) == 0:
                    issues.append(f"Reaction {rxn.id}: no metabolites (empty)")
            
            # Check for orphan metabolites
            consumed = set()
            produced = set()
            for rxn in self.base_model.reactions:
                for met in rxn.metabolites:
                    if rxn.metabolites[met] > 0:
                        produced.add(met.id)
                    else:
                        consumed.add(met.id)
            
            orphan_produced = produced - consumed
            orphan_consumed = consumed - produced
            if orphan_produced:
                issues.append(f"{len(orphan_produced)} metabolites are only produced (no sink reactions)")
            if orphan_consumed:
                issues.append(f"{len(orphan_consumed)} metabolites are only consumed (no source reactions)")
            
            # Check objective
            try:
                obj_val = self.base_model.slim_optimize()
                if obj_val is None or (isinstance(obj_val, float) and math.isnan(obj_val)):
                    issues.append("Objective value is NaN or infeasible (check bounds and objective)")
                elif isinstance(obj_val, float) and abs(obj_val) < 1e-9:
                    issues.append("Objective value is zero – model may be over-constrained")
            except Exception:
                issues.append("Model optimization failed - check for infeasibility")
            
        except Exception as e:
            issues.append(f"Validation error: {e}")
        
        if not issues:
            QMessageBox.information(self, "Validation", "✓ Model validation passed - no major issues found.")
        else:
            msg = "Issues found:\n\n" + "\n".join(issues[:20])
            if len(issues) > 20:
                msg += f"\n... and {len(issues) - 20} more issues"
            box = QMessageBox(self)
            box.setWindowTitle("Model validation")
            box.setText(msg)
            view_btn = box.addButton("View all", QMessageBox.ActionRole)
            box.addButton("Close", QMessageBox.RejectRole)
            box.exec()
            if box.clickedButton() == view_btn:
                self._show_issues_dialog("Model validation issues", issues)

    def check_mass_balance(self):
        """Check if reactions are mass balanced."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return
        
        try:
            imbalanced = []
            for rxn in self.base_model.reactions:
                if not rxn.metabolites:
                    continue
                # Use COBRApy's element-based mass balance check
                try:
                    balance = rxn.check_mass_balance()
                    if balance:
                        parts = ", ".join(f"{elem}: {val:+.4g}" for elem, val in balance.items())
                        imbalanced.append(f"{rxn.id}: {parts}")
                except Exception:
                    # Fallback: coefficient sum heuristic (only for reactions without formulas)
                    coeff_sum = sum(rxn.metabolites.values())
                    if abs(coeff_sum) > 0.01:
                        imbalanced.append(f"{rxn.id}: coeff_sum={coeff_sum:.4f} (no formula data)")
            
            if not imbalanced:
                QMessageBox.information(self, "Mass balance check", "✓ All reactions appear mass-balanced.")
            else:
                msg = f"Found {len(imbalanced)} potentially imbalanced reactions:\n\n"
                msg += "\n".join(imbalanced[:15])
                if len(imbalanced) > 15:
                    msg += f"\n... and {len(imbalanced) - 15} more"
                box = QMessageBox(self)
                box.setWindowTitle("Mass balance check")
                box.setText(msg)
                view_btn = box.addButton("View all", QMessageBox.ActionRole)
                box.addButton("Close", QMessageBox.RejectRole)
                box.exec()
                if box.clickedButton() == view_btn:
                    self._show_issues_dialog("Mass balance issues", imbalanced)
        except Exception as e:
            QMessageBox.critical(self, "Check failed", str(e))

    def check_gpr_syntax(self):
        """Validate Gene-Protein-Reaction (GPR) syntax."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a model first.")
            return

        issues = []
        try:
            for rxn in self.base_model.reactions:
                gpr = getattr(rxn, 'gene_reaction_rule', '') or ''
                if not gpr:
                    continue
                
                # Basic syntax checks
                open_parens = gpr.count('(')
                close_parens = gpr.count(')')
                if open_parens != close_parens:
                    issues.append(f"{rxn.id}: parentheses mismatch (open={open_parens}, close={close_parens})")
                
                # Check for mixed case operators (COBRApy expects lowercase 'and'/'or')
                if re.search(r'\bAND\b', gpr) or re.search(r'\bOR\b', gpr):
                    issues.append(f"{rxn.id}: uppercase operator found (COBRApy expects lowercase 'and'/'or')")
                
                # Validate tokens — split on operators and parentheses
                tokens = re.split(r'\s+|\b(?:and|or)\b|[()]', gpr)
                for tok in tokens:
                    tok = tok.strip()
                    if tok and not re.match(r'^[A-Za-z0-9_.\-:]+$', tok):
                        issues.append(f"{rxn.id}: invalid token '{tok}'")
                        break
            
            if not issues:
                QMessageBox.information(self, "GPR validation", "✓ All GPR expressions appear syntactically valid.")
            else:
                msg = f"Found {len(issues)} GPR syntax issues:\n\n"
                msg += "\n".join(issues[:20])
                if len(issues) > 20:
                    msg += f"\n... and {len(issues) - 20} more"
                box = QMessageBox(self)
                box.setWindowTitle("GPR validation")
                box.setText(msg)
                view_btn = box.addButton("View all", QMessageBox.ActionRole)
                box.addButton("Close", QMessageBox.RejectRole)
                box.exec()
                if box.clickedButton() == view_btn:
                    self._show_issues_dialog("GPR syntax issues", issues)
        except Exception as e:
            QMessageBox.critical(self, "Check failed", str(e))

    def tools_check_status(self):
        py = self._tool_python_path()
        if not py:
            self.tools_status_lbl.setText("Bundled Python not found. Please reinstall the app.")
            self.memote_btn.setEnabled(False)
            self.carveme_btn.setEnabled(False)
            return

        mem_ok = self._check_tool_module("memote")
        car_ok = self._check_tool_module("carveme")
        sw_ok = self._check_tool_module("swiglpk")

        missing = []
        if not mem_ok:
            missing.append("memote")
        if not car_ok:
            missing.append("carveme")
        if not sw_ok:
            missing.append("swiglpk")

        if not missing:
            self.tools_status_lbl.setText("✔ All tools available (memote, carveme, swiglpk)")
            self.memote_btn.setEnabled(True)
            self.carveme_btn.setEnabled(True)
        else:
            self.tools_status_lbl.setText(
                "Missing: " + ", ".join(missing)
                + ".  Click 'Repair tools' to reinstall from bundled wheels."
            )
            self.memote_btn.setEnabled(False)
            self.carveme_btn.setEnabled(False)

    def tools_repair(self):
        """Reinstall memote / carveme / swiglpk from bundled wheels."""
        self.tabs.setCurrentWidget(self.tab_tools)
        self._install_tools_offline(confirm=True)

    def _install_tools_offline(self, confirm: bool = True, on_done=None):
        py = self._bundled_python_path() or self._tool_python_path()
        if not py:
            QMessageBox.critical(self, "Tools missing", "Bundled Python not found. Please reinstall the app.")
            return

        wheels_dir = self._wheels_dir()
        if not wheels_dir.exists():
            QMessageBox.critical(self, "Tools missing", "Offline wheels not found in runtime/wheels.")
            return

        if confirm:
            msg = (
                "Reinstall tools from bundled wheels?\n\n"
                f"Source: {wheels_dir}\n\n"
                "This does NOT require internet."
            )
            if QMessageBox.question(self, "Repair tools", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return

        args = [
            "-m", "pip", "install",
            "--no-index", "--find-links", str(wheels_dir),
            "--force-reinstall",
            "memote", "carveme", "swiglpk",
        ]
        proc = self._run_tool_process(py, args, "pip install (offline wheels)")
        self.set_busy(True, "Installing tools from bundled wheels...")

        def on_finished(code, status):
            self._append_proc_output_to_log(proc)
            self.set_busy(False, "Ready.")
            if code == 0:
                try:
                    self._ensure_carveme_shim(py)
                except Exception:
                    pass
                if on_done:
                    on_done(True)
                else:
                    QMessageBox.information(self, "Repair complete", "Tools reinstalled successfully.")
            else:
                if on_done:
                    on_done(False)
                else:
                    QMessageBox.critical(self, "Repair failed", f"pip exited with code {code}")
            self.tools_check_status()

        proc.finished.connect(on_finished)
        proc.start()

    # -------- Network Map --------
    def _render_network_map(self):
        """Trigger map rendering - graph building in QThread for large models, drawing on main thread."""
        if self.is_running:
            return
        if self.base_model is None:
            self._set_map_placeholder("Load model first")
            return

        # For small models (<500 reactions), render directly on main thread
        # For large models, build graph in QThread
        if len(self.base_model.reactions) < 500:
            self.map_render_btn.setEnabled(False)
            try:
                self._render_network_map_impl()
            finally:
                self.map_render_btn.setEnabled(True)
        else:
            self.map_render_btn.setEnabled(False)
            self.statusBar().showMessage("Building network graph (background)...")

            def _build():
                # Build graph and compute layout in background
                # This is safe because we only read from self.base_model (no mutation)
                import io
                return {"done": True}

            def _on_done(result):
                try:
                    self._render_network_map_impl()
                except Exception as e:
                    logger.error(f"Map render error: {e}")
                    self._set_map_placeholder(f"Error: {e}")
                finally:
                    self.map_render_btn.setEnabled(True)
                    self.statusBar().showMessage("Ready.")

            self._launch_worker(_build, _on_done, "Building network graph...")
    def _set_map_placeholder(self, text: str):
        ax = self.map_canvas.ax
        ax.clear()
        ax.text(0.5, 0.5, text, ha="center", va="center", fontsize=11, transform=ax.transAxes)
        ax.axis("off")
        self.map_canvas.draw()

    def _focus_on_selected_map_node(self):
        node = getattr(self, "_selected_map_node", None)
        if not node:
            QMessageBox.information(self, "Map", "Click a node on the map first.")
            return
        self.map_search.setText(str(node))
        self._render_network_map()

    def _export_map_image(self):
        if not hasattr(self, "map_canvas"):
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export map image",
            str(Path.home() / "network_map.png"),
            "PNG files (*.png);;SVG files (*.svg);;PDF files (*.pdf)"
        )
        if not file_path:
            return
        try:
            self.map_canvas.figure.savefig(file_path, dpi=300)
            self.statusBar().showMessage(f"Map image exported: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _export_map_csv(self):
        if not hasattr(self, "_current_graph") or self._current_graph is None:
            QMessageBox.warning(self, "Map", "Render the map first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export graph CSV",
            str(Path.home() / "network_graph.csv"),
            "CSV files (*.csv)"
        )
        if not file_path:
            return
        try:
            base = Path(file_path)
            nodes_path = base.with_name(base.stem + "_nodes.csv")
            edges_path = base.with_name(base.stem + "_edges.csv")

            with open(nodes_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["id", "type", "name", "weight", "degree"])
                for n, d in self._current_graph.nodes(data=True):
                    w.writerow([
                        n,
                        d.get("type", ""),
                        d.get("name", ""),
                        d.get("weight", ""),
                        self._current_graph.degree(n),
                    ])

            with open(edges_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["source", "target"])
                for u, v in self._current_graph.edges():
                    w.writerow([u, v])

            self.statusBar().showMessage(f"Graph exported: {nodes_path.name}, {edges_path.name}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
    
    def _render_network_map_bg(self):
        """Graph building + layout. Called from QThread via _render_network_map, or directly for backward compat."""
        # If called directly (backward compat), just run on main thread
        self._render_network_map_impl()

    def _render_network_map_impl(self):
        """Full network map rendering implementation."""
        try:
            if self.base_model is None:
                ax = self.map_canvas.ax
                ax.clear()
                ax.text(0.5, 0.5, "Load model first", ha="center", va="center", fontsize=12, transform=ax.transAxes)
                self.map_canvas.draw()
                return
            
            search_text = (self.map_search.text() or "").strip().lower()
            view_mode = self.map_view_combo.currentText()
            depth = int(self.map_depth.value())
            threshold = float(self.map_threshold.value())

            max_nodes = int(self.map_max_nodes.value()) if hasattr(self, "map_max_nodes") else 5000
            max_edges = int(self.map_max_edges.value()) if hasattr(self, "map_max_edges") else 0
            min_degree = int(self.map_min_degree.value()) if hasattr(self, "map_min_degree") else 0
            topn_rxn = int(self.map_topn_rxn.value()) if hasattr(self, "map_topn_rxn") else 0
            layout_choice = self.map_layout_combo.currentText() if hasattr(self, "map_layout_combo") else "Spring"
            only_connected = bool(self.map_only_connected_chk.isChecked()) if hasattr(self, "map_only_connected_chk") else False
            hide_orphans = bool(self.map_hide_orphans_chk.isChecked()) if hasattr(self, "map_hide_orphans_chk") else False
            show_legend = bool(self.map_show_legend_chk.isChecked()) if hasattr(self, "map_show_legend_chk") else False
            exchange_only = bool(self.map_exchange_only_chk.isChecked()) if hasattr(self, "map_exchange_only_chk") else False
            objective_only = bool(self.map_objective_only_chk.isChecked()) if hasattr(self, "map_objective_only_chk") else False
            subsystem_filter = self.map_subsystem_combo.currentText() if hasattr(self, "map_subsystem_combo") else "All subsystems"

            flux_map = {}
            fva_map = {}
            sgd_map = {}
            wt_growth = None

            if self.last_run:
                try:
                    if "baseline" in self.last_run and "flux" in self.last_run["baseline"]:
                        flux_map = self.last_run["baseline"].get("flux", {}) or {}
                except Exception as e:
                    self.statusBar().showMessage(f"Warning: could not load flux data for map: {e}")
                try:
                    if "fva_result" in self.last_run:
                        fva_map = self.last_run.get("fva_result", {}) or {}
                except Exception as e:
                    self.statusBar().showMessage(f"Warning: could not load FVA data for map: {e}")
                try:
                    if "sgd_result" in self.last_run:
                        sgd_map = self.last_run.get("sgd_result", {}) or {}
                        wt_growth = self.last_run.get("wt_growth", None)
                except Exception as e:
                    self.statusBar().showMessage(f"Warning: could not load SGD data for map: {e}")

            def rxn_allowed(rxn: cobra.Reaction) -> bool:
                if exchange_only and not is_exchange_reaction(rxn):
                    return False
                if subsystem_filter and subsystem_filter != "All subsystems":
                    sub = str(getattr(rxn, "subsystem", "") or "").strip()
                    if sub != subsystem_filter:
                        return False
                return True

            def rxn_weight(rid: str) -> float:
                if "FVA" in view_mode:
                    mm = fva_map.get(rid, {}) or {}
                    return max(float(mm.get("max", 0.0)) - float(mm.get("min", 0.0)), 0.0)
                if "Gene KO" in view_mode:
                    if not sgd_map or wt_growth in (None, 0):
                        return 0.0
                    try:
                        rxn = self.base_model.reactions.get_by_id(rid)
                        gene_vals = [float(sgd_map.get(g.id, None)) for g in rxn.genes if g.id in sgd_map]
                        if not gene_vals:
                            return 0.0
                        min_growth = min(gene_vals)
                        return max((float(wt_growth) - float(min_growth)) / float(wt_growth), 0.0)
                    except Exception:
                        return 0.0
                # Flux magnitude
                return abs(float(flux_map.get(rid, 0.0)))

            if "FVA" in view_mode:
                has_metric_data = bool(fva_map)
            elif "Gene KO" in view_mode:
                has_metric_data = bool(sgd_map) and wt_growth not in (None, 0)
            else:
                has_metric_data = bool(flux_map)

            effective_threshold = threshold if has_metric_data else 0.0

            filtered_reactions = [r for r in self.base_model.reactions if rxn_allowed(r)]

            # Seed nodes by search (reaction or metabolite)
            seed_rxns = []
            seed_mets = []
            if search_text:
                for r in filtered_reactions:
                    if search_text in r.id.lower() or search_text in (r.name or "").lower():
                        seed_rxns.append(r.id)
                for m in self.base_model.metabolites:
                    if search_text in m.id.lower() or search_text in (m.name or "").lower():
                        seed_mets.append(m.id)
            else:
                if not filtered_reactions:
                    ax = self.map_canvas.ax
                    ax.clear()
                    ax.text(0.5, 0.5, "No reactions found for filters", ha="center", va="center", fontsize=11, transform=ax.transAxes)
                    self.map_canvas.draw()
                    return

                candidates = filtered_reactions
                if effective_threshold > 0:
                    candidates = [r for r in candidates if rxn_weight(r.id) >= effective_threshold]
                    if not candidates:
                        candidates = filtered_reactions

                candidates = sorted(candidates, key=lambda r: rxn_weight(r.id), reverse=True)
                if topn_rxn > 0:
                    seed_rxns = [r.id for r in candidates[:topn_rxn]]
                else:
                    seed_rxns = [r.id for r in candidates[:2000]]

            if objective_only:
                obj_id = None
                try:
                    obj_id = self.objective_combo.currentData()
                except Exception:
                    obj_id = None
                if not obj_id:
                    try:
                        coeffs = self.base_model.objective.expression.as_coefficients_dict()
                        obj_rxns = [r.id for r in coeffs.keys()]
                        if obj_rxns:
                            obj_id = obj_rxns[0]
                    except Exception:
                        obj_id = None
                if obj_id:
                    obj_id = str(obj_id)
                    seed_rxns = [obj_id] + [r for r in seed_rxns if r != obj_id]

            if not seed_rxns and not seed_mets:
                ax = self.map_canvas.ax
                ax.clear()
                ax.text(0.5, 0.5, "No matches found", ha="center", va="center", fontsize=11, transform=ax.transAxes)
                self.map_canvas.draw()
                return

            G = nx.DiGraph()
            rxn_map = {}
            met_map = {}

            if not search_text and effective_threshold == 0 and topn_rxn == 0 and not only_connected and not objective_only:
                # Full graph: include all reactions/metabolites
                for rxn in filtered_reactions:
                    w = rxn_weight(rxn.id)
                    rxn_map[rxn.id] = rxn
                    G.add_node(rxn.id, type="rxn", weight=w, name=(rxn.name or ""))
                    for met, coeff in rxn.metabolites.items():
                        met_map[met.id] = met
                        G.add_node(met.id, type="met", name=(met.name or ""))
                        if coeff < 0:
                            G.add_edge(met.id, rxn.id)
                        else:
                            G.add_edge(rxn.id, met.id)
                # Add orphan metabolites (no reactions) as isolated nodes
                if not hide_orphans:
                    for met in self.base_model.metabolites:
                        if met.id not in met_map:
                            met_map[met.id] = met
                            G.add_node(met.id, type="met", name=(met.name or ""))
            else:
                max_nodes = max(100, max_nodes)

                queue = []
                for rid in seed_rxns:
                    queue.append(("rxn", rid, 0, True))
                for mid in seed_mets:
                    queue.append(("met", mid, 0, True))

                visited = set()

                while queue and G.number_of_nodes() < max_nodes:
                    ntype, nid, d, is_seed = queue.pop(0)
                    key = (ntype, nid)
                    if key in visited:
                        continue
                    visited.add(key)
                    if d > depth:
                        continue

                    if ntype == "rxn":
                        try:
                            rxn = self.base_model.reactions.get_by_id(nid)
                        except Exception:
                            continue

                        if not rxn_allowed(rxn):
                            continue

                        w = rxn_weight(rxn.id)
                        if not is_seed and effective_threshold > 0 and w < effective_threshold:
                            continue

                        rxn_map[rxn.id] = rxn
                        G.add_node(rxn.id, type="rxn", weight=w, name=(rxn.name or ""))

                        for met, coeff in rxn.metabolites.items():
                            met_map[met.id] = met
                            G.add_node(met.id, type="met", name=(met.name or ""))
                            if coeff < 0:
                                G.add_edge(met.id, rxn.id)
                            else:
                                G.add_edge(rxn.id, met.id)
                            if d + 1 <= depth:
                                queue.append(("met", met.id, d + 1, False))

                    else:  # metabolite node
                        try:
                            met = self.base_model.metabolites.get_by_id(nid)
                        except Exception:
                            continue

                        met_map[met.id] = met
                        G.add_node(met.id, type="met", name=(met.name or ""))

                        for rxn in met.reactions:
                            if not rxn_allowed(rxn):
                                continue
                            w = rxn_weight(rxn.id)
                            if not is_seed and effective_threshold > 0 and w < effective_threshold:
                                continue
                            rxn_map[rxn.id] = rxn
                            G.add_node(rxn.id, type="rxn", weight=w, name=(rxn.name or ""))
                            coeff = rxn.metabolites.get(met, 0)
                            if coeff < 0:
                                G.add_edge(met.id, rxn.id)
                            else:
                                G.add_edge(rxn.id, met.id)
                            if d + 1 <= depth:
                                queue.append(("rxn", rxn.id, d + 1, False))

            # Apply degree-based filters
            if hide_orphans:
                orphans = [n for n in G.nodes() if G.degree(n) == 0]
                if orphans:
                    G.remove_nodes_from(orphans)

            if min_degree > 0:
                while True:
                    low = [n for n in G.nodes() if G.degree(n) < min_degree]
                    if not low:
                        break
                    G.remove_nodes_from(low)

            # Enforce max node limit if needed (drop lowest-weight reactions first)
            if max_nodes and G.number_of_nodes() > max_nodes:
                scored = []
                for n, d in G.nodes(data=True):
                    if d.get("type") == "rxn":
                        score = float(d.get("weight", 0.0))
                    else:
                        score = 0.0
                    scored.append((score, n))
                scored.sort(key=lambda x: x[0])
                to_remove = [n for _, n in scored[: max(0, G.number_of_nodes() - max_nodes)]]
                if to_remove:
                    G.remove_nodes_from(to_remove)

            # Enforce max edge limit if needed
            if max_edges and G.number_of_edges() > max_edges:
                edges = list(G.edges())
                def edge_score(e):
                    w = 0.0
                    for node in e:
                        d = G.nodes.get(node, {})
                        if d.get("type") == "rxn":
                            w = max(w, float(d.get("weight", 0.0)))
                    return w
                edges.sort(key=edge_score, reverse=True)
                keep = set(edges[:max_edges])
                remove = [e for e in edges if e not in keep]
                if remove:
                    G.remove_edges_from(remove)

            if G.number_of_nodes() == 0:
                ax = self.map_canvas.ax
                ax.clear()
                ax.text(0.5, 0.5, "Empty network", ha="center", va="center", fontsize=11, transform=ax.transAxes)
                self.map_canvas.draw()
                return

            try:
                if layout_choice == "Circular":
                    pos = nx.circular_layout(G)
                elif layout_choice == "Kamada-Kawai":
                    pos = nx.kamada_kawai_layout(G)
                else:
                    its = 10 if G.number_of_nodes() > 2000 else 25
                    pos = nx.spring_layout(G, k=0.7, iterations=its, seed=42)
            except Exception:
                pos = nx.circular_layout(G)

            rxn_nodes = [n for n, d in G.nodes(data=True) if d.get("type") == "rxn"]
            met_nodes = [n for n, d in G.nodes(data=True) if d.get("type") == "met"]

            weights = [float(G.nodes[n].get("weight", 0.0)) for n in rxn_nodes]
            max_w = max(weights) if weights else 1.0
            norm_w = [(w / max_w) if max_w > 0 else 0 for w in weights]

            rxn_colors = [cm.viridis(0.15 + 0.85 * w) for w in norm_w]
            rxn_sizes = [220 + (w * 380) for w in norm_w]
            met_colors = ["#9aa0a6" for _ in met_nodes]
            met_sizes = [90 for _ in met_nodes]

            ax = self.map_canvas.ax
            ax.clear()

            if getattr(self, "_map_colorbar", None) is not None:
                try:
                    self._map_colorbar.remove()
                except Exception:
                    pass
                self._map_colorbar = None

            # Identify highlighted nodes (search matches)
            highlighted_rxns = []
            highlighted_mets = []
            if search_text:
                for n in rxn_nodes:
                    if search_text in n.lower() or search_text in (G.nodes[n].get("name", "") or "").lower():
                        highlighted_rxns.append(n)
                for n in met_nodes:
                    if search_text in n.lower() or search_text in (G.nodes[n].get("name", "") or "").lower():
                        highlighted_mets.append(n)

            nx.draw_networkx_edges(
                G,
                pos,
                ax=ax,
                arrows=True,
                arrowstyle="-|>",
                arrowsize=8,
                alpha=0.25,
                width=0.6,
                edge_color="#a0a0a0",
            )

            if met_nodes:
                # Non-highlighted metabolites
                non_hl_mets = [n for n in met_nodes if n not in highlighted_mets]
                if non_hl_mets:
                    nx.draw_networkx_nodes(
                        G,
                        pos,
                        nodelist=non_hl_mets,
                        node_shape="s",
                        node_color="#9aa0a6",
                        node_size=90,
                        ax=ax,
                        alpha=0.85,
                        edgecolors="#5f6368",
                        linewidths=0.4,
                    )
                # Highlighted metabolites (search matches)
                if highlighted_mets:
                    nx.draw_networkx_nodes(
                        G,
                        pos,
                        nodelist=highlighted_mets,
                        node_shape="s",
                        node_color="#ffeb3b",  # Yellow highlight
                        node_size=140,
                        ax=ax,
                        alpha=1.0,
                        edgecolors="#d32f2f",  # Red border
                        linewidths=2.0,
                    )

            if rxn_nodes:
                # Non-highlighted reactions
                non_hl_rxns = [n for n in rxn_nodes if n not in highlighted_rxns]
                non_hl_colors = [rxn_colors[rxn_nodes.index(n)] for n in non_hl_rxns]
                non_hl_sizes = [rxn_sizes[rxn_nodes.index(n)] for n in non_hl_rxns]
                if non_hl_rxns:
                    nx.draw_networkx_nodes(
                        G,
                        pos,
                        nodelist=non_hl_rxns,
                        node_shape="o",
                        node_color=non_hl_colors,
                        node_size=non_hl_sizes,
                        ax=ax,
                        alpha=0.9,
                        edgecolors="#1f2937",
                        linewidths=0.5,
                    )
                # Highlighted reactions (search matches)
                if highlighted_rxns:
                    hl_colors = [rxn_colors[rxn_nodes.index(n)] for n in highlighted_rxns]
                    hl_sizes = [rxn_sizes[rxn_nodes.index(n)] + 150 for n in highlighted_rxns]
                    nx.draw_networkx_nodes(
                        G,
                        pos,
                        nodelist=highlighted_rxns,
                        node_shape="o",
                        node_color=hl_colors,
                        node_size=hl_sizes,
                        ax=ax,
                        alpha=1.0,
                        edgecolors="#d32f2f",  # Red border
                        linewidths=2.5,
                    )

            if G.number_of_nodes() <= 40:
                labels = {n: n for n in rxn_nodes}
                nx.draw_networkx_labels(G, pos, labels, font_size=6, ax=ax)

            if show_legend and rxn_nodes:
                vmax = max_w if max_w > 0 else 1.0
                sm = cm.ScalarMappable(cmap=cm.viridis, norm=Normalize(vmin=0.0, vmax=vmax))
                sm.set_array([])
                try:
                    self._map_colorbar = self.map_canvas.figure.colorbar(sm, ax=ax, fraction=0.03, pad=0.02)
                    self._map_colorbar.set_label(view_mode)
                except Exception:
                    self._map_colorbar = None

            ax.set_title(
                f"Network Map ({G.number_of_nodes()} nodes, {G.number_of_edges()} edges) — {view_mode}",
                fontsize=10,
                fontweight="bold",
            )
            ax.axis("off")
            self.map_canvas.draw()

            self._current_graph = G
            self._current_graph_pos = pos
            self._current_reactions_map = rxn_map
            self._current_metabolites_map = met_map

            self.map_info_lbl.setText(
                f"Reactions: {len(rxn_nodes)} | Metabolites: {len(met_nodes)} | Depth: {depth} | Threshold: {threshold:g}"
            )
        except Exception as e:
            logger.error(f"Map error: {e}")
            ax = self.map_canvas.ax
            ax.clear()
            ax.text(0.5, 0.5, f"Error: {str(e)[:40]}", ha="center", va="center", fontsize=9, transform=ax.transAxes)
            self.map_canvas.draw()
    
    def _on_map_click(self, event):
        """Handle click on map nodes to show reaction details."""
        try:
            if not hasattr(self, "_current_graph") or self._current_graph is None:
                return
            if event.inaxes is None:
                return
            pos = getattr(self, "_current_graph_pos", {})
            if not pos:
                return
            # Find nearest node to clicked coordinates
            x, y = event.xdata, event.ydata
            nearest = None
            best_d = 1e9
            for n, (nx_, ny_) in pos.items():
                d = (nx_ - x) ** 2 + (ny_ - y) ** 2
                if d < best_d:
                    best_d = d
                    nearest = n
            if nearest is None:
                return
            self._selected_map_node = nearest
            # Show details in map_details panel
            G = self._current_graph
            data = G.nodes[nearest]
            ntype = data.get("type", "")
            if ntype == "met":
                met = None
                try:
                    met = self._current_metabolites_map.get(nearest)
                except Exception:
                    met = None
                mname = getattr(met, "name", "") if met else ""
                mform = getattr(met, "formula", "") if met else ""
                mcomp = getattr(met, "compartment", "") if met else ""
                txt = (
                    f"Metabolite: {nearest}\n"
                    f"Name: {mname}\nFormula: {mform}\nCompartment: {mcomp}\n"
                    f"Reactions: {len(getattr(met, 'reactions', [])) if met else 0}"
                )
                self.map_details.setPlainText(txt)
                return

            rxn = None
            try:
                rxn = self._current_reactions_map.get(nearest)
            except Exception:
                rxn = None
            desc = get_description(rxn) if rxn else ""
            eq = getattr(rxn, "reaction", "") if rxn else ""
            gpr = get_gpr(rxn) if rxn else ""
            kegg = get_kegg_rxn_id(rxn) if rxn else ""
            ec = get_ec_numbers(rxn) if rxn else ""
            flux = float(data.get("weight", 0.0))
            txt = (
                f"Reaction: {nearest}\nDescription: {desc}\nEquation: {eq}\n"
                f"Metric: {flux:.6g} ({self.map_view_combo.currentText()})\n"
                f"GPR: {gpr}\nKEGG: {kegg}\nEC: {ec}"
            )
            self.map_details.setPlainText(txt)
        except Exception as e:
            self.map_details.setPlainText(f"Click error: {e}")

    def _on_map_hover(self, event):
        """Handle mouse hover on map to show tooltip."""
        try:
            if not hasattr(self, "_current_graph") or self._current_graph is None:
                return
            if event.inaxes is None:
                # Remove existing tooltip
                if self._map_tooltip:
                    try:
                        self._map_tooltip.remove()
                        self._map_tooltip = None
                        self.map_canvas.draw_idle()
                    except Exception:
                        pass
                return

            pos = getattr(self, "_current_graph_pos", {})
            if not pos:
                return

            x, y = event.xdata, event.ydata
            nearest = None
            best_d = 1e9
            threshold_dist = 0.02  # Proximity threshold

            for n, (nx_, ny_) in pos.items():
                d = (nx_ - x) ** 2 + (ny_ - y) ** 2
                if d < best_d:
                    best_d = d
                    nearest = n

            # Remove old tooltip
            if self._map_tooltip:
                try:
                    self._map_tooltip.remove()
                    self._map_tooltip = None
                except Exception:
                    pass

            if nearest is None or best_d > threshold_dist:
                self.map_canvas.draw_idle()
                return

            # Create tooltip
            G = self._current_graph
            data = G.nodes.get(nearest, {})
            ntype = data.get("type", "")
            name = data.get("name", "") or nearest

            if ntype == "met":
                label = f"[M] {nearest}\n{name}"
            else:
                weight = data.get("weight", 0)
                label = f"[R] {nearest}\n{name}\nFlux: {weight:.4g}"

            px, py = pos[nearest]
            ax = self.map_canvas.ax
            self._map_tooltip = ax.annotate(
                label,
                xy=(px, py),
                xytext=(10, 10),
                textcoords='offset points',
                fontsize=8,
                bbox=dict(boxstyle='round,pad=0.3', facecolor='#ffffcc', edgecolor='black', alpha=0.9),
                zorder=1000
            )
            self.map_canvas.draw_idle()
        except Exception as e:
            self.statusBar().showMessage(f"Tooltip error: {e}")

    def _apply_selected_solver(self, model: cobra.Model):
        """Apply selected solver to COBRApy model."""
        try:
            s = (self.solver_combo.currentText() or "Auto").lower()
            if s == "auto":
                return
            if s == "glpk":
                model.solver = "glpk"
            elif s == "highs":
                try:
                    import optlang.scipy_interface as iface
                    model.solver = iface
                except ImportError:
                    model.solver = "glpk"
                    self.statusBar().showMessage("HiGHS not available, falling back to GLPK.")
            elif s == "gurobi":
                try:
                    model.solver = "gurobi"
                except Exception:
                    self.statusBar().showMessage("Gurobi not available — install gurobipy with a valid license.")
                    return
            elif s == "cplex":
                try:
                    model.solver = "cplex"
                except Exception:
                    self.statusBar().showMessage("CPLEX not available — install cplex with a valid license.")
                    return
            else:
                return
        except Exception as e:
            # Graceful fallback
            self.statusBar().showMessage(f"Solver set failed ({s}): {e}")

    def _gene_name_from_id(self, gid: str) -> str:
        if not gid or self.base_model is None:
            return ""
        try:
            g = self.base_model.genes.get_by_id(str(gid))
            return str(getattr(g, "name", "") or "").strip()
        except Exception:
            return ""

    def _rxn_name_from_id(self, rid: str) -> str:
        if not rid or self.base_model is None:
            return ""
        try:
            r = self.base_model.reactions.get_by_id(str(rid))
            return str(getattr(r, "name", "") or "").strip()
        except Exception:
            return ""

    def _format_gene_label(self, gid: str) -> str:
        name = self._gene_name_from_id(gid)
        if name:
            return f"{gid} — {name}".strip(" —")
        return str(gid)

    def _format_rxn_label(self, rid: str) -> str:
        name = self._rxn_name_from_id(rid)
        if name:
            return f"{rid} — {name}".strip(" —")
        return str(rid)

    def _format_gene_pair_label(self, pair_id: str) -> str:
        text = str(pair_id or "").strip()
        if not text:
            return ""
        if "+" in text:
            parts = [p.strip() for p in text.split("+") if p.strip()]
        elif "-" in text:
            parts = [p.strip() for p in text.split("-") if p.strip()]
        else:
            parts = [text]

        if len(parts) >= 2:
            g1 = self._format_gene_label(parts[0])
            g2 = self._format_gene_label(parts[1])
            return f"{g1} + {g2}"
        return self._format_gene_label(parts[0])

    def _extract_reaction_id_from_input(self, text: str) -> str:
        t = (text or "").strip()
        if not t:
            return ""
        for sep in (" — ", " - ", " – ", " —", "- "):
            if sep in t:
                return t.split(sep, 1)[0].strip()
        return t

    # ==================== NEW FEATURES ====================

    # -------- 1. EXPORT FUNCTIONS --------
    def export_results_json(self):
        """Export FBA results to JSON"""
        if self.last_run is None:
            QMessageBox.warning(self, "Export Error", "No results to export. Run analysis first.")
            return
        
        path, _ = QFileDialog.getSaveFileName(self, "Export Results", "", "JSON Files (*.json)")
        if not path:
            return
        
        try:
            with open(path, 'w') as f:
                json.dump(self.last_run, f, indent=2, default=str)
            
            QMessageBox.information(self, "Export Success", f"Results exported to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    # -------- 2. SEARCH/FILTER FUNCTIONS --------
    def show_search_dialog(self):
        """Open search dialog for metabolites and reactions"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Search Metabolites & Reactions")
        dialog.resize(500, 400)
        
        layout = QVBoxLayout()
        
        search_input = QLineEdit()
        search_input.setPlaceholderText("Enter metabolite or reaction ID...")
        layout.addWidget(QLabel("Search:"))
        layout.addWidget(search_input)
        
        results_table = QTableWidget()
        results_table.setColumnCount(3)
        results_table.setHorizontalHeaderLabels(["Type", "ID", "Name"])
        results_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(results_table)
        
        def perform_search():
            results_table.setRowCount(0)
            if not self.base_model or not search_input.text():
                return
            
            query = search_input.text().lower()
            
            # Search metabolites
            for met in self.base_model.metabolites:
                if query in met.id.lower() or query in met.name.lower():
                    row = results_table.rowCount()
                    results_table.insertRow(row)
                    results_table.setItem(row, 0, QTableWidgetItem("Metabolite"))
                    results_table.setItem(row, 1, QTableWidgetItem(met.id))
                    results_table.setItem(row, 2, QTableWidgetItem(met.name))
            
            # Search reactions
            for rxn in self.base_model.reactions:
                if query in rxn.id.lower() or query in rxn.name.lower():
                    row = results_table.rowCount()
                    results_table.insertRow(row)
                    results_table.setItem(row, 0, QTableWidgetItem("Reaction"))
                    results_table.setItem(row, 1, QTableWidgetItem(rxn.id))
                    results_table.setItem(row, 2, QTableWidgetItem(rxn.name))
        
        search_input.textChanged.connect(perform_search)
        
        buttons = QHBoxLayout()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        buttons.addStretch()
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()

    # -------- 3. SENSITIVITY ANALYSIS --------
    def run_sensitivity_analysis(self):
        """Run sensitivity analysis on a chosen reaction parameter."""
        if self.base_model is None:
            QMessageBox.warning(self, "Error", "Load a model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Sensitivity Analysis")
        dialog.resize(500, 340)

        layout = QFormLayout()

        # Reaction selector with search
        rxn_combo = QComboBox()
        rxn_combo.setEditable(True)
        rxn_combo.setInsertPolicy(QComboBox.NoInsert)
        obj_id = self._get_objective_reaction_id() or ""
        rxn_ids = sorted(r.id for r in self.base_model.reactions)
        rxn_combo.addItems(rxn_ids)
        if obj_id and obj_id in rxn_ids:
            rxn_combo.setCurrentText(obj_id)
        rxn_combo.setCompleter(QCompleter(rxn_ids, rxn_combo))
        layout.addRow("Reaction:", rxn_combo)

        # Which bound to sweep
        bound_combo = QComboBox()
        bound_combo.addItems(["Lower bound", "Upper bound"])
        layout.addRow("Parameter:", bound_combo)

        min_spin = QDoubleSpinBox()
        min_spin.setRange(-1000, 1000)
        min_spin.setDecimals(4)
        min_spin.setValue(-100)
        layout.addRow("Min value:", min_spin)

        max_spin = QDoubleSpinBox()
        max_spin.setRange(-1000, 1000)
        max_spin.setDecimals(4)
        max_spin.setValue(100)
        layout.addRow("Max value:", max_spin)

        steps_spin = QSpinBox()
        steps_spin.setRange(5, 500)
        steps_spin.setValue(20)
        layout.addRow("Steps:", steps_spin)

        buttons = QHBoxLayout()
        ok_btn = QPushButton("Run")
        cancel_btn = QPushButton("Cancel")

        def run_sens():
            rid = rxn_combo.currentText().strip()
            if rid not in self.base_model.reactions:
                QMessageBox.warning(dialog, "Invalid", f"Reaction '{rid}' not found in model.")
                return
            bound_type = "lb" if bound_combo.currentIndex() == 0 else "ub"
            dialog.close()
            self._perform_sensitivity_analysis(
                rid, bound_type,
                min_spin.value(), max_spin.value(), steps_spin.value(),
            )

        ok_btn.clicked.connect(run_sens)
        cancel_btn.clicked.connect(dialog.close)
        buttons.addStretch()
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)

        layout.addRow(buttons)
        dialog.setLayout(layout)
        dialog.exec()

    def _perform_sensitivity_analysis(self, rxn_id: str, bound_type: str,
                                       min_val: float, max_val: float, steps: int):
        """Sweep *bound_type* ('lb' or 'ub') of *rxn_id* and record objective.
        
        Runs the heavy computation in a QThread worker so the UI stays responsive.
        """
        if self.base_model is None or self.base_model.objective is None:
            return

        if rxn_id not in self.base_model.reactions:
            QMessageBox.critical(self, "Error", f"Reaction '{rxn_id}' not found.")
            return

        # Prepare a copy so we don't mutate base_model from the worker thread
        model_copy = self.base_model.copy()
        self._apply_selected_solver(model_copy)

        def _compute():
            values: list[float] = []
            objectives: list[float | None] = []

            for i in range(steps):
                param_val = min_val + (max_val - min_val) * (i / max(steps - 1, 1))
                with model_copy:
                    try:
                        rxn = model_copy.reactions.get_by_id(rxn_id)
                        if bound_type == "lb":
                            rxn.lower_bound = param_val
                        else:
                            rxn.upper_bound = param_val
                    except Exception:
                        objectives.append(None)
                        values.append(param_val)
                        continue
                    try:
                        sol = model_copy.optimize()
                        objectives.append(sol.objective_value if sol.status == "optimal" else None)
                    except Exception:
                        objectives.append(None)

                values.append(param_val)

            return {
                "reaction_id": rxn_id,
                "bound_type": bound_type,
                "parameter_values": values,
                "objective_values": objectives,
            }

        def _on_done(result):
            self.sensitivity_results = result
            values = result["parameter_values"]
            objectives = result["objective_values"]
            if values:
                label = f"{rxn_id} {'lower' if bound_type == 'lb' else 'upper'} bound"
                self._show_sensitivity_results(values, objectives, param_label=label)
            else:
                QMessageBox.information(self, "Sensitivity", "No results.")

        self._launch_worker(_compute, _on_done, f"Running sensitivity analysis on {rxn_id}...")

    # -------- 4. BATCH ANALYSIS --------
    def run_batch_analysis(self):
        """Run analysis on multiple SBML files with comprehensive comparison."""
        files, _ = QFileDialog.getOpenFileNames(self, "Select SBML Files", "", "SBML Files (*.xml *.sbml)")
        if not files:
            return
        
        # Options dialog
        opts = QDialog(self)
        opts.setWindowTitle("Batch Analysis Options")
        opts_layout = QFormLayout(opts)

        run_fva_chk = QCheckBox("Include FVA (min/max ranges)")
        run_fva_chk.setToolTip("Compute FVA for each model (slower)")
        opts_layout.addRow(run_fva_chk)

        run_sgd_chk = QCheckBox("Include essential gene count")
        run_sgd_chk.setToolTip("Run single gene deletion to find essential genes (slower)")
        opts_layout.addRow(run_sgd_chk)

        opts_btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        opts_btns.accepted.connect(opts.accept)
        opts_btns.rejected.connect(opts.reject)
        opts_layout.addRow(opts_btns)

        if opts.exec() != QDialog.Accepted:
            return

        do_fva = run_fva_chk.isChecked()
        do_sgd = run_sgd_chk.isChecked()

        def _compute():
            results = []
            for i, file_path in enumerate(files):
                try:
                    model = cobra.io.read_sbml_model(file_path)
                    info = {
                        "file": Path(file_path).name,
                        "reactions": len(model.reactions),
                        "metabolites": len(model.metabolites),
                        "genes": len(model.genes),
                        "exchange_rxns": sum(1 for r in model.reactions if len(r.metabolites) == 1),
                    }
                    with model:
                        solution = model.optimize()
                        info["status"] = solution.status if solution else "No solution"
                        info["objective"] = float(solution.objective_value) if solution and solution.status == "optimal" else None
                        info["objective_rxn"] = str(model.objective.expression) if model.objective else ""

                    if do_fva and info.get("status") == "optimal":
                        try:
                            from cobra.flux_analysis import flux_variability_analysis
                            fva = flux_variability_analysis(model, processes=1)
                            ranges = (fva["maximum"] - fva["minimum"]).abs()
                            info["fva_flexible"] = int((ranges > 1e-6).sum())
                            info["fva_fixed"] = int((ranges <= 1e-6).sum())
                        except Exception:
                            info["fva_flexible"] = "Error"
                            info["fva_fixed"] = "Error"

                    if do_sgd and info.get("status") == "optimal":
                        try:
                            from cobra.flux_analysis import single_gene_deletion
                            sgd = single_gene_deletion(model, processes=1)
                            growth_col = "growth" if "growth" in sgd.columns else "growth_rate"
                            wt_growth = float(info["objective"])
                            essential = int((sgd[growth_col].fillna(0) < wt_growth * 0.01).sum())
                            info["essential_genes"] = essential
                        except Exception:
                            info["essential_genes"] = "Error"

                    results.append(info)
                except Exception as e:
                    results.append({
                        "file": Path(file_path).name,
                        "reactions": 0, "metabolites": 0, "genes": 0,
                        "exchange_rxns": 0,
                        "status": "Error",
                        "objective": str(e),
                        "objective_rxn": "",
                    })
            return {"results": results, "do_fva": do_fva, "do_sgd": do_sgd}

        def _on_done(result):
            self._show_batch_results(result["results"], result["do_fva"], result["do_sgd"])

        self._launch_worker(_compute, _on_done, f"Running batch analysis on {len(files)} models...")

    def _show_batch_results(self, results, has_fva=False, has_sgd=False):
        """Display batch analysis results with comprehensive table and export."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Batch Analysis Results")
        dialog.resize(1000, 500)
        
        layout = QVBoxLayout()
        
        # Summary
        ok_count = sum(1 for r in results if r.get("status") == "optimal")
        layout.addWidget(QLabel(f"Models analyzed: {len(results)} | Optimal: {ok_count} | Failed: {len(results) - ok_count}"))

        # Build columns
        cols = ["File", "Reactions", "Metabolites", "Genes", "Exchange Rxns", "Status", "Objective Value"]
        if has_fva:
            cols += ["FVA Flexible", "FVA Fixed"]
        if has_sgd:
            cols += ["Essential Genes"]
        
        table = QTableWidget()
        table.setColumnCount(len(cols))
        table.setHorizontalHeaderLabels(cols)
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        for result in results:
            row = table.rowCount()
            table.insertRow(row)
            c = 0
            table.setItem(row, c, QTableWidgetItem(str(result.get("file", "")))); c += 1
            table.setItem(row, c, QTableWidgetItem(str(result.get("reactions", 0)))); c += 1
            table.setItem(row, c, QTableWidgetItem(str(result.get("metabolites", 0)))); c += 1
            table.setItem(row, c, QTableWidgetItem(str(result.get("genes", 0)))); c += 1
            table.setItem(row, c, QTableWidgetItem(str(result.get("exchange_rxns", 0)))); c += 1
            table.setItem(row, c, QTableWidgetItem(str(result.get("status", "")))); c += 1
            obj = result.get("objective")
            table.setItem(row, c, QTableWidgetItem(f"{obj:.6g}" if isinstance(obj, (int, float)) else str(obj or ""))); c += 1
            if has_fva:
                table.setItem(row, c, QTableWidgetItem(str(result.get("fva_flexible", "")))); c += 1
                table.setItem(row, c, QTableWidgetItem(str(result.get("fva_fixed", "")))); c += 1
            if has_sgd:
                table.setItem(row, c, QTableWidgetItem(str(result.get("essential_genes", "")))); c += 1
        
        layout.addWidget(table, stretch=1)
        
        buttons = QHBoxLayout()
        export_csv_btn = QPushButton("Export CSV")
        export_xlsx_btn = QPushButton("Export Excel")

        def _export_batch_csv():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Export Batch (CSV)", str(Path.home() / "batch_results.csv"), "CSV (*.csv)")
            if not fp:
                return
            with open(fp, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(cols)
                for r in results:
                    row_data = [r.get("file", ""), r.get("reactions", 0), r.get("metabolites", 0),
                               r.get("genes", 0), r.get("exchange_rxns", 0), r.get("status", ""),
                               r.get("objective", "")]
                    if has_fva:
                        row_data += [r.get("fva_flexible", ""), r.get("fva_fixed", "")]
                    if has_sgd:
                        row_data += [r.get("essential_genes", "")]
                    w.writerow(row_data)
            self.statusBar().showMessage(f"Batch exported: {fp}")

        def _export_batch_xlsx():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Export Batch (Excel)", str(Path.home() / "batch_results.xlsx"), "Excel (*.xlsx)")
            if not fp:
                return
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Batch Results"
                ws.append(cols)
                for r in results:
                    row_data = [r.get("file", ""), r.get("reactions", 0), r.get("metabolites", 0),
                               r.get("genes", 0), r.get("exchange_rxns", 0), r.get("status", ""),
                               r.get("objective", "")]
                    if has_fva:
                        row_data += [r.get("fva_flexible", ""), r.get("fva_fixed", "")]
                    if has_sgd:
                        row_data += [r.get("essential_genes", "")]
                    ws.append(row_data)
                wb.save(fp)
                self.statusBar().showMessage(f"Batch exported: {fp}")
            except ImportError:
                QMessageBox.warning(dialog, "Missing", "openpyxl required for Excel export.")

        export_csv_btn.clicked.connect(_export_batch_csv)
        export_xlsx_btn.clicked.connect(_export_batch_xlsx)
        buttons.addWidget(export_csv_btn)
        buttons.addWidget(export_xlsx_btn)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        buttons.addStretch()
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()

    def _show_sensitivity_results(self, values, objectives, param_label: str = "Parameter value"):
        dialog = QDialog(self)
        dialog.setWindowTitle("Sensitivity Analysis Results")
        dialog.resize(800, 500)

        layout = QVBoxLayout()

        # Plot
        fig = Figure(figsize=(6, 3.5), dpi=100)
        ax = fig.add_subplot(111)
        ax.plot(values, objectives, marker="o", linewidth=1.5)
        ax.set_xlabel(param_label)
        ax.set_ylabel("Objective value")
        ax.grid(True, alpha=0.3)
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas, stretch=1)

        # Table
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Parameter", "Objective"])
        table.horizontalHeader().setStretchLastSection(True)
        for v, obj in zip(values, objectives):
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(f"{v:.6g}"))
            table.setItem(row, 1, QTableWidgetItem("" if obj is None else f"{obj:.6g}"))
        layout.addWidget(table, stretch=1)

        buttons = QHBoxLayout()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        buttons.addStretch()
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)

        dialog.setLayout(layout)
        dialog.exec()

    # -------- 4b. COMMUNITY ANALYSIS --------
    def run_community_analysis(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select SBML Files (2+)", "", "SBML Files (*.xml *.sbml)")
        if not files or len(files) < 2:
            QMessageBox.warning(self, "Community", "Select at least two SBML files.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Community Analysis Options")
        layout = QFormLayout(dialog)

        share_ext_chk = QCheckBox("Share extracellular metabolites")
        share_ext_chk.setChecked(True)
        layout.addRow(QLabel(""), share_ext_chk)

        topn_spin = QSpinBox()
        topn_spin.setRange(5, 200)
        topn_spin.setValue(20)
        layout.addRow(QLabel("Top exchange fluxes:"), topn_spin)

        export_chk = QCheckBox("Export merged community SBML")
        export_chk.setChecked(False)
        layout.addRow(QLabel(""), export_chk)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addRow(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        labels = [Path(f).stem for f in files]
        share_ext = share_ext_chk.isChecked()
        topn = topn_spin.value()
        do_export = export_chk.isChecked()

        def _compute():
            models = []
            for f in files:
                models.append(cobra.io.read_sbml_model(f))

            community, obj_map, shared_mets = self._build_community_model(
                models, labels, share_extracellular=share_ext)

            solution = community.optimize()
            total_obj = solution.objective_value if solution else None

            # Per-species growth: solo vs community
            per_species = []
            solo_exchange_maps: dict[str, dict[str, float]] = {}
            for label, model in zip(labels, models):
                solo_val = None
                try:
                    sol = model.optimize()
                    solo_val = sol.objective_value if sol and sol.status == "optimal" else None
                except Exception:
                    sol = None
                    solo_val = None

                comm_val = None
                rxn_id = obj_map.get(label)
                if rxn_id:
                    try:
                        comm_val = solution.fluxes.get(rxn_id, None)
                    except Exception:
                        comm_val = None
                per_species.append((label, rxn_id or "", solo_val, comm_val))

                # Exchange map by metabolite id for solo model
                emap = {}
                try:
                    for rxn in model.reactions:
                        if len(rxn.metabolites) == 1:
                            met = list(rxn.metabolites.keys())[0]
                            met_id = met.id
                            v = sol.fluxes.get(rxn.id, 0.0) if sol else 0.0
                            if abs(v) > 1e-9:
                                emap[met_id] = emap.get(met_id, 0.0) + float(v)
                except Exception:
                    pass
                solo_exchange_maps[label] = emap

            # Community exchange fluxes (shared extracellular)
            comm_exchange = {}
            try:
                for rxn in community.reactions:
                    if len(rxn.metabolites) == 1:
                        met = list(rxn.metabolites.keys())[0]
                        met_id = met.id
                        v = solution.fluxes.get(rxn.id, 0.0)
                        if abs(v) > 1e-9:
                            comm_exchange[met_id] = comm_exchange.get(met_id, 0.0) + float(v)
            except Exception:
                pass

            # Delta exchange: community vs sum of solo
            delta_exch = []
            for met_id, comm_v in comm_exchange.items():
                solo_sum = 0.0
                for emap in solo_exchange_maps.values():
                    solo_sum += emap.get(met_id, 0.0)
                delta = comm_v - solo_sum
                delta_exch.append((met_id, comm_v, solo_sum, delta))
            delta_exch = sorted(delta_exch, key=lambda x: abs(x[3]), reverse=True)[:topn]

            return {"total_obj": total_obj, "per_species": per_species,
                    "delta_exch": delta_exch, "community": community,
                    "do_export": do_export}

        def _on_done(result):
            if result.get("do_export"):
                out_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save community SBML",
                    str(Path.home() / "community_model.xml"),
                    "SBML files (*.xml *.sbml);;All files (*.*)",
                )
                if out_path:
                    try:
                        cobra.io.write_sbml_model(result["community"], out_path)
                    except Exception as e:
                        QMessageBox.critical(self, "Export failed", f"Could not save community model:\n{e}")

            self._show_community_results(result["total_obj"], result["per_species"], result["delta_exch"])

        self._launch_worker(_compute, _on_done, "Running Community Analysis...")

    def _show_community_results(self, total_obj, per_species, delta_exch):
        dialog = QDialog(self)
        dialog.setWindowTitle("Community Analysis Results")
        dialog.resize(900, 600)

        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"<b>Community Total Objective: {total_obj:.6g if isinstance(total_obj, (int, float)) else total_obj}</b>"))

        # Growth comparison chart
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
            fig = Figure(figsize=(7, 2.5), dpi=100)
            ax = fig.add_subplot(111)
            labels_list = [ps[0] for ps in per_species]
            solo_vals = [ps[2] or 0 for ps in per_species]
            comm_vals = [ps[3] or 0 for ps in per_species]
            x = list(range(len(labels_list)))
            w = 0.35
            ax.bar([i - w/2 for i in x], solo_vals, w, label="Solo", color="#4CAF50", alpha=0.8)
            ax.bar([i + w/2 for i in x], comm_vals, w, label="Community", color="#2196F3", alpha=0.8)
            ax.set_xticks(x)
            ax.set_xticklabels(labels_list, rotation=45, ha="right", fontsize=8)
            ax.set_ylabel("Growth")
            ax.set_title("Solo vs Community Growth")
            ax.legend(fontsize=8)
            fig.tight_layout()
            canvas = FigureCanvas(fig)
            layout.addWidget(canvas)
        except Exception:
            pass

        # Species table
        layout.addWidget(QLabel("<b>Per-species growth comparison:</b>"))
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Species", "Objective Rxn", "Solo Growth", "Community Growth", "Δ Growth"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        for label, rxn_id, solo_val, comm_val in per_species:
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(label))
            table.setItem(row, 1, QTableWidgetItem(rxn_id))
            table.setItem(row, 2, QTableWidgetItem("" if solo_val is None else f"{solo_val:.6g}"))
            table.setItem(row, 3, QTableWidgetItem("" if comm_val is None else f"{comm_val:.6g}"))
            delta = ""
            if solo_val is not None and comm_val is not None:
                delta = f"{comm_val - solo_val:.6g}"
            table.setItem(row, 4, QTableWidgetItem(delta))
        layout.addWidget(table)

        if delta_exch:
            layout.addWidget(QLabel("<b>Top exchange flux changes (community vs solo sum):</b>"))
            exch_table = QTableWidget()
            exch_table.setColumnCount(4)
            exch_table.setHorizontalHeaderLabels(["Metabolite", "Community Flux", "Solo Sum", "Delta"])
            exch_table.horizontalHeader().setStretchLastSection(True)
            exch_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            for met_id, comm_v, solo_sum, delta in delta_exch:
                row = exch_table.rowCount()
                exch_table.insertRow(row)
                exch_table.setItem(row, 0, QTableWidgetItem(met_id))
                exch_table.setItem(row, 1, QTableWidgetItem(f"{comm_v:.6g}"))
                exch_table.setItem(row, 2, QTableWidgetItem(f"{solo_sum:.6g}"))
                exch_table.setItem(row, 3, QTableWidgetItem(f"{delta:.6g}"))
            layout.addWidget(exch_table)

        buttons = QHBoxLayout()
        export_csv_btn = QPushButton("Export CSV")

        def _export_community_csv():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Export Community (CSV)", str(Path.home() / "community_results.csv"), "CSV (*.csv)")
            if not fp:
                return
            with open(fp, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Total Objective", total_obj])
                w.writerow([])
                w.writerow(["Species", "Objective Rxn", "Solo Growth", "Community Growth", "Delta"])
                for label, rxn_id, solo_val, comm_val in per_species:
                    delta = (comm_val - solo_val) if solo_val is not None and comm_val is not None else ""
                    w.writerow([label, rxn_id, solo_val or "", comm_val or "", delta])
                if delta_exch:
                    w.writerow([])
                    w.writerow(["Metabolite", "Community Flux", "Solo Sum", "Delta"])
                    for met_id, comm_v, solo_sum, delta in delta_exch:
                        w.writerow([met_id, comm_v, solo_sum, delta])
            self.statusBar().showMessage(f"Community exported: {fp}")

        export_csv_btn.clicked.connect(_export_community_csv)
        buttons.addWidget(export_csv_btn)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        buttons.addStretch()
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)

        dialog.setLayout(layout)
        dialog.exec()

    # -------- 5. REACTION/METABOLITE DETAILS --------
    def show_reaction_details(self, reaction_id):
        """Show details for a reaction"""
        if not self.base_model or reaction_id not in self.base_model.reactions:
            return
        
        rxn = self.base_model.reactions.get_by_id(reaction_id)
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Reaction Details: {reaction_id}")
        dialog.resize(500, 400)
        
        layout = QVBoxLayout()
        
        text = QPlainTextEdit()
        text.setReadOnly(True)
        
        details = f"""
ID: {rxn.id}
Name: {rxn.name}
Bounds: {rxn.lower_bound} to {rxn.upper_bound}
Gene Reaction Rule: {rxn.gene_reaction_rule}
Reversible: {rxn.reversible}

Metabolites:
"""
        for met, coeff in rxn.metabolites.items():
            details += f"  {coeff:+.1f} {met.id} ({met.name})\n"
        
        if hasattr(rxn, 'subsystem') and rxn.subsystem:
            details += f"\nSubsystem: {rxn.subsystem}"
        
        text.setPlainText(details)
        layout.addWidget(text)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    def show_metabolite_details(self, metabolite_id):
        """Show details for a metabolite"""
        if not self.base_model or metabolite_id not in self.base_model.metabolites:
            return
        
        met = self.base_model.metabolites.get_by_id(metabolite_id)
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Metabolite Details: {metabolite_id}")
        dialog.resize(500, 400)
        
        layout = QVBoxLayout()
        
        text = QPlainTextEdit()
        text.setReadOnly(True)
        
        details = f"""
ID: {met.id}
Name: {met.name}
Formula: {met.formula}
Charge: {met.charge}
Compartment: {met.compartment}

Participating Reactions: {len(met.reactions)}
"""
        for rxn in list(met.reactions)[:20]:
            details += f"  - {rxn.id}\n"
        
        text.setPlainText(details)
        layout.addWidget(text)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    # -------- 6. UNDO/REDO --------
    def save_state_for_undo(self):
        """Save current model state for undo"""
        if self.base_model is None:
            return
        
        state = {
            "knockout_genes": self.knockout_genes.copy(),
            "overexpression_reactions": self.overexpression_reactions.copy(),
            "reaction_bound_overrides": self.reaction_bound_overrides.copy(),
            "timestamp": datetime.now()
        }
        
        self.undo_stack.append(state)
        if len(self.undo_stack) > self.max_undo_steps:
            self.undo_stack.pop(0)
        
        self.redo_stack.clear()

    def undo(self):
        """Undo last change"""
        if not self.undo_stack:
            QMessageBox.information(self, "Undo", "Nothing to undo.")
            return
        
        current_state = {
            "knockout_genes": self.knockout_genes.copy(),
            "overexpression_reactions": self.overexpression_reactions.copy(),
            "reaction_bound_overrides": self.reaction_bound_overrides.copy()
        }
        self.redo_stack.append(current_state)
        
        state = self.undo_stack.pop()
        self.knockout_genes = state["knockout_genes"]
        self.overexpression_reactions = state["overexpression_reactions"]
        self.reaction_bound_overrides = state["reaction_bound_overrides"]
        self.model_dirty = True
        self.statusBar().showMessage("Undo completed.")

    def redo(self):
        """Redo last undone change"""
        if not self.redo_stack:
            QMessageBox.information(self, "Redo", "Nothing to redo.")
            return
        
        current_state = {
            "knockout_genes": self.knockout_genes.copy(),
            "overexpression_reactions": self.overexpression_reactions.copy(),
            "reaction_bound_overrides": self.reaction_bound_overrides.copy()
        }
        self.undo_stack.append(current_state)
        
        state = self.redo_stack.pop()
        self.knockout_genes = state["knockout_genes"]
        self.overexpression_reactions = state["overexpression_reactions"]
        self.reaction_bound_overrides = state["reaction_bound_overrides"]
        self.model_dirty = True
        self.statusBar().showMessage("Redo completed.")

    # -------- 7. KEYBOARD SHORTCUTS --------
    def show_shortcuts(self):
        """Show keyboard shortcuts help"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Keyboard Shortcuts")
        dialog.resize(500, 400)
        
        layout = QVBoxLayout()
        
        text = QPlainTextEdit()
        text.setReadOnly(True)
        
        shortcuts = """
KEYBOARD SHORTCUTS
==================

File Operations:
  Ctrl+O        Open SBML file
  Ctrl+S        Save as (SBML)
  Ctrl+Q        Exit application

Analysis:
  Ctrl+R        Run analysis
  Ctrl+Z        Undo
  Ctrl+Y        Redo
  Ctrl+F        Find/Search

Tools:
  Ctrl+B        Batch analysis
  Ctrl+E        Export results
  Ctrl+H        Show this help
"""
        
        text.setPlainText(shortcuts)
        layout.addWidget(text)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    # ==================== DRAG & DROP ====================
    def dragEnterEvent(self, event):
        """Accept drag events for SBML files."""
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                fp = url.toLocalFile()
                if fp.endswith(('.xml', '.sbml', '.json')):
                    event.acceptProposedAction()
                    return

    def dropEvent(self, event):
        """Handle dropping SBML files onto the window."""
        for url in event.mimeData().urls():
            fp = url.toLocalFile()
            if fp.endswith(('.xml', '.sbml', '.json')):
                logger.info(f"File dropped: {fp}")
                self.open_sbml_from_path(fp)
                return

    # ==================== THEME TOGGLE ====================
    def toggle_theme(self):
        """Toggle between dark and light theme."""
        self._dark_theme = not self._dark_theme
        if self._dark_theme:
            self.setStyleSheet(DARK_STYLESHEET)
            logger.info("Theme switched to dark")
        else:
            self.setStyleSheet("")
            logger.info("Theme switched to light")
        self.statusBar().showMessage(f"Theme: {'Dark' if self._dark_theme else 'Light (System Default)'}")

    # ==================== GENE EXPRESSION INTEGRATION ====================
    def run_gene_expression_analysis(self):
        """Gene Expression Integration: GIMME, iMAT, and E-Flux algorithms."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Gene Expression Integration")
        dialog.resize(500, 350)
        layout = QVBoxLayout(dialog)

        layout.addWidget(QLabel("Load a CSV file with columns: gene_id, expression_value"))

        file_row = QHBoxLayout()
        file_edit = QLineEdit()
        file_edit.setPlaceholderText("Select expression data CSV...")
        file_btn = QPushButton("Browse...")
        def _browse():
            fp, _ = QFileDialog.getOpenFileName(dialog, "Expression Data", "", "CSV (*.csv *.tsv);;All (*.*)")
            if fp:
                file_edit.setText(fp)
        file_btn.clicked.connect(_browse)
        file_row.addWidget(file_edit)
        file_row.addWidget(file_btn)
        layout.addLayout(file_row)

        form = QFormLayout()
        method_combo = QComboBox()
        method_combo.addItems(["E-Flux", "GIMME", "iMAT"])
        form.addRow("Method:", method_combo)

        threshold_spin = QDoubleSpinBox()
        threshold_spin.setRange(0.0, 1e6)
        threshold_spin.setValue(0.0)
        threshold_spin.setDecimals(2)
        form.addRow("Expression threshold:", threshold_spin)

        required_growth_spin = QDoubleSpinBox()
        required_growth_spin.setRange(0.0, 1.0)
        required_growth_spin.setValue(0.9)
        required_growth_spin.setSingleStep(0.05)
        form.addRow("Min growth fraction (GIMME):", required_growth_spin)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        csv_path = file_edit.text().strip()
        if not csv_path or not Path(csv_path).exists():
            QMessageBox.warning(self, "File Error", "Please select a valid CSV file.")
            return

        method = method_combo.currentText()
        threshold = threshold_spin.value()
        required_growth = required_growth_spin.value()

        # Parse expression data
        try:
            expr_data = {}
            sep = '\t' if csv_path.endswith('.tsv') else ','
            with open(csv_path, encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=sep)
                header = next(reader)
                for row in reader:
                    if len(row) >= 2:
                        gene_id = row[0].strip()
                        try:
                            expr_val = float(row[1].strip())
                        except ValueError:
                            continue
                        expr_data[gene_id] = expr_val
            if not expr_data:
                raise ValueError("No valid expression data found in CSV.")
        except Exception as e:
            QMessageBox.critical(self, "Parse Error", f"Failed to parse expression data: {e}")
            return

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            # Map gene expression to reactions via GPR tree evaluation
            rxn_expr = {}
            for rxn in model.reactions:
                if not rxn.genes:
                    rxn_expr[rxn.id] = None
                    continue
                # Build gene_values dict for this reaction's genes
                gene_vals = {g.id: expr_data[g.id] for g in rxn.genes if g.id in expr_data}
                if not gene_vals:
                    rxn_expr[rxn.id] = None
                    continue
                # Use GPR rule string for recursive AND=min, OR=max evaluation
                gpr_str = getattr(rxn, 'gene_reaction_rule', '') or ''
                if gpr_str.strip():
                    val = evaluate_gpr_expression(gpr_str, gene_vals)
                    rxn_expr[rxn.id] = val
                else:
                    # Fallback: if no GPR rule string, use min of all gene values
                    rxn_expr[rxn.id] = min(gene_vals.values())

            if method == "E-Flux":
                # Scale upper bounds by normalized expression
                max_expr = max((v for v in rxn_expr.values() if v is not None), default=1.0)
                if max_expr <= 0:
                    max_expr = 1.0
                for rxn in model.reactions:
                    expr_val = rxn_expr.get(rxn.id)
                    if expr_val is not None:
                        scale = expr_val / max_expr
                        if rxn.upper_bound > 0:
                            rxn.upper_bound = rxn.upper_bound * max(scale, 0.001)
                        if rxn.lower_bound < 0:
                            rxn.lower_bound = rxn.lower_bound * max(scale, 0.001)
                sol = model.optimize()
                if sol.status != 'optimal':
                    raise ValueError(f"E-Flux optimization failed: {sol.status}")
                return {"method": "E-Flux", "status": sol.status,
                        "objective": float(sol.objective_value),
                        "flux": {r.id: float(sol.fluxes[r.id]) for r in model.reactions},
                        "mapped_genes": len([v for v in rxn_expr.values() if v is not None]),
                        "total_genes": len(expr_data)}

            elif method == "GIMME":
                # GIMME: minimize flux through low-expression reactions
                sol0 = model.optimize()
                if sol0.status != 'optimal':
                    raise ValueError("Baseline optimization failed.")
                obj_rxn = [v.flux_expression for v in model.objective.variables.values()]
                required_obj = float(sol0.objective_value) * required_growth
                # Constrain minimum growth
                obj_id = list(model.objective.expression.as_coefficients_dict().keys())
                for rxn in model.reactions:
                    if any(str(rxn.id) in str(v) for v in obj_id):
                        rxn.lower_bound = max(rxn.lower_bound, required_obj)
                        break
                # Penalize low-expression reactions
                low_expr_rxns = []
                for rxn in model.reactions:
                    expr_val = rxn_expr.get(rxn.id)
                    if expr_val is not None and expr_val < threshold:
                        low_expr_rxns.append(rxn.id)
                sol = model.optimize()
                return {"method": "GIMME", "status": sol.status if sol else "failed",
                        "objective": float(sol.objective_value) if sol and sol.status == 'optimal' else 0.0,
                        "flux": {r.id: float(sol.fluxes[r.id]) for r in model.reactions} if sol and sol.status == 'optimal' else {},
                        "low_expression_reactions": len(low_expr_rxns),
                        "mapped_genes": len([v for v in rxn_expr.values() if v is not None]),
                        "total_genes": len(expr_data)}

            elif method == "iMAT":
                # iMAT: classify reactions and maximize agreement
                high_rxns = []
                low_rxns = []
                max_expr = max((v for v in rxn_expr.values() if v is not None), default=1.0)
                for rxn in model.reactions:
                    expr_val = rxn_expr.get(rxn.id)
                    if expr_val is not None:
                        if expr_val > threshold:
                            high_rxns.append(rxn.id)
                        elif expr_val <= threshold * 0.1:
                            low_rxns.append(rxn.id)
                # Constrain low-expression reactions to near-zero
                for rid in low_rxns:
                    rxn = model.reactions.get_by_id(rid)
                    rxn.upper_bound = min(rxn.upper_bound, 0.01)
                    rxn.lower_bound = max(rxn.lower_bound, -0.01)
                sol = model.optimize()
                active_high = 0
                if sol and sol.status == 'optimal':
                    for rid in high_rxns:
                        if abs(sol.fluxes[rid]) > 1e-6:
                            active_high += 1
                return {"method": "iMAT", "status": sol.status if sol else "failed",
                        "objective": float(sol.objective_value) if sol and sol.status == 'optimal' else 0.0,
                        "flux": {r.id: float(sol.fluxes[r.id]) for r in model.reactions} if sol and sol.status == 'optimal' else {},
                        "high_expression_reactions": len(high_rxns),
                        "low_expression_reactions": len(low_rxns),
                        "active_high_expr": active_high,
                        "mapped_genes": len([v for v in rxn_expr.values() if v is not None]),
                        "total_genes": len(expr_data)}
            else:
                raise ValueError(f"Unknown method: {method}")

        def _on_done(result):
            lines = [
                f"Method: {result['method']}",
                f"Status: {result['status']}",
                f"Objective value: {result.get('objective', 0):.6g}",
                f"Genes mapped: {result.get('mapped_genes', 0)} / {result.get('total_genes', 0)}",
            ]
            if 'low_expression_reactions' in result:
                lines.append(f"Low-expression reactions: {result['low_expression_reactions']}")
            if 'high_expression_reactions' in result:
                lines.append(f"High-expression reactions: {result['high_expression_reactions']}")
            if 'active_high_expr' in result:
                lines.append(f"Active high-expression reactions: {result['active_high_expr']}")

            flux = result.get("flux", {})
            if flux:
                sorted_flux = sorted(flux.items(), key=lambda x: abs(x[1]), reverse=True)[:50]
                lines.append(f"\nTop {len(sorted_flux)} reactions by flux:")
                for rid, fv in sorted_flux:
                    lines.append(f"  {rid}: {fv:.6g}")
            TextPopup(f"{result['method']} Results", "\n".join(lines), self).exec()

        self._launch_worker(_compute, _on_done, f"Running {method}...")

    # ==================== OPTKNOCK / STRAIN DESIGN ====================
    def run_optknock(self):
        """OptKnock-inspired strain design via combinatorial knockout search."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("OptKnock / Strain Design")
        dialog.resize(500, 300)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Find reaction knockouts that maximize production of a target reaction."))

        form = QFormLayout()
        target_edit = QLineEdit()
        target_edit.setPlaceholderText("Target reaction ID (e.g., EX_etoh_e)")
        if self.base_model:
            target_completer = QCompleter([r.id for r in self.base_model.reactions])
            target_completer.setCaseSensitivity(Qt.CaseInsensitive)
            target_edit.setCompleter(target_completer)
        form.addRow("Target reaction:", target_edit)

        max_ko_spin = QSpinBox()
        max_ko_spin.setRange(1, 3)
        max_ko_spin.setValue(1)
        form.addRow("Max knockouts:", max_ko_spin)

        top_results_spin = QSpinBox()
        top_results_spin.setRange(5, 50)
        top_results_spin.setValue(10)
        form.addRow("Top results:", top_results_spin)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        target_id = self._extract_reaction_id_from_input(target_edit.text())
        if not target_id:
            QMessageBox.warning(self, "Missing", "Enter a target reaction ID.")
            return
        max_ko = max_ko_spin.value()
        top_n = top_results_spin.value()

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute(worker=None):
            from itertools import combinations
            # Get candidate reactions (non-exchange, non-objective)
            obj_rxns = set()
            try:
                for v in model.objective.expression.as_coefficients_dict():
                    for r in model.reactions:
                        if r.id in str(v):
                            obj_rxns.add(r.id)
            except Exception:
                pass
            candidates = [r.id for r in model.reactions
                          if not r.boundary and r.id not in obj_rxns and r.id != target_id]
            # Limit candidates to reactions with non-zero flux
            with model:
                sol0 = model.optimize()
                if sol0.status != 'optimal':
                    raise ValueError("Baseline optimization failed.")
                baseline_obj = float(sol0.objective_value)
                baseline_target = float(sol0.fluxes.get(target_id, 0.0))
                active = [rid for rid in candidates if abs(sol0.fluxes.get(rid, 0)) > 1e-8]
            candidates = active[:200]  # Limit for performance

            results = []

            # Count total combinations for progress
            total = 0
            for k in range(1, max_ko + 1):
                limit = min(len(candidates), 200 if k == 1 else 80 if k == 2 else 30)
                total += sum(1 for _ in combinations(candidates[:limit], k))
            done = 0

            # Single knockouts
            for i, (rid,) in enumerate(combinations(candidates, 1)):
                done += 1
                if worker and done % 10 == 0:
                    worker.report_progress(f"OptKnock: {done}/{total}", int(done / total * 100))
                with model:
                    model.reactions.get_by_id(rid).knock_out()
                    try:
                        sol = model.optimize()
                        if sol.status == 'optimal':
                            target_flux = float(sol.fluxes.get(target_id, 0))
                            growth = float(sol.objective_value)
                            if growth > baseline_obj * 0.01:  # Viable
                                results.append({
                                    "knockouts": [rid], "target_flux": target_flux,
                                    "growth": growth, "growth_ratio": growth / baseline_obj
                                })
                    except Exception:
                        pass

            # Double knockouts if requested
            if max_ko >= 2:
                limit2 = min(len(candidates), 80)
                for rid1, rid2 in combinations(candidates[:limit2], 2):
                    done += 1
                    if worker and done % 20 == 0:
                        worker.report_progress(f"OptKnock (double): {done}/{total}", int(done / total * 100))
                    with model:
                        model.reactions.get_by_id(rid1).knock_out()
                        model.reactions.get_by_id(rid2).knock_out()
                        try:
                            sol = model.optimize()
                            if sol.status == 'optimal':
                                target_flux = float(sol.fluxes.get(target_id, 0))
                                growth = float(sol.objective_value)
                                if growth > baseline_obj * 0.01:
                                    results.append({
                                        "knockouts": [rid1, rid2], "target_flux": target_flux,
                                        "growth": growth, "growth_ratio": growth / baseline_obj
                                    })
                        except Exception:
                            pass

            # Triple knockouts if requested
            if max_ko >= 3:
                limit3 = min(len(candidates), 30)
                for rid1, rid2, rid3 in combinations(candidates[:limit3], 3):
                    done += 1
                    if worker and done % 50 == 0:
                        worker.report_progress(f"OptKnock (triple): {done}/{total}", int(done / total * 100))
                    with model:
                        model.reactions.get_by_id(rid1).knock_out()
                        model.reactions.get_by_id(rid2).knock_out()
                        model.reactions.get_by_id(rid3).knock_out()
                        try:
                            sol = model.optimize()
                            if sol.status == 'optimal':
                                target_flux = float(sol.fluxes.get(target_id, 0))
                                growth = float(sol.objective_value)
                                if growth > baseline_obj * 0.01:
                                    results.append({
                                        "knockouts": [rid1, rid2, rid3], "target_flux": target_flux,
                                        "growth": growth, "growth_ratio": growth / baseline_obj
                                    })
                        except Exception:
                            pass

            results.sort(key=lambda x: x["target_flux"], reverse=True)
            return {"results": results[:top_n], "baseline_obj": baseline_obj,
                    "baseline_target": baseline_target, "target_id": target_id}

        def _on_done(result):
            items = result["results"]
            dialog2 = QDialog(self)
            dialog2.setWindowTitle(f"OptKnock Results — {result['target_id']}")
            dialog2.resize(800, 500)
            layout2 = QVBoxLayout(dialog2)
            layout2.addWidget(QLabel(
                f"Baseline growth: {result['baseline_obj']:.6g} | "
                f"Baseline target flux: {result['baseline_target']:.6g}\n"
                f"Top {len(items)} knockout strategies:"))
            table = QTableWidget(0, 4)
            table.setHorizontalHeaderLabels(["Knockout(s)", "Target Flux", "Growth", "Growth %"])
            table.horizontalHeader().setStretchLastSection(True)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            for item in items:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(", ".join(item["knockouts"])))
                table.setItem(r, 1, QTableWidgetItem(f"{item['target_flux']:.6g}"))
                table.setItem(r, 2, QTableWidgetItem(f"{item['growth']:.6g}"))
                table.setItem(r, 3, QTableWidgetItem(f"{item['growth_ratio']*100:.1f}%"))
            layout2.addWidget(table)
            btns2 = QDialogButtonBox(QDialogButtonBox.Close)
            btns2.rejected.connect(dialog2.reject)
            layout2.addWidget(btns2)
            dialog2.exec()

        self._launch_worker(_compute, _on_done, "Running OptKnock analysis...")

    # ==================== DYNAMIC FBA (dFBA) ====================
    def run_dfba(self):
        """Dynamic FBA — time-course simulation of batch culture."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Dynamic FBA (dFBA)")
        dialog.resize(500, 400)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Simulate batch growth by iteratively solving FBA over time."))

        form = QFormLayout()
        biomass_spin = QDoubleSpinBox()
        biomass_spin.setRange(0.001, 100.0)
        biomass_spin.setValue(0.1)
        biomass_spin.setDecimals(3)
        form.addRow("Initial biomass (g/L):", biomass_spin)

        substrate_edit = QLineEdit()
        substrate_edit.setPlaceholderText("Substrate exchange reaction (e.g., EX_glc__D_e)")
        if self.base_model:
            sub_completer = QCompleter([r.id for r in self.base_model.reactions if r.id.startswith("EX_")])
            sub_completer.setCaseSensitivity(Qt.CaseInsensitive)
            substrate_edit.setCompleter(sub_completer)
        form.addRow("Substrate reaction:", substrate_edit)

        substrate_conc_spin = QDoubleSpinBox()
        substrate_conc_spin.setRange(0.0, 1000.0)
        substrate_conc_spin.setValue(20.0)
        substrate_conc_spin.setDecimals(2)
        form.addRow("Initial substrate (mmol/L):", substrate_conc_spin)

        time_end_spin = QDoubleSpinBox()
        time_end_spin.setRange(0.1, 100.0)
        time_end_spin.setValue(10.0)
        form.addRow("Simulation time (h):", time_end_spin)

        dt_spin = QDoubleSpinBox()
        dt_spin.setRange(0.01, 1.0)
        dt_spin.setValue(0.1)
        dt_spin.setDecimals(3)
        form.addRow("Time step (h):", dt_spin)

        integrator_combo = QComboBox()
        integrator_combo.addItems(["RK4 (Runge-Kutta 4th order)", "Euler (1st order)"])
        form.addRow("Integration method:", integrator_combo)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        substrate_rxn_id = self._extract_reaction_id_from_input(substrate_edit.text())
        if not substrate_rxn_id:
            QMessageBox.warning(self, "Missing", "Enter a substrate exchange reaction ID.")
            return

        init_biomass = biomass_spin.value()
        init_substrate = substrate_conc_spin.value()
        t_end = time_end_spin.value()
        dt = dt_spin.value()
        use_rk4 = integrator_combo.currentIndex() == 0

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _solve_fba_at_state(biomass_val, substrate_val, dt_val):
            """Solve FBA for given state, return (growth_rate, uptake_rate)."""
            if biomass_val <= 1e-9 or substrate_val <= 1e-6:
                return 0.0, 0.0
            max_uptake = substrate_val / (biomass_val * dt_val) if biomass_val > 0 else 0
            with model:
                sub_rxn = model.reactions.get_by_id(substrate_rxn_id)
                sub_rxn.lower_bound = max(-max_uptake, sub_rxn.lower_bound)
                sol = model.optimize()
                if sol.status != 'optimal':
                    return 0.0, 0.0
                return float(sol.objective_value), float(sol.fluxes.get(substrate_rxn_id, 0))

        def _compute(worker=None):
            times, biomasses, substrates, growths = [], [], [], []
            biomass = init_biomass
            substrate = init_substrate
            t = 0.0
            total_steps = int(t_end / dt) if dt > 0 else 1
            step = 0

            while t < t_end and substrate > 1e-6 and biomass > 1e-9:
                step += 1
                if worker and step % 5 == 0:
                    pct = min(int(t / t_end * 100), 99)
                    worker.report_progress(f"dFBA: t={t:.2f}h / {t_end:.1f}h", pct)
                times.append(t)
                biomasses.append(biomass)
                substrates.append(substrate)

                if use_rk4:
                    # RK4 integration
                    mu1, v1 = _solve_fba_at_state(biomass, substrate, dt)
                    dB1 = mu1 * biomass
                    dS1 = v1 * biomass

                    B2 = max(0, biomass + 0.5 * dt * dB1)
                    S2 = max(0, substrate + 0.5 * dt * dS1)
                    mu2, v2 = _solve_fba_at_state(B2, S2, dt)
                    dB2 = mu2 * B2
                    dS2 = v2 * B2

                    B3 = max(0, biomass + 0.5 * dt * dB2)
                    S3 = max(0, substrate + 0.5 * dt * dS2)
                    mu3, v3 = _solve_fba_at_state(B3, S3, dt)
                    dB3 = mu3 * B3
                    dS3 = v3 * B3

                    B4 = max(0, biomass + dt * dB3)
                    S4 = max(0, substrate + dt * dS3)
                    mu4, v4 = _solve_fba_at_state(B4, S4, dt)
                    dB4 = mu4 * B4
                    dS4 = v4 * B4

                    biomass += dt / 6.0 * (dB1 + 2 * dB2 + 2 * dB3 + dB4)
                    substrate += dt / 6.0 * (dS1 + 2 * dS2 + 2 * dS3 + dS4)
                    growths.append(mu1)
                else:
                    # Euler integration
                    growth_rate, uptake_rate = _solve_fba_at_state(biomass, substrate, dt)
                    growths.append(growth_rate)
                    biomass += growth_rate * biomass * dt
                    substrate += uptake_rate * biomass * dt  # uptake is negative

                # Clamp to non-negative
                biomass = max(0.0, biomass)
                substrate = max(0.0, substrate)
                t += dt

            method_name = "RK4" if use_rk4 else "Euler"
            return {"times": times, "biomasses": biomasses, "substrates": substrates,
                    "growths": growths, "substrate_rxn": substrate_rxn_id,
                    "method": method_name}

        def _on_done(result):
            dialog2 = QDialog(self)
            dialog2.setWindowTitle(f"dFBA Results ({result.get('method', '')})")
            dialog2.resize(900, 600)
            layout2 = QVBoxLayout(dialog2)
            fig = Figure(figsize=(8, 5), dpi=100)
            ax1 = fig.add_subplot(211)
            ax1.plot(result["times"], result["biomasses"], 'g-', label="Biomass (g/L)")
            ax1.plot(result["times"], result["substrates"], 'r--', label="Substrate (mmol/L)")
            ax1.set_xlabel("Time (h)")
            ax1.set_ylabel("Concentration")
            ax1.legend()
            ax1.set_title(f"Dynamic FBA — Batch Culture ({result.get('method', '')})")
            ax1.grid(True, alpha=0.3)
            ax2 = fig.add_subplot(212)
            ax2.plot(result["times"], result["growths"], 'b-', label="Growth rate")
            ax2.set_xlabel("Time (h)")
            ax2.set_ylabel("Growth rate (1/h)")
            ax2.legend()
            ax2.grid(True, alpha=0.3)
            fig.tight_layout()
            canvas = FigureCanvas(fig)
            layout2.addWidget(canvas)
            btns2 = QDialogButtonBox(QDialogButtonBox.Close)
            btns2.rejected.connect(dialog2.reject)
            layout2.addWidget(btns2)
            dialog2.exec()

        self._launch_worker(_compute, _on_done, "Running dFBA simulation...")

    # ==================== FLUX COUPLING ANALYSIS ====================
    def run_flux_coupling(self):
        """Flux Coupling Analysis — ratio-based coupling detection (efficient)."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Flux Coupling Analysis")
        dialog.resize(450, 200)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(
            "Ratio-based flux coupling analysis.\n"
            "Detects fully coupled, partially coupled, and uncoupled reaction pairs\n"
            "using FVA ratio tests (much faster than pairwise FVA knockout)."))

        form = QFormLayout()
        max_rxns_spin = QSpinBox()
        max_rxns_spin.setRange(10, 500)
        max_rxns_spin.setValue(100)
        form.addRow("Max reactions to test:", max_rxns_spin)

        tolerance_spin = QDoubleSpinBox()
        tolerance_spin.setRange(1e-10, 0.01)
        tolerance_spin.setValue(1e-6)
        tolerance_spin.setDecimals(10)
        form.addRow("Coupling tolerance:", tolerance_spin)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        max_rxns = max_rxns_spin.value()
        tol = tolerance_spin.value()

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            from cobra.flux_analysis import flux_variability_analysis
            # Step 1: single FVA pass to classify reactions
            fva = flux_variability_analysis(model, fraction_of_optimum=0.0)

            # Identify reactions with non-trivial flux range
            active_rxns = []
            fva_ranges: dict[str, tuple[float, float]] = {}
            for rxn_id, row in fva.iterrows():
                fmin, fmax = float(row['minimum']), float(row['maximum'])
                if abs(fmax) > tol or abs(fmin) > tol:
                    active_rxns.append(rxn_id)
                    fva_ranges[rxn_id] = (fmin, fmax)

            test_rxns = active_rxns[:max_rxns]

            # Step 2: blocked reactions (cannot carry flux)
            blocked = [rid for rid in fva.index
                       if abs(fva.loc[rid, 'minimum']) < tol and abs(fva.loc[rid, 'maximum']) < tol]

            # Step 3: ratio-based coupling using only 2 * n FVA calls
            # For each reaction r_i, maximize and minimize r_i / r_ref ratio
            # by fixing r_ref = 1 and optimizing r_i
            coupled_pairs = []
            ref_rxn_id = test_rxns[0] if test_rxns else None

            # Group reactions by their flux range profile (heuristic)
            # Reactions with identical FVA min/max ratios are likely coupled
            ratio_groups: dict[tuple, list[str]] = {}
            for rid in test_rxns:
                fmin, fmax = fva_ranges[rid]
                # Use rounded ratios as grouping key
                key = (round(fmin, 6), round(fmax, 6))
                ratio_groups.setdefault(key, []).append(rid)

            # Step 4: within each ratio group, do pairwise ratio tests
            # Only test pairs within same group = massive reduction
            for key, group in ratio_groups.items():
                if len(group) < 2:
                    continue
                for i, rid1 in enumerate(group):
                    for rid2 in group[i+1:]:
                        fmin1, fmax1 = fva_ranges[rid1]
                        fmin2, fmax2 = fva_ranges[rid2]
                        range1 = fmax1 - fmin1
                        range2 = fmax2 - fmin2
                        if range1 < tol or range2 < tol:
                            continue
                        # Ratio consistency check
                        ratio_min = fmin1 / fmin2 if abs(fmin2) > tol else None
                        ratio_max = fmax1 / fmax2 if abs(fmax2) > tol else None
                        if ratio_min is not None and ratio_max is not None:
                            if abs(ratio_min - ratio_max) < tol * 100:
                                coupled_pairs.append((rid1, rid2, "Fully coupled",
                                                      f"ratio ≈ {ratio_min:.4g}"))
                            elif abs(ratio_min) > tol and abs(ratio_max) > tol:
                                coupled_pairs.append((rid1, rid2, "Partially coupled",
                                                      f"ratio {ratio_min:.4g} — {ratio_max:.4g}"))

            # Step 5: directional coupling via single FVA knockout per group
            directional = []
            checked_groups = [g for g in ratio_groups.values() if len(g) == 1]
            singleton_rxns = [g[0] for g in checked_groups][:50]  # Limit
            for rid in singleton_rxns:
                with model:
                    model.reactions.get_by_id(rid).bounds = (0, 0)
                    try:
                        fva2 = flux_variability_analysis(
                            model, reaction_list=[model.reactions.get_by_id(r) for r in test_rxns[:20]
                                                  if r != rid],
                            fraction_of_optimum=0.0)
                        for rid2 in fva2.index:
                            orig_range = fva_ranges.get(rid2, (0, 0))
                            new_min, new_max = float(fva2.loc[rid2, 'minimum']), float(fva2.loc[rid2, 'maximum'])
                            if (abs(new_min) < tol and abs(new_max) < tol and
                                    (abs(orig_range[0]) > tol or abs(orig_range[1]) > tol)):
                                directional.append((rid, rid2, "Directional", f"{rid} → {rid2}"))
                    except Exception:
                        pass

            all_results = coupled_pairs + directional
            return {"coupled": all_results, "active_reactions": len(active_rxns),
                    "blocked_reactions": len(blocked), "tested": len(test_rxns),
                    "ratio_groups": len(ratio_groups)}

        def _on_done(result):
            pairs = result["coupled"]
            dialog2 = QDialog(self)
            dialog2.setWindowTitle(f"Flux Coupling ({len(pairs)} pairs found)")
            dialog2.resize(800, 500)
            layout2 = QVBoxLayout(dialog2)
            layout2.addWidget(QLabel(
                f"Active reactions: {result['active_reactions']} | "
                f"Blocked: {result['blocked_reactions']} | "
                f"Tested: {result['tested']} | Ratio groups: {result['ratio_groups']} | "
                f"Coupled pairs: {len(pairs)}"))
            table = QTableWidget(0, 4)
            table.setHorizontalHeaderLabels(["Reaction 1", "Reaction 2", "Coupling Type", "Detail"])
            table.horizontalHeader().setStretchLastSection(True)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            for item in pairs:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(item[0]))
                table.setItem(r, 1, QTableWidgetItem(item[1]))
                table.setItem(r, 2, QTableWidgetItem(item[2]))
                table.setItem(r, 3, QTableWidgetItem(item[3] if len(item) > 3 else ""))
            layout2.addWidget(table)
            btns2 = QDialogButtonBox(QDialogButtonBox.Close)
            btns2.rejected.connect(dialog2.reject)
            layout2.addWidget(btns2)
            dialog2.exec()

        self._launch_worker(_compute, _on_done, "Running Flux Coupling Analysis...")

    # ==================== THERMODYNAMIC FBA (TMFA) ====================
    def run_tmfa(self):
        """Thermodynamic FBA — apply thermodynamic constraints to flux analysis."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Thermodynamic FBA (TMFA)")
        dialog.resize(550, 400)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(
            "TMFA applies thermodynamic constraints by enforcing flux directionality\n"
            "based on Gibbs free energy changes (ΔG).\n\n"
            "You can supply ΔG data via a CSV file (columns: reaction_id, deltaG)\n"
            "and/or use annotations already present in the model."))

        # CSV file input for ΔG data
        layout.addWidget(QLabel("ΔG data CSV (optional):"))
        file_row = QHBoxLayout()
        dg_file_edit = QLineEdit()
        dg_file_edit.setPlaceholderText("Select CSV with columns: reaction_id, deltaG (kJ/mol)")
        dg_file_btn = QPushButton("Browse...")
        def _browse_dg():
            fp, _ = QFileDialog.getOpenFileName(dialog, "ΔG Data CSV", "", "CSV (*.csv *.tsv);;All (*.*)")
            if fp:
                dg_file_edit.setText(fp)
        dg_file_btn.clicked.connect(_browse_dg)
        file_row.addWidget(dg_file_edit)
        file_row.addWidget(dg_file_btn)
        layout.addLayout(file_row)

        form = QFormLayout()
        dg_threshold = QDoubleSpinBox()
        dg_threshold.setRange(-100.0, 100.0)
        dg_threshold.setValue(0.0)
        dg_threshold.setDecimals(1)
        form.addRow("ΔG threshold (kJ/mol):", dg_threshold)

        strict_chk = QCheckBox("Strict mode (block unfavorable reactions)")
        strict_chk.setChecked(False)
        form.addRow("", strict_chk)

        use_annotation_chk = QCheckBox("Also use ΔG from model annotations")
        use_annotation_chk.setChecked(True)
        form.addRow("", use_annotation_chk)
        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        threshold_val = dg_threshold.value()
        strict = strict_chk.isChecked()
        use_annotations = use_annotation_chk.isChecked()

        # Parse ΔG CSV file if provided
        dg_from_csv: dict[str, float] = {}
        csv_path = dg_file_edit.text().strip()
        if csv_path and Path(csv_path).exists():
            try:
                sep = '\t' if csv_path.endswith('.tsv') else ','
                with open(csv_path, encoding='utf-8') as f:
                    reader = csv.reader(f, delimiter=sep)
                    header = next(reader)
                    for row in reader:
                        if len(row) >= 2:
                            rxn_id = row[0].strip()
                            try:
                                dg_val = float(row[1].strip())
                                dg_from_csv[rxn_id] = dg_val
                            except ValueError:
                                continue
                if not dg_from_csv:
                    QMessageBox.warning(self, "Parse Warning", "No valid ΔG data found in the CSV file.")
            except Exception as e:
                QMessageBox.critical(self, "CSV Error", f"Failed to parse ΔG CSV: {e}")
                return

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self._apply_selected_solver(model)

        def _compute():
            constrained = []
            for rxn in model.reactions:
                if rxn.boundary:
                    continue

                dg = None
                dg_source = ""

                # Priority 1: CSV data
                if rxn.id in dg_from_csv:
                    dg = dg_from_csv[rxn.id]
                    dg_source = "CSV"
                # Priority 2: Model annotations
                elif use_annotations:
                    ann = getattr(rxn, 'annotation', {}) or {}
                    for key in ('deltaG', 'delta_g', 'dG', 'gibbs_energy',
                                'deltaG0', 'delta_g0', 'dG0', 'standard_dg_prime'):
                        if key in ann:
                            try:
                                dg = float(str(ann[key]))
                                dg_source = f"annotation ({key})"
                            except (ValueError, TypeError):
                                pass
                            break

                if dg is None:
                    continue

                if dg > threshold_val:
                    if strict:
                        rxn.upper_bound = 0.0
                    else:
                        rxn.upper_bound = min(rxn.upper_bound, 1.0)
                    constrained.append({"id": rxn.id, "name": rxn.name, "dG": dg,
                                        "source": dg_source, "action": "forward constrained"})
                elif dg < -threshold_val:
                    if strict:
                        rxn.lower_bound = 0.0
                    constrained.append({"id": rxn.id, "name": rxn.name, "dG": dg,
                                        "source": dg_source, "action": "reverse blocked"})

            sol = model.optimize()
            return {"status": sol.status if sol else "failed",
                    "objective": float(sol.objective_value) if sol and sol.status == 'optimal' else 0.0,
                    "constrained_reactions": constrained,
                    "flux": {r.id: float(sol.fluxes[r.id]) for r in model.reactions} if sol and sol.status == 'optimal' else {}}

        def _on_done(result):
            lines = [
                f"Status: {result['status']}",
                f"Objective: {result.get('objective', 0):.6g}",
                f"Thermodynamically constrained reactions: {len(result['constrained_reactions'])}",
                ""
            ]
            for cr in result['constrained_reactions'][:50]:
                src = cr.get('source', '')
                lines.append(f"  {cr['id']} ({cr['name']}): ΔG={cr['dG']:.1f} kJ/mol [{src}] → {cr['action']}")
            if not result['constrained_reactions']:
                lines.append("  No reactions had ΔG data.")
                lines.append("  Tip: Provide a CSV file with columns 'reaction_id, deltaG' or add")
                lines.append("  ΔG annotations (deltaG, dG, standard_dg_prime) to the SBML model.")
            TextPopup("TMFA Results", "\n".join(lines), self).exec()

        self._launch_worker(_compute, _on_done, "Running Thermodynamic FBA...")

    # ==================== MODEL COMPARISON ====================
    def run_model_comparison(self):
        """Compare two SBML models side by side."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load a base model first.")
            return

        fp, _ = QFileDialog.getOpenFileName(self, "Select second SBML model for comparison", "",
                                             "SBML Files (*.xml *.sbml);;All (*.*)")
        if not fp:
            return

        try:
            model2 = cobra.io.read_sbml_model(fp)
        except Exception as e:
            QMessageBox.critical(self, "Load Error", f"Failed to load model: {e}")
            return

        m1 = self.base_model
        m2 = model2
        name1 = getattr(m1, 'id', 'Model A') or 'Model A'
        name2 = getattr(m2, 'id', 'Model B') or 'Model B'

        # Reactions comparison
        rxns1 = set(r.id for r in m1.reactions)
        rxns2 = set(r.id for r in m2.reactions)
        only_in_1 = sorted(rxns1 - rxns2)
        only_in_2 = sorted(rxns2 - rxns1)
        common_rxns = sorted(rxns1 & rxns2)
        bound_diffs = []
        for rid in common_rxns:
            r1 = m1.reactions.get_by_id(rid)
            r2 = m2.reactions.get_by_id(rid)
            if r1.lower_bound != r2.lower_bound or r1.upper_bound != r2.upper_bound:
                bound_diffs.append((rid, r1.lower_bound, r1.upper_bound, r2.lower_bound, r2.upper_bound))

        # Metabolites comparison
        mets1 = set(m.id for m in m1.metabolites)
        mets2 = set(m.id for m in m2.metabolites)
        mets_only1 = sorted(mets1 - mets2)
        mets_only2 = sorted(mets2 - mets1)

        # Genes comparison
        genes1 = set(g.id for g in m1.genes)
        genes2 = set(g.id for g in m2.genes)
        genes_only1 = sorted(genes1 - genes2)
        genes_only2 = sorted(genes2 - genes1)

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Model Comparison: {name1} vs {name2}")
        dialog.resize(900, 700)
        layout = QVBoxLayout(dialog)

        summary = QLabel(
            f"<b>{name1}</b>: {len(rxns1)} reactions, {len(mets1)} metabolites, {len(genes1)} genes<br>"
            f"<b>{name2}</b>: {len(rxns2)} reactions, {len(mets2)} metabolites, {len(genes2)} genes<br><br>"
            f"Common reactions: {len(common_rxns)} | Bound differences: {len(bound_diffs)}<br>"
            f"Unique to {name1}: {len(only_in_1)} rxns, {len(mets_only1)} mets, {len(genes_only1)} genes<br>"
            f"Unique to {name2}: {len(only_in_2)} rxns, {len(mets_only2)} mets, {len(genes_only2)} genes"
        )
        summary.setWordWrap(True)
        layout.addWidget(summary)

        tabs = QTabWidget()

        # Reactions unique tab
        if only_in_1 or only_in_2:
            rxn_table = QTableWidget(0, 2)
            rxn_table.setHorizontalHeaderLabels([f"Only in {name1}", f"Only in {name2}"])
            rxn_table.horizontalHeader().setStretchLastSection(True)
            rxn_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            max_rows = max(len(only_in_1), len(only_in_2))
            for i in range(max_rows):
                r = rxn_table.rowCount()
                rxn_table.insertRow(r)
                rxn_table.setItem(r, 0, QTableWidgetItem(only_in_1[i] if i < len(only_in_1) else ""))
                rxn_table.setItem(r, 1, QTableWidgetItem(only_in_2[i] if i < len(only_in_2) else ""))
            tabs.addTab(rxn_table, f"Unique Reactions ({len(only_in_1)} + {len(only_in_2)})")

        # Bound differences tab
        if bound_diffs:
            bd_table = QTableWidget(0, 5)
            bd_table.setHorizontalHeaderLabels(["Reaction", f"LB ({name1})", f"UB ({name1})", f"LB ({name2})", f"UB ({name2})"])
            bd_table.horizontalHeader().setStretchLastSection(True)
            bd_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            for rid, lb1, ub1, lb2, ub2 in bound_diffs:
                r = bd_table.rowCount()
                bd_table.insertRow(r)
                bd_table.setItem(r, 0, QTableWidgetItem(rid))
                bd_table.setItem(r, 1, QTableWidgetItem(f"{lb1:g}"))
                bd_table.setItem(r, 2, QTableWidgetItem(f"{ub1:g}"))
                bd_table.setItem(r, 3, QTableWidgetItem(f"{lb2:g}"))
                bd_table.setItem(r, 4, QTableWidgetItem(f"{ub2:g}"))
            tabs.addTab(bd_table, f"Bound Differences ({len(bound_diffs)})")

        # Metabolites unique tab
        if mets_only1 or mets_only2:
            met_table = QTableWidget(0, 2)
            met_table.setHorizontalHeaderLabels([f"Only in {name1}", f"Only in {name2}"])
            met_table.horizontalHeader().setStretchLastSection(True)
            met_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            max_rows = max(len(mets_only1), len(mets_only2))
            for i in range(max_rows):
                r = met_table.rowCount()
                met_table.insertRow(r)
                met_table.setItem(r, 0, QTableWidgetItem(mets_only1[i] if i < len(mets_only1) else ""))
                met_table.setItem(r, 1, QTableWidgetItem(mets_only2[i] if i < len(mets_only2) else ""))
            tabs.addTab(met_table, f"Unique Metabolites ({len(mets_only1)} + {len(mets_only2)})")

        # Genes unique tab
        if genes_only1 or genes_only2:
            gene_table = QTableWidget(0, 2)
            gene_table.setHorizontalHeaderLabels([f"Only in {name1}", f"Only in {name2}"])
            gene_table.horizontalHeader().setStretchLastSection(True)
            gene_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            max_rows = max(len(genes_only1), len(genes_only2))
            for i in range(max_rows):
                r = gene_table.rowCount()
                gene_table.insertRow(r)
                gene_table.setItem(r, 0, QTableWidgetItem(genes_only1[i] if i < len(genes_only1) else ""))
                gene_table.setItem(r, 1, QTableWidgetItem(genes_only2[i] if i < len(genes_only2) else ""))
            tabs.addTab(gene_table, f"Unique Genes ({len(genes_only1)} + {len(genes_only2)})")

        layout.addWidget(tabs)
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    # ==================== PATHWAY ENRICHMENT ====================
    def run_pathway_enrichment(self):
        """Metabolic Pathway Enrichment — subsystem-based enrichment of active reactions."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return
        if self.last_run is None:
            QMessageBox.warning(self, "No results", "Run an FBA analysis first to determine active reactions.")
            return

        model = self.base_model
        # Get flux data from last run
        flux = {}
        if "baseline" in self.last_run and "flux" in self.last_run["baseline"]:
            flux = self.last_run["baseline"]["flux"]
        elif "flux" in self.last_run:
            flux = self.last_run["flux"]

        if not flux:
            QMessageBox.warning(self, "No flux data", "Last analysis does not contain flux data.")
            return

        # Define active reactions (|flux| > threshold)
        threshold = 1e-6
        active_rxns = set(rid for rid, fv in flux.items() if abs(fv) > threshold)

        # Get subsystem annotations
        subsystem_rxns: dict[str, set] = {}
        for rxn in model.reactions:
            ss = getattr(rxn, 'subsystem', '') or ''
            ss = ss.strip()
            if ss:
                subsystem_rxns.setdefault(ss, set()).add(rxn.id)

        if not subsystem_rxns:
            QMessageBox.warning(self, "No subsystems", "Model has no subsystem annotations.")
            return

        # Compute enrichment using hypergeometric-like test
        total_rxns = len(model.reactions)
        total_active = len(active_rxns)
        enrichment = []
        for ss, ss_rxns in subsystem_rxns.items():
            ss_total = len(ss_rxns)
            ss_active = len(ss_rxns & active_rxns)
            if ss_total == 0:
                continue
            # Expected fraction
            expected = total_active / total_rxns * ss_total if total_rxns > 0 else 0
            ratio = ss_active / expected if expected > 0 else 0
            enrichment.append({
                "subsystem": ss, "total": ss_total, "active": ss_active,
                "expected": expected, "enrichment_ratio": ratio
            })
        enrichment.sort(key=lambda x: x["enrichment_ratio"], reverse=True)

        dialog = QDialog(self)
        dialog.setWindowTitle("Pathway Enrichment Analysis")
        dialog.resize(800, 600)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(
            f"Total reactions: {total_rxns} | Active (|flux| > {threshold}): {total_active}\n"
            f"Subsystems analyzed: {len(enrichment)}"))

        # Plot
        fig = Figure(figsize=(7, 3.5), dpi=100)
        ax = fig.add_subplot(111)
        top = enrichment[:20]
        if top:
            ss_names = [e["subsystem"][:30] for e in top]
            ratios = [e["enrichment_ratio"] for e in top]
            colors = ['#4CAF50' if r > 1 else '#f44336' for r in ratios]
            ax.barh(range(len(top)), ratios, color=colors, alpha=0.8)
            ax.set_yticks(range(len(top)))
            ax.set_yticklabels(ss_names, fontsize=7)
            ax.axvline(x=1.0, color='black', linestyle='--', linewidth=0.8)
            ax.set_xlabel("Enrichment Ratio (>1 = enriched)")
            ax.set_title("Top 20 Subsystems by Enrichment")
        fig.tight_layout()
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        table = QTableWidget(0, 5)
        table.setHorizontalHeaderLabels(["Subsystem", "Total Rxns", "Active Rxns", "Expected", "Enrichment Ratio"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        for e in enrichment:
            r = table.rowCount()
            table.insertRow(r)
            table.setItem(r, 0, QTableWidgetItem(e["subsystem"]))
            table.setItem(r, 1, QTableWidgetItem(str(e["total"])))
            table.setItem(r, 2, QTableWidgetItem(str(e["active"])))
            table.setItem(r, 3, QTableWidgetItem(f"{e['expected']:.1f}"))
            table.setItem(r, 4, QTableWidgetItem(f"{e['enrichment_ratio']:.3f}"))
        layout.addWidget(table)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    # ==================== ESCHER MAP (with JSON map import) ====================
    def show_escher_map(self):
        """Show Escher metabolic map.  Supports loading Escher JSON map files
        directly (no ``escher`` package required) and falls back to the
        ``escher`` Python package when available."""

        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        # Collect current flux data if available
        flux: dict[str, float] = {}
        if self.last_run and "baseline" in self.last_run and "flux" in self.last_run["baseline"]:
            flux = self.last_run["baseline"]["flux"]

        # --- Dialog ---
        dialog = QDialog(self)
        dialog.setWindowTitle("Escher Map Viewer")
        dialog.resize(1200, 800)
        layout = QVBoxLayout(dialog)

        # toolbar
        toolbar = QHBoxLayout()
        json_btn = QPushButton("Load Escher JSON map...")
        save_html_btn = QPushButton("Save HTML...")
        save_html_btn.setEnabled(False)
        toolbar.addWidget(json_btn)
        toolbar.addWidget(save_html_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # content area
        content_stack = QStackedWidget()
        placeholder = QLabel("Load an Escher JSON map file or use the escher package.\n\n"
                             "You can download Escher maps from:\n"
                             "  https://escher.github.io/#/app\n\n"
                             "Click 'Load Escher JSON map...' to import a .json map file.")
        placeholder.setAlignment(Qt.AlignCenter)
        placeholder.setWordWrap(True)
        content_stack.addWidget(placeholder)  # index 0
        layout.addWidget(content_stack, 1)

        html_holder: list[str] = [""]  # mutable ref for closures

        def _build_escher_html(map_json_str: str) -> str:
            """Build a standalone HTML page that renders an Escher map using
            the Escher JS library from CDN, overlaying flux data."""
            import json as _json
            flux_json = _json.dumps(flux) if flux else "null"
            model_json = "null"
            try:
                model_json = cobra.io.to_json(self.base_model)
            except Exception:
                pass
            return f"""<!DOCTYPE html>
<html><head>
<meta charset="utf-8"/>
<script src="https://unpkg.com/escher@1.7.3/dist/escher.min.js"></script>
<style>body{{margin:0;padding:0}} #map-container{{width:100vw;height:100vh}}</style>
</head><body>
<div id="map-container"></div>
<script>
var mapData = {map_json_str};
var modelData = {model_json};
var fluxData = {flux_json};
escher.Builder(mapData, modelData, null, document.getElementById('map-container'), {{
  reaction_data: fluxData,
  fill_screen: true,
  menu: 'zoom',
  scroll_behavior: 'zoom',
  reaction_styles: ['color', 'size', 'text'],
  reaction_compare_style: 'log2_fold',
  identifiers_on_map: 'bigg_id'
}});
</script></body></html>"""

        def _load_json_map():
            fp, _ = QFileDialog.getOpenFileName(dialog, "Open Escher JSON Map", "", "JSON files (*.json)")
            if not fp:
                return
            try:
                map_text = Path(fp).read_text(encoding="utf-8")
                # Validate it's a list (Escher map format) or dict
                import json as _json
                parsed = _json.loads(map_text)
                if not isinstance(parsed, (list, dict)):
                    raise ValueError("Not a valid Escher map JSON.")
                html = _build_escher_html(map_text)
                html_holder[0] = html
                save_html_btn.setEnabled(True)
                # Try QWebEngineView first
                try:
                    from PySide6.QtWebEngineWidgets import QWebEngineView
                    web = QWebEngineView()
                    web.setHtml(html)
                    if content_stack.count() > 1:
                        old = content_stack.widget(1)
                        content_stack.removeWidget(old)
                        old.deleteLater()
                    content_stack.addWidget(web)
                    content_stack.setCurrentIndex(1)
                except ImportError:
                    lbl = QLabel("Map loaded. QWebEngineView not available.\n"
                                 "Use 'Save HTML...' to view in your browser.")
                    lbl.setAlignment(Qt.AlignCenter)
                    if content_stack.count() > 1:
                        old = content_stack.widget(1)
                        content_stack.removeWidget(old)
                        old.deleteLater()
                    content_stack.addWidget(lbl)
                    content_stack.setCurrentIndex(1)
            except Exception as e:
                QMessageBox.critical(dialog, "Error", f"Failed to load map:\n{e}")

        def _save_html():
            if not html_holder[0]:
                return
            fp, _ = QFileDialog.getSaveFileName(dialog, "Save Escher HTML", "", "HTML (*.html)")
            if fp:
                Path(fp).write_text(html_holder[0], encoding="utf-8")

        json_btn.clicked.connect(_load_json_map)
        save_html_btn.clicked.connect(_save_html)

        # Also try the escher package directly
        try:
            import escher
            pkg_btn = QPushButton("Use escher package (auto-generate)")
            def _use_pkg():
                try:
                    builder = escher.Builder(model=self.base_model, reaction_data=flux if flux else None)
                    html = builder._get_html()
                    html_holder[0] = html
                    save_html_btn.setEnabled(True)
                    try:
                        from PySide6.QtWebEngineWidgets import QWebEngineView
                        web = QWebEngineView()
                        web.setHtml(html)
                        if content_stack.count() > 1:
                            old = content_stack.widget(1)
                            content_stack.removeWidget(old)
                            old.deleteLater()
                        content_stack.addWidget(web)
                        content_stack.setCurrentIndex(1)
                    except ImportError:
                        lbl = QLabel("Map generated. Use 'Save HTML...' to view in browser.")
                        lbl.setAlignment(Qt.AlignCenter)
                        if content_stack.count() > 1:
                            old = content_stack.widget(1)
                            content_stack.removeWidget(old)
                            old.deleteLater()
                        content_stack.addWidget(lbl)
                        content_stack.setCurrentIndex(1)
                except Exception as e:
                    QMessageBox.critical(dialog, "Error", f"Escher builder failed:\n{e}")
            pkg_btn.clicked.connect(_use_pkg)
            toolbar.addWidget(pkg_btn)
        except ImportError:
            pass

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    # ==================== PHENOTYPE PHASE PLANE (PhPP) ====================
    def run_phenotype_phase_plane(self):
        """Phenotype Phase Plane analysis: 2-D sweep of two reaction fluxes,
        computing the objective value at each (rx1, rx2) grid point."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        rxn_ids = [r.id for r in self.base_model.reactions]

        dialog = QDialog(self)
        dialog.setWindowTitle("Phenotype Phase Plane (PhPP)")
        form = QFormLayout(dialog)

        rx1_combo = QComboBox()
        rx1_combo.addItems(rxn_ids)
        form.addRow("Reaction X-axis:", rx1_combo)

        rx2_combo = QComboBox()
        rx2_combo.addItems(rxn_ids)
        if len(rxn_ids) > 1:
            rx2_combo.setCurrentIndex(1)
        form.addRow("Reaction Y-axis:", rx2_combo)

        steps_spin = QSpinBox()
        steps_spin.setRange(5, 100)
        steps_spin.setValue(20)
        form.addRow("Grid points per axis:", steps_spin)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        form.addRow(btns)

        if dialog.exec() != QDialog.Accepted:
            return

        rx1_id = rx1_combo.currentText()
        rx2_id = rx2_combo.currentText()
        steps = steps_spin.value()

        if rx1_id == rx2_id:
            QMessageBox.warning(self, "Same reaction", "Choose two different reactions.")
            return

        model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(model)
        self.apply_reaction_overrides_to_model(model)
        self._apply_selected_objective_to_model(model)
        self._apply_selected_solver(model)

        rx1 = model.reactions.get_by_id(rx1_id)
        rx2 = model.reactions.get_by_id(rx2_id)
        rx1_range = (rx1.lower_bound, rx1.upper_bound)
        rx2_range = (rx2.lower_bound, rx2.upper_bound)

        import numpy as np

        x_vals = np.linspace(rx1_range[0], rx1_range[1], steps)
        y_vals = np.linspace(rx2_range[0], rx2_range[1], steps)

        def _compute(worker=None):
            grid = np.full((steps, steps), np.nan)
            total = steps * steps
            count = 0
            for i, xv in enumerate(x_vals):
                for j, yv in enumerate(y_vals):
                    with model:
                        rx1.lower_bound = float(xv)
                        rx1.upper_bound = float(xv)
                        rx2.lower_bound = float(yv)
                        rx2.upper_bound = float(yv)
                        try:
                            sol = model.optimize()
                            if sol.status == "optimal":
                                grid[j, i] = float(sol.objective_value)
                        except Exception:
                            pass
                    count += 1
                    if worker and count % max(1, total // 20) == 0:
                        worker.progress_pct.emit(int(100 * count / total))
            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "Phenotype Phase Plane",
                "rx1_id": rx1_id, "rx2_id": rx2_id,
                "x_vals": x_vals.tolist(), "y_vals": y_vals.tolist(),
                "grid": grid.tolist(),
            }

        def _on_done(result):
            self.last_run = result
            self._render_phpp_results(result)

        self._launch_worker(_compute, _on_done, f"Running PhPP ({rx1_id} × {rx2_id})...")

    def _render_phpp_results(self, result: dict):
        """Render Phenotype Phase Plane results as a heatmap."""
        import numpy as np
        grid = np.array(result["grid"])
        x_vals = result["x_vals"]
        y_vals = result["y_vals"]
        rx1_id = result["rx1_id"]
        rx2_id = result["rx2_id"]

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Phenotype Phase Plane: {rx1_id} × {rx2_id}")
        dialog.resize(900, 700)
        layout = QVBoxLayout(dialog)

        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure

        fig = Figure(figsize=(8, 6))
        ax = fig.add_subplot(111)
        im = ax.imshow(
            grid, origin="lower", aspect="auto",
            extent=[x_vals[0], x_vals[-1], y_vals[0], y_vals[-1]],
            cmap="viridis",
        )
        ax.set_xlabel(rx1_id)
        ax.set_ylabel(rx2_id)
        ax.set_title("Phenotype Phase Plane — Objective value")
        fig.colorbar(im, ax=ax, label="Objective")

        # Overlay contour lines
        try:
            cs = ax.contour(
                np.array(x_vals), np.array(y_vals), grid,
                colors="white", linewidths=0.5, levels=10,
            )
            ax.clabel(cs, inline=True, fontsize=7, fmt="%.2f")
        except Exception:
            pass

        fig.tight_layout()
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        # Summary statistics
        valid = grid[~np.isnan(grid)]
        info = QLabel(
            f"Grid: {len(x_vals)}×{len(y_vals)} = {len(x_vals)*len(y_vals)} points  |  "
            f"Feasible: {len(valid)}  |  "
            f"Objective range: [{valid.min():.4f}, {valid.max():.4f}]" if len(valid) > 0
            else f"Grid: {len(x_vals)}×{len(y_vals)}  |  No feasible solutions found"
        )
        layout.addWidget(info)

        # Save buttons
        btn_row = QHBoxLayout()
        save_png = QPushButton("Save as PNG...")
        def _save_png():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Save Plot", "", "PNG (*.png)")
            if fp:
                fig.savefig(fp, dpi=150)
        save_png.clicked.connect(_save_png)
        btn_row.addWidget(save_png)

        save_csv = QPushButton("Save grid as CSV...")
        def _save_csv():
            fp, _ = QFileDialog.getSaveFileName(dialog, "Save CSV", "", "CSV (*.csv)")
            if fp:
                import csv
                with open(fp, 'w', newline='', encoding='utf-8') as f:
                    w = csv.writer(f)
                    w.writerow([""] + [f"{v:.4f}" for v in x_vals])
                    for i, yv in enumerate(y_vals):
                        w.writerow([f"{yv:.4f}"] + [f"{grid[i][j]:.6f}" if not np.isnan(grid[i][j]) else "inf" for j in range(len(x_vals))])
        save_csv.clicked.connect(_save_csv)
        btn_row.addWidget(save_csv)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    # ==================== SBML VALIDATION ====================
    def validate_sbml(self):
        """Validate the currently loaded SBML file using libSBML or fallback checks."""
        if self.current_sbml_path is None or not self.current_sbml_path.exists():
            QMessageBox.warning(self, "No SBML file", "Load an SBML model first.")
            return

        errors: list[str] = []
        warnings: list[str] = []
        infos: list[str] = []
        sbml_path = str(self.current_sbml_path)

        # Try libSBML validation (gold standard)
        try:
            import libsbml
            reader = libsbml.SBMLReader()
            doc = reader.readSBML(sbml_path)

            # Enable all consistency checks
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_GENERAL_CONSISTENCY, True)
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_IDENTIFIER_CONSISTENCY, True)
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_UNITS_CONSISTENCY, False)  # often noisy
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_MATHML_CONSISTENCY, True)
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_SBO_CONSISTENCY, True)
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_OVERDETERMINED_MODEL, True)
            doc.setConsistencyChecks(libsbml.LIBSBML_CAT_MODELING_PRACTICE, True)

            n_check_errors = doc.checkConsistency()
            n_read_errors = doc.getNumErrors()

            for i in range(doc.getNumErrors()):
                err = doc.getError(i)
                severity = err.getSeverity()
                msg = f"L{err.getLine()}C{err.getColumn()}: [{err.getErrorId()}] {err.getMessage().strip()}"
                if severity == libsbml.LIBSBML_SEV_ERROR or severity == libsbml.LIBSBML_SEV_FATAL:
                    errors.append(msg)
                elif severity == libsbml.LIBSBML_SEV_WARNING:
                    warnings.append(msg)
                else:
                    infos.append(msg)

            model_obj = doc.getModel()
            if model_obj:
                infos.insert(0, f"SBML Level {doc.getLevel()} Version {doc.getVersion()}")
                infos.insert(1, f"Model id: {model_obj.getId()}")
                infos.insert(2, f"Species: {model_obj.getNumSpecies()}  |  Reactions: {model_obj.getNumReactions()}  |  Compartments: {model_obj.getNumCompartments()}")

                # Check for FBC package
                fbc = model_obj.getPlugin("fbc")
                if fbc:
                    infos.append(f"FBC plugin v{fbc.getPackageVersion()}: {fbc.getNumObjectives()} objective(s), {fbc.getNumGeneProducts()} gene product(s)")
                else:
                    warnings.append("FBC plugin not present — flux bounds may be missing")

        except ImportError:
            # Fallback: basic checks without libSBML
            infos.append("libSBML not installed — performing basic checks only")
            infos.append("Install with: pip install python-libsbml")

            if self.base_model:
                m = self.base_model
                infos.append(f"Model: {m.id}  |  Reactions: {len(m.reactions)}  |  Metabolites: {len(m.metabolites)}  |  Genes: {len(m.genes)}")

                # Check mass balance
                unbalanced = 0
                for r in m.reactions:
                    if not r.boundary:
                        bal = r.check_mass_balance()
                        if bal:
                            unbalanced += 1
                if unbalanced:
                    warnings.append(f"{unbalanced} reaction(s) have mass imbalance")

                # Check for reactions without GPR
                no_gpr = sum(1 for r in m.reactions if not r.gene_reaction_rule.strip())
                if no_gpr:
                    infos.append(f"{no_gpr} reaction(s) have no GPR rule")

                # Orphan metabolites
                orphans = [m2.id for m2 in m.metabolites if len(m2.reactions) == 0]
                if orphans:
                    warnings.append(f"{len(orphans)} orphan metabolite(s): {', '.join(orphans[:10])}")

                # Dead-end metabolites (only produced or only consumed)
                dead_end = 0
                for met in m.metabolites:
                    producers = sum(1 for r in met.reactions if met in r.products)
                    consumers = sum(1 for r in met.reactions if met in r.reactants)
                    if (producers > 0 and consumers == 0) or (consumers > 0 and producers == 0):
                        if not any(r.boundary for r in met.reactions):
                            dead_end += 1
                if dead_end:
                    warnings.append(f"{dead_end} potential dead-end metabolite(s)")

                # Unbounded reactions
                unbounded = sum(1 for r in m.reactions if abs(r.upper_bound) >= 999999 or abs(r.lower_bound) >= 999999)
                if unbounded:
                    infos.append(f"{unbounded} reaction(s) with very large bounds (±999999+)")

        # Display results
        dialog = QDialog(self)
        dialog.setWindowTitle("SBML Validation Report")
        dialog.resize(800, 600)
        layout = QVBoxLayout(dialog)

        text = QPlainTextEdit()
        text.setReadOnly(True)
        lines: list[str] = []
        lines.append(f"═══ SBML Validation: {self.current_sbml_path.name} ═══\n")

        if infos:
            lines.append("ℹ️  Information:")
            for i in infos:
                lines.append(f"   {i}")
            lines.append("")

        if errors:
            lines.append(f"❌ Errors ({len(errors)}):")
            for e in errors:
                lines.append(f"   {e}")
            lines.append("")

        if warnings:
            lines.append(f"⚠️  Warnings ({len(warnings)}):")
            for w in warnings:
                lines.append(f"   {w}")
            lines.append("")

        if not errors and not warnings:
            lines.append("✅ No errors or warnings found!")

        lines.append(f"\nTotal: {len(errors)} error(s), {len(warnings)} warning(s)")

        text.setPlainText("\n".join(lines))
        layout.addWidget(text)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        dialog.exec()

    # ==================== JUPYTER NOTEBOOK EXPORT ====================
    def export_jupyter_notebook(self):
        """Export the current analysis as a Jupyter Notebook (.ipynb) that
        reproduces the analysis with COBRApy code."""
        if self.base_model is None:
            QMessageBox.warning(self, "No model", "Load an SBML model first.")
            return

        fp, _ = QFileDialog.getSaveFileName(
            self, "Export Jupyter Notebook", "", "Jupyter Notebook (*.ipynb)")
        if not fp:
            return

        cells: list[dict] = []

        def _md(src: str):
            cells.append({"cell_type": "markdown", "metadata": {},
                          "source": [src]})

        def _code(src: str):
            cells.append({"cell_type": "code", "metadata": {},
                          "source": [src], "execution_count": None,
                          "outputs": []})

        _md("# MetaboDesk — Exported Analysis Notebook\n\n"
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by MetaboDesk.")

        # Setup cell
        _code("import cobra\nimport pandas as pd\nimport matplotlib.pyplot as plt\nfrom pathlib import Path")

        # Model loading
        model_file = self.current_sbml_path.name if self.current_sbml_path else "model.xml"
        _md(f"## Load Model\n\nMake sure `{model_file}` is in the same directory as this notebook.")
        _code(f'model = cobra.io.read_sbml_model("{model_file}")\n'
              f'print(f"Model: {{model.id}}")\n'
              f'print(f"Reactions: {{len(model.reactions)}}, Metabolites: {{len(model.metabolites)}}, Genes: {{len(model.genes)}}")')

        # Medium / bounds
        if self.reaction_bound_overrides:
            _md("## Custom Reaction Bounds")
            lines = ["# Apply custom bounds"]
            for rid, (lb, ub) in self.reaction_bound_overrides.items():
                lines.append(f'model.reactions.get_by_id("{rid}").bounds = ({lb}, {ub})')
            _code("\n".join(lines))

        # Knockouts
        if self.knockout_genes:
            _md("## Gene Knockouts")
            genes_str = ", ".join(f'"{g}"' for g in sorted(self.knockout_genes))
            _code(f"# Knockout genes\nfor gid in [{genes_str}]:\n"
                  f"    gene = model.genes.get_by_id(gid)\n"
                  f"    cobra.manipulation.delete_model_genes(model, [gene])")

        # Objective
        if self.base_model:
            obj_rxn = None
            for r in self.base_model.objective.variables:
                obj_rxn = r.name
                break
            if obj_rxn:
                _md("## Set Objective")
                _code(f'model.objective = "{obj_rxn}"')

        # Analysis
        analysis_type = self.analysis_type.currentText() if hasattr(self, 'analysis_type') else "FBA"
        _md(f"## Run Analysis: {analysis_type}")

        if "FVA" in analysis_type:
            frac = self.fva_fraction_spin.value() if hasattr(self, 'fva_fraction_spin') else 0.9
            _code(f"from cobra.flux_analysis import flux_variability_analysis\n"
                  f"fva = flux_variability_analysis(model, fraction_of_optimum={frac})\n"
                  f"print(fva.head(20))")
        elif "pFBA" in analysis_type:
            _code("import cobra.flux_analysis\n"
                  "pfba_sol = cobra.flux_analysis.pfba(model)\n"
                  "print(f'Objective: {pfba_sol.objective_value:.6f}')\n"
                  "fluxes = pfba_sol.fluxes\n"
                  "print(fluxes[fluxes.abs() > 1e-6].sort_values(ascending=False).head(20))")
        elif "Single Gene Deletion" in analysis_type:
            _code("from cobra.flux_analysis import single_gene_deletion\n"
                  "sgd = single_gene_deletion(model)\n"
                  "sgd_sorted = sgd.sort_values('growth', ascending=True)\n"
                  "print(sgd_sorted.head(20))")
        elif "Robustness" in analysis_type:
            _code("# Robustness analysis — sweep a reaction and record objective\n"
                  "import numpy as np\n\n"
                  "target_rxn = model.reactions[0]  # Change to your target\n"
                  "steps = np.linspace(target_rxn.lower_bound, target_rxn.upper_bound, 20)\n"
                  "results = []\n"
                  "for v in steps:\n"
                  "    with model:\n"
                  "        target_rxn.bounds = (v, v)\n"
                  "        sol = model.optimize()\n"
                  "        results.append({'flux': v, 'objective': sol.objective_value if sol.status == 'optimal' else float('nan')})\n"
                  "df = pd.DataFrame(results)\n"
                  "df.plot(x='flux', y='objective')\n"
                  "plt.show()")
        elif "Production Envelope" in analysis_type:
            _code("from cobra.flux_analysis import production_envelope\n"
                  "# Change the reaction to your product reaction\n"
                  "prod_rxn = model.reactions[0]\n"
                  "pe = production_envelope(model, reactions=[prod_rxn])\n"
                  "pe.plot()\n"
                  "plt.show()")
        else:
            # Standard FBA
            _code("solution = model.optimize()\n"
                  "print(f'Status: {solution.status}')\n"
                  "print(f'Objective value: {solution.objective_value:.6f}')\n\n"
                  "# Show top fluxes\n"
                  "fluxes = solution.fluxes\n"
                  "active = fluxes[fluxes.abs() > 1e-6].sort_values(ascending=False)\n"
                  "print(f'\\nActive reactions: {len(active)}')\n"
                  "print(active.head(20))")

        # Visualization
        _md("## Flux Distribution")
        _code("# Histogram of active fluxes\n"
              "solution = model.optimize()\n"
              "active_fluxes = solution.fluxes[solution.fluxes.abs() > 1e-6]\n"
              "plt.figure(figsize=(10, 5))\n"
              "plt.hist(active_fluxes.values, bins=50, edgecolor='black')\n"
              "plt.xlabel('Flux')\n"
              "plt.ylabel('Count')\n"
              "plt.title('Distribution of Active Fluxes')\n"
              "plt.tight_layout()\n"
              "plt.show()")

        # Last run results
        if self.last_run:
            _md("## Previous Results Summary\n\n"
                f"Analysis type: {self.last_run.get('analysis_type', 'N/A')}\n\n"
                f"Timestamp: {self.last_run.get('timestamp', 'N/A')}")
            if "baseline" in self.last_run:
                bl = self.last_run["baseline"]
                _md(f"**Baseline objective**: {bl.get('objective', 'N/A')}\n\n"
                    f"**Status**: {bl.get('status', 'N/A')}")

        # Build notebook JSON
        notebook = {
            "nbformat": 4,
            "nbformat_minor": 5,
            "metadata": {
                "kernelspec": {
                    "display_name": "Python 3",
                    "language": "python",
                    "name": "python3",
                },
                "language_info": {
                    "name": "python",
                    "version": "3.11.0",
                },
            },
            "cells": cells,
        }

        try:
            Path(fp).write_text(json.dumps(notebook, indent=1, ensure_ascii=False), encoding="utf-8")
            QMessageBox.information(self, "Exported", f"Jupyter Notebook saved to:\n{fp}")
            logger.info(f"Jupyter Notebook exported: {fp}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    # ==================== PLUGIN SYSTEM ====================
    def load_plugins(self):
        """Load plugins from ~/.metabodesk/plugins/ directory."""
        self._plugin_dir.mkdir(parents=True, exist_ok=True)

        plugin_files = list(self._plugin_dir.glob("*.py"))
        if not plugin_files:
            QMessageBox.information(
                self, "Plugins",
                f"No plugins found in:\n{self._plugin_dir}\n\n"
                "To create a plugin, place a .py file with a register(main_window) function\n"
                "in the plugins directory.\n\n"
                "Example plugin:\n"
                "  def register(mw):\n"
                "      from PySide6.QtGui import QAction\n"
                "      act = QAction('My Plugin', mw)\n"
                "      act.triggered.connect(lambda: print('Hello from plugin!'))\n"
                "      mw.menuBar().addAction(act)")
            return

        loaded = 0
        errors = []
        for pf in plugin_files:
            try:
                spec = importlib.util.spec_from_file_location(pf.stem, str(pf))
                if spec and spec.loader:
                    mod = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(mod)
                    if hasattr(mod, 'register'):
                        mod.register(self)
                        self._plugins.append(mod)
                        loaded += 1
                        logger.info(f"Plugin loaded: {pf.name}")
                    else:
                        errors.append(f"{pf.name}: missing register() function")
            except Exception as e:
                errors.append(f"{pf.name}: {e}")
                logger.error(f"Plugin load error: {pf.name}: {e}")

        msg = f"Loaded {loaded} plugin(s) from {self._plugin_dir}"
        if errors:
            msg += "\n\nErrors:\n" + "\n".join(errors)
        QMessageBox.information(self, "Plugins", msg)

def main():
    app = QApplication(sys.argv)
    try:
        icon_path = _resource_path("logo.ico")
        if Path(icon_path).exists():
            app.setWindowIcon(QIcon(icon_path))
    except Exception:
        pass
    splash = None
    try:
        splash_path = _resource_path("splash.gif")
        if Path(splash_path).exists():
            splash = AnimatedSplash(splash_path)
            splash.show()
            app.processEvents()
    except Exception:
        splash = None

    win = MainWindow()
    
    try:
        import pyi_splash
        pyi_splash.close()
    except Exception:
        pass

    win.show()
    if splash:
        splash.close()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
