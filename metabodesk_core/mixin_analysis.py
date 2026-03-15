"""Core analysis mixin for MetaboDesk.

Implements the nine standard analysis types available from the toolbar:
FBA, pFBA, FVA, Single Gene Deletion (SGD), Double Gene Deletion (DGD),
Single Reaction Deletion (SRD), Robustness Analysis, Production Envelope,
and Flux Sampling (ACHR / OptGP).  Each analysis runs in a background
``QThread`` via :meth:`_launch_worker` to keep the UI responsive.

Also provides result rendering (tables + matplotlib charts), export
(CSV, Excel, LaTeX), and the FBA result cache.
"""

from __future__ import annotations

import sys
import csv
import logging
from typing import Any, TypedDict

import cobra
from datetime import datetime
from itertools import combinations
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem

from metabodesk_core.widgets import AnalysisWorker

logger = logging.getLogger("MetaboDesk")


# ── Result TypedDicts ──────────────────────────────────────────────────

class FVAResultEntry(TypedDict):
    """FVA result for a single reaction."""
    min: float
    max: float


class PFBAResult(TypedDict):
    """Parsimonious FBA result."""
    status: str
    objective: float
    flux: dict[str, float]


class DeletionResult(TypedDict, total=False):
    """Gene / reaction deletion result  — maps id → growth rate."""


class RobustnessResult(TypedDict):
    """Robustness analysis result."""
    values: list[float]
    objectives: list[float]


class ProductionEnvelopeResult(TypedDict):
    """Production envelope result."""
    growth: list[float]
    product: list[float]


class SamplingResultEntry(TypedDict):
    """Flux-sampling statistics for a single reaction."""
    mean: float
    stdev: float
    min: float
    max: float


class FBABaselineResult(TypedDict, total=False):
    """FBA baseline result stored in ``last_run``."""
    status: str
    objective: float
    flux: dict[str, float]


class AnalysisResult(TypedDict, total=False):
    """Top-level analysis result dict stored as ``self.last_run``."""
    timestamp: str
    analysis_type: str
    baseline: FBABaselineResult
    compared: FBABaselineResult | None
    fva: dict[str, FVAResultEntry]
    pfba: PFBAResult
    sgd: dict[str, float]
    dgd: dict[str, float]
    srd: dict[str, float]
    robustness: RobustnessResult
    envelope: ProductionEnvelopeResult
    sampling: dict[str, SamplingResultEntry]

class AnalysisMixin:
    """Mixin providing analysis functionality."""

    def compute_fva(self, model: cobra.Model) -> dict[str, FVAResultEntry]:
        """
        Compute Flux Variability Analysis (FVA).
        Returns a dict with min/max flux for each reaction.
        """
        try:
            from cobra.flux_analysis import flux_variability_analysis
            fva_result = flux_variability_analysis(model, processes=1)
            result = {}
            for rid in fva_result.index:
                result[rid] = {
                    "min": float(fva_result.loc[rid, "minimum"]),
                    "max": float(fva_result.loc[rid, "maximum"]),
                }
            return result
        except Exception as e:
            raise ValueError(f"FVA computation failed: {e}")

    def compute_pfba(self, model: cobra.Model) -> PFBAResult:
        """
        Compute Parsimonious FBA (pFBA).
        First optimizes the objective, then minimizes total flux.
        """
        try:
            from cobra.flux_analysis import pfba
            solution = pfba(model)
            result = {
                "status": str(solution.status),
                "objective": float(solution.objective_value),
                "flux": solution.fluxes.to_dict(),
            }
            return result
        except Exception as e:
            raise ValueError(f"pFBA computation failed: {e}")

    def compute_sgd(self, model: cobra.Model, worker: AnalysisWorker | None = None) -> dict[str, float]:
        """Single Gene Deletion analysis with progress reporting.

        Iterates gene-by-gene instead of COBRApy's bulk function so
        that granular progress can be reported for large models.
        """
        try:
            genes = list(model.genes)
            if not genes:
                return {}
            result = {}
            total = len(genes)
            for i, gene in enumerate(genes):
                if worker and i % max(1, total // 100) == 0:
                    worker.report_progress(
                        f"SGD: gene {i + 1}/{total} ({gene.id})",
                        int((i + 1) / total * 100),
                    )
                with model:
                    gene.knock_out()
                    sol = model.optimize()
                    growth = float(sol.objective_value) if sol.status == "optimal" else 0.0
                result[str(gene.id)] = growth
            return result
        except Exception as e:
            raise ValueError(f"SGD computation failed: {e}")

    def compute_dgd(self, model: cobra.Model, genes: list[str], worker: AnalysisWorker | None = None) -> dict[str, float]:
        """Double Gene Deletion analysis with progress reporting.

        Iterates pair-by-pair instead of COBRApy's bulk function so
        that granular progress can be reported.
        """
        try:
            if not genes or len(genes) < 2:
                return {}
            # Validate gene IDs exist in model
            valid_genes = []
            for gid in genes:
                try:
                    model.genes.get_by_id(gid)
                    valid_genes.append(gid)
                except KeyError:
                    pass
            if len(valid_genes) < 2:
                return {}
            result = {}
            pairs = list(combinations(valid_genes, 2))
            total = len(pairs)
            for i, (gid1, gid2) in enumerate(pairs):
                if worker and i % max(1, total // 100) == 0:
                    worker.report_progress(
                        f"DGD: pair {i + 1}/{total}",
                        int((i + 1) / total * 100),
                    )
                with model:
                    model.genes.get_by_id(gid1).knock_out()
                    model.genes.get_by_id(gid2).knock_out()
                    sol = model.optimize()
                    growth = float(sol.objective_value) if sol.status == "optimal" else 0.0
                result[f"{gid1} + {gid2}"] = growth
            return result
        except Exception as e:
            raise ValueError(f"DGD computation failed: {e}")

    def compute_srd(self, model: cobra.Model, worker: AnalysisWorker | None = None) -> dict[str, float]:
        """Single Reaction Deletion analysis with progress reporting.

        Iterates reaction-by-reaction instead of COBRApy's bulk function
        so that granular progress can be reported for large models.
        """
        try:
            reactions = list(model.reactions)
            if not reactions:
                return {}
            result = {}
            total = len(reactions)
            for i, rxn in enumerate(reactions):
                if worker and i % max(1, total // 100) == 0:
                    worker.report_progress(
                        f"SRD: reaction {i + 1}/{total} ({rxn.id})",
                        int((i + 1) / total * 100),
                    )
                with model:
                    rxn.knock_out()
                    sol = model.optimize()
                    growth = float(sol.objective_value) if sol.status == "optimal" else 0.0
                result[str(rxn.id)] = growth
            return result
        except Exception as e:
            raise ValueError(f"SRD computation failed: {e}")

    def compute_robustness(self, model: cobra.Model, rxn_id: str, min_val: float, max_val: float, steps: int, bound_type: str = "ub", worker: AnalysisWorker | None = None) -> RobustnessResult:
        """Robustness analysis: sweep a reaction's bound and measure objective."""
        result = {"values": [], "objectives": []}
        try:
            if steps < 2:
                raise ValueError("Steps must be >= 2.")
            model.reactions.get_by_id(rxn_id)  # validate rxn exists
            sweep_vals = list(dict.fromkeys([min_val + i * (max_val - min_val) / (steps - 1) for i in range(steps)]))
            for idx, val in enumerate(sweep_vals):
                if worker:
                    pct = int((idx + 1) / len(sweep_vals) * 100)
                    worker.report_progress(f"Robustness: step {idx+1}/{len(sweep_vals)}", pct)
                # Use an isolated per-step model copy for solver stability.
                step_model = model.copy()
                with step_model:
                    rxn_ctx = step_model.reactions.get_by_id(rxn_id)
                    if bound_type.lower() == "ub":
                        # Ensure LB ≤ new UB to avoid infeasible bounds
                        if rxn_ctx.lower_bound > val:
                            rxn_ctx.lower_bound = val
                        rxn_ctx.upper_bound = val
                    else:  # lb
                        # Ensure UB ≥ new LB to avoid infeasible bounds
                        if rxn_ctx.upper_bound < val:
                            rxn_ctx.upper_bound = val
                        rxn_ctx.lower_bound = val
                    
                    sol = step_model.optimize()
                    if str(sol.status) == "optimal" and sol.objective_value is not None:
                        obj = float(sol.objective_value)
                        if obj == obj and obj != float("inf") and obj != float("-inf"):
                            result["values"].append(float(val))
                            result["objectives"].append(obj)
            
            return result
        except Exception as e:
            raise ValueError(f"Robustness computation failed: {e}")

    def compute_production_envelope(self, model: cobra.Model, product_id: str, steps: int) -> ProductionEnvelopeResult:
        """Production envelope: measure trade-off between growth and product formation.
        
        Uses FVA to determine actual max production, then sweeps from 0 to max.
        """
        result = {"growth": [], "product": []}
        try:
            if steps < 2:
                raise ValueError("Steps must be >= 2.")
            if model.objective is None:
                raise ValueError("No objective reaction found")

            product_rxn = model.reactions.get_by_id(product_id)

            # Find actual max production via temporary objective swap
            with model:
                model.objective = product_rxn
                max_sol = model.optimize()
                max_production = float(max_sol.objective_value) if max_sol.status == "optimal" else product_rxn.upper_bound

            if max_production <= 0:
                max_production = product_rxn.upper_bound if product_rxn.upper_bound > 0 else 10.0

            for i in range(steps):
                bound = i * max_production / max(steps - 1, 1)
                with model:
                    prod_rxn = model.reactions.get_by_id(product_id)
                    prod_rxn.lower_bound = bound
                    prod_rxn.upper_bound = bound

                    sol = model.optimize()
                    if str(sol.status) == "optimal":
                        result["product"].append(float(bound))
                        result["growth"].append(float(sol.objective_value))

            return result
        except Exception as e:
            raise ValueError(f"Production envelope computation failed: {e}")

    def compute_flux_sampling(self, model: cobra.Model, sample_size: int = 500) -> tuple[dict[str, SamplingResultEntry], Any]:
        """Flux sampling using ACHR method."""
        try:
            from cobra.sampling import ACHRSampler
            sampler = ACHRSampler(model)
            samples = sampler.sample(sample_size)
            
            result = {}
            for rxn_id in samples.columns:
                flux_vals = samples[rxn_id].values
                result[rxn_id] = {
                    "mean": float(flux_vals.mean()),
                    "stdev": float(flux_vals.std()),
                    "min": float(flux_vals.min()),
                    "max": float(flux_vals.max()),
                }
            return result, samples
        except Exception as e:
            raise ValueError(f"Flux sampling failed: {e}")

    # ---------------- Plotting & Results ----------------

    def _clear_chart(self) -> None:
        """Reset the main results chart to its empty placeholder state."""
        if not hasattr(self, "canvas") or self.canvas is None:
            return
        ax = self.canvas.ax
        if ax is None:
            return
        ax.clear()
        ax.set_title("Run FBA to see flux comparison")
        ax.set_xlabel("Reaction")
        ax.set_ylabel("Flux")
        self.canvas.draw()

    def _plot_flux_comparison(self, rxn_ids, base_flux, compared_flux, base_label: str, compared_label: str, title: str) -> None:
        """Draw a bar/histogram/waterfall chart comparing two flux distributions.

        Delegates to one of three specialised sub-methods based on the
        user's chart-type selection.
        """
        ax = self.canvas.ax
        ax.clear()
        if not rxn_ids:
            ax.set_title("No flux changes above threshold")
            self.canvas.draw()
            return

        chart_type = self.results_chart_type.currentText()

        if chart_type == "Flux histogram":
            self._plot_flux_histogram(ax, base_flux, compared_flux)
        elif chart_type == "Waterfall (top changes)":
            self._plot_flux_waterfall(ax, rxn_ids, base_flux, compared_flux, base_label, compared_label)
        else:
            self._plot_flux_bar_comparison(ax, rxn_ids, base_flux, compared_flux, base_label, compared_label, title)

        self.canvas.figure.tight_layout()
        self.canvas.draw()

    # -- Chart sub-renderers ---------------------------------------------------

    @staticmethod
    def _plot_flux_histogram(ax, base_flux: dict, compared_flux: dict) -> None:
        """Render a histogram of all non-zero flux values on *ax*."""
        all_flux = [f for f in list(base_flux.values()) + list(compared_flux.values()) if abs(f) > 1e-9]
        if all_flux:
            ax.hist(all_flux, bins=50, edgecolor="black", alpha=0.7)
            ax.set_title("Flux Distribution (all non-zero fluxes)")
            ax.set_xlabel("Flux value")
            ax.set_ylabel("Count")
        else:
            ax.set_title("No non-zero fluxes to display")

    @staticmethod
    def _plot_flux_waterfall(ax, rxn_ids: list[str], base_flux: dict, compared_flux: dict,
                             base_label: str, compared_label: str) -> None:
        """Render a horizontal waterfall chart of the top-20 flux deltas on *ax*."""
        deltas = [(rid, compared_flux.get(rid, 0) - base_flux.get(rid, 0)) for rid in rxn_ids]
        deltas.sort(key=lambda x: x[1], reverse=True)
        deltas = deltas[:20]
        rids = [d[0] for d in deltas]
        vals = [d[1] for d in deltas]
        colors = ["green" if v > 0 else "red" for v in vals]
        ax.barh(range(len(rids)), vals, color=colors)
        ax.set_yticks(range(len(rids)))
        ax.set_yticklabels(rids, fontsize=8)
        ax.axvline(x=0, color="black", linewidth=0.5)
        ax.set_title(f"Top flux changes ({base_label} → {compared_label})")
        ax.set_xlabel("Delta flux")
        ax.invert_yaxis()

    @staticmethod
    def _plot_flux_bar_comparison(ax, rxn_ids: list[str], base_flux: dict, compared_flux: dict,
                                  base_label: str, compared_label: str, title: str) -> None:
        """Render a side-by-side bar chart comparing two flux sets on *ax*."""
        bvals = [base_flux.get(r, 0.0) for r in rxn_ids]
        cvals = [compared_flux.get(r, 0.0) for r in rxn_ids]
        x = list(range(len(rxn_ids)))
        width = 0.4
        ax.bar([i - width / 2 for i in x], bvals, width=width, label=base_label)
        ax.bar([i + width / 2 for i in x], cvals, width=width, label=compared_label)
        ax.set_title(title)
        ax.set_xlabel("Reaction")
        ax.set_ylabel("Flux")
        ax.set_xticks(x)
        ax.set_xticklabels(rxn_ids, rotation=90, fontsize=8)
        ax.legend()

    def _refresh_results_view_from_last_run(self, *_) -> None:
        """Re-render results after a compare-mode or top-N change."""
        if not self.last_run:
            return
        self.last_run["compare_mode"] = self.compare_mode.currentText()
        self._recompute_flux_rows_for_compare_mode()
        self._render_results_from_last_run()

    def _set_flux_table_headers_for_mode(self, mode: str) -> None:
        """Set flux table column headers based on the active compare mode."""
        if mode == "Original vs Baseline":
            self.flux_tbl.setHorizontalHeaderLabels(["Reaction", "Original flux", "Baseline flux", "Delta (baseline - original)"])
        else:
            self.flux_tbl.setHorizontalHeaderLabels(["Reaction", "Baseline flux", "Compared flux", "Delta (compared - baseline)"])

    def _recompute_flux_rows_for_compare_mode(self) -> None:
        """Rebuild ``last_run['flux_rows']`` for the current compare mode and top-N."""
        assert self.last_run is not None
        mode = self.last_run.get("compare_mode", "Gene knockout only")

        if mode == "Original vs Baseline":
            base_flux = self.last_run["original"]["flux"]
            cmp_flux = self.last_run["baseline"]["flux"]
        elif mode == "Gene knockout only":
            base_flux = self.last_run["baseline"]["flux"]
            cmp_flux = self.last_run["gene_knockout_only"]["flux"]
        else:
            base_flux = self.last_run["baseline"]["flux"]
            cmp_flux = self.last_run["overexpression_only"]["flux"]

        rows = []
        for rxn_id, b in base_flux.items():
            c = cmp_flux.get(rxn_id, 0.0)
            d = c - b
            if abs(d) > 1e-9:
                rows.append((rxn_id, b, c, d))

        rows.sort(key=lambda t: abs(t[3]), reverse=True)
        rows = rows[: int(self.topn_spin.value())]

        self.last_run["flux_rows"] = [
            {"reaction": rid, "baseline": float(b), "compared": float(c), "delta": float(d)}
            for rid, b, c, d in rows
        ]

    def _render_results_from_last_run(self):
        """Render FBA comparison results: summary label, table, chart."""
        assert self.last_run is not None
        self._ensure_tab_built("Results")
        mode = self.last_run.get("compare_mode", "Gene knockout only")

        original = self.last_run.get("original") or self.last_run["baseline"]
        baseline = self.last_run["baseline"]
        gene_knockout_only = self.last_run["gene_knockout_only"]
        overexpression_only = self.last_run["overexpression_only"]

        self._update_results_summary_label(mode, original, baseline, gene_knockout_only, overexpression_only)
        rxn_ids = self._populate_flux_table(mode)
        base_flux, cmp_flux, base_label, cmp_label, title = self._resolve_compare_fluxes(
            mode, original, baseline, gene_knockout_only, overexpression_only
        )

        self._plot_flux_comparison(rxn_ids, base_flux, cmp_flux, base_label, cmp_label, title)

        # Refresh the All-Fluxes browser if it's open
        if hasattr(self, 'allflux_container') and self.allflux_container.isVisible():
            self._populate_allflux_table()

        self.tabs.setCurrentWidget(self.tab_results)

    # -- FBA result sub-helpers ------------------------------------------------

    def _update_results_summary_label(self, mode, original, baseline, gene_knockout_only, overexpression_only):
        """Set the Results tab summary label with objective values and biomass info."""
        biomass_info = ""
        if self.base_model:
            for rxn in self.base_model.reactions:
                if 'biomass' in rxn.id.lower() or 'biomass' in rxn.name.lower():
                    bid = rxn.id
                    biomass_info = (
                        f"\nBiomass ({bid}):\n"
                        f"  Original: {original['flux'].get(bid, 0.0):.6g}"
                        f" | Baseline: {baseline['flux'].get(bid, 0.0):.6g}"
                        f" | Knockout: {gene_knockout_only['flux'].get(bid, 0.0):.6g}"
                        f" | Overexpression: {overexpression_only['flux'].get(bid, 0.0):.6g}"
                    )
                    break

        self.results_lbl.setText(
            f"Original (no changes): {original['status']} | obj={original['objective']}\n"
            f"Baseline: {baseline['status']} | obj={baseline['objective']}\n"
            f"Gene knockout only: {gene_knockout_only['status']} | obj={gene_knockout_only['objective']}\n"
            f"Overexpression only: {overexpression_only['status']} | obj={overexpression_only['objective']}"
            f"{biomass_info}\n"
            f"\nCompare mode: {mode} | Top N={int(self.topn_spin.value())}\n"
            f"Reaction overrides: {len(self.reaction_bound_overrides)}"
        )

    def _populate_flux_table(self, mode: str) -> list[str]:
        """Fill the flux comparison table and return reaction IDs for plotting."""
        self._set_flux_table_headers_for_mode(mode)
        rows = self.last_run.get("flux_rows", [])
        self.flux_tbl.setRowCount(0)
        rxn_ids: list[str] = []
        for row in rows:
            rid = row["reaction"]
            r = self.flux_tbl.rowCount()
            self.flux_tbl.insertRow(r)
            self.flux_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.flux_tbl.setItem(r, 1, QTableWidgetItem(f"{row['baseline']:.6g}"))
            self.flux_tbl.setItem(r, 2, QTableWidgetItem(f"{row['compared']:.6g}"))
            self.flux_tbl.setItem(r, 3, QTableWidgetItem(f"{row['delta']:.6g}"))
            rxn_ids.append(rid)
        self.filter_flux_table()
        return rxn_ids

    @staticmethod
    def _resolve_compare_fluxes(mode, original, baseline, gene_knockout_only, overexpression_only):
        """Return (base_flux, cmp_flux, base_label, cmp_label, title) for the selected compare mode."""
        if mode == "Original vs Baseline":
            return original["flux"], baseline["flux"], "Original", "Baseline", "Flux comparison (Original vs Baseline)"
        if mode == "Gene knockout only":
            return baseline["flux"], gene_knockout_only["flux"], "Baseline", "Gene knockout only", "Flux comparison (Baseline vs Gene knockout only)"
        return baseline["flux"], overexpression_only["flux"], "Baseline", "Overexpression only", "Flux comparison (Baseline vs Overexpression only)"

    # ── All-Fluxes Browser ────────────────────────────────────

    def _toggle_allflux_panel(self):
        """Show/hide the All Fluxes browser panel."""
        visible = not self.allflux_container.isVisible()
        self.allflux_container.setVisible(visible)
        self.allflux_toggle_btn.setText("▲ Hide" if visible else "▼ Show")
        if visible and self.last_run:
            self._populate_allflux_table()

    def _populate_allflux_table(self):
        """Fill the all-fluxes table from the latest FBA baseline solution."""
        if not self.last_run or not self.base_model:
            return

        baseline = self.last_run.get("baseline", self.last_run)
        flux_dict = baseline.get("flux", {})
        if not flux_dict:
            self.allflux_info_lbl.setText("No flux data available. Run FBA first.")
            return

        # Build subsystem filter list
        subsystems = set()
        for rxn in self.base_model.reactions:
            ss = getattr(rxn, "subsystem", "") or ""
            if ss:
                subsystems.add(ss)
        self.allflux_subsystem_combo.blockSignals(True)
        current_ss = self.allflux_subsystem_combo.currentText()
        self.allflux_subsystem_combo.clear()
        self.allflux_subsystem_combo.addItem("All")
        for ss in sorted(subsystems):
            self.allflux_subsystem_combo.addItem(ss)
        idx = self.allflux_subsystem_combo.findText(current_ss)
        if idx >= 0:
            self.allflux_subsystem_combo.setCurrentIndex(idx)
        self.allflux_subsystem_combo.blockSignals(False)

        # Populate table — pre-allocate all rows for speed
        reactions = list(self.base_model.reactions)
        n = len(reactions)
        self.allflux_tbl.setSortingEnabled(False)
        self.allflux_tbl.blockSignals(True)
        self.allflux_tbl.setRowCount(n)

        nonzero_count = 0
        for r, rxn in enumerate(reactions):
            flux_val = flux_dict.get(rxn.id, 0.0)
            if flux_val != 0:
                nonzero_count += 1

            self.allflux_tbl.setItem(r, 0, QTableWidgetItem(rxn.id))
            self.allflux_tbl.setItem(r, 1, QTableWidgetItem(rxn.name or ""))

            # Use a sortable numeric item
            flux_item = QTableWidgetItem()
            flux_item.setData(Qt.DisplayRole, round(float(flux_val), 9))
            self.allflux_tbl.setItem(r, 2, flux_item)

            self.allflux_tbl.setItem(r, 3, QTableWidgetItem(
                f"[{rxn.lower_bound:.6g}, {rxn.upper_bound:.6g}]"
            ))
            self.allflux_tbl.setItem(r, 4, QTableWidgetItem(
                getattr(rxn, "subsystem", "") or ""
            ))

        self.allflux_tbl.blockSignals(False)
        self.allflux_tbl.setSortingEnabled(True)
        self.allflux_info_lbl.setText(
            f"Total reactions: {self.allflux_tbl.rowCount()}  |  "
            f"Non-zero flux: {nonzero_count}  |  "
            f"Objective: {baseline.get('objective', '?')}"
        )
        self._filter_allflux_table()

    def _filter_allflux_table(self, *_):
        """Filter the all-fluxes table by search text, min |flux|, non-zero, subsystem."""
        text = (self.allflux_search.text() or "").strip().lower()
        min_abs = float(self.allflux_min_abs.value())
        nonzero = self.allflux_nonzero_chk.isChecked()
        subsystem = self.allflux_subsystem_combo.currentText()

        visible_count = 0
        for row in range(self.allflux_tbl.rowCount()):
            rid = (self.allflux_tbl.item(row, 0).text() if self.allflux_tbl.item(row, 0) else "").lower()
            rname = (self.allflux_tbl.item(row, 1).text() if self.allflux_tbl.item(row, 1) else "").lower()
            try:
                flux_val = float(self.allflux_tbl.item(row, 2).data(Qt.DisplayRole))
            except Exception:
                flux_val = 0.0
            rss = (self.allflux_tbl.item(row, 4).text() if self.allflux_tbl.item(row, 4) else "")

            hide = False
            if text and text not in rid and text not in rname:
                hide = True
            if abs(flux_val) < min_abs:
                hide = True
            if nonzero and flux_val == 0.0:
                hide = True
            if subsystem != "All" and rss != subsystem:
                hide = True

            self.allflux_tbl.setRowHidden(row, hide)
            if not hide:
                visible_count += 1

        # Update info label with filter count
        total = self.allflux_tbl.rowCount()
        if visible_count < total:
            current_text = self.allflux_info_lbl.text()
            base = current_text.split("|")[0].strip() if "|" in current_text else current_text
            self.allflux_info_lbl.setText(
                f"Showing {visible_count} / {total} reactions  |  "
                + "  |  ".join(current_text.split("|")[1:]).strip()
                if "|" in current_text else f"Showing {visible_count} / {total} reactions"
            )

    def _export_allflux_csv(self):
        """Export all visible fluxes to CSV."""
        if self.allflux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No flux data to export. Run FBA first.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export All Fluxes (CSV)", "", "CSV files (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                headers = [self.allflux_tbl.horizontalHeaderItem(c).text()
                           for c in range(self.allflux_tbl.columnCount())]
                f.write(",".join(headers) + "\n")
                for row in range(self.allflux_tbl.rowCount()):
                    if self.allflux_tbl.isRowHidden(row):
                        continue
                    cells = []
                    for col in range(self.allflux_tbl.columnCount()):
                        item = self.allflux_tbl.item(row, col)
                        val = str(item.data(Qt.DisplayRole) if item else "")
                        if "," in val:
                            val = f'"{val}"'
                        cells.append(val)
                    f.write(",".join(cells) + "\n")
            QMessageBox.information(self, "Export", f"All fluxes exported to:\n{path}")
        except Exception as e:
            self._show_error("Export Error", f"Could not export CSV: {e}")

    def _export_allflux_excel(self):
        """Export all visible fluxes to Excel."""
        if self.allflux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No flux data to export. Run FBA first.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export All Fluxes (Excel)", "", "Excel files (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "All Fluxes"
            for c in range(self.allflux_tbl.columnCount()):
                ws.cell(row=1, column=c + 1, value=self.allflux_tbl.horizontalHeaderItem(c).text())
            out_row = 2
            for row in range(self.allflux_tbl.rowCount()):
                if self.allflux_tbl.isRowHidden(row):
                    continue
                for col in range(self.allflux_tbl.columnCount()):
                    item = self.allflux_tbl.item(row, col)
                    val = item.data(Qt.DisplayRole) if item else ""
                    ws.cell(row=out_row, column=col + 1, value=val)
                out_row += 1
            wb.save(path)
            QMessageBox.information(self, "Export", f"All fluxes exported to:\n{path}")
        except Exception as e:
            self._show_error("Export Error", f"Could not export Excel: {e}")

    # ---------------- Results Export Functions ----------------

    def _export_results_csv(self) -> None:
        """Export the visible flux comparison table rows to a CSV file."""
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results to CSV", "", "CSV files (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                headers = [self.flux_tbl.horizontalHeaderItem(c).text() for c in range(self.flux_tbl.columnCount())]
                f.write(",".join(headers) + "\n")
                for row in range(self.flux_tbl.rowCount()):
                    if self.flux_tbl.isRowHidden(row):
                        continue
                    vals = []
                    for col in range(self.flux_tbl.columnCount()):
                        item = self.flux_tbl.item(row, col)
                        vals.append(item.text() if item else "")
                    f.write(",".join(vals) + "\n")
            QMessageBox.information(self, "Export", f"Results exported to:\n{path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export CSV results.", e)

    def _export_results_excel(self) -> None:
        """Export the visible flux comparison table rows to an Excel (.xlsx) file."""
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results to Excel", "", "Excel files (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Flux Results"
            # Headers
            for c in range(self.flux_tbl.columnCount()):
                ws.cell(row=1, column=c + 1, value=self.flux_tbl.horizontalHeaderItem(c).text())
            # Data
            row_idx = 2
            for row in range(self.flux_tbl.rowCount()):
                if self.flux_tbl.isRowHidden(row):
                    continue
                for col in range(self.flux_tbl.columnCount()):
                    item = self.flux_tbl.item(row, col)
                    val = item.text() if item else ""
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                    ws.cell(row=row_idx, column=col + 1, value=val)
                row_idx += 1
            wb.save(path)
            QMessageBox.information(self, "Export", f"Results exported to:\n{path}")
        except ImportError:
            self._show_error("Export failed", "openpyxl is required for Excel export.")
        except Exception as e:
            self._show_error("Export failed", "Could not export Excel results.", e)

    def _export_results_chart(self) -> None:
        """Save the current results chart as PNG, SVG, or PDF."""
        path, _ = QFileDialog.getSaveFileName(self, "Export Chart", "", "PNG (*.png);;SVG (*.svg);;PDF (*.pdf)")
        if not path:
            return
        try:
            self.canvas.figure.savefig(path, dpi=150, bbox_inches='tight')
            QMessageBox.information(self, "Export", f"Chart exported to:\n{path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export chart.", e)

    def _export_results_latex(self):
        """Export the current flux table as a LaTeX table for academic publications."""
        if self.flux_tbl.rowCount() == 0:
            QMessageBox.information(self, "Export", "No results to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Results as LaTeX", "", "TeX files (*.tex)")
        if not path:
            return
        try:
            headers = [self.flux_tbl.horizontalHeaderItem(c).text() for c in range(self.flux_tbl.columnCount())]
            ncols = len(headers)
            col_spec = "l" + "r" * (ncols - 1)

            lines = []
            lines.append("\\begin{table}[htbp]")
            lines.append("\\centering")

            # Caption with analysis type
            analysis_type = self.last_run.get("analysis_type", "Flux") if self.last_run else "Flux"
            lines.append(f"\\caption{{{analysis_type} Analysis Results}}")
            lines.append(f"\\label{{tab:{analysis_type.lower().replace(' ', '_')}_results}}")
            lines.append(f"\\begin{{tabular}}{{{col_spec}}}")
            lines.append("\\hline")
            lines.append(" & ".join(f"\\textbf{{{h}}}" for h in headers) + " \\\\")
            lines.append("\\hline")

            for row in range(self.flux_tbl.rowCount()):
                if self.flux_tbl.isRowHidden(row):
                    continue
                vals = []
                for col in range(ncols):
                    item = self.flux_tbl.item(row, col)
                    text = item.text() if item else ""
                    # Escape LaTeX special characters
                    text = text.replace("_", "\\_").replace("&", "\\&").replace("%", "\\%")
                    vals.append(text)
                lines.append(" & ".join(vals) + " \\\\")

            lines.append("\\hline")
            lines.append("\\end{tabular}")
            lines.append("\\end{table}")

            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            QMessageBox.information(self, "Export", f"LaTeX table exported to:\n{path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export LaTeX table.", e)

    def _launch_worker(self, func, on_done, message="Running analysis..."):
        """Launch an analysis function in a background QThread.

        The compute function *func* can accept an optional ``worker`` keyword
        argument.  If it does, the AnalysisWorker instance is passed in so the
        function can call ``worker.report_progress(msg, pct)`` during
        long-running loops.
        """
        self.set_busy(True, message)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)

        def _on_finished(result):
            self._worker = None
            self.progress_bar.setVisible(False)
            try:
                on_done(result)
            except Exception as e:
                self._show_error("Render Error", "Could not render analysis results.", e)
            self.set_busy(False, "Ready.")

        def _on_error(err_msg):
            self._worker = None
            self.progress_bar.setVisible(False)
            self.set_busy(False, "Analysis failed.")
            self._show_error("Analysis Error", str(err_msg))

        def _on_progress(msg):
            self.status_lbl.setText(msg)
            self.statusBar().showMessage(msg)

        def _on_progress_pct(pct):
            self.progress_bar.setValue(min(pct, 100))

        worker = AnalysisWorker(func)
        worker.finished.connect(_on_finished)
        worker.error.connect(_on_error)
        worker.progress.connect(_on_progress)
        worker.progress_pct.connect(_on_progress_pct)
        self._worker = worker
        worker.start()

    # Dispatch table: analysis combo-box text → runner method name.
    # Using a class-level dict avoids rebuilding on every call and makes
    # it trivial to register new analysis types.
    _ANALYSIS_DISPATCH: dict[str, str] = {
        "FVA": "run_fva",
        "pFBA": "run_pfba",
        "Single Gene Deletion (SGD)": "run_sgd",
        "Double Gene Deletion (DGD)": "run_dgd",
        "Single Reaction Deletion (SRD)": "run_srd",
        "Robustness Analysis": "run_robustness",
        "Production Envelope": "run_production_envelope",
        "Flux Sampling": "run_flux_sampling",
    }

    def run_fba(self):
        if self.base_model is None or self.is_running:
            return

        analysis_type = self.analysis_type.currentText()

        # Look up the runner via the dispatch table; fall back to
        # substring matching for forward-compatibility, then default
        # to standard FBA.
        runner_name = self._ANALYSIS_DISPATCH.get(analysis_type)
        if runner_name is None:
            for key, name in self._ANALYSIS_DISPATCH.items():
                if key in analysis_type:
                    runner_name = name
                    break

        if runner_name is not None:
            getattr(self, runner_name)()
        else:
            self.run_standard_fba()

    def run_standard_fba(self):
        """Run standard FBA with baseline, knockout, and overexpression comparisons."""
        # Prepare all models on main thread (thread-safe copies)
        original_model = (self.original_model_snapshot.copy() if self.original_model_snapshot is not None else self.base_model.copy())
        self._apply_selected_objective_to_model(original_model)
        self._apply_selected_solver(original_model)

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        knockout_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(knockout_model)
        self.apply_reaction_overrides_to_model(knockout_model)
        self.apply_knockouts_to_model(knockout_model)
        self._apply_selected_objective_to_model(knockout_model)
        self._apply_selected_solver(knockout_model)

        overexpression_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(overexpression_model)
        self.apply_reaction_overrides_to_model(overexpression_model)
        self.apply_temp_upper_bound_overrides(overexpression_model)
        self.apply_overexpression_to_model(overexpression_model)
        self._apply_selected_objective_to_model(overexpression_model)
        self._apply_selected_solver(overexpression_model)

        use_loopless = self.loopless_chk.isChecked()
        compare_mode = self.compare_mode.currentText()

        # --- FBA cache check ---
        cache_key = self._model_state_hash(baseline_model) + (":loopless" if use_loopless else "")
        cached = self._cache_get(cache_key)
        if cached is not None:
            logger.info("FBA cache hit – reusing previous result")
            self.last_run = dict(cached)
            self.last_run["compare_mode"] = compare_mode
            self.last_run["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "  (cached)"
            self._recompute_flux_rows_for_compare_mode()
            self._render_results_from_last_run()
            return

        _cache_key_ref = cache_key  # closure capture

        def _compute():
            if use_loopless:
                from cobra.flux_analysis import loopless_solution
                original_sol = loopless_solution(original_model)
                baseline_sol = loopless_solution(baseline_model)
                knockout_sol = loopless_solution(knockout_model)
                overexpression_sol = loopless_solution(overexpression_model)
            else:
                original_sol = original_model.optimize()
                baseline_sol = baseline_model.optimize()
                knockout_sol = knockout_model.optimize()
                overexpression_sol = overexpression_model.optimize()

            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "FBA",
                "original": {"status": str(original_sol.status), "objective": float(original_sol.objective_value), "flux": original_sol.fluxes.to_dict(),
                              "shadow_prices": original_sol.shadow_prices.to_dict() if hasattr(original_sol, 'shadow_prices') and original_sol.shadow_prices is not None else {},
                              "reduced_costs": original_sol.reduced_costs.to_dict() if hasattr(original_sol, 'reduced_costs') and original_sol.reduced_costs is not None else {}},
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict(),
                              "shadow_prices": baseline_sol.shadow_prices.to_dict() if hasattr(baseline_sol, 'shadow_prices') and baseline_sol.shadow_prices is not None else {},
                              "reduced_costs": baseline_sol.reduced_costs.to_dict() if hasattr(baseline_sol, 'reduced_costs') and baseline_sol.reduced_costs is not None else {}},
                "gene_knockout_only": {"status": str(knockout_sol.status), "objective": float(knockout_sol.objective_value), "flux": knockout_sol.fluxes.to_dict()},
                "overexpression_only": {"status": str(overexpression_sol.status), "objective": float(overexpression_sol.objective_value), "flux": overexpression_sol.fluxes.to_dict()},
                "compare_mode": compare_mode,
                "flux_rows": [],
            }

        def _on_done(result):
            self._cache_put(_cache_key_ref, result)
            self.last_run = result
            self._recompute_flux_rows_for_compare_mode()
            self._render_results_from_last_run()

        label = "Running FBA (Loopless)..." if use_loopless else "Running FBA..."
        self._launch_worker(_compute, _on_done, label)

    def run_fva(self):
        """Run Flux Variability Analysis (FVA)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)
        topn = int(self.topn_spin.value())
        fva_fraction = float(self.fva_fraction_spin.value())
        fva_processes = int(self.fva_processes_spin.value())
        # Frozen exe (PyInstaller) cannot use multiprocessing safely
        if getattr(sys, 'frozen', False) and fva_processes > 1:
            fva_processes = 1

        def _compute(worker=None):
            from cobra.flux_analysis import flux_variability_analysis
            import pandas as pd
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            # Chunk reactions for progress reporting on large models
            all_rxns = list(baseline_model.reactions)
            chunk_size = max(50, len(all_rxns) // 20)
            chunks = [all_rxns[i:i + chunk_size] for i in range(0, len(all_rxns), chunk_size)]
            fva_parts = []
            for cidx, chunk in enumerate(chunks):
                if worker:
                    worker.report_progress(
                        f"FVA: chunk {cidx + 1}/{len(chunks)}",
                        int((cidx + 1) / len(chunks) * 90),
                    )
                part = flux_variability_analysis(
                    baseline_model,
                    reaction_list=chunk,
                    fraction_of_optimum=fva_fraction,
                    processes=fva_processes,
                )
                fva_parts.append(part)
            fva_result = pd.concat(fva_parts)
            fva_dict = {}
            for rid in fva_result.index:
                fva_dict[rid] = {"min": float(fva_result.loc[rid, "minimum"]), "max": float(fva_result.loc[rid, "maximum"])}
            rows = []
            for rxn_id, minmax in fva_dict.items():
                flux_range = abs(minmax["max"] - minmax["min"])
                if flux_range > 1e-9:
                    rows.append({"reaction": rxn_id, "min_flux": minmax["min"], "max_flux": minmax["max"], "range": flux_range})
            rows.sort(key=lambda x: x["range"], reverse=True)
            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "FVA",
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict()},
                "fva_result": fva_dict,
                "flux_rows": rows[:topn],
            }

        def _on_done(result):
            self.last_run = result
            self._render_fva_results()

        self._launch_worker(_compute, _on_done, "Running FVA (Flux Variability Analysis)...")

    def run_pfba(self):
        """Run Parsimonious FBA (pFBA)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        baseline_model_std = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model_std)
        self.apply_reaction_overrides_to_model(baseline_model_std)
        self._apply_selected_objective_to_model(baseline_model_std)
        self._apply_selected_solver(baseline_model_std)
        topn = int(self.topn_spin.value())

        def _compute():
            from cobra.flux_analysis import pfba
            pfba_sol = pfba(baseline_model)
            pfba_result = {"status": str(pfba_sol.status), "objective": float(pfba_sol.objective_value), "flux": pfba_sol.fluxes.to_dict()}
            baseline_sol = baseline_model_std.optimize()
            rows = []
            for rxn_id, bf in baseline_sol.fluxes.items():
                pf = pfba_result["flux"].get(rxn_id, 0.0)
                d = abs(bf) - abs(pf)
                if abs(d) > 1e-9:
                    rows.append({"reaction": rxn_id, "fba_flux": float(bf), "pfba_flux": float(pf), "delta": float(d)})
            rows.sort(key=lambda x: abs(x["delta"]), reverse=True)
            return {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "analysis_type": "pFBA",
                "baseline": {"status": str(baseline_sol.status), "objective": float(baseline_sol.objective_value), "flux": baseline_sol.fluxes.to_dict()},
                "pfba": pfba_result,
                "compare_mode": "pFBA vs FBA",
                "flux_rows": rows[:topn],
            }

        def _on_done(result):
            self.last_run = result
            self._render_pfba_results()

        self._launch_worker(_compute, _on_done, "Running pFBA (Parsimonious FBA)...")

    def run_sgd(self):
        """Run Single Gene Deletion analysis."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        def _compute(worker=None):
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            sgd_result = self.compute_sgd(baseline_model, worker=worker)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "SGD", "wt_growth": wt_growth, "sgd_result": sgd_result}

        def _on_done(result):
            self.last_run = result
            self._render_sgd_results()

        self._launch_worker(_compute, _on_done, "Running Single Gene Deletion (SGD)...")

    def _get_selected_genes_for_dgd(self) -> list[str]:
        """Collect selected gene IDs from the gene list for double-gene deletion.

        Falls back to the current knockout gene set if fewer than 2 genes
        are selected in the UI.
        """
        genes: list[str] = []

        try:
            items = list(self.gene_list.selectedItems())
            genes.extend([str(i.text()).strip() for i in items if i and str(i.text()).strip()])

            if not genes:
                sel_model = self.gene_list.selectionModel()
                if sel_model is not None:
                    for idx in sel_model.selectedIndexes():
                        it = self.gene_list.item(idx.row())
                        if it:
                            t = str(it.text()).strip()
                            if t:
                                genes.append(t)

            cur = self.gene_list.currentItem()
            if cur:
                t = str(cur.text()).strip()
                if t:
                    genes.append(t)
        except Exception:
            pass

        if len(genes) < 2 and self.knockout_genes:
            genes.extend(sorted(self.knockout_genes))

        out: list[str] = []
        seen = set()
        for g in genes:
            if g and g not in seen:
                seen.add(g)
                out.append(g)

        return out

    def run_dgd(self):
        """Run Double Gene Deletion analysis."""
        genes = self._get_selected_genes_for_dgd()
        if len(genes) < 2:
            QMessageBox.warning(self, "DGD", "Select at least 2 genes in Gene knockout tab for Double Gene Deletion.")
            return

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)
        genes_copy = list(genes)

        def _compute(worker=None):
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            dgd_result = self.compute_dgd(baseline_model, genes_copy, worker=worker)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "DGD", "wt_growth": wt_growth, "dgd_result": dgd_result, "dgd_genes": genes_copy}

        def _on_done(result):
            self.last_run = result
            self._render_dgd_results()

        self._launch_worker(_compute, _on_done, "Running Double Gene Deletion (DGD)...")

    def run_srd(self):
        """Run Single Reaction Deletion analysis."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        def _compute(worker=None):
            baseline_sol = baseline_model.optimize()
            if str(baseline_sol.status) != "optimal":
                raise ValueError(f"Model optimization failed: {baseline_sol.status}")
            wt_growth = float(baseline_sol.objective_value)
            srd_result = self.compute_srd(baseline_model, worker=worker)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "SRD", "wt_growth": wt_growth, "srd_result": srd_result}

        def _on_done(result):
            self.last_run = result
            self._render_srd_results()

        self._launch_worker(_compute, _on_done, "Running Single Reaction Deletion (SRD)...")

    def run_robustness(self):
        """Run Robustness Analysis."""
        rxn_id = self._extract_reaction_id_from_input(self.robustness_rxn.text())
        if not rxn_id:
            QMessageBox.warning(self, "Missing", "Enter a reaction ID for robustness analysis.")
            return
        sweep_ub = self.robustness_bound_ub.isChecked()
        sweep_lb = self.robustness_bound_lb.isChecked()
        if not sweep_ub and not sweep_lb:
            QMessageBox.warning(self, "Missing", "Select at least one bound type (UB or LB).")
            return
        bound_type = "ub" if sweep_ub else "lb"
        min_val = float(self.robustness_min.value())
        max_val = float(self.robustness_max.value())
        steps = int(self.robustness_steps.value())

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)

        def _compute(worker=None):
            robustness_result = self.compute_robustness(baseline_model, rxn_id, min_val, max_val, steps, bound_type, worker=worker)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "Robustness", "rxn_id": rxn_id, "bound_type": bound_type, "robustness_result": robustness_result}

        def _on_done(result):
            self.last_run = result
            self._render_robustness_results()

        self._launch_worker(_compute, _on_done, f"Running Robustness Analysis for {rxn_id} ({bound_type.upper()})...")

    def run_production_envelope(self):
        """Run Production Envelope Analysis using COBRApy's native function."""
        product_id = self._extract_reaction_id_from_input(self.envelope_product.text())
        if not product_id:
            QMessageBox.warning(self, "Missing", "Enter a product reaction ID for production envelope.")
            return

        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        self._apply_selected_solver(baseline_model)
        steps = int(self.envelope_steps.value())

        def _compute():
            try:
                from cobra.flux_analysis import production_envelope as pe_func
                prod_rxn = baseline_model.reactions.get_by_id(product_id)
                pe_df = pe_func(baseline_model, reactions=[prod_rxn], points=steps)

                # COBRApy production_envelope returns a DataFrame with columns:
                #   carbon_source | flux_minimum | carbon_yield_min | mass_yield_min
                #   flux_maximum  | carbon_yield_max | mass_yield_max  | <reaction_id>
                # X axis = product flux (last column, named after the reaction)
                # Y axis = flux_minimum / flux_maximum (growth rate bounds)

                # Product column: named after the reaction ID (always last)
                prod_col = None
                for col in pe_df.columns:
                    if product_id in col:
                        prod_col = col
                        break
                if prod_col is None:
                    prod_col = pe_df.columns[-1]  # fallback: last column

                product_vals = pe_df[prod_col].tolist()

                # Growth rate columns
                growth_max = pe_df["flux_maximum"].tolist() if "flux_maximum" in pe_df.columns else []
                growth_min = pe_df["flux_minimum"].tolist() if "flux_minimum" in pe_df.columns else []

                # Replace NaN with 0
                import math
                growth_max = [0.0 if (isinstance(v, float) and math.isnan(v)) else float(v) for v in growth_max]
                growth_min = [0.0 if (isinstance(v, float) and math.isnan(v)) else float(v) for v in growth_min]
                product_vals = [0.0 if (isinstance(v, float) and math.isnan(v)) else float(v) for v in product_vals]

                result = {"product": product_vals, "growth_max": growth_max, "growth_min": growth_min,
                          "growth": growth_max}  # "growth" kept for backward compat
            except Exception:
                # Fallback to manual sweep if cobra.flux_analysis.production_envelope is unavailable
                result = self.compute_production_envelope(baseline_model, product_id, steps)
            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "analysis_type": "Production Envelope", "product_id": product_id, "envelope_result": result}

        def _on_done(result):
            self.last_run = result
            self._render_envelope_results()

        self._launch_worker(_compute, _on_done, f"Running Production Envelope for {product_id}...")

    def run_flux_sampling(self):
        """Run Flux Sampling (ACHR or OptGP)."""
        baseline_model = self.base_model.copy()
        self.apply_medium_table_bounds_to_model(baseline_model)
        self.apply_reaction_overrides_to_model(baseline_model)
        self._apply_selected_objective_to_model(baseline_model)
        sample_size = int(self.sampling_size.value())
        sampler_type = self.sampler_type_combo.currentText()

        def _compute():
            if sampler_type == "OptGP":
                try:
                    from cobra.sampling import OptGPSampler
                    sampler = OptGPSampler(baseline_model, processes=1)
                except ImportError:
                    from cobra.sampling import ACHRSampler
                    sampler = ACHRSampler(baseline_model)
            else:
                from cobra.sampling import ACHRSampler
                sampler = ACHRSampler(baseline_model)
            samples = sampler.sample(sample_size)
            result = {}
            for rxn_id in samples.columns:
                fv = samples[rxn_id].values
                result[rxn_id] = {"mean": float(fv.mean()), "stdev": float(fv.std()), "min": float(fv.min()), "max": float(fv.max())}

            # Pre-compute diagnostics in worker thread (avoid main-thread stall)
            import numpy as np
            sorted_rxns = sorted(result.items(), key=lambda x: abs(x[1]["mean"]), reverse=True)
            top_rxn_ids = [rid for rid, _ in sorted_rxns[:20]]

            # Geweke convergence Z-scores
            geweke_scores = {}
            for rid in top_rxn_ids:
                if rid in samples.columns:
                    vals = samples[rid].values
                    n = len(vals)
                    if n < 20:
                        continue
                    first_10 = vals[:n // 10]
                    last_50 = vals[n // 2:]
                    mean_a, mean_b = first_10.mean(), last_50.mean()
                    var_a, var_b = first_10.var(), last_50.var()
                    denom = np.sqrt(var_a / len(first_10) + var_b / len(last_50))
                    z = float((mean_a - mean_b) / denom) if denom > 1e-15 else 0.0
                    geweke_scores[rid] = z

            # Pairwise correlation matrix (top 10)
            corr_rxns = [r for r in top_rxn_ids[:10] if r in samples.columns]
            corr_matrix = None
            if len(corr_rxns) >= 2:
                corr_matrix = samples[corr_rxns].corr().values.tolist()

            # Pre-compute KDE histogram bins in worker (avoid main-thread numpy)
            kde_histograms: list[dict] = []
            top_kde_ids = top_rxn_ids[:5]
            for rid in top_kde_ids:
                if rid in samples.columns:
                    vals = samples[rid].values
                    counts, edges = np.histogram(vals, bins=40, density=True)
                    kde_histograms.append({
                        "rid": rid,
                        "counts": counts.tolist(),
                        "edges": edges.tolist(),
                    })

            return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "analysis_type": "Flux Sampling", "sampler_type": sampler_type,
                    "sampling_result": result, "samples_df": samples,
                    "geweke_scores": geweke_scores,
                    "corr_matrix": corr_matrix, "corr_rxns": corr_rxns,
                    "kde_histograms": kde_histograms}

        def _on_done(result):
            self.last_run = result
            self._render_sampling_results()

        self._launch_worker(_compute, _on_done, f"Running Flux Sampling ({sampler_type})...")

    def filter_fva_table(self):
        """Filter FVA table by reaction ID."""
        text = (self.fva_flux_search.text() or "").strip().lower()
        for row in range(self.fva_tbl.rowCount()):
            rid = (self.fva_tbl.item(row, 0).text() if self.fva_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.fva_tbl.setRowHidden(row, hide)

    def _render_fva_results(self):
        """Render FVA results in the FVA tab."""
        assert self.last_run is not None
        self._ensure_tab_built("FVA")
        baseline = self.last_run.get("baseline", {})

        self.fva_info_lbl.setText(
            f"Baseline: {baseline.get('status', 'N/A')} | Objective: {baseline.get('objective', 'N/A'):.6g}\n"
            f"Fraction of optimum: {self.fva_fraction_spin.value():.2f} | "
            f"Processes: {self.fva_processes_spin.value()} | "
            f"Top N reactions by flux range: {int(self.topn_spin.value())}"
        )
        
        # Populate flux table
        rows = self.last_run.get("flux_rows", [])
        self.fva_tbl.setRowCount(0)
        rxn_ids_for_plot = []
        
        for row in rows:
            rid = row["reaction"]
            min_f = row["min_flux"]
            max_f = row["max_flux"]
            flux_range = row["range"]
            
            r = self.fva_tbl.rowCount()
            self.fva_tbl.insertRow(r)
            self.fva_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.fva_tbl.setItem(r, 1, QTableWidgetItem(f"{min_f:.6g}"))
            self.fva_tbl.setItem(r, 2, QTableWidgetItem(f"{max_f:.6g}"))
            self.fva_tbl.setItem(r, 3, QTableWidgetItem(f"{flux_range:.6g}"))
            rxn_ids_for_plot.append(rid)
        
        # Apply filters
        self.filter_fva_table()
        
        # Plot FVA ranges
        ax = self.fva_canvas.ax
        ax.clear()
        if not rxn_ids_for_plot:
            ax.set_title("No reactions with flux variability found")
            self.fva_canvas.draw()
        else:
            # Plot min/max ranges for each reaction
            x = list(range(len(rxn_ids_for_plot)))
            mins = [self.last_run["fva_result"].get(rid, {}).get("min", 0.0) for rid in rxn_ids_for_plot]
            maxs = [self.last_run["fva_result"].get(rid, {}).get("max", 0.0) for rid in rxn_ids_for_plot]
            
            ax.barh(x, [m - n for m, n in zip(maxs, mins)], left=mins, label="Flux range", color="steelblue", alpha=0.7)
            ax.axvline(x=0, color="black", linestyle="-", linewidth=0.5)
            ax.set_title("FVA: Flux Variability Ranges")
            ax.set_ylabel("Reaction")
            ax.set_xlabel("Flux")
            ax.set_yticks(x)
            ax.set_yticklabels(rxn_ids_for_plot, fontsize=8)
            ax.legend()
            self.fva_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_fva)

    def export_fva_csv(self):
        if not self.last_run or not self.last_run.get("fva_result"):
            QMessageBox.warning(self, "FVA", "Run FVA first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export FVA (CSV)", str(Path.home() / "fva_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction", "min", "max", "range"]) 
                for rid, mm in self.last_run["fva_result"].items():
                    rng = float(mm["max"]) - float(mm["min"]) 
                    w.writerow([rid, mm["min"], mm["max"], rng])
            self.statusBar().showMessage(f"FVA exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export FVA results to CSV.", e)

    def _render_pfba_results(self):
        """Render pFBA results in the pFBA tab."""
        assert self.last_run is not None
        self._ensure_tab_built("pFBA")
        baseline = self.last_run.get("baseline", {})
        pfba = self.last_run.get("pfba", {})
        
        self.pfba_info_lbl.setText(
            f"pFBA (Parsimonious FBA) Results\n"
            f"Standard FBA: {baseline.get('status', 'N/A')} | obj={baseline.get('objective', 'N/A')}\n"
            f"pFBA: {pfba.get('status', 'N/A')} | obj={pfba.get('objective', 'N/A')}\n"
            f"Top N reactions by flux difference: {int(self.topn_spin.value())}"
        )
        
        # Populate pFBA table
        rows = self.last_run.get("flux_rows", [])
        self.pfba_tbl.setHorizontalHeaderLabels(["Reaction", "FBA flux", "pFBA flux", "Difference"])
        self.pfba_tbl.setRowCount(0)
        rxn_ids_for_plot = []
        
        for row in rows:
            rid = row["reaction"]
            fba_f = row["fba_flux"]
            pfba_f = row["pfba_flux"]
            diff = row["delta"]
            
            r = self.pfba_tbl.rowCount()
            self.pfba_tbl.insertRow(r)
            self.pfba_tbl.setItem(r, 0, QTableWidgetItem(rid))
            self.pfba_tbl.setItem(r, 1, QTableWidgetItem(f"{fba_f:.6g}"))
            self.pfba_tbl.setItem(r, 2, QTableWidgetItem(f"{pfba_f:.6g}"))
            self.pfba_tbl.setItem(r, 3, QTableWidgetItem(f"{diff:.6g}"))
            rxn_ids_for_plot.append(rid)
        
        # Apply filters
        self.filter_pfba_table()
        
        # Plot pFBA comparison
        ax = self.pfba_canvas.ax
        ax.clear()
        if not rxn_ids_for_plot:
            ax.set_title("FBA and pFBA results are identical")
        else:
            baseline_flux = self.last_run["baseline"]["flux"]
            pfba_flux = pfba["flux"]
            
            fba_vals = [abs(baseline_flux.get(r, 0.0)) for r in rxn_ids_for_plot]
            pfba_vals = [abs(pfba_flux.get(r, 0.0)) for r in rxn_ids_for_plot]
            
            x = list(range(len(rxn_ids_for_plot)))
            width = 0.4
            ax.bar([i - width / 2 for i in x], fba_vals, width=width, label="FBA", alpha=0.7)
            ax.bar([i + width / 2 for i in x], pfba_vals, width=width, label="pFBA", alpha=0.7)
            
            ax.set_title("pFBA vs Standard FBA: Flux Comparison")
            ax.set_xlabel("Reaction")
            ax.set_ylabel("Absolute Flux")
            ax.set_xticks(x)
            ax.set_xticklabels(rxn_ids_for_plot, rotation=90, fontsize=8)
            ax.legend()
        self.pfba_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_pfba)

    def export_pfba_csv(self):
        if not self.last_run or not self.last_run.get("pfba"):
            QMessageBox.warning(self, "pFBA", "Run pFBA first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export pFBA (CSV)", str(Path.home() / "pfba_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            rows = self.last_run.get("flux_rows", [])
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction", "fba_flux", "pfba_flux", "difference"]) 
                for r in rows:
                    w.writerow([r["reaction"], r["fba_flux"], r["pfba_flux"], r["delta"]])
            self.statusBar().showMessage(f"pFBA exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export pFBA results to CSV.", e)

    def filter_deletion_table(self):
        """Filter deletion analysis table with search text and optimal/non-optimal filters."""
        text = (self.deletion_search.text() or "").strip().lower()
        show_optimal_only = self.deletion_filter_optimal.isChecked()
        show_nonoptimal_only = self.deletion_filter_nonoptimal.isChecked()
        for row in range(self.deletion_tbl.rowCount()):
            rid = (self.deletion_tbl.item(row, 0).text() if self.deletion_tbl.item(row, 0) else "").lower()
            growth_text = (self.deletion_tbl.item(row, 2).text() if self.deletion_tbl.item(row, 2) else "0")
            try:
                growth_val = float(growth_text)
            except ValueError:
                growth_val = 0.0
            hide = False
            if text and text not in rid:
                hide = True
            if show_optimal_only and growth_val < 1e-6:
                hide = True
            if show_nonoptimal_only and growth_val >= 1e-6:
                hide = True
            self.deletion_tbl.setRowHidden(row, hide)

    def _render_sgd_results(self):
        """Render SGD results."""
        assert self.last_run is not None
        self._ensure_tab_built("Deletion Analysis")
        wt_growth = self.last_run.get("wt_growth", 0.0)
        sgd_result = self.last_run.get("sgd_result", {})
        
        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Total: {len(sgd_result)} genes"
        )
        
        # Populate table with ALL results
        self.deletion_tbl.setHorizontalHeaderLabels(["Gene Name", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        genes_for_plot = []
        growth_vals = []
        
        sorted_genes = sorted(sgd_result.items(), key=lambda x: x[1])
        
        for gene_id, growth in sorted_genes:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            gene_label = self._format_gene_label(str(gene_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(gene_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))
        
        # Chart uses top N only
        for gene_id, growth in sorted_genes[:int(self.topn_spin.value())]:
            gene_label = self._format_gene_label(str(gene_id))
            genes_for_plot.append(gene_label)
            growth_vals.append(growth)
        
        self.filter_deletion_table()
        
        # Plot with value labels
        ax = self.deletion_canvas.ax
        ax.clear()
        if genes_for_plot:
            bars = ax.barh(genes_for_plot, growth_vals, color="coral", alpha=0.7, edgecolor="darkred", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Single Gene Deletion - Growth on Deletion")
        self.deletion_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_sgd_csv(self):
        if not self.last_run or not self.last_run.get("sgd_result"):
            QMessageBox.warning(self, "SGD", "Run SGD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export SGD (CSV)", str(Path.home() / "sgd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["gene_name", "gene_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for gid, growth in sorted(self.last_run["sgd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    gname = self._format_gene_label(str(gid))
                    w.writerow([gname, gid, wt, growth, delta_pct])
            self.statusBar().showMessage(f"SGD exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export SGD results to CSV.", e)

    def _render_dgd_results(self):
        """Render DGD results."""
        assert self.last_run is not None
        self._ensure_tab_built("Deletion Analysis")
        wt_growth = self.last_run.get("wt_growth", 0.0)
        dgd_result = self.last_run.get("dgd_result", {})

        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Total: {len(dgd_result)} gene pairs"
        )

        self.deletion_tbl.setHorizontalHeaderLabels(["Gene Pair (Names)", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        pairs_for_plot = []
        growth_vals = []

        sorted_pairs = sorted(dgd_result.items(), key=lambda x: x[1])

        for pair_id, growth in sorted_pairs:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            pair_label = self._format_gene_pair_label(str(pair_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(pair_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))

        # Chart uses top N only
        for pair_id, growth in sorted_pairs[:int(self.topn_spin.value())]:
            pair_label = self._format_gene_pair_label(str(pair_id))
            pairs_for_plot.append(pair_label)
            growth_vals.append(growth)

        self.filter_deletion_table()

        ax = self.deletion_canvas.ax
        ax.clear()
        if pairs_for_plot:
            bars = ax.barh(pairs_for_plot, growth_vals, color="#8ecae6", alpha=0.7, edgecolor="#219ebc", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Double Gene Deletion - Growth on Deletion")
        self.deletion_canvas.draw()

        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_dgd_csv(self):
        if not self.last_run or not self.last_run.get("dgd_result"):
            QMessageBox.warning(self, "DGD", "Run DGD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export DGD (CSV)", str(Path.home() / "dgd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["gene_pair_name", "gene_pair_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for pair, growth in sorted(self.last_run["dgd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    pname = self._format_gene_pair_label(str(pair))
                    w.writerow([pname, pair, wt, growth, delta_pct])
            self.statusBar().showMessage(f"DGD exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export DGD results to CSV.", e)

    def _render_srd_results(self):
        """Render SRD results."""
        assert self.last_run is not None
        self._ensure_tab_built("Deletion Analysis")
        wt_growth = self.last_run.get("wt_growth", 0.0)
        srd_result = self.last_run.get("srd_result", {})
        
        self.deletion_info_lbl.setText(
            f"WT Growth: {wt_growth:.4f}\n"
            f"Total: {len(srd_result)} reactions"
        )
        
        # Populate table with ALL results
        self.deletion_tbl.setHorizontalHeaderLabels(["Reaction Name", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
        self.deletion_tbl.setRowCount(0)
        rxns_for_plot = []
        growth_vals = []
        
        sorted_rxns = sorted(srd_result.items(), key=lambda x: x[1])
        
        for rxn_id, growth in sorted_rxns:
            r = self.deletion_tbl.rowCount()
            self.deletion_tbl.insertRow(r)
            rxn_label = self._format_rxn_label(str(rxn_id))
            self.deletion_tbl.setItem(r, 0, QTableWidgetItem(rxn_label))
            self.deletion_tbl.setItem(r, 1, QTableWidgetItem(f"{wt_growth:.4f}"))
            self.deletion_tbl.setItem(r, 2, QTableWidgetItem(f"{growth:.4f}"))
            delta_pct = (0.0 if wt_growth == 0 else ((growth - wt_growth) / wt_growth) * 100.0)
            self.deletion_tbl.setItem(r, 3, QTableWidgetItem(f"{delta_pct:.2f}"))
            if growth < 1e-6:
                cat = "Lethal"
            elif delta_pct < -75:
                cat = "Severe"
            elif delta_pct < -25:
                cat = "Moderate"
            else:
                cat = "Mild"
            self.deletion_tbl.setItem(r, 4, QTableWidgetItem(cat))
        
        # Chart uses top N only
        for rxn_id, growth in sorted_rxns[:int(self.topn_spin.value())]:
            rxn_label = self._format_rxn_label(str(rxn_id))
            rxns_for_plot.append(rxn_label)
            growth_vals.append(growth)
        
        self.filter_deletion_table()
        
        # Plot with value labels
        ax = self.deletion_canvas.ax
        ax.clear()
        if rxns_for_plot:
            bars = ax.barh(rxns_for_plot, growth_vals, color="lightgreen", alpha=0.7, edgecolor="darkgreen", linewidth=0.8)
            for bar, val in zip(bars, growth_vals):
                ax.text(val, bar.get_y() + bar.get_height()/2, f"{val:.3g}", fontsize=7, va="center", ha="left", style="italic")
            ax.set_xlabel("Growth Rate")
            ax.set_title("Single Reaction Deletion - Growth on Deletion")
        self.deletion_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_deletion)

    def export_srd_csv(self):
        if not self.last_run or not self.last_run.get("srd_result"):
            QMessageBox.warning(self, "SRD", "Run SRD first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export SRD (CSV)", str(Path.home() / "srd_results.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            wt = float(self.last_run.get("wt_growth", 0.0))
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction_name", "reaction_id", "wt_growth", "growth_on_deletion", "delta_pct"])
                for rid, growth in sorted(self.last_run["srd_result"].items(), key=lambda x: x[1]):
                    delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                    rname = self._format_rxn_label(str(rid))
                    w.writerow([rname, rid, wt, growth, delta_pct])
            self.statusBar().showMessage(f"SRD exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export SRD results to CSV.", e)

    # -------- Excel export helpers --------

    def _export_deletion_excel(self, data_key: str, id_formatter, default_name: str, label: str):
        """Generic Excel export for deletion analyses."""
        if not self.last_run or not self.last_run.get(data_key):
            QMessageBox.warning(self, label, f"Run {label} first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, f"Export {label} (Excel)",
                                                    str(Path.home() / default_name), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = label
            wt = float(self.last_run.get("wt_growth", 0.0))
            ws.append(["Name", "ID", "WT Growth", "Growth on Deletion", "Δ Growth (%)", "Category"])
            for item_id, growth in sorted(self.last_run[data_key].items(), key=lambda x: x[1]):
                delta_pct = (0.0 if wt == 0 else ((growth - wt) / wt) * 100.0)
                name = id_formatter(str(item_id))
                if growth < 1e-6:
                    cat = "Lethal"
                elif delta_pct < -75:
                    cat = "Severe"
                elif delta_pct < -25:
                    cat = "Moderate"
                else:
                    cat = "Mild"
                ws.append([name, str(item_id), wt, growth, delta_pct, cat])
            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            wb.save(file_path)
            self.statusBar().showMessage(f"{label} exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export. Install it with: pip install openpyxl")
        except Exception as e:
            self._show_error("Export failed", "Could not export deletion results to Excel.", e)

    def export_sgd_excel(self):
        self._export_deletion_excel("sgd_result", self._format_gene_label, "sgd_results.xlsx", "SGD")

    def export_dgd_excel(self):
        self._export_deletion_excel("dgd_result", self._format_gene_pair_label, "dgd_results.xlsx", "DGD")

    def export_srd_excel(self):
        self._export_deletion_excel("srd_result", self._format_rxn_label, "srd_results.xlsx", "SRD")

    def export_robustness_csv(self):
        if not self.last_run or not self.last_run.get("robustness_result"):
            QMessageBox.warning(self, "Robustness", "Run Robustness Analysis first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Robustness (CSV)",
                                                    str(Path.home() / "robustness_results.csv"), "CSV files (*.csv)")
        if not file_path:
            return
        try:
            result = self.last_run["robustness_result"]
            rxn_id = self.last_run.get("rxn_id", "reaction")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([f"{rxn_id}_bound_value", "objective_value"])
                for val, obj in zip(result.get("values", []), result.get("objectives", [])):
                    w.writerow([val, obj])
            self.statusBar().showMessage(f"Robustness exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export robustness results to CSV.", e)

    def export_robustness_excel(self):
        if not self.last_run or not self.last_run.get("robustness_result"):
            QMessageBox.warning(self, "Robustness", "Run Robustness Analysis first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Robustness (Excel)",
                                                    str(Path.home() / "robustness_results.xlsx"), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Robustness"
            result = self.last_run["robustness_result"]
            rxn_id = self.last_run.get("rxn_id", "reaction")
            ws.append([f"{rxn_id} bound value", "Objective value"])
            for val, obj in zip(result.get("values", []), result.get("objectives", [])):
                ws.append([val, obj])
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            wb.save(file_path)
            self.statusBar().showMessage(f"Robustness exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export.")
        except Exception as e:
            self._show_error("Export failed", "Could not export robustness results to Excel.", e)

    def export_envelope_csv(self):
        if not self.last_run or not self.last_run.get("envelope_result"):
            QMessageBox.warning(self, "Envelope", "Run Production Envelope first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Envelope (CSV)",
                                                    str(Path.home() / "envelope_results.csv"), "CSV files (*.csv)")
        if not file_path:
            return
        try:
            result = self.last_run["envelope_result"]
            product_id = self.last_run.get("product_id", "product")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([f"{product_id}_bound", "growth_rate"])
                for pv, gv in zip(result.get("product", []), result.get("growth", [])):
                    w.writerow([pv, gv])
            self.statusBar().showMessage(f"Envelope exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export envelope results to CSV.", e)

    def export_envelope_excel(self):
        if not self.last_run or not self.last_run.get("envelope_result"):
            QMessageBox.warning(self, "Envelope", "Run Production Envelope first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Envelope (Excel)",
                                                    str(Path.home() / "envelope_results.xlsx"), "Excel files (*.xlsx)")
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Envelope"
            result = self.last_run["envelope_result"]
            product_id = self.last_run.get("product_id", "product")
            ws.append([f"{product_id} bound", "Growth rate"])
            for pv, gv in zip(result.get("product", []), result.get("growth", [])):
                ws.append([pv, gv])
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            wb.save(file_path)
            self.statusBar().showMessage(f"Envelope exported: {file_path}")
        except ImportError:
            QMessageBox.warning(self, "Missing package", "openpyxl is required for Excel export.")
        except Exception as e:
            self._show_error("Export failed", "Could not export envelope results to Excel.", e)

    def _render_robustness_results(self):
        """Render Robustness analysis results."""
        assert self.last_run is not None
        self._ensure_tab_built("Robustness")
        rxn_id = self.last_run.get("rxn_id", "")
        bound_type = self.last_run.get("bound_type", "ub").upper()
        robustness_result = self.last_run.get("robustness_result", {})
        
        raw_values = robustness_result.get("values", [])
        raw_objectives = robustness_result.get("objectives", [])
        points = []
        for x, y in zip(raw_values, raw_objectives):
            try:
                xf = float(x)
                yf = float(y)
            except Exception:
                continue
            if xf == xf and yf == yf and xf != float("inf") and xf != float("-inf") and yf != float("inf") and yf != float("-inf"):
                points.append((xf, yf))
        values = [p[0] for p in points]
        objectives = [p[1] for p in points]
        
        self.robustness_info_lbl.setText(
            f"Robustness Analysis for {rxn_id} ({bound_type})\n"
            f"Points computed: {len(values)}"
        )
        
        # Populate data table
        self.robustness_tbl.setHorizontalHeaderLabels([f"{rxn_id} {bound_type} Value", "Objective Value"])
        self.robustness_tbl.setRowCount(0)
        for val, obj in zip(values, objectives):
            r = self.robustness_tbl.rowCount()
            self.robustness_tbl.insertRow(r)
            self.robustness_tbl.setItem(r, 0, QTableWidgetItem(f"{val:.6g}"))
            self.robustness_tbl.setItem(r, 1, QTableWidgetItem(f"{obj:.6g}"))
        
        # Plot with clear markers and (x,y) labels
        ax = self.robustness_canvas.ax
        ax.clear()
        if values and objectives:
            ax.plot(values, objectives, marker="o", linestyle="-", color="steelblue",
                    linewidth=2, markersize=8, markerfacecolor="orange",
                    markeredgecolor="steelblue", markeredgewidth=1.5, zorder=5)
            # Add (x, y) coordinate labels at each point
            for i, (x, y) in enumerate(zip(values, objectives)):
                offset = 8 if (i % 2 == 0) else -12
                ax.annotate(f"({x:.2g}, {y:.4g})", (x, y),
                           textcoords="offset points", xytext=(0, offset),
                           fontsize=6.5, ha="center", color="#333")
            ax.set_xlabel(f"{rxn_id} bound value")
            ax.set_ylabel("Objective value")
            ax.set_title(f"Robustness: {rxn_id} ({bound_type})")
            ax.grid(True, alpha=0.3)
        else:
            ax.set_title("No feasible robustness points computed")
            ax.set_xlabel(f"{rxn_id} bound value")
            ax.set_ylabel("Objective value")
            ax.grid(True, alpha=0.3)
        self.robustness_canvas.draw()
        
        self.tabs.setCurrentWidget(self.tab_robustness)

    def _render_envelope_results(self):
        """Render Production Envelope results."""
        assert self.last_run is not None
        self._ensure_tab_built("Envelope")
        product_id = self.last_run.get("product_id", "")
        envelope_result = self.last_run.get("envelope_result", {})

        product_vals = [float(v) for v in envelope_result.get("product", [])]
        growth_max = [float(v) for v in envelope_result.get("growth_max", envelope_result.get("growth", []))]
        growth_min = [float(v) for v in envelope_result.get("growth_min", [0.0] * len(product_vals))]

        self.envelope_info_lbl.setText(
            f"Production Envelope: {product_id} vs Growth\n"
            f"Points computed: {len(product_vals)}"
        )

        # Populate data table (3 columns: product, growth min, growth max)
        self.envelope_tbl.setColumnCount(3)
        self.envelope_tbl.setHorizontalHeaderLabels(
            [f"{product_id} Flux", "Growth Min", "Growth Max"]
        )
        self.envelope_tbl.setRowCount(0)
        for pv, gmin, gmax in zip(product_vals, growth_min, growth_max):
            r = self.envelope_tbl.rowCount()
            self.envelope_tbl.insertRow(r)
            self.envelope_tbl.setItem(r, 0, QTableWidgetItem(f"{pv:.6g}"))
            self.envelope_tbl.setItem(r, 1, QTableWidgetItem(f"{gmin:.6g}"))
            self.envelope_tbl.setItem(r, 2, QTableWidgetItem(f"{gmax:.6g}"))

        # Plot the envelope
        ax = self.envelope_canvas.ax
        ax.clear()
        if product_vals and growth_max:
            # Upper bound line (maximum growth)
            ax.plot(product_vals, growth_max, color="#7B2FBE", linewidth=2,
                    label="Max growth", zorder=4)
            ax.scatter(product_vals, growth_max, s=30, color="#7B2FBE",
                       edgecolors="white", linewidth=0.5, zorder=5)

            if growth_min and any(v > 0 for v in growth_min):
                # Lower bound line (minimum growth)
                ax.plot(product_vals, growth_min, color="#D4A5FF", linewidth=2,
                        label="Min growth", zorder=4)
                ax.scatter(product_vals, growth_min, s=30, color="#D4A5FF",
                           edgecolors="white", linewidth=0.5, zorder=5)
                # Fill between min and max
                ax.fill_between(product_vals, growth_min, growth_max,
                                alpha=0.20, color="#9B59B6", label="Feasible region")
            else:
                # Only upper bound — fill from zero
                ax.fill_between(product_vals, 0, growth_max,
                                alpha=0.15, color="#9B59B6", label="Feasible region")

            ax.set_xlabel(f"{product_id} production rate (mmol/gDW/h)")
            ax.set_ylabel("Growth rate (1/h)")
            ax.set_title(f"Production Envelope: {product_id}")
            ax.legend(loc="best", fontsize=8)
            ax.grid(True, alpha=0.3)
        self.envelope_canvas.draw()

        self.tabs.setCurrentWidget(self.tab_envelope)

    def _render_sampling_results(self):
        """Render Flux Sampling results with diagnostics.

        Delegates each subplot to a dedicated helper so the overall flow
        stays readable.
        """
        assert self.last_run is not None
        self._ensure_tab_built("Sampling")
        sampling_result = self.last_run.get("sampling_result", {})
        samples_df = self.last_run.get("samples_df", None)

        self.sampling_info_lbl.setText(
            f"Flux Sampling ({self.last_run.get('sampler_type', 'ACHR')})\n"
            f"Sample size: {len(samples_df) if samples_df is not None else 0}\n"
            f"Reactions analyzed: {len(sampling_result)}"
        )

        rxns_for_plot, means = self._populate_sampling_table(sampling_result)
        self._plot_sampling_diagnostics(rxns_for_plot, means, samples_df)
        self.tabs.setCurrentWidget(self.tab_sampling)

    # -- Sampling sub-helpers --------------------------------------------------

    def _populate_sampling_table(self, sampling_result: dict) -> tuple[list[str], list[float]]:
        """Fill the sampling table and return (rxn_ids, means) for plotting."""
        self.sampling_tbl.setRowCount(0)
        rxns: list[str] = []
        means: list[float] = []
        sorted_rxns = sorted(sampling_result.items(), key=lambda x: abs(x[1]["mean"]), reverse=True)
        for rxn_id, stats in sorted_rxns[:int(self.topn_spin.value())]:
            r = self.sampling_tbl.rowCount()
            self.sampling_tbl.insertRow(r)
            self.sampling_tbl.setItem(r, 0, QTableWidgetItem(rxn_id))
            self.sampling_tbl.setItem(r, 1, QTableWidgetItem(f"{stats['mean']:.6g}"))
            self.sampling_tbl.setItem(r, 2, QTableWidgetItem(f"{stats['stdev']:.6g}"))
            self.sampling_tbl.setItem(r, 3, QTableWidgetItem(f"[{stats['min']:.3g}, {stats['max']:.3g}]"))
            rxns.append(rxn_id)
            means.append(stats["mean"])
        return rxns, means

    def _plot_sampling_diagnostics(self, rxns: list[str], means: list[float], samples_df) -> None:
        """Draw sampling diagnostic panels (bar, KDE, Geweke, correlation)."""
        ax = self.sampling_canvas.ax
        ax.clear()

        if samples_df is not None and len(rxns) >= 2:
            fig = self.sampling_canvas.figure
            fig.clear()
            self._plot_sampling_mean_bars(fig.add_subplot(221), rxns, means)
            self._plot_sampling_kde(fig.add_subplot(222), rxns, samples_df)
            self._plot_sampling_geweke(fig.add_subplot(223))
            self._plot_sampling_correlation(fig.add_subplot(224), fig)
            fig.tight_layout()
            self.sampling_canvas.draw()
        elif rxns:
            colors = ["green" if m > 0 else "red" for m in means]
            ax.barh(rxns, means, color=colors, alpha=0.6)
            ax.set_xlabel("Mean flux")
            ax.set_title("Flux Sampling - Mean Flux Distribution")
            self.sampling_canvas.draw()

    def _plot_sampling_mean_bars(self, ax, rxns: list[str], means: list[float]) -> None:
        """Top-left subplot: horizontal bar chart of mean fluxes."""
        top_n = min(15, len(rxns))
        colors = ["green" if m > 0 else "red" for m in means[:top_n]]
        ax.barh(rxns[:top_n], means[:top_n], color=colors, alpha=0.6)
        ax.set_xlabel("Mean flux")
        ax.set_title("Mean Flux", fontsize=9)
        ax.tick_params(axis='y', labelsize=7)

    def _plot_sampling_kde(self, ax, rxns: list[str], samples_df) -> None:
        """Top-right subplot: KDE / histogram for top reactions."""
        kde_histograms = self.last_run.get("kde_histograms", [])
        if kde_histograms:
            for h in kde_histograms:
                edges = h["edges"]
                centers = [(edges[i] + edges[i + 1]) / 2 for i in range(len(edges) - 1)]
                ax.bar(centers, h["counts"], width=(edges[1] - edges[0]) * 0.9,
                       alpha=0.4, label=h["rid"])
        else:
            for rid in rxns[:5]:
                if rid in samples_df.columns:
                    vals = samples_df[rid].values
                    ax.hist(vals, bins=40, alpha=0.4, density=True, label=rid)
        ax.set_xlabel("Flux")
        ax.set_ylabel("Density")
        ax.set_title("KDE — Top reactions", fontsize=9)
        ax.legend(fontsize=6)

    def _plot_sampling_geweke(self, ax) -> None:
        """Bottom-left subplot: Geweke convergence Z-scores."""
        geweke_scores = self.last_run.get("geweke_scores", {})
        if geweke_scores:
            g_rxns = list(geweke_scores.keys())
            g_vals = list(geweke_scores.values())
            bar_colors = ["#2ca02c" if abs(z) < 2 else "#d62728" for z in g_vals]
            ax.barh(g_rxns, g_vals, color=bar_colors, alpha=0.7)
            ax.axvline(x=-2, color='gray', linestyle='--', linewidth=0.8)
            ax.axvline(x=2, color='gray', linestyle='--', linewidth=0.8)
            ax.set_xlabel("Geweke Z-score")
            ax.set_title("Convergence (|Z|<2 = OK)", fontsize=9)
            ax.tick_params(axis='y', labelsize=7)
        else:
            ax.set_title("Convergence: insufficient data", fontsize=9)

    def _plot_sampling_correlation(self, ax, fig) -> None:
        """Bottom-right subplot: pairwise correlation heatmap."""
        corr_matrix = self.last_run.get("corr_matrix", None)
        corr_rxns = self.last_run.get("corr_rxns", [])
        if corr_matrix is not None and len(corr_rxns) >= 2:
            import numpy as np
            cm = np.array(corr_matrix)
            im = ax.imshow(cm, cmap='RdBu_r', vmin=-1, vmax=1, aspect='auto')
            ax.set_xticks(range(len(corr_rxns)))
            ax.set_yticks(range(len(corr_rxns)))
            ax.set_xticklabels(corr_rxns, rotation=90, fontsize=6)
            ax.set_yticklabels(corr_rxns, fontsize=6)
            ax.set_title("Pairwise Correlation", fontsize=9)
            fig.colorbar(im, ax=ax, fraction=0.046)
        else:
            ax.set_title("Correlation: need ≥2 rxns", fontsize=9)

    def export_sampling_csv(self):
        if not self.last_run or not self.last_run.get("sampling_result"):
            QMessageBox.warning(self, "Sampling", "Run sampling first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Sampling Stats (CSV)", str(Path.home() / "sampling_stats.csv"), "CSV files (*.csv);")
        if not file_path:
            return
        try:
            sampling_result = self.last_run["sampling_result"]
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["reaction_id", "mean_flux", "stdev", "min_flux", "max_flux"])
                for rxn_id, stats in sorted(sampling_result.items(), key=lambda x: abs(x[1]["mean"]), reverse=True):
                    w.writerow([rxn_id, stats["mean"], stats["stdev"], stats["min"], stats["max"]])
            self.statusBar().showMessage(f"Sampling exported: {file_path}")
        except Exception as e:
            self._show_error("Export failed", "Could not export sampling results to CSV.", e)

    def filter_pfba_table(self):
        """Filter pFBA table by reaction ID."""
        text = (self.pfba_flux_search.text() or "").strip().lower()
        for row in range(self.pfba_tbl.rowCount()):
            rid = (self.pfba_tbl.item(row, 0).text() if self.pfba_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.pfba_tbl.setRowHidden(row, hide)

    def filter_sampling_table(self):
        """Filter sampling table."""
        text = (self.sampling_search.text() or "").strip().lower()
        for row in range(self.sampling_tbl.rowCount()):
            rid = (self.sampling_tbl.item(row, 0).text() if self.sampling_tbl.item(row, 0) else "").lower()
            hide = bool(text) and text not in rid
            self.sampling_tbl.setRowHidden(row, hide)

    def _parse_float(self, val, allow_comma: bool = True) -> float:
        """Robust float parser for user inputs. Supports comma decimal.
        Raises ValueError if parsing fails.
        """
        if isinstance(val, (int, float)):
            return float(val)
        if val is None:
            raise ValueError("Missing numeric value")
        s = str(val).strip()
        if allow_comma:
            s = s.replace(",", ".")
        return float(s)
    # ---------------- Scenario export/import ----------------

